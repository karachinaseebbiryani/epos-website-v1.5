from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Form
from fastapi.responses import RedirectResponse, HTMLResponse
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, logging, bcrypt, jwt, secrets, json, hmac
from pydantic import BaseModel, field_validator
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from collections import defaultdict, deque
from bson import ObjectId
import smtplib, ssl, asyncio
from email.message import EmailMessage
import httpx
import pytz
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

mongo_url = os.environ.get('MONGO_URL')
db_name = os.environ.get('DB_NAME')
if not mongo_url or not db_name:
    import sys as _sys
    _sys.stderr.write(
        f"[FATAL] Required env vars missing: "
        f"MONGO_URL={'SET' if mongo_url else 'MISSING'}, "
        f"DB_NAME={'SET' if db_name else 'MISSING'}\n"
    )
    _sys.exit(1)
# serverSelectionTimeoutMS=5000 makes motor fail fast if Mongo is unreachable,
# instead of hanging the default 30s on every DB call.
client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=5000)
db = client[db_name]
app = FastAPI()
api_router = APIRouter(prefix="/api")

# ---------------------------------------------------------------------------
# Read-mostly menu cache (perf).
# The customer site (/menu), the admin POS (/menu-items + /categories) and
# every page transition between Home → Menu → Home repeatedly request the
# same large JSON payload. Without a cache layer Mongo gets hammered + the
# whole document set is re-serialized for every single hit.
#
# This in-memory TTL cache (per backend process) plus ETag/Cache-Control on
# the responses gives:
#   • DB hit ratio drops to ~1 per 30 s instead of 1 per request.
#   • Repeat browser navigations short-circuit at 304 Not Modified (no body).
#   • Cache is busted instantly on any menu/category mutation so admins never
#     see stale data after their own edits — TTL only matters cross-process.
#
# Functionality, response shape and DB schema are untouched.
# ---------------------------------------------------------------------------
import time as _time
import hashlib as _hashlib
_MENU_CACHE_TTL = float(os.environ.get("MENU_CACHE_TTL_SEC", "30"))
_menu_cache: Dict[str, dict] = {}

def _cache_get(key: str):
    e = _menu_cache.get(key)
    if not e:
        return None
    if _time.time() - e["ts"] > _MENU_CACHE_TTL:
        _menu_cache.pop(key, None)
        return None
    return e

def _cache_set(key: str, value):
    try:
        etag = '"' + _hashlib.md5(json.dumps(value, default=str, sort_keys=True).encode()).hexdigest() + '"'
    except Exception:
        etag = '"' + str(_time.time()) + '"'
    entry = {"ts": _time.time(), "value": value, "etag": etag}
    _menu_cache[key] = entry
    return entry

def _cache_bust(*keys: str):
    if not keys:
        _menu_cache.clear()
        return
    for k in keys:
        _menu_cache.pop(k, None)

def _menu_cache_bust_all():
    """Invalidate every cached menu/category/items response. Called from every
    mutation endpoint that can change what the menu looks like."""
    _cache_bust("menu", "menu-items", "categories")

# Liveness/readiness probe for Fly.io health checks.
# Intentionally performs NO database I/O so the app stays "healthy" from the
# load-balancer's perspective even when MongoDB is degraded. The rest of the
# routes will still return their own errors. Used by fly.toml [[http_service.checks]].
@api_router.get("/health")
async def health():
    return {"ok": True}
JWT_ALGORITHM = "HS256"
# Cross-domain deploys (e.g. Vercel frontend + Fly.io backend) need samesite=none + secure=true.
# Defaults preserve existing same-origin behavior (lax / not-secure).
COOKIE_SECURE = os.environ.get("COOKIE_SECURE", "false").lower() == "true"
COOKIE_SAMESITE = os.environ.get("COOKIE_SAMESITE", "lax")

# --- Security constants ---
# Maximum accepted password length. bcrypt only uses the first 72 bytes, but accepting
# unbounded passwords lets an attacker DoS the server with multi-MB payloads (each must be
# JSON-parsed and bcrypt-checked). 128 chars is well above any realistic user password.
MAX_PASSWORD_LEN = 128
# How long after order creation a guest can still submit/replace a bank-payment reference
# or upload a payment screenshot. Beyond this window the action is rejected to prevent
# tampering / payment hijacking on long-lived orders.
PAYMENT_SUBMIT_WINDOW_SEC = int(os.environ.get("PAYMENT_SUBMIT_WINDOW_SEC", "86400"))  # 24h
# Login brute-force throttling. In-memory; per (ip,email) key. Resets on backend restart.
LOGIN_MAX_ATTEMPTS = int(os.environ.get("LOGIN_MAX_ATTEMPTS", "10"))
LOGIN_WINDOW_SEC = int(os.environ.get("LOGIN_WINDOW_SEC", "300"))  # 5 min
LOGIN_LOCKOUT_SEC = int(os.environ.get("LOGIN_LOCKOUT_SEC", "900"))  # 15 min
_login_attempts: "dict[str, deque]" = defaultdict(deque)
_login_lockouts: "dict[str, float]" = {}

# --- Web Push (browser notification) configuration ---
# VAPID keys are stored in MongoDB (app_config collection) so they survive every redeploy
# without any manual env var management. Env vars still override when set (dev/CI use).
# Migration path: on first boot the old vapid_keys.json is read and saved to MongoDB,
# after which the file is no longer needed.
VAPID_EMAIL = os.environ.get("VAPID_EMAIL", "mailto:karachinaseebbiryani@gmail.com")
VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY")
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY")
_VAPID_KEYS_PATH = Path(__file__).parent / "vapid_keys.json"
# Module-level sync fallback: load from file so the key is available before startup().
# MongoDB load (async) happens in startup() and overwrites these if a DB record exists.
if not (VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY):
    try:
        if _VAPID_KEYS_PATH.exists():
            _kj = json.loads(_VAPID_KEYS_PATH.read_text())
            VAPID_PUBLIC_KEY = _kj.get("public_key")
            VAPID_PRIVATE_KEY = _kj.get("private_key")
    except Exception as _e:
        logger_init_err = _e  # captured for later; logger may not exist yet


async def _load_vapid_keys_from_db() -> bool:
    """Load VAPID keys from MongoDB app_config. Returns True if keys were found."""
    global VAPID_PUBLIC_KEY, VAPID_PRIVATE_KEY
    try:
        doc = await db.app_config.find_one({"key": "vapid_keys"})
        if doc and doc.get("public_key") and doc.get("private_key"):
            VAPID_PUBLIC_KEY = doc["public_key"]
            VAPID_PRIVATE_KEY = doc["private_key"]
            return True
    except Exception as e:
        logger.warning(f"Could not load VAPID keys from MongoDB: {e}")
    return False


async def _save_vapid_keys_to_db(pub: str, priv: str) -> None:
    """Persist VAPID keys to MongoDB so they survive redeploys."""
    try:
        await db.app_config.update_one(
            {"key": "vapid_keys"},
            {"$set": {"key": "vapid_keys", "public_key": pub, "private_key": priv,
                      "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True,
        )
    except Exception as e:
        logger.warning(f"Could not save VAPID keys to MongoDB: {e}")


def _generate_vapid_keypair_raw() -> tuple:
    """Generate a fresh VAPID keypair. Returns (public_key_url_b64, private_key_pem)."""
    from py_vapid import Vapid  # type: ignore
    from cryptography.hazmat.primitives.serialization import Encoding, PrivateFormat, NoEncryption, PublicFormat
    import base64 as _b64
    _v = Vapid()
    _v.generate_keys()
    priv_pem = _v.private_key.private_bytes(Encoding.PEM, PrivateFormat.PKCS8, NoEncryption()).decode()
    pub_raw = _v.public_key.public_bytes(Encoding.X962, PublicFormat.UncompressedPoint)
    return _b64.urlsafe_b64encode(pub_raw).rstrip(b"=").decode(), priv_pem


def _normalize_vapid_pem(priv: str) -> str:
    """Best-effort repair of a VAPID private key string that came out of a deployment
    env var. Handles every corruption I've seen in the wild:

    1. Literal backslash-n (``\\n``, 2 chars) instead of real newlines — happens when
       the user runs ``flyctl secrets set VAPID_PRIVATE_KEY="-----BEGIN...\\n..."`` in
       a shell, OR pastes the key into a UI that escapes line breaks.
    2. Wrapping single/double quotes copied with the value.
    3. Leading / trailing whitespace.
    4. PEM headers present but no internal newlines (the body is one long line glued
       to the BEGIN / END markers) — rewrap to canonical 64-char lines.

    Returns a PEM string that `cryptography.serialization.load_pem_private_key` can
    actually parse. Idempotent: a clean PEM is returned unchanged."""
    if not priv:
        return priv
    p = priv.strip()
    # Strip wrapping quotes if the operator pasted ``"…"`` literally.
    if len(p) >= 2 and ((p[0] == '"' and p[-1] == '"') or (p[0] == "'" and p[-1] == "'")):
        p = p[1:-1].strip()
    # Convert literal backslash-n / backslash-r sequences into real newlines/CR.
    if "\\n" in p:
        p = p.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\\r", "\r")
    # Drop any stray \r so the body is pure LF.
    p = p.replace("\r", "")
    # If PEM markers are present but there are no real newlines anywhere, rebuild
    # the canonical block (header / 64-char-wrapped body / footer).
    if "BEGIN" in p and "\n" not in p:
        for begin, end in (
            ("-----BEGIN PRIVATE KEY-----", "-----END PRIVATE KEY-----"),
            ("-----BEGIN EC PRIVATE KEY-----", "-----END EC PRIVATE KEY-----"),
        ):
            if begin in p and end in p:
                head, _, tail = p.partition(begin)
                body, _, foot_tail = tail.partition(end)
                body = "".join(body.split())  # strip ALL whitespace from the base64 body
                wrapped = "\n".join(body[i : i + 64] for i in range(0, len(body), 64))
                p = f"{head}{begin}\n{wrapped}\n{end}{foot_tail}".strip()
                break
    # pywebpush 2.x requires SEC1 format (BEGIN EC PRIVATE KEY) not PKCS8
    # (BEGIN PRIVATE KEY). Convert if needed — the underlying key material is identical.
    if "BEGIN PRIVATE KEY" in p and "BEGIN EC PRIVATE KEY" not in p:
        try:
            from cryptography.hazmat.primitives.serialization import (
                load_pem_private_key, Encoding, PrivateFormat, NoEncryption)
            _key = load_pem_private_key(p.encode(), password=None)
            p = _key.private_bytes(Encoding.PEM, PrivateFormat.TraditionalOpenSSL,
                                   NoEncryption()).decode()
        except Exception:
            pass  # leave as-is if conversion fails; let pywebpush surface the real error
    # Make sure the PEM ends with a newline (cryptography is tolerant but pywebpush /
    # py-vapid sometimes feed the string through openssl which is pickier).
    if "BEGIN" in p and not p.endswith("\n"):
        p = p + "\n"
    return p


def _generate_vapid_keys() -> tuple[str, str]:
    """Generate a fresh VAPID keypair and return (public_key_url_b64, private_key_pem).
    Used by the admin "Regenerate VAPID keys" button when production keys are malformed.
    Caller is responsible for persisting + wiping stale push_subscriptions."""
    from py_vapid import Vapid  # type: ignore
    from cryptography.hazmat.primitives.serialization import Encoding, PrivateFormat, NoEncryption, PublicFormat
    import base64 as _b64
    _v = Vapid()
    _v.generate_keys()
    priv_pem = _v.private_key.private_bytes(Encoding.PEM, PrivateFormat.PKCS8, NoEncryption()).decode()
    pub_raw = _v.public_key.public_bytes(Encoding.X962, PublicFormat.UncompressedPoint)
    return _b64.urlsafe_b64encode(pub_raw).rstrip(b"=").decode(), priv_pem


def _vapid_key_health() -> dict:
    """Verify the VAPID PEM private key can actually be parsed by `cryptography`.
    Returns a structured payload the admin UI can render so the operator instantly
    knows *why* push notifications are failing in prod (90% of the time: env var
    pasted with stripped newlines)."""
    info = {
        "public_key_set": bool(VAPID_PUBLIC_KEY),
        "private_key_set": bool(VAPID_PRIVATE_KEY),
        "public_key_preview": (VAPID_PUBLIC_KEY[:14] + "…") if VAPID_PUBLIC_KEY else None,
        "private_key_is_pem": bool(VAPID_PRIVATE_KEY and "BEGIN" in (VAPID_PRIVATE_KEY or "")),
        "private_key_has_newlines": bool(VAPID_PRIVATE_KEY and "\n" in (VAPID_PRIVATE_KEY or "")),
        "private_key_has_literal_backslash_n": bool(VAPID_PRIVATE_KEY and "\\n" in (VAPID_PRIVATE_KEY or "")),
        "source": "env" if (os.environ.get("VAPID_PUBLIC_KEY") and os.environ.get("VAPID_PRIVATE_KEY")) else "file",
        "parsable": False,
        "parse_error": None,
        "normalized": False,
    }
    if not (VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY):
        info["parse_error"] = "Public or private key missing"
        return info
    try:
        normalized = _normalize_vapid_pem(VAPID_PRIVATE_KEY)
        info["normalized"] = normalized != VAPID_PRIVATE_KEY
        # Verify with BOTH parsers: cryptography's load_pem_private_key (always lenient)
        # AND py-vapid's Vapid01.from_pem — the latter is what pywebpush actually uses
        # at send-time. If we only checked the former, a PEM that crashes pywebpush
        # would report parsable:true here (the original bug that hid this issue for
        # weeks). Both must succeed for the diagnostic to be trustworthy.
        from cryptography.hazmat.primitives import serialization as _ser
        _ser.load_pem_private_key(normalized.encode(), password=None)
        from py_vapid import Vapid01  # type: ignore
        Vapid01.from_pem(normalized.encode())
        info["parsable"] = True
    except Exception as e:
        info["parse_error"] = f"{type(e).__name__}: {e}"
    return info


def _client_ip(request: Request) -> str:
    """Best-effort client IP. Trusts X-Forwarded-For when set by our edge (Fly proxy)."""
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

def _login_throttle_check(ip: str, email: str):
    """Raise 429 if the (ip,email) pair is currently locked out."""
    key = f"{ip}|{(email or '').lower().strip()}"
    now = datetime.now(timezone.utc).timestamp()
    locked_until = _login_lockouts.get(key)
    if locked_until and now < locked_until:
        retry = int(locked_until - now)
        raise HTTPException(status_code=429, detail=f"Too many failed attempts. Try again in {retry}s.")

def _login_record_failure(ip: str, email: str):
    key = f"{ip}|{(email or '').lower().strip()}"
    now = datetime.now(timezone.utc).timestamp()
    dq = _login_attempts[key]
    dq.append(now)
    # Evict attempts older than the rolling window
    while dq and (now - dq[0]) > LOGIN_WINDOW_SEC:
        dq.popleft()
    if len(dq) >= LOGIN_MAX_ATTEMPTS:
        _login_lockouts[key] = now + LOGIN_LOCKOUT_SEC
        dq.clear()

def _login_record_success(ip: str, email: str):
    key = f"{ip}|{(email or '').lower().strip()}"
    _login_attempts.pop(key, None)
    _login_lockouts.pop(key, None)

def _validate_password_length(pw: str):
    """Reject absurdly long passwords (bcrypt only uses first 72 bytes anyway).
    Prevents long-password DoS amplification on register / login."""
    if pw is None:
        raise HTTPException(status_code=400, detail="Password is required")
    if len(pw) > MAX_PASSWORD_LEN or len(pw.encode("utf-8", errors="ignore")) > MAX_PASSWORD_LEN * 4:
        raise HTTPException(status_code=400, detail=f"Password too long (max {MAX_PASSWORD_LEN} chars)")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- Helpers ---
import uuid as _uuid
INSTANCE_ID = _uuid.uuid4().hex  # rotates on every backend restart → invalidates old JWTs

def hash_password(p): return bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
def verify_password(p, h): return bcrypt.checkpw(p.encode("utf-8"), h.encode("utf-8"))
def get_jwt_secret(): return os.environ["JWT_SECRET"]

def create_access_token(uid, email, role):
    return jwt.encode({"sub": uid, "email": email, "role": role, "exp": datetime.now(timezone.utc) + timedelta(hours=8), "type": "access", "iid": INSTANCE_ID}, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(uid):
    return jwt.encode({"sub": uid, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh", "iid": INSTANCE_ID}, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        ah = request.headers.get("Authorization", "")
        if ah.startswith("Bearer "): token = ah[7:]
    if not token: raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access": raise HTTPException(status_code=401, detail="Invalid token")
        # Force re-login on backend restart: any token issued by a previous instance is invalid
        if payload.get("iid") != INSTANCE_ID:
            raise HTTPException(status_code=401, detail="Session expired — please log in again")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user: raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError: raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError: raise HTTPException(status_code=401, detail="Invalid token")

ALL_PERMISSIONS = [
    # POS Operations (legacy)
    "dashboard", "pos", "menu", "menu_edit", "inventory", "reports_x", "reports_z",
    "orders_history", "settings", "expenses", "vendors", "reprint_invoices", "refunds",
    # Online Store modules (gated independently in AdminLayout sidebar)
    "online_dashboard", "online_orders", "online_menu", "online_offers",
    "online_events", "online_settings",
]
ADMIN_PERMISSIONS = ALL_PERMISSIONS.copy()


def _has_perm(user: dict, perm: str) -> bool:
    """Authorization helper used by every admin-shell endpoint.

    Returns True if the caller is the master admin (role=admin, always full
    access) OR if the named permission is in their per-user permissions list.

    Why it exists: dozens of endpoints used to check `role != "admin"` and
    return 403 — which broke the entire "give a staff user this module's
    permission" feature. With this helper a non-admin user that has, say,
    `online_orders` in their permissions list can now actually list / accept /
    reject orders, while users without that permission still get a 403.
    """
    if not user:
        return False
    if user.get("role") == "admin":
        return True
    return perm in (user.get("permissions") or [])

# --- Models ---
class LoginRequest(BaseModel):
    email: str
    password: str

class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str
    role: str = "cashier"

class CategoryCreate(BaseModel):
    name: str
    color: Optional[str] = None

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None

class MenuItemVariation(BaseModel):
    name: str
    price: float

class PosVariation(BaseModel):
    """POS-only size/portion variation. Independent from the online-store
    `variations` list so the counter menu can differ from the website. `color`
    paints the picker button — lets staff who can't read match by colour
    (e.g. red = Half Rs 200, blue = Full Rs 300)."""
    name: str
    price: float
    color: Optional[str] = None

class ModifierOption(BaseModel):
    id: Optional[str] = None
    name: str
    price: float = 0

class ModifierGroup(BaseModel):
    id: Optional[str] = None
    name: str
    type: str = "multi"
    required: bool = False
    min_select: int = 0
    max_select: int = 0
    active: bool = True
    options: List[ModifierOption] = []

class MenuIngredient(BaseModel):
    id: Optional[str] = None
    name: str
    removable: bool = True

class MenuItemCreate(BaseModel):
    name: str
    price: float
    price_fp1: Optional[float] = None  # FoodPanda 1 price (overrides price when payment_type=foodpanda1)
    price_fp2: Optional[float] = None  # FoodPanda 2 price (overrides price when payment_type=foodpanda2)
    category_id: str
    stock: int = 100
    low_stock_threshold: int = 10
    color: Optional[str] = None
    variations: Optional[List[MenuItemVariation]] = None
    variations_active: Optional[bool] = True
    pos_variations: Optional[List[PosVariation]] = None   # POS-only, colour-coded
    pos_allergies: Optional[List[str]] = None             # POS "leave out" list (allergy/preference)
    voice_aliases: Optional[List[str]] = None             # spoken names ("adhi deg", "بریانی") for the voice assistant
    modifier_groups: Optional[List[ModifierGroup]] = None
    ingredients: Optional[List[MenuIngredient]] = None
    discount_type: Optional[str] = None
    discount_value: Optional[float] = 0
    is_bestseller: Optional[bool] = False
    is_popular: Optional[bool] = False
    image_url: Optional[str] = ""
    image_type: Optional[str] = None
    description: Optional[str] = ""
    related_item_ids: Optional[List[str]] = None  # F7: explicit upsell suggestions
    # Vendor-linked outsourced products (e.g. Pepsi 500ml from "Khokha")
    is_outsourced: Optional[bool] = False
    outsourced_vendor_id: Optional[str] = None
    outsourced_unit_cost: Optional[float] = None  # cost per unit owed to vendor; falls back to item price if None

class MenuItemUpdate(BaseModel):
    name: Optional[str] = None
    price: Optional[float] = None
    price_fp1: Optional[float] = None
    price_fp2: Optional[float] = None
    category_id: Optional[str] = None
    stock: Optional[int] = None
    low_stock_threshold: Optional[int] = None
    color: Optional[str] = None
    variations: Optional[List[MenuItemVariation]] = None
    variations_active: Optional[bool] = None
    pos_variations: Optional[List[PosVariation]] = None
    pos_allergies: Optional[List[str]] = None
    voice_aliases: Optional[List[str]] = None
    modifier_groups: Optional[List[ModifierGroup]] = None
    ingredients: Optional[List[MenuIngredient]] = None
    discount_type: Optional[str] = None
    discount_value: Optional[float] = None
    is_bestseller: Optional[bool] = None
    is_popular: Optional[bool] = None
    image_url: Optional[str] = None
    image_type: Optional[str] = None
    description: Optional[str] = None
    related_item_ids: Optional[List[str]] = None
    is_outsourced: Optional[bool] = None
    outsourced_vendor_id: Optional[str] = None
    outsourced_unit_cost: Optional[float] = None

class OrderItemInput(BaseModel):
    item_id: str
    name: str
    price: float
    original_price: Optional[float] = None
    quantity: int
    # Ingredients the customer asked to leave out for this line (allergy/preference).
    removed_ingredients: Optional[List[str]] = None

class OrderCreate(BaseModel):
    items: List[OrderItemInput]
    payment_type: str
    subtotal: float
    tax: float
    total: float
    discount_type: Optional[str] = None
    discount_value: Optional[float] = 0
    discount_amount: Optional[float] = 0
    order_type: Optional[str] = None   # dine_in / takeaway / delivery (defaults to takeaway)
    table_id: Optional[str] = None

# --- Dine-in tables + open (unpaid) order models (ported from Marhaba) ---
class TableCreate(BaseModel):
    name: str
    section: Optional[str] = "Main Hall"
    capacity: Optional[int] = 4
    status: Optional[str] = "available"  # available / occupied / reserved / cleaning

class TableUpdate(BaseModel):
    name: Optional[str] = None
    section: Optional[str] = None
    capacity: Optional[int] = None
    status: Optional[str] = None

class OpenOrderItem(BaseModel):
    uid: Optional[str] = None          # stable per-line id (frontend generated)
    item_id: str
    name: str
    price: float
    original_price: Optional[float] = None
    quantity: int
    removed_ingredients: Optional[List[str]] = None
    kitchen_status: Optional[str] = "new"  # new / sent

class OpenOrderCreate(BaseModel):
    table_id: str

class OpenOrderItemsUpdate(BaseModel):
    items: List[OpenOrderItem]
    subtotal: Optional[float] = 0
    tax: Optional[float] = 0
    total: Optional[float] = 0
    discount_type: Optional[str] = None
    discount_value: Optional[float] = 0
    discount_amount: Optional[float] = 0

class OpenOrderPay(BaseModel):
    payment_type: str
    subtotal: float
    tax: float
    total: float
    discount_type: Optional[str] = None
    discount_value: Optional[float] = 0
    discount_amount: Optional[float] = 0

class StockUpdate(BaseModel):
    stock: int

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "cashier"
    permissions: Optional[List[str]] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None
    permissions: Optional[List[str]] = None

class SettingsUpdate(BaseModel):
    tax_rate: Optional[float] = None
    online_tax_rate: Optional[float] = None
    foodpanda1_tax_rate: Optional[float] = None
    foodpanda2_tax_rate: Optional[float] = None
    currency: Optional[str] = None
    restaurant_name: Optional[str] = None
    restaurant_address: Optional[str] = None
    restaurant_phone: Optional[str] = None
    restaurant_email: Optional[str] = None
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_from: Optional[str] = None
    smtp_use_tls: Optional[bool] = None
    email_recipients: Optional[List[Dict[str, Any]]] = None
    auto_email_on_z_close: Optional[bool] = None
    # Daily auto-send schedule
    daily_report_time: Optional[str] = None  # "HH:MM" 24h
    daily_report_timezone: Optional[str] = None  # IANA tz e.g., "Asia/Karachi"
    auto_email_daily: Optional[bool] = None
    auto_whatsapp_daily: Optional[bool] = None
    daily_report_type: Optional[str] = None  # "yesterday" | "today"
    # WhatsApp
    whatsapp_service_url: Optional[str] = None
    whatsapp_recipients: Optional[List[Dict[str, Any]]] = None
    auto_whatsapp_on_z_close: Optional[bool] = None
    # Cloudflare Tunnel
    tunnel_log_path: Optional[str] = None
    tunnel_notify_on_change: Optional[bool] = None
    # Receipt formatting
    receipt_font_family: Optional[str] = None
    receipt_base_size: Optional[int] = None
    receipt_header_size: Optional[int] = None
    receipt_total_size: Optional[int] = None
    receipt_bold_all: Optional[bool] = None
    receipt_bold_total: Optional[bool] = None
    receipt_show_logo: Optional[bool] = None
    receipt_footer_text: Optional[str] = None
    receipt_paper_width: Optional[int] = None  # in mm or pixels (CSS)
    receipt_show_tax_line: Optional[bool] = None
    # Order alert ring (POS + admin). Staff read these via GET /settings (any
    # signed-in user); only admins can PUT. Preset key from ALERT_SOUNDS on the
    # frontend, or a custom URL. Volume is 0.0-1.0.
    order_alert_sound: Optional[str] = None
    order_alert_volume: Optional[float] = None
    # Branding – custom logo (base64 data URL, e.g. "data:image/png;base64,...")
    restaurant_logo: Optional[str] = None

class ExpenseCreate(BaseModel):
    description: str
    amount: float
    category: Optional[str] = "general"

class ExpenseUpdate(BaseModel):
    description: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None

class VendorCreate(BaseModel):
    name: str
    contact: Optional[str] = ""
    items_supplied: Optional[str] = ""

class VendorUpdate(BaseModel):
    name: Optional[str] = None
    contact: Optional[str] = None
    items_supplied: Optional[str] = None

class VendorTransactionCreate(BaseModel):
    vendor_id: str
    items: List[dict]  # [{name, quantity, unit_price}]
    total: float
    notes: Optional[str] = ""

class VendorPaymentCreate(BaseModel):
    vendor_id: str
    amount: float
    notes: Optional[str] = ""

class RefundCreate(BaseModel):
    order_id: str
    reason: str
    amount: float
    items: Optional[List[dict]] = None

# --- Loyalty / Diamond Reward System Models ---
class LoyaltySettingsUpdate(BaseModel):
    enabled: Optional[bool] = True
    earning_rate: Optional[float] = 10.0  # Diamonds per Rs spent
    min_order_for_points: Optional[float] = 0.0
    points_expiry_days: Optional[int] = None  # null = never expire

class LoyaltyRewardCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    cost_diamonds: int
    reward_type: str  # "free_item", "discount_percent", "discount_fixed"
    reward_value: str  # menu_item_id for free_item, or number string for discounts
    is_active: Optional[bool] = True
    max_redemptions_per_customer: Optional[int] = None

class LoyaltyRewardUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    cost_diamonds: Optional[int] = None
    reward_type: Optional[str] = None
    reward_value: Optional[str] = None
    is_active: Optional[bool] = None
    max_redemptions_per_customer: Optional[int] = None

class LoyaltyRedeemRequest(BaseModel):
    reward_id: str

class LoyaltyAdjustRequest(BaseModel):
    customer_id: str
    diamonds: int  # Can be positive or negative
    notes: str

# --- Auth ---
@api_router.post("/auth/login")
async def login(req: LoginRequest, request: Request, response: Response):
    email = req.email.lower().strip()
    ip = _client_ip(request)
    _login_throttle_check(ip, email)
    _validate_password_length(req.password)
    user = await db.users.find_one({"email": email})
    if not user:
        _login_record_failure(ip, email)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not verify_password(req.password, user["password_hash"]):
        _login_record_failure(ip, email)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    _login_record_success(ip, email)
    uid = str(user["_id"])
    at = create_access_token(uid, email, user.get("role", "cashier"))
    rt = create_refresh_token(uid)
    response.set_cookie(key="access_token", value=at, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=28800, path="/")
    response.set_cookie(key="refresh_token", value=rt, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")
    perms = user.get("permissions", ADMIN_PERMISSIONS if user.get("role") == "admin" else ["pos"])
    return {"id": uid, "email": user["email"], "name": user.get("name", ""), "role": user.get("role", "cashier"), "permissions": perms, "token": at}

@api_router.post("/auth/register")
async def register(req: RegisterRequest, response: Response):
    email = req.email.lower().strip()
    _validate_password_length(req.password)
    if await db.users.find_one({"email": email}): raise HTTPException(status_code=400, detail="Email already registered")
    hashed = hash_password(req.password)
    perms = ["pos"]
    doc = {"email": email, "password_hash": hashed, "name": req.name, "role": req.role, "permissions": perms, "created_at": datetime.now(timezone.utc).isoformat()}
    result = await db.users.insert_one(doc)
    uid = str(result.inserted_id)
    at = create_access_token(uid, email, req.role)
    rt = create_refresh_token(uid)
    response.set_cookie(key="access_token", value=at, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=28800, path="/")
    response.set_cookie(key="refresh_token", value=rt, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")
    return {"id": uid, "email": email, "name": req.name, "role": req.role, "permissions": perms, "token": at}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    perms = user.get("permissions", ADMIN_PERMISSIONS if user.get("role") == "admin" else ["pos"])
    return {"id": user["_id"], "email": user["email"], "name": user.get("name", ""), "role": user.get("role", "cashier"), "permissions": perms}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

# --- Users ---
@api_router.get("/users")
async def list_users(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    users = await db.users.find({}, {"password_hash": 0}).to_list(500)
    return [{"id": str(u["_id"]), "email": u["email"], "name": u.get("name", ""), "role": u.get("role", "cashier"), "permissions": u.get("permissions", ADMIN_PERMISSIONS if u.get("role") == "admin" else ["pos"]), "created_at": u.get("created_at", "")} for u in users]

@api_router.post("/users")
async def create_user(req: UserCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    email = req.email.lower().strip()
    if await db.users.find_one({"email": email}): raise HTTPException(status_code=400, detail="Email already exists")
    perms = req.permissions if req.permissions else (ADMIN_PERMISSIONS if req.role == "admin" else ["pos"])
    doc = {"email": email, "password_hash": hash_password(req.password), "name": req.name, "role": req.role, "permissions": perms, "created_at": datetime.now(timezone.utc).isoformat()}
    result = await db.users.insert_one(doc)
    return {"id": str(result.inserted_id), "email": email, "name": req.name, "role": req.role, "permissions": perms}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, req: UserUpdate, request: Request):
    current = await get_current_user(request)
    if current.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    update_data = {}
    if req.name is not None: update_data["name"] = req.name
    if req.role is not None: update_data["role"] = req.role
    if req.permissions is not None: update_data["permissions"] = req.permissions
    if req.password and req.password.strip(): update_data["password_hash"] = hash_password(req.password)
    if update_data: await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update_data})
    updated = await db.users.find_one({"_id": ObjectId(user_id)}, {"password_hash": 0})
    if not updated: raise HTTPException(status_code=404, detail="User not found")
    return {"id": str(updated["_id"]), "email": updated["email"], "name": updated.get("name", ""), "role": updated.get("role", "cashier"), "permissions": updated.get("permissions", [])}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request):
    current = await get_current_user(request)
    if current.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    if current["_id"] == user_id: raise HTTPException(status_code=400, detail="Cannot delete yourself")
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

# --- Settings ---
DEFAULT_SETTINGS = {"tax_rate": 5.0, "online_tax_rate": 0.0, "foodpanda1_tax_rate": 0.0, "foodpanda2_tax_rate": 0.0, "currency": "Rs", "restaurant_name": "KARACHI NASEEB BIRYANI AND MURG PULAO", "restaurant_address": "68 Chatri Chowk, Punjab Small Industry, D Block, Lahore", "restaurant_phone": "+923004928411", "restaurant_email": "karachinaseebbiryani599@gmail.com", "smtp_host": "smtp.gmail.com", "smtp_port": 587, "smtp_user": "", "smtp_password": "", "smtp_from": "", "smtp_use_tls": True, "email_recipients": [], "auto_email_on_z_close": False, "daily_report_time": "02:15", "daily_report_timezone": "Asia/Karachi", "auto_email_daily": False, "auto_whatsapp_daily": False, "daily_report_type": "yesterday", "whatsapp_service_url": "http://127.0.0.1:3030", "whatsapp_recipients": [], "auto_whatsapp_on_z_close": False, "tunnel_log_path": "", "tunnel_notify_on_change": True, "receipt_font_family": "Courier New", "receipt_base_size": 12, "receipt_header_size": 16, "receipt_total_size": 16, "receipt_bold_all": False, "receipt_bold_total": True, "receipt_show_logo": False, "receipt_footer_text": "Thank you for your order!", "receipt_paper_width": 300, "receipt_show_tax_line": True, "order_alert_sound": "classic", "order_alert_volume": 1.0, "restaurant_logo": ""}

@api_router.get("/settings")
async def get_settings(request: Request):
    await get_current_user(request)
    s = await db.settings.find_one({"key": "global"}, {"_id": 0})
    if not s: return {**DEFAULT_SETTINGS, "restaurant_logo": ""}
    # Migrate base64 logo to a file (one-shot, idempotent) so it never ships inline again
    logo_val = s.get("restaurant_logo") or ""
    if isinstance(logo_val, str) and logo_val.startswith("data:"):
        new_url = _persist_data_url_image(logo_val, kind="logo")
        if new_url.startswith("/api/uploads/"):
            await db.settings.update_one({"key": "global"}, {"$set": {"restaurant_logo": new_url}})
            s["restaurant_logo"] = new_url
    return {k: (s.get(k, v) if k != "restaurant_logo" else "") for k, v in DEFAULT_SETTINGS.items()}

@api_router.put("/settings")
async def update_settings(req: SettingsUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    ud = {k: v for k, v in req.model_dump().items() if v is not None}
    # If admin uploaded a new logo as a base64 data URL, persist it to disk first
    # so we never store inline image bytes in MongoDB.
    if "restaurant_logo" in ud and isinstance(ud["restaurant_logo"], str) and ud["restaurant_logo"].startswith("data:"):
        new_url = _persist_data_url_image(ud["restaurant_logo"], kind="logo")
        if new_url.startswith("/api/uploads/"):
            ud["restaurant_logo"] = new_url
    if ud: await db.settings.update_one({"key": "global"}, {"$set": ud}, upsert=True)
    s = await db.settings.find_one({"key": "global"}, {"_id": 0})
    schedule_keys = ("daily_report_time", "daily_report_timezone", "auto_email_daily", "auto_whatsapp_daily")
    if any(k in ud for k in schedule_keys):
        logger.info(f"Settings: schedule fields changed: {[k for k in schedule_keys if k in ud]}")
        try: await _reschedule_daily_job()
        except Exception as e: logger.exception(f"Reschedule failed: {e}")
    if not s: return {**DEFAULT_SETTINGS, "restaurant_logo": ""}
    return {k: (s.get(k, v) if k != "restaurant_logo" else "") for k, v in DEFAULT_SETTINGS.items()}
@api_router.get("/settings/logo")
async def get_settings_logo(request: Request):
    """Return only the restaurant logo URL/data. Separated from /settings so
    the main settings payload stays small (was ~411 KB when the logo was
    stored as a base64 data URL)."""
    await get_current_user(request)
    s = await db.settings.find_one({"key": "global"}, {"_id": 0}) or {}
    return {"restaurant_logo": s.get("restaurant_logo") or ""}
# --- Categories ---
def _can_edit_menu(user):
    return user.get("role") == "admin" or "menu_edit" in (user.get("permissions") or [])

@api_router.get("/categories")
async def get_categories(request: Request, response: Response):
    cached = _cache_get("categories")
    inm = request.headers.get("if-none-match", "")
    if cached:
        if inm and inm == cached["etag"]:
            return Response(status_code=304, headers={"ETag": cached["etag"], "Cache-Control": "public, max-age=30"})
        response.headers["ETag"] = cached["etag"]
        response.headers["Cache-Control"] = "public, max-age=30"
        return cached["value"]
    cats = await db.categories.find({}).sort([("sort_order", 1), ("created_at", 1)]).to_list(100)
    out = [{"id": str(c["_id"]), "name": c["name"], "color": c.get("color"), "sort_order": c.get("sort_order", 0)} for c in cats]
    entry = _cache_set("categories", out)
    response.headers["ETag"] = entry["etag"]
    response.headers["Cache-Control"] = "public, max-age=30"
    return out

@api_router.post("/categories")
async def create_category(cat: CategoryCreate, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    count = await db.categories.count_documents({})
    result = await db.categories.insert_one({"name": cat.name, "color": cat.color, "sort_order": count, "created_at": datetime.now(timezone.utc).isoformat()})
    _menu_cache_bust_all()
    return {"id": str(result.inserted_id), "name": cat.name, "color": cat.color, "sort_order": count}

@api_router.put("/categories/{cat_id}")
async def update_category(cat_id: str, cat: CategoryUpdate, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    ud = {k: v for k, v in cat.model_dump().items() if v is not None}
    if ud: await db.categories.update_one({"_id": ObjectId(cat_id)}, {"$set": ud})
    updated = await db.categories.find_one({"_id": ObjectId(cat_id)}, {"_id": 0})
    if not updated: raise HTTPException(status_code=404, detail="Not found")
    _menu_cache_bust_all()
    return {"id": cat_id, "name": updated.get("name"), "color": updated.get("color"), "sort_order": updated.get("sort_order", 0)}

@api_router.delete("/categories/{cat_id}")
async def delete_category(cat_id: str, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    await db.categories.delete_one({"_id": ObjectId(cat_id)})
    await db.menu_items.delete_many({"category_id": cat_id})
    _menu_cache_bust_all()
    return {"message": "Deleted"}

@api_router.post("/categories/reorder")
async def reorder_categories(payload: dict, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    order = payload.get("order") or []
    for idx, cid in enumerate(order):
        try:
            await db.categories.update_one({"_id": ObjectId(cid)}, {"$set": {"sort_order": idx}})
        except Exception:
            pass
    _menu_cache_bust_all()
    return {"message": "Reordered", "count": len(order)}

# --- Menu Items ---
@api_router.get("/menu-items")
async def get_menu_items(request: Request, response: Response):
    cached = _cache_get("menu-items")
    inm = request.headers.get("if-none-match", "")
    if cached:
        if inm and inm == cached["etag"]:
            return Response(status_code=304, headers={"ETag": cached["etag"], "Cache-Control": "public, max-age=30"})
        response.headers["ETag"] = cached["etag"]
        response.headers["Cache-Control"] = "public, max-age=30"
        return cached["value"]
    items = await db.menu_items.find({}).sort([("sort_order", 1), ("created_at", 1)]).to_list(500)
    out = [{
        "id": str(i["_id"]), "name": i["name"], "price": i["price"],
        "price_fp1": i.get("price_fp1"), "price_fp2": i.get("price_fp2"),
        "category_id": i["category_id"], "stock": i.get("stock", 0),
        "low_stock_threshold": i.get("low_stock_threshold", 10), "color": i.get("color"),
        "sort_order": i.get("sort_order", 0), "variations": i.get("variations", []),
        # POS customise dialog (ported from Marhaba): variations are offered only
        # when active, and removable ingredients drive the "leave out" checklist.
        # Additive fields — older clients simply ignore them.
        "variations_active": i.get("variations_active", True),
        "ingredients": i.get("ingredients", []),
        # POS-only variation/allergy lists (independent from online-store
        # `variations`/`ingredients`; colour-coded for the counter).
        "pos_variations": i.get("pos_variations", []),
        "pos_allergies": i.get("pos_allergies", []),
        "voice_aliases": i.get("voice_aliases", []),
        "discount_type": i.get("discount_type"), "discount_value": i.get("discount_value", 0),
        "is_bestseller": i.get("is_bestseller", False), "is_popular": i.get("is_popular", False),
        "image_url": i.get("image_url", ""), "image_type": i.get("image_type", "url"),
        "description": i.get("description", ""),
        "related_item_ids": i.get("related_item_ids", []),
        "is_outsourced": bool(i.get("is_outsourced", False)),
        "outsourced_vendor_id": i.get("outsourced_vendor_id"),
        "outsourced_unit_cost": i.get("outsourced_unit_cost"),
    } for i in items]
    entry = _cache_set("menu-items", out)
    response.headers["ETag"] = entry["etag"]
    response.headers["Cache-Control"] = "public, max-age=30"
    return out

def _norm_modifier_groups(groups):
    """Assign stable ids to modifier groups + their options and coerce types."""
    out = []
    for g in (groups or []):
        gid = (str(g.get("id") or "").strip()) or ("g_" + _uuid.uuid4().hex[:8])
        opts = []
        for o in (g.get("options") or []):
            oid = (str(o.get("id") or "").strip()) or ("o_" + _uuid.uuid4().hex[:8])
            opts.append({"id": oid, "name": str(o.get("name", "")).strip(),
                         "price": float(o.get("price", 0) or 0)})
        out.append({
            "id": gid,
            "name": str(g.get("name", "")).strip(),
            "type": "single" if g.get("type") == "single" else "multi",
            "required": bool(g.get("required", False)),
            "min_select": int(g.get("min_select", 0) or 0),
            "max_select": int(g.get("max_select", 0) or 0),
            "active": bool(g.get("active", True)),
            "options": opts,
        })
    return out

def _norm_ingredients(ings):
    out = []
    for ing in (ings or []):
        iid = (str(ing.get("id") or "").strip()) or ("i_" + _uuid.uuid4().hex[:8])
        out.append({"id": iid, "name": str(ing.get("name", "")).strip(),
                    "removable": bool(ing.get("removable", True))})
    return out

@api_router.post("/menu-items")
async def create_menu_item(item: MenuItemCreate, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    count = await db.menu_items.count_documents({})
    variations = [v.model_dump() for v in (item.variations or [])]
    # If the admin uploaded the photo as a base64 data: URL, persist it to disk
    # immediately so the document keeps only a small URL — never inline bytes.
    persisted_image_url = _persist_data_url_image(item.image_url or "", kind="menu")
    doc = {
        "name": item.name, "price": item.price, "price_fp1": item.price_fp1, "price_fp2": item.price_fp2,
        "category_id": item.category_id, "stock": item.stock, "low_stock_threshold": item.low_stock_threshold,
        "color": item.color, "variations": variations,
        "variations_active": True if item.variations_active is None else bool(item.variations_active),
        # POS-only counterparts — never read by the customer website/app.
        "pos_variations": [v.model_dump() for v in (item.pos_variations or [])],
        "pos_allergies": [str(a).strip() for a in (item.pos_allergies or []) if str(a).strip()],
        "voice_aliases": [str(a).strip() for a in (item.voice_aliases or []) if str(a).strip()],
        "modifier_groups": _norm_modifier_groups([g.model_dump() for g in (item.modifier_groups or [])]),
        "ingredients": _norm_ingredients([ing.model_dump() for ing in (item.ingredients or [])]),
        "discount_type": item.discount_type, "discount_value": item.discount_value or 0,
        "is_bestseller": bool(item.is_bestseller), "is_popular": bool(item.is_popular),
        "image_url": persisted_image_url, "image_type": "url" if persisted_image_url.startswith("/api/uploads/") else (item.image_type or ("upload" if (item.image_url or "").startswith("data:") else "url")),
        "description": item.description or "",
        "related_item_ids": list(item.related_item_ids or []),
        "is_outsourced": bool(item.is_outsourced),
        "outsourced_vendor_id": item.outsourced_vendor_id or None,
        "outsourced_unit_cost": item.outsourced_unit_cost,
        "sort_order": count, "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.menu_items.insert_one(doc)
    _menu_cache_bust_all()
    return {**{k: v for k, v in doc.items() if k not in ("created_at", "_id")}, "id": str(result.inserted_id)}

@api_router.put("/menu-items/{item_id}")
async def update_menu_item(item_id: str, item: MenuItemUpdate, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    ud = {k: v for k, v in item.model_dump().items() if v is not None}
    if "modifier_groups" in ud: ud["modifier_groups"] = _norm_modifier_groups(ud["modifier_groups"])
    if "ingredients" in ud: ud["ingredients"] = _norm_ingredients(ud["ingredients"])
    if "pos_allergies" in ud: ud["pos_allergies"] = [str(a).strip() for a in ud["pos_allergies"] if str(a).strip()]
    if "voice_aliases" in ud: ud["voice_aliases"] = [str(a).strip() for a in ud["voice_aliases"] if str(a).strip()]
    # If the admin re-uploaded a photo as a base64 data: URL, persist it to disk
    # before saving, so we never store inline image bytes in MongoDB.
    if "image_url" in ud and isinstance(ud["image_url"], str) and ud["image_url"].startswith("data:"):
        new_url = _persist_data_url_image(ud["image_url"], kind="menu")
        if new_url.startswith("/api/uploads/"):
            ud["image_url"] = new_url
            ud["image_type"] = "url"
    # variations should be replaced wholesale, including with empty list — model_dump() drops only None
    if ud: await db.menu_items.update_one({"_id": ObjectId(item_id)}, {"$set": ud})
    updated = await db.menu_items.find_one({"_id": ObjectId(item_id)}, {"_id": 0})
    if not updated: raise HTTPException(status_code=404, detail="Not found")
    updated["id"] = item_id
    updated.setdefault("variations", [])
    _menu_cache_bust_all()
    return updated

@api_router.delete("/menu-items/{item_id}")
async def delete_menu_item(item_id: str, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    await db.menu_items.delete_one({"_id": ObjectId(item_id)})
    _menu_cache_bust_all()
    return {"message": "Deleted"}

@api_router.post("/menu-items/reorder")
async def reorder_menu_items(payload: dict, request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user): raise HTTPException(status_code=403, detail="Menu edit permission required")
    order = payload.get("order") or []
    for idx, iid in enumerate(order):
        try:
            await db.menu_items.update_one({"_id": ObjectId(iid)}, {"$set": {"sort_order": idx}})
        except Exception:
            pass
    _menu_cache_bust_all()
    return {"message": "Reordered", "count": len(order)}

# --- Menu bulk import / export (Excel / CSV) ---
# Lets a restaurant set up its whole menu from a spreadsheet instead of adding items
# one-by-one. Download the .xlsx template, fill it in, upload it back. Import upserts by
# item name (case-insensitive) and auto-creates any categories it hasn't seen before.
MENU_IMPORT_HEADERS = ["name", "category", "price", "description", "variations",
                       "discount_type", "discount_value", "stock", "is_popular", "is_bestseller"]

def _parse_bool_cell(v) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in ("1", "yes", "y", "true", "t", "on")

def _parse_variations_cell(v):
    """Parse a variations cell like 'Half=250; Full=450' into [{name, price}, ...].
    Accepts ';', ',' or newline separators and '=' or ':' between name and price.
    Malformed pairs are skipped rather than failing the whole row."""
    import re as _re_var
    out = []
    if v is None:
        return out
    text = str(v).strip()
    if not text:
        return out
    for part in _re_var.split(r"[;\n,]", text):
        part = part.strip()
        if not part:
            continue
        if "=" in part:
            name, _, price = part.partition("=")
        elif ":" in part:
            name, _, price = part.partition(":")
        else:
            continue
        name = name.strip()
        try:
            price_f = float(str(price).strip())
        except (TypeError, ValueError):
            continue
        if name:
            out.append({"name": name, "price": price_f})
    return out

@api_router.get("/menu-items/template")
async def download_menu_template(request: Request):
    user = await get_current_user(request)
    if not _can_edit_menu(user):
        raise HTTPException(status_code=403, detail="Menu edit permission required")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    header_fill = PatternFill(start_color="1E3F20", end_color="1E3F20", fill_type="solid")
    for col, h in enumerate(MENU_IMPORT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
    # Two example rows so staff can see the expected format (delete before importing).
    for r in [
        ["Chicken Biryani", "Biryani", 400, "Our signature dish", "Half=250; Full=450", "percentage", 10, 100, "yes", "yes"],
        ["Mineral Water 500ml", "Drinks", 80, "", "", "", "", 200, "no", "no"],
    ]:
        ws.append(r)
    for i, w in enumerate([22, 16, 8, 30, 24, 14, 14, 8, 12, 14], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    notes = wb.create_sheet("Instructions")
    for i, line in enumerate([
        "How to use this template",
        "1. Fill one row per menu item on the 'Menu' sheet.",
        "2. Required columns: name, category, price. The rest are optional.",
        "3. category: just type the category name — new categories are created automatically.",
        "4. variations: optional sizes. Format 'Half=250; Full=450'. Leave blank for none.",
        "5. discount_type: 'percentage' or 'fixed' (or blank). discount_value: the number.",
        "6. is_popular / is_bestseller: type yes or no.",
        "7. Delete the two example rows before importing your own menu.",
        "8. Save the file, then upload it with 'Import Excel' on the Menu page.",
        "Note: importing updates any item whose name already exists, and adds the rest.",
    ], start=1):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions["A"].width = 95
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="menu_template.xlsx"'},
    )

@api_router.post("/menu-items/import")
async def import_menu_items(request: Request, file: UploadFile = File(...)):
    import re as _re_imp
    user = await get_current_user(request)
    if not _can_edit_menu(user):
        raise HTTPException(status_code=403, detail="Menu edit permission required")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    fname = (file.filename or "").lower()

    rows = []  # list of {header: value}
    if fname.endswith(".csv"):
        import csv as _csv
        text = raw.decode("utf-8-sig", errors="replace")
        for r in _csv.DictReader(io.StringIO(text)):
            rows.append({(k or "").strip().lower(): v for k, v in r.items()})
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(status_code=400, detail="Could not read the file. Please upload the .xlsx template or a .csv file.")
        ws = wb["Menu"] if "Menu" in wb.sheetnames else wb.active
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                continue
            if row is None or all(c is None for c in row):
                continue
            rec = {}
            for i, c in enumerate(row):
                if i < len(header) and header[i]:
                    rec[header[i]] = c
            rows.append(rec)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found. Fill in the template and try again.")
    if len(rows) > 1000:
        raise HTTPException(status_code=400, detail="Too many rows (max 1000 per import).")

    cats = await db.categories.find({}).to_list(500)
    cat_by_name = {str(c.get("name", "")).strip().lower(): str(c["_id"]) for c in cats}
    cat_count = await db.categories.count_documents({})

    created = updated = cats_created = 0
    errors = []

    for idx, rec in enumerate(rows, start=2):  # row 2 = first data row (row 1 is the header)
        try:
            name = str(rec.get("name", "") or "").strip()
            cat_name = str(rec.get("category", "") or "").strip()
            if not name:
                errors.append({"row": idx, "message": "Missing name"}); continue
            if not cat_name:
                errors.append({"row": idx, "message": f"'{name}': missing category"}); continue
            try:
                price = float(rec.get("price"))
            except (TypeError, ValueError):
                errors.append({"row": idx, "message": f"'{name}': invalid or missing price"}); continue
            if price < 0:
                errors.append({"row": idx, "message": f"'{name}': price cannot be negative"}); continue

            key = cat_name.lower()
            cat_id = cat_by_name.get(key)
            if not cat_id:
                res = await db.categories.insert_one({"name": cat_name, "color": None, "sort_order": cat_count,
                                                      "created_at": datetime.now(timezone.utc).isoformat()})
                cat_id = str(res.inserted_id)
                cat_by_name[key] = cat_id
                cat_count += 1
                cats_created += 1

            d_type = str(rec.get("discount_type", "") or "").strip().lower() or None
            if d_type not in ("percentage", "fixed"):
                d_type = None
            try:
                d_val = float(rec.get("discount_value") or 0)
            except (TypeError, ValueError):
                d_val = 0.0
            try:
                stock = int(float(rec.get("stock"))) if rec.get("stock") not in (None, "") else 100
            except (TypeError, ValueError):
                stock = 100

            fields = {
                "name": name,
                "price": price,
                "category_id": cat_id,
                "description": str(rec.get("description", "") or "").strip(),
                "variations": _parse_variations_cell(rec.get("variations")),
                "discount_type": d_type,
                "discount_value": d_val,
                "stock": stock,
                "is_popular": _parse_bool_cell(rec.get("is_popular")),
                "is_bestseller": _parse_bool_cell(rec.get("is_bestseller")),
            }

            existing = await db.menu_items.find_one({"name": {"$regex": f"^{_re_imp.escape(name)}$", "$options": "i"}})
            if existing:
                await db.menu_items.update_one({"_id": existing["_id"]}, {"$set": fields})
                updated += 1
            else:
                sort_order = await db.menu_items.count_documents({})
                await db.menu_items.insert_one({
                    **fields,
                    "price_fp1": None, "price_fp2": None,
                    "low_stock_threshold": 10, "color": None,
                    "image_url": "", "image_type": "url",
                    "related_item_ids": [], "is_outsourced": False,
                    "outsourced_vendor_id": None, "outsourced_unit_cost": None,
                    "sort_order": sort_order, "created_at": datetime.now(timezone.utc).isoformat(),
                })
                created += 1
        except Exception as e:
            errors.append({"row": idx, "message": f"Unexpected error ({type(e).__name__})"})

    _menu_cache_bust_all()
    return {
        "created": created,
        "updated": updated,
        "categories_created": cats_created,
        "total_rows": len(rows),
        "error_count": len(errors),
        "errors": errors[:50],
    }

@api_router.get("/menu-items/export")
async def export_menu_items(request: Request):
    """Export the whole live menu as an .xlsx in the exact same column layout as the
    import template, so staff can download it, tweak prices/items in Excel, and upload
    it back to update everything at once (import upserts by name)."""
    user = await get_current_user(request)
    if not _can_edit_menu(user):
        raise HTTPException(status_code=403, detail="Menu edit permission required")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    cats = await db.categories.find({}).to_list(500)
    cat_name_by_id = {str(c["_id"]): str(c.get("name", "")) for c in cats}
    items = await db.menu_items.find({}).sort("sort_order", 1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    header_fill = PatternFill(start_color="1E3F20", end_color="1E3F20", fill_type="solid")
    for col, h in enumerate(MENU_IMPORT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill

    for it in items:
        variations = it.get("variations") or []
        var_str = "; ".join(
            f"{v.get('name')}={v.get('price')}" for v in variations
            if isinstance(v, dict) and v.get("name") is not None
        )
        ws.append([
            it.get("name", ""),
            cat_name_by_id.get(str(it.get("category_id", "")), ""),
            it.get("price", 0),
            it.get("description", "") or "",
            var_str,
            it.get("discount_type") or "",
            it.get("discount_value") or 0,
            it.get("stock") if it.get("stock") is not None else 100,
            "yes" if it.get("is_popular") else "no",
            "yes" if it.get("is_bestseller") else "no",
        ])

    for i, w in enumerate([22, 16, 8, 30, 24, 14, 14, 8, 12, 14], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="menu_export.xlsx"'},
    )

# --- Inventory ---
@api_router.get("/inventory")
async def get_inventory(request: Request):
    await get_current_user(request)
    items = await db.menu_items.find({}).to_list(500)
    # Fix N+1: load ALL categories once and index by id, instead of one
    # find_one() per menu item (was scaling linearly with menu size).
    cats = await db.categories.find({}).to_list(500)
    cat_by_id = {str(c["_id"]): c for c in cats}
    result = []
    for i in items:
        cid = i.get("category_id")
        cat = cat_by_id.get(cid) if cid else None
        result.append({"id": str(i["_id"]), "name": i["name"], "price": i["price"], "category_name": cat["name"] if cat else "Uncategorized", "stock": i.get("stock", 0), "low_stock_threshold": i.get("low_stock_threshold", 10), "is_low_stock": i.get("stock", 0) <= i.get("low_stock_threshold", 10)})
    return result

@api_router.put("/inventory/{item_id}")
async def update_stock(item_id: str, su: StockUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    await db.menu_items.update_one({"_id": ObjectId(item_id)}, {"$set": {"stock": su.stock}})
    _menu_cache_bust_all()
    return {"message": "Stock updated", "stock": su.stock}

# --- Orders ---
async def _finalize_pos_order(items, payment_type, subtotal, tax, total,
                              discount_type, discount_value, discount_amount, user,
                              order_type="takeaway", table_id=None):
    """Shared finalisation for a completed POS sale. Decrements stock, records the
    order in db.orders (the collection every report reads), and posts vendor
    transactions for any outsourced items. Used by both the quick-sale /orders
    endpoint and the dine-in table 'pay' endpoint so their behaviour is identical.

    `items` is a list of plain dicts: {item_id, name, price, original_price,
    quantity, removed_ingredients}. Any extra keys (e.g. kitchen_status) are ignored.
    """
    # Track outsourced items so we can post one vendor_transaction per vendor.
    outsourced_by_vendor = {}  # {vendor_id: {"items": [...], "total": float}}
    for oi in items:
        try:
            item = await db.menu_items.find_one({"_id": ObjectId(oi["item_id"])})
            if item:
                await db.menu_items.update_one({"_id": ObjectId(oi["item_id"])}, {"$set": {"stock": max(0, item.get("stock", 0) - oi["quantity"])}})
                # Outsourced -> accumulate vendor payable
                if item.get("is_outsourced") and item.get("outsourced_vendor_id"):
                    vid = item["outsourced_vendor_id"]
                    unit_cost = item.get("outsourced_unit_cost")
                    if unit_cost is None:
                        unit_cost = float(oi.get("original_price") or oi.get("price") or item.get("price", 0))
                    line_total = round(float(unit_cost) * int(oi["quantity"]), 2)
                    bucket = outsourced_by_vendor.setdefault(vid, {"items": [], "total": 0.0})
                    bucket["items"].append({
                        "name": oi["name"],
                        "quantity": int(oi["quantity"]),
                        "unit_price": float(unit_cost),
                        "menu_item_id": oi["item_id"],
                    })
                    bucket["total"] = round(bucket["total"] + line_total, 2)
        except Exception:
            pass
    now = datetime.now(timezone.utc)
    # Stock just changed on the items above — bust the menu cache so the next
    # /menu or /menu-items call sees the new stock figures instead of stale.
    if items:
        _menu_cache_bust_all()
    doc = {"items": [{"item_id": oi["item_id"], "name": oi["name"], "price": oi["price"], "original_price": oi.get("original_price") or oi["price"], "quantity": oi["quantity"], "removed_ingredients": [str(x).strip() for x in (oi.get("removed_ingredients") or []) if str(x).strip()]} for oi in items], "payment_type": payment_type, "subtotal": subtotal, "tax": tax, "total": total, "discount_type": discount_type, "discount_value": discount_value or 0, "discount_amount": discount_amount or 0, "order_type": order_type, "table_id": table_id, "status": "paid", "cashier_id": user["_id"], "cashier_name": user.get("name", ""), "created_at": now.isoformat(), "date": now.strftime("%Y-%m-%d")}
    result = await db.orders.insert_one(doc)
    order_id = str(result.inserted_id)
    order_receipt_no = order_id[-6:].upper()
    
    # Auto-create vendor transactions for outsourced items AND prepare vendor tickets
    vendor_tickets = []
    for vid, bucket in outsourced_by_vendor.items():
        try:
            # Get vendor info for ticket
            vendor = await db.vendors.find_one({"_id": ObjectId(vid)})
            vendor_name = vendor.get("name", "Unknown Vendor") if vendor else "Unknown Vendor"
            
            # Generate vendor ticket number
            vt_count = await db.vendor_transactions.count_documents({}) + 1
            ticket_no = f"VT-{vt_count:05d}"
            
            # Create vendor transaction
            await db.vendor_transactions.insert_one({
                "vendor_id": vid,
                "ticket_no": ticket_no,
                "items": bucket["items"],
                "total": bucket["total"],
                "notes": f"Auto: order #{order_receipt_no} (outsourced sale)",
                "auto_source": "order",
                "source_order_id": order_id,
                "created_by": user["_id"],
                "created_by_name": user.get("name", ""),
                "date": now.strftime("%Y-%m-%d"),
                "created_at": now.isoformat(),
            })
            
            # Prepare vendor ticket data for auto-printing
            vendor_tickets.append({
                "vendor_id": vid,
                "vendor_name": vendor_name,
                "ticket_no": ticket_no,
                "order_receipt_no": order_receipt_no,
                "items": bucket["items"],
                "total": bucket["total"],
                "date": now.strftime("%Y-%m-%d"),
                "time": now.strftime("%H:%M:%S"),
                "cashier_name": user.get("name", ""),
            })
        except Exception as e:
            logger.warning(f"Failed to create vendor transaction for vendor {vid}: {e}")
            pass
    
    return {
        "id": order_id,
        "items": doc["items"],
        "payment_type": payment_type,
        "subtotal": subtotal,
        "tax": tax,
        "total": total,
        "discount_amount": discount_amount or 0,
        "order_type": order_type,
        "table_id": table_id,
        "cashier_name": user.get("name", ""),
        "created_at": now.isoformat(),
        "vendor_tickets": vendor_tickets  # NEW: vendor tickets for auto-printing
    }

@api_router.post("/orders")
async def create_order(order: OrderCreate, request: Request):
    user = await get_current_user(request)
    items = [{"item_id": oi.item_id, "name": oi.name, "price": oi.price, "original_price": oi.original_price, "quantity": oi.quantity, "removed_ingredients": oi.removed_ingredients or []} for oi in order.items]
    return await _finalize_pos_order(
        items, order.payment_type, order.subtotal, order.tax, order.total,
        order.discount_type, order.discount_value, order.discount_amount, user,
        order_type=getattr(order, "order_type", None) or "takeaway", table_id=getattr(order, "table_id", None),
    )

# =============================================================================
# DINE-IN: restaurant tables (floor) + open (unpaid) orders  (ported from Marhaba)
# =============================================================================
def _serialize_table(t):
    return {"id": str(t["_id"]), "name": t.get("name", ""), "section": t.get("section", "Main Hall"),
            "capacity": t.get("capacity", 4), "status": t.get("status", "available")}

def _serialize_open_order(o):
    return {"id": str(o["_id"]), "order_type": o.get("order_type", "dine_in"), "table_id": o.get("table_id"),
            "status": o.get("status", "open"), "items": o.get("items", []),
            "subtotal": o.get("subtotal", 0), "tax": o.get("tax", 0), "total": o.get("total", 0),
            "discount_type": o.get("discount_type"), "discount_value": o.get("discount_value", 0),
            "discount_amount": o.get("discount_amount", 0), "created_at": o.get("created_at", "")}

@api_router.get("/tables")
async def list_tables(request: Request):
    await get_current_user(request)
    tables = await db.restaurant_tables.find({}).to_list(1000)
    tables.sort(key=lambda t: (t.get("section", ""), t.get("name", "")))
    # Attach a live-order summary per table so the floor can flag an unpaid running
    # tab INDEPENDENTLY of the table's status colour (a mis-tapped status must never
    # hide the fact that money is still owed on that table).
    open_orders = await db.open_orders.find({"status": "open"}).to_list(5000)
    by_table = {}
    for o in open_orders:
        tid = o.get("table_id")
        if not tid:
            continue
        items = o.get("items", [])
        by_table[tid] = {
            "open_order_id": str(o["_id"]),
            "item_count": sum(int(it.get("quantity", 0)) for it in items),
            "has_items": len(items) > 0,
            "unsent_count": sum(1 for it in items if (it.get("kitchen_status") or "new") == "new"),
            "total": o.get("total", 0),
            "created_at": o.get("created_at", ""),
        }
    out = []
    for t in tables:
        st = _serialize_table(t)
        st["open_order"] = by_table.get(st["id"])  # None when no live tab
        out.append(st)
    return out

@api_router.post("/tables")
async def create_table(body: TableCreate, request: Request):
    await get_current_user(request)
    doc = {"name": body.name.strip(), "section": (body.section or "Main Hall").strip(),
           "capacity": int(body.capacity or 4), "status": body.status or "available",
           "created_at": datetime.now(timezone.utc).isoformat()}
    res = await db.restaurant_tables.insert_one(doc)
    doc["_id"] = res.inserted_id
    return _serialize_table(doc)

@api_router.put("/tables/{table_id}")
async def update_table(table_id: str, body: TableUpdate, request: Request):
    await get_current_user(request)
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if "name" in updates: updates["name"] = updates["name"].strip()
    if "capacity" in updates: updates["capacity"] = int(updates["capacity"])
    if updates:
        await db.restaurant_tables.update_one({"_id": ObjectId(table_id)}, {"$set": updates})
    t = await db.restaurant_tables.find_one({"_id": ObjectId(table_id)})
    if not t: raise HTTPException(status_code=404, detail="Table not found")
    return _serialize_table(t)

@api_router.delete("/tables/{table_id}")
async def delete_table(table_id: str, request: Request):
    await get_current_user(request)
    # Guard: don't delete a table that still has an open order.
    open_o = await db.open_orders.find_one({"table_id": table_id, "status": "open"})
    if open_o:
        raise HTTPException(status_code=400, detail="Table has an open order — settle it first")
    await db.restaurant_tables.delete_one({"_id": ObjectId(table_id)})
    return {"message": "Table deleted"}

@api_router.post("/open-orders")
async def create_open_order(body: OpenOrderCreate, request: Request):
    """Open a dine-in tab for a table. If one already exists (occupied table),
    return it — enforcing one open order per table. Marks the table occupied."""
    user = await get_current_user(request)
    table = await db.restaurant_tables.find_one({"_id": ObjectId(body.table_id)})
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    existing = await db.open_orders.find_one({"table_id": body.table_id, "status": "open"})
    if existing:
        return _serialize_open_order(existing)
    now = datetime.now(timezone.utc)
    doc = {"order_type": "dine_in", "table_id": body.table_id, "status": "open",
           "items": [], "subtotal": 0, "tax": 0, "total": 0,
           "discount_type": None, "discount_value": 0, "discount_amount": 0,
           "cashier_id": user["_id"], "cashier_name": user.get("name", ""),
           "created_at": now.isoformat()}
    res = await db.open_orders.insert_one(doc)
    doc["_id"] = res.inserted_id
    await db.restaurant_tables.update_one({"_id": ObjectId(body.table_id)}, {"$set": {"status": "occupied"}})
    return _serialize_open_order(doc)

@api_router.get("/open-orders/by-table/{table_id}")
async def get_open_order_by_table(table_id: str, request: Request):
    await get_current_user(request)
    o = await db.open_orders.find_one({"table_id": table_id, "status": "open"})
    if not o: raise HTTPException(status_code=404, detail="No open order for this table")
    return _serialize_open_order(o)

@api_router.get("/open-orders/{open_order_id}")
async def get_open_order(open_order_id: str, request: Request):
    await get_current_user(request)
    o = await db.open_orders.find_one({"_id": ObjectId(open_order_id)})
    if not o: raise HTTPException(status_code=404, detail="Open order not found")
    return _serialize_open_order(o)

@api_router.put("/open-orders/{open_order_id}")
async def update_open_order_items(open_order_id: str, body: OpenOrderItemsUpdate, request: Request):
    """Persist the current cart to the tab. Item kitchen_status is taken as-is from
    the client (new lines default to 'new'; already-sent lines stay 'sent')."""
    await get_current_user(request)
    o = await db.open_orders.find_one({"_id": ObjectId(open_order_id)})
    if not o: raise HTTPException(status_code=404, detail="Open order not found")
    if o.get("status") != "open":
        raise HTTPException(status_code=400, detail="Order is not open")
    items = [{"uid": it.uid, "item_id": it.item_id, "name": it.name, "price": it.price,
              "original_price": it.original_price if it.original_price is not None else it.price,
              "quantity": it.quantity,
              "removed_ingredients": [str(x).strip() for x in (it.removed_ingredients or []) if str(x).strip()],
              "kitchen_status": it.kitchen_status or "new"} for it in body.items]
    await db.open_orders.update_one({"_id": ObjectId(open_order_id)}, {"$set": {
        "items": items, "subtotal": body.subtotal or 0, "tax": body.tax or 0, "total": body.total or 0,
        "discount_type": body.discount_type, "discount_value": body.discount_value or 0,
        "discount_amount": body.discount_amount or 0,
        "updated_at": datetime.now(timezone.utc).isoformat()}})
    o = await db.open_orders.find_one({"_id": ObjectId(open_order_id)})
    return _serialize_open_order(o)

@api_router.post("/open-orders/{open_order_id}/send-kitchen")
async def send_open_order_to_kitchen(open_order_id: str, request: Request):
    """Return the items whose kitchen_status is 'new', then mark them 'sent'.
    Already-sent items are never returned again (no duplicate sends)."""
    await get_current_user(request)
    o = await db.open_orders.find_one({"_id": ObjectId(open_order_id)})
    if not o: raise HTTPException(status_code=404, detail="Open order not found")
    items = o.get("items", [])
    new_items = [it for it in items if (it.get("kitchen_status") or "new") == "new"]
    if not new_items:
        return {"new_items": [], "message": "No new items to send"}
    for it in items:
        if (it.get("kitchen_status") or "new") == "new":
            it["kitchen_status"] = "sent"
    await db.open_orders.update_one({"_id": ObjectId(open_order_id)}, {"$set": {
        "items": items, "kitchen_sent_at": datetime.now(timezone.utc).isoformat()}})
    return {"new_items": new_items, "message": f"Sent {len(new_items)} item(s) to kitchen"}

@api_router.post("/open-orders/{open_order_id}/pay")
async def pay_open_order(open_order_id: str, body: OpenOrderPay, request: Request):
    """Close a dine-in tab: finalise it as a normal paid order (stock, vendor
    tickets, reports), mark the tab paid, and free the table."""
    user = await get_current_user(request)
    o = await db.open_orders.find_one({"_id": ObjectId(open_order_id)})
    if not o: raise HTTPException(status_code=404, detail="Open order not found")
    if o.get("status") != "open":
        raise HTTPException(status_code=400, detail="Order is not open")
    items = o.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="Cannot pay an empty order")
    result = await _finalize_pos_order(
        items, body.payment_type, body.subtotal, body.tax, body.total,
        body.discount_type, body.discount_value, body.discount_amount, user,
        order_type="dine_in", table_id=o.get("table_id"),
    )
    await db.open_orders.update_one({"_id": ObjectId(open_order_id)}, {"$set": {
        "status": "paid", "final_order_id": result["id"],
        "paid_at": datetime.now(timezone.utc).isoformat()}})
    if o.get("table_id"):
        await db.restaurant_tables.update_one({"_id": ObjectId(o["table_id"])}, {"$set": {"status": "available"}})
    return result

@api_router.post("/open-orders/{open_order_id}/cancel")
async def cancel_open_order(open_order_id: str, request: Request):
    """Cancel a dine-in tab that has NO items (customer left before ordering) and
    free the table — no payment required. Refuses if the tab already has items,
    which must go through the normal payment workflow instead."""
    await get_current_user(request)
    o = await db.open_orders.find_one({"_id": ObjectId(open_order_id)})
    if not o: raise HTTPException(status_code=404, detail="Open order not found")
    if o.get("status") != "open":
        raise HTTPException(status_code=400, detail="Order is not open")
    if o.get("items"):
        raise HTTPException(status_code=400, detail="Order has items — settle it via payment instead")
    await db.open_orders.update_one({"_id": ObjectId(open_order_id)}, {"$set": {
        "status": "cancelled", "cancelled_at": datetime.now(timezone.utc).isoformat()}})
    if o.get("table_id"):
        await db.restaurant_tables.update_one({"_id": ObjectId(o["table_id"])}, {"$set": {"status": "available"}})
    return {"message": "Order cancelled, table cleared", "table_id": o.get("table_id")}

@api_router.get("/orders/today")
async def get_today_orders(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    orders = await db.orders.find({"date": today}).sort("created_at", -1).to_list(1000)
    out = []
    for o in orders:
        oid = str(o.pop("_id"))
        o["id"] = oid
        o["receipt_no"] = oid[-6:].upper()
        out.append(o)
    return out

@api_router.get("/orders/history")
async def get_orders_history(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, payment_type: Optional[str] = None, q: Optional[str] = None, limit: int = 500):
    user = await get_current_user(request)
    perms = user.get("permissions") or []
    if user.get("role") != "admin" and "orders_history" not in perms:
        raise HTTPException(status_code=403, detail="Orders history permission required")
    query = {}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    if payment_type and payment_type != "all":
        query["payment_type"] = payment_type
    orders = await db.orders.find(query).sort("created_at", -1).to_list(max(1, min(limit, 5000)))
    out = []
    for o in orders:
        oid = str(o.pop("_id"))
        o["id"] = oid
        o["receipt_no"] = oid[-6:].upper()
        if q:
            ql = q.lower().strip()
            hay = f"{o['receipt_no']} {o.get('cashier_name','')} {o.get('date','')}".lower()
            items_hay = " ".join([str(i.get("name","")) for i in o.get("items", [])]).lower()
            if ql not in hay and ql not in items_hay:
                continue
        out.append(o)
    return out

@api_router.get("/orders/search/{receipt_id}")
async def search_order(receipt_id: str, request: Request):
    await get_current_user(request)
    # Search by last 6 chars of _id
    orders = await db.orders.find({}).to_list(50000)
    for o in orders:
        oid = str(o["_id"])
        if oid.endswith(receipt_id.lower()) or oid[-6:].upper() == receipt_id.upper():
            o.pop("_id", None)
            o["id"] = oid
            return o
    raise HTTPException(status_code=404, detail="Receipt not found")

# --- Voice Assistant (Urdu / Punjabi) ---
# Pipeline: audio → Whisper STT → GPT-4o parser → structured items + Urdu confirmation + TTS audio
import io, base64, json as _json

_EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY", "")

def _voice_ready():
    return bool(_EMERGENT_KEY)

async def _stt_transcribe(audio_bytes: bytes, filename: str, language: Optional[str] = None) -> str:
    from emergentintegrations.llm.openai.speech_to_text import OpenAISpeechToText
    stt = OpenAISpeechToText(api_key=_EMERGENT_KEY)
    bio = io.BytesIO(audio_bytes); bio.name = filename or "audio.webm"
    res = await stt.transcribe(bio, "whisper-1", "json", None, language)
    if isinstance(res, dict):
        return res.get("text", "")
    if hasattr(res, "text"):
        return res.text
    return str(res)

async def _llm_parse_order(transcript: str, menu_names: List[str]) -> dict:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    system = (
        "You parse restaurant POS orders spoken in Urdu, Punjabi, Hindi, or English. "
        "Return ONLY compact JSON – no prose. Schema: "
        '{"intent":"order"|"expense"|"unknown","items":[{"name":str,"qty":int}],"expense":{"description":str,"amount":float}|null,"confidence":0..1}. '
        f"Available menu item names (match case-insensitively; pick the closest one): {menu_names}. "
        "Urdu/Punjabi numbers: ek=1, do=2, tin=3, char=4, paanch=5, chha=6, saat=7, aath=8, nau=9, dus=10. "
        "If user says something like 'kharch', 'expense', 'spent', 'kharcha', treat as expense and extract amount + description. "
        "If nothing matches confidently, set intent='unknown' and items=[]."
    )
    chat = LlmChat(api_key=_EMERGENT_KEY, session_id=f"voice-{datetime.now(timezone.utc).timestamp()}", system_message=system).with_model("openai", "gpt-4o-mini")
    reply = await chat.send_message(UserMessage(text=f"Transcript: {transcript}\nReturn only JSON."))
    text = reply if isinstance(reply, str) else getattr(reply, "content", str(reply))
    # strip code fences
    t = text.strip()
    if t.startswith("```"):
        t = t.strip("`")
        if t.lower().startswith("json"): t = t[4:]
    try:
        return _json.loads(t)
    except Exception:
        # try to find first {...}
        start = t.find("{"); end = t.rfind("}")
        if start >= 0 and end > start:
            try: return _json.loads(t[start:end+1])
            except Exception: pass
        return {"intent": "unknown", "items": [], "expense": None, "confidence": 0.0}

async def _tts_speak(text: str, voice: str = "alloy") -> bytes:
    from emergentintegrations.llm.openai.text_to_speech import OpenAITextToSpeech
    tts = OpenAITextToSpeech(api_key=_EMERGENT_KEY)
    return await tts.generate_speech(text=text, model="tts-1", voice=voice, speed=1.0, response_format="mp3")

def _build_urdu_confirmation(parsed: dict, currency: str, items_resolved: list, total: float) -> str:
    """
    Build a confirmation that keeps NUMBERS and ITEM NAMES in English (TTS pronounces them clearly)
    while connector words stay in Urdu so it still feels native.
    """
    intent = parsed.get("intent", "unknown")
    if intent == "expense" and parsed.get("expense"):
        e = parsed["expense"]
        return f"خرچہ — {e.get('description','')} — {currency} {int(e.get('amount',0))}۔ کیا آپ تصدیق کرتے ہیں؟"
    if intent == "order" and items_resolved:
        parts = []
        for it in items_resolved:
            q = int(it.get("quantity", 1))
            # English digits + English item name = clear TTS, Urdu structure around it
            parts.append(f"{q} {it['name']}")
        items_text = " اور ".join(parts)  # Urdu "and"
        return f"آپ کا آرڈر — {items_text}۔ کل {currency} {int(total)}۔ کیا آپ تصدیق کرتے ہیں؟"
    return "معاف کیجیے، میں سمجھ نہیں سکا۔ براہ کرم دوبارہ کہیں۔"

@api_router.get("/voice/status")
async def voice_status(request: Request):
    await get_current_user(request)
    return {"enabled": _voice_ready()}

@api_router.post("/voice/parse")
async def voice_parse(request: Request, audio: UploadFile = File(...), language: Optional[str] = Form(None)):
    """Accept an audio blob, transcribe, parse against current menu, and return parsed order + Urdu TTS."""
    await get_current_user(request)
    if not _voice_ready():
        raise HTTPException(status_code=503, detail="Voice assistant not configured (EMERGENT_LLM_KEY missing)")
    data = await audio.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty audio")
    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Audio too large (max 25MB)")

    try:
        transcript = await _stt_transcribe(data, audio.filename or "audio.webm", language)
    except Exception as e:
        logger.exception("STT failed")
        raise HTTPException(status_code=502, detail=f"Transcription failed: {str(e)[:200]}")
    if not transcript or not transcript.strip():
        raise HTTPException(status_code=422, detail="Could not transcribe any speech. Please try again.")

    menu_docs = await db.menu_items.find({}).to_list(500)
    menu_index = {m["name"].lower(): {"id": str(m["_id"]), "name": m["name"], "price": m.get("price", 0), "stock": m.get("stock", 0)} for m in menu_docs}
    menu_names = list(menu_index.keys())

    try:
        parsed = await _llm_parse_order(transcript, menu_names)
    except Exception as e:
        logger.exception("LLM parse failed")
        raise HTTPException(status_code=502, detail=f"Parser failed: {str(e)[:200]}")

    # Resolve items against menu
    items_resolved = []
    subtotal = 0.0
    for it in (parsed.get("items") or []):
        nm = str(it.get("name", "")).lower().strip()
        qty = int(it.get("qty") or 1)
        if not nm: continue
        m = menu_index.get(nm)
        if not m:
            # fuzzy: find a menu name containing the token or vice versa
            for k, v in menu_index.items():
                if nm in k or k in nm:
                    m = v; break
        if m:
            items_resolved.append({"item_id": m["id"], "name": m["name"], "price": m["price"], "quantity": qty})
            subtotal += m["price"] * qty

    s = await db.settings.find_one({"key": "global"}, {"_id": 0}) or {}
    currency = s.get("currency", "Rs")
    urdu_text = _build_urdu_confirmation(parsed, currency, items_resolved, subtotal)

    try:
        audio_bytes = await _tts_speak(urdu_text)
        audio_b64 = base64.b64encode(audio_bytes).decode()
    except Exception as e:
        logger.warning(f"TTS failed (non-fatal): {e}")
        audio_b64 = ""

    return {
        "transcript": transcript,
        "parsed": parsed,
        "items": items_resolved,
        "subtotal": round(subtotal, 2),
        "expense": parsed.get("expense"),
        "intent": parsed.get("intent", "unknown"),
        "confirmation_text": urdu_text,
        "audio_base64": audio_b64,
        "currency": currency,
    }

# --- Refunds ---
@api_router.post("/refunds")
async def create_refund(refund: RefundCreate, request: Request):
    user = await get_current_user(request)
    perms = user.get("permissions", [])
    if user.get("role") != "admin" and "refunds" not in perms:
        raise HTTPException(status_code=403, detail="Refund permission required")
    now = datetime.now(timezone.utc)
    refund_no = f"RF-{await db.refunds.count_documents({}) + 1:05d}"
    # Reverse vendor payables for any outsourced items being refunded
    reversed_by_vendor = {}
    for ri in (refund.items or []):
        try:
            mid = ri.get("item_id") or ri.get("menu_item_id")
            qty = int(ri.get("quantity", 0))
            if not mid or qty <= 0:
                continue
            item = await db.menu_items.find_one({"_id": ObjectId(mid)})
            if item and item.get("is_outsourced") and item.get("outsourced_vendor_id"):
                vid = item["outsourced_vendor_id"]
                unit_cost = item.get("outsourced_unit_cost")
                if unit_cost is None:
                    unit_cost = float(ri.get("price") or item.get("price", 0))
                line_total = round(float(unit_cost) * qty, 2)
                bucket = reversed_by_vendor.setdefault(vid, {"items": [], "total": 0.0})
                bucket["items"].append({
                    "name": ri.get("name") or item.get("name"),
                    "quantity": qty,
                    "unit_price": float(unit_cost),
                    "menu_item_id": mid,
                })
                bucket["total"] = round(bucket["total"] + line_total, 2)
        except Exception:
            pass
    doc = {
        "refund_no": refund_no,
        "order_id": refund.order_id,
        "reason": refund.reason,
        "amount": refund.amount,
        "items": refund.items or [],
        "refunded_by": user["_id"],
        "refunded_by_name": user.get("name", ""),
        "date": now.strftime("%Y-%m-%d"),
        "created_at": now.isoformat(),
    }
    result = await db.refunds.insert_one(doc)
    refund_id = str(result.inserted_id)
    # Insert NEGATIVE vendor transactions to reverse the payable
    for vid, bucket in reversed_by_vendor.items():
        try:
            await db.vendor_transactions.insert_one({
                "vendor_id": vid,
                "items": bucket["items"],
                "total": -bucket["total"],
                "notes": f"Auto reversal: refund {refund_no} (outsourced)",
                "auto_source": "refund",
                "source_refund_id": refund_id,
                "source_order_id": refund.order_id,
                "created_by": user["_id"],
                "created_by_name": user.get("name", ""),
                "date": now.strftime("%Y-%m-%d"),
                "created_at": now.isoformat(),
            })
        except Exception:
            pass
    return {"id": refund_id, "refund_no": refund_no, "amount": refund.amount, "reason": refund.reason, "date": now.strftime("%Y-%m-%d"), "created_at": now.isoformat(), "refunded_by_name": user.get("name", "")}

@api_router.get("/refunds/today")
async def get_today_refunds(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    refunds = await db.refunds.find({"date": today}).sort("created_at", -1).to_list(500)
    return [{"id": str(r["_id"]), "refund_no": r.get("refund_no", ""), "order_id": r.get("order_id", ""), "reason": r.get("reason", ""), "amount": r.get("amount", 0), "items": r.get("items", []), "refunded_by_name": r.get("refunded_by_name", ""), "date": r.get("date", ""), "created_at": r.get("created_at", "")} for r in refunds]

@api_router.get("/refunds/summary")
async def get_refund_summary(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    refunds = await db.refunds.find({"date": today}).to_list(500)
    total = sum(r.get("amount", 0) for r in refunds)
    return {"date": today, "total_refunds": round(total, 2), "count": len(refunds)}

# --- Expenses ---
@api_router.get("/expenses")
async def get_expenses(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    exps = await db.expenses.find({"date": today}).sort("created_at", -1).to_list(500)
    return [{"id": str(e["_id"]), "description": e["description"], "amount": e["amount"], "category": e.get("category", "general"), "created_at": e.get("created_at", ""), "date": e.get("date", today)} for e in exps]

@api_router.post("/expenses")
async def create_expense(exp: ExpenseCreate, request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    doc = {"description": exp.description, "amount": exp.amount, "category": exp.category or "general", "created_by": user["_id"], "created_by_name": user.get("name", ""), "created_at": now.isoformat(), "date": now.strftime("%Y-%m-%d")}
    result = await db.expenses.insert_one(doc)
    return {"id": str(result.inserted_id), "description": exp.description, "amount": exp.amount, "category": exp.category, "created_at": now.isoformat()}

@api_router.delete("/expenses/{exp_id}")
async def delete_expense(exp_id: str, request: Request):
    await get_current_user(request)
    await db.expenses.delete_one({"_id": ObjectId(exp_id)})
    return {"message": "Deleted"}

@api_router.get("/expenses/summary")
async def get_expenses_summary(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    exps = await db.expenses.find({"date": today}).to_list(500)
    total = sum(e.get("amount", 0) for e in exps)
    return {"date": today, "total_expenses": round(total, 2), "count": len(exps)}

# --- Vendors ---
@api_router.get("/vendors")
async def list_vendors(request: Request):
    await get_current_user(request)
    vendors = await db.vendors.find({}).to_list(500)
    result = []
    for v in vendors:
        vid = str(v["_id"])
        txns = await db.vendor_transactions.find({"vendor_id": vid}).to_list(10000)
        pmts = await db.vendor_payments.find({"vendor_id": vid}).to_list(10000)
        total_billed = sum(t.get("total", 0) for t in txns)
        total_paid = sum(p.get("amount", 0) for p in pmts)
        balance = round(total_billed - total_paid, 2)
        result.append({"id": vid, "name": v["name"], "contact": v.get("contact", ""), "items_supplied": v.get("items_supplied", ""), "products": v.get("products", []), "total_billed": round(total_billed, 2), "total_paid": round(total_paid, 2), "balance": balance})
    return result

@api_router.post("/vendors")
async def create_vendor(v: VendorCreate, request: Request):
    user = await get_current_user(request)
    doc = {"name": v.name, "contact": v.contact, "items_supplied": v.items_supplied, "products": [], "created_at": datetime.now(timezone.utc).isoformat()}
    result = await db.vendors.insert_one(doc)
    return {"id": str(result.inserted_id), "name": v.name, "contact": v.contact, "items_supplied": v.items_supplied, "products": [], "total_billed": 0, "total_paid": 0, "balance": 0}

@api_router.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, v: VendorUpdate, request: Request):
    await get_current_user(request)
    ud = {k: val for k, val in v.model_dump().items() if val is not None}
    if ud: await db.vendors.update_one({"_id": ObjectId(vendor_id)}, {"$set": ud})
    return {"message": "Updated"}

@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str, request: Request):
    await get_current_user(request)
    await db.vendors.delete_one({"_id": ObjectId(vendor_id)})
    return {"message": "Deleted"}

@api_router.get("/vendors/{vendor_id}/products")
async def get_vendor_products(vendor_id: str, request: Request):
    await get_current_user(request)
    vendor = await db.vendors.find_one({"_id": ObjectId(vendor_id)})
    if not vendor: raise HTTPException(status_code=404, detail="Vendor not found")
    return vendor.get("products", [])

@api_router.post("/vendors/{vendor_id}/products")
async def add_vendor_product(vendor_id: str, request: Request):
    await get_current_user(request)
    body = await request.json()
    name = body.get("name", "").strip()
    default_price = body.get("default_price", 0)
    if not name: raise HTTPException(status_code=400, detail="Product name required")
    await db.vendors.update_one({"_id": ObjectId(vendor_id)}, {"$push": {"products": {"name": name, "default_price": default_price}}})
    return {"message": "Product added", "name": name, "default_price": default_price}

@api_router.delete("/vendors/{vendor_id}/products/{product_name}")
async def remove_vendor_product(vendor_id: str, product_name: str, request: Request):
    await get_current_user(request)
    await db.vendors.update_one({"_id": ObjectId(vendor_id)}, {"$pull": {"products": {"name": product_name}}})
    return {"message": "Product removed"}

@api_router.get("/vendors/{vendor_id}/transactions")
async def get_vendor_transactions(vendor_id: str, request: Request):
    await get_current_user(request)
    txns = await db.vendor_transactions.find({"vendor_id": vendor_id}).sort("created_at", -1).to_list(500)
    return [{"id": str(t["_id"]), "ticket_no": t.get("ticket_no", ""), "vendor_id": t["vendor_id"], "items": t.get("items", []), "total": t.get("total", 0), "notes": t.get("notes", ""), "date": t.get("date", ""), "created_at": t.get("created_at", "")} for t in txns]

@api_router.post("/vendors/{vendor_id}/transactions")
async def create_vendor_transaction(vendor_id: str, txn: VendorTransactionCreate, request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    # Generate ticket number
    count = await db.vendor_transactions.count_documents({}) + 1
    ticket_no = f"VT-{count:05d}"
    doc = {"vendor_id": vendor_id, "ticket_no": ticket_no, "items": txn.items, "total": txn.total, "notes": txn.notes, "created_by": user["_id"], "date": now.strftime("%Y-%m-%d"), "created_at": now.isoformat()}
    result = await db.vendor_transactions.insert_one(doc)
    return {"id": str(result.inserted_id), "ticket_no": ticket_no, "vendor_id": vendor_id, "items": txn.items, "total": txn.total, "notes": txn.notes, "date": now.strftime("%Y-%m-%d"), "created_at": now.isoformat()}

@api_router.get("/vendors/{vendor_id}/payments")
async def get_vendor_payments(vendor_id: str, request: Request):
    await get_current_user(request)
    pmts = await db.vendor_payments.find({"vendor_id": vendor_id}).sort("created_at", -1).to_list(500)
    return [{"id": str(p["_id"]), "vendor_id": p["vendor_id"], "amount": p.get("amount", 0), "notes": p.get("notes", ""), "date": p.get("date", ""), "created_at": p.get("created_at", "")} for p in pmts]

@api_router.post("/vendors/{vendor_id}/payments")
async def create_vendor_payment(vendor_id: str, pmt: VendorPaymentCreate, request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    doc = {"vendor_id": vendor_id, "amount": pmt.amount, "notes": pmt.notes, "created_by": user["_id"], "date": now.strftime("%Y-%m-%d"), "created_at": now.isoformat()}
    result = await db.vendor_payments.insert_one(doc)
    return {"id": str(result.inserted_id), "amount": pmt.amount, "notes": pmt.notes, "date": now.strftime("%Y-%m-%d")}

@api_router.get("/vendors/{vendor_id}/today")
async def get_vendor_today(vendor_id: str, request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    txns = await db.vendor_transactions.find({"vendor_id": vendor_id, "date": today}).to_list(500)
    pmts = await db.vendor_payments.find({"vendor_id": vendor_id, "date": today}).to_list(500)
    billed = sum(t.get("total", 0) for t in txns)
    paid = sum(p.get("amount", 0) for p in pmts)
    items = []
    for t in txns:
        for i in t.get("items", []):
            items.append(i)
    return {"date": today, "items": items, "total_billed": round(billed, 2), "total_paid": round(paid, 2), "balance": round(billed - paid, 2), "transactions": len(txns)}

@api_router.get("/vendors/{vendor_id}/sales-summary")
async def vendor_sales_summary(vendor_id: str, request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Aggregated outsourced-product sales for a vendor over a date range.
    Surfaces what was auto-billed via outsourced order flow vs. payments made.
    """
    await get_current_user(request)
    q = {"vendor_id": vendor_id}
    if start_date and end_date:
        q["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        q["date"] = {"$gte": start_date}
    elif end_date:
        q["date"] = {"$lte": end_date}
    txns = await db.vendor_transactions.find(q).to_list(5000)
    pmts = await db.vendor_payments.find(q).to_list(5000)
    auto_billed = 0.0  # from auto outsourced sales
    auto_reversed = 0.0  # from refunds reversing outsourced sales
    manual_billed = 0.0
    by_product = {}
    for t in txns:
        amount = float(t.get("total", 0))
        src = t.get("auto_source")
        if src == "order":
            auto_billed += amount
        elif src == "refund":
            auto_reversed += amount  # already negative
        else:
            manual_billed += amount
        for it in t.get("items", []):
            key = it.get("name", "Unknown")
            entry = by_product.setdefault(key, {"name": key, "quantity": 0, "total": 0.0})
            entry["quantity"] += int(it.get("quantity", 0)) * (1 if amount >= 0 else -1)
            entry["total"] = round(entry["total"] + (float(it.get("unit_price", 0)) * int(it.get("quantity", 0))) * (1 if amount >= 0 else -1), 2)
    total_paid = sum(float(p.get("amount", 0)) for p in pmts)
    total_billed = round(auto_billed + auto_reversed + manual_billed, 2)
    return {
        "vendor_id": vendor_id,
        "start_date": start_date,
        "end_date": end_date,
        "auto_billed_from_orders": round(auto_billed, 2),
        "auto_reversed_from_refunds": round(auto_reversed, 2),
        "manual_billed": round(manual_billed, 2),
        "total_billed": total_billed,
        "total_paid": round(total_paid, 2),
        "balance": round(total_billed - total_paid, 2),
        "products": sorted(by_product.values(), key=lambda x: x["total"], reverse=True),
        "transactions_count": len(txns),
        "payments_count": len(pmts),
    }

# --- Reports ---
def calc_report(orders, expenses=None, refunds=None):
    total_sales = sum(o.get("total", 0) for o in orders)
    cash = sum(o.get("total", 0) for o in orders if o.get("payment_type") == "cash")
    credit = sum(o.get("total", 0) for o in orders if o.get("payment_type") == "credit")
    fp1 = sum(o.get("total", 0) for o in orders if o.get("payment_type") == "foodpanda1")
    fp2 = sum(o.get("total", 0) for o in orders if o.get("payment_type") == "foodpanda2")
    online = fp1 + fp2
    total_orders = len(orders)
    total_items = sum(sum(i.get("quantity", 0) for i in o.get("items", [])) for o in orders)
    ic = {}
    for o in orders:
        for i in o.get("items", []):
            k = i.get("name", "Unknown")
            ic[k] = ic.get(k, 0) + i.get("quantity", 0)
    top = sorted(ic.items(), key=lambda x: x[1], reverse=True)[:10]
    total_exp = sum(e.get("amount", 0) for e in (expenses or []))
    total_ref = sum(r.get("amount", 0) for r in (refunds or []))
    return {"total_sales": round(total_sales, 2), "cash_sales": round(cash, 2), "credit_sales": round(credit, 2), "foodpanda1_sales": round(fp1, 2), "foodpanda2_sales": round(fp2, 2), "online_sales": round(online, 2), "total_orders": total_orders, "total_items_sold": total_items, "top_items": [{"name": n, "quantity": q} for n, q in top], "total_expenses": round(total_exp, 2), "total_refunds": round(total_ref, 2), "net_revenue": round(total_sales - total_exp - total_ref, 2)}

@api_router.get("/reports/x")
async def get_x_report(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    orders = await db.orders.find({"date": today}).to_list(10000)
    expenses = await db.expenses.find({"date": today}).to_list(500)
    refunds = await db.refunds.find({"date": today}).to_list(500)
    r = calc_report(orders, expenses, refunds)
    r.update({"date": today, "report_type": "X", "generated_at": datetime.now(timezone.utc).isoformat()})
    return r

@api_router.get("/reports/z")
async def get_z_report(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin" and "reports_z" not in (user.get("permissions") or []):
        raise HTTPException(status_code=403, detail="Z Report permission required")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    orders = await db.orders.find({"date": today}).to_list(10000)
    expenses = await db.expenses.find({"date": today}).to_list(500)
    refunds = await db.refunds.find({"date": today}).to_list(500)
    r = calc_report(orders, expenses, refunds)
    r.update({"date": today, "report_type": "Z", "generated_at": datetime.now(timezone.utc).isoformat()})
    return r

@api_router.post("/reports/z/close")
async def close_z_report(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin" and "reports_z" not in (user.get("permissions") or []):
        raise HTTPException(status_code=403, detail="Z Report permission required")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if await db.z_reports.find_one({"date": today}): raise HTTPException(status_code=400, detail="Z Report already closed for today")
    orders = await db.orders.find({"date": today}).to_list(10000)
    expenses = await db.expenses.find({"date": today}).to_list(500)
    refunds = await db.refunds.find({"date": today}).to_list(500)
    r = calc_report(orders, expenses, refunds)
    r.update({"date": today, "report_type": "Z", "closed_at": datetime.now(timezone.utc).isoformat(), "closed_by": user["_id"]})
    await db.z_reports.insert_one(r)
    response = {"message": "Z Report closed and archived", "report": {k: v for k, v in r.items() if k != "_id"}}

    # Auto-email if configured
    try:
        s = await _get_settings_doc()
        if s.get("auto_email_on_z_close") and s.get("smtp_host") and s.get("smtp_user") and s.get("smtp_password"):
            recipients = _filter_recipients(s.get("email_recipients", []), "Z")
            if recipients:
                rep = {**r}
                rep.pop("_id", None)
                rep["report_type"] = "Z"
                subject, plain, html = _format_report_email(rep, s)
                await asyncio.to_thread(_send_email_sync, s["smtp_host"], s["smtp_port"], s.get("smtp_user",""), s.get("smtp_password",""), bool(s.get("smtp_use_tls", True)), s.get("smtp_from") or s.get("smtp_user"), recipients, subject, plain, html)
                response["emailed_to"] = recipients
    except Exception as e:
        logger.error(f"Auto-email on Z close failed: {e}")
        response["email_error"] = str(e)[:200]

    return response

@api_router.get("/reports/history")
async def get_report_history(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None):
    user = await get_current_user(request)
    if user.get("role") != "admin" and "reports_z" not in (user.get("permissions") or []):
        raise HTTPException(status_code=403, detail="Z Report permission required")
    query = {}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    else:
        two_months_ago = (datetime.now(timezone.utc) - timedelta(days=60)).strftime("%Y-%m-%d")
        query["date"] = {"$gte": two_months_ago}
    reports = await db.z_reports.find(query, {"_id": 0}).sort("date", -1).to_list(100)
    return reports

@api_router.get("/reports/export/csv")
async def export_report_csv(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    orders = await db.orders.find({"date": today}, {"_id": 0}).to_list(10000)
    rows = []
    for o in orders:
        for item in o.get("items", []):
            rows.append({"date": o.get("date"), "time": o.get("created_at", ""), "item_name": item.get("name"), "quantity": item.get("quantity", 0), "unit_price": item.get("price", 0), "line_total": round(item.get("price", 0) * item.get("quantity", 0), 2), "payment_type": o.get("payment_type"), "discount": o.get("discount_amount", 0), "order_total": o.get("total", 0), "cashier": o.get("cashier_name", "")})
    return rows

@api_router.get("/reports/history/export")
async def export_history_csv(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin" and "reports_z" not in (user.get("permissions") or []):
        raise HTTPException(status_code=403, detail="Z Report permission required")
    two_months_ago = (datetime.now(timezone.utc) - timedelta(days=60)).strftime("%Y-%m-%d")
    return await db.z_reports.find({"date": {"$gte": two_months_ago}}, {"_id": 0}).sort("date", -1).to_list(100)

# --- Email Reports ---
def _format_report_email(report: dict, settings: dict) -> tuple:
    """Returns (subject, plain_body, html_body) for a report email."""
    rname = settings.get("restaurant_name", "RestoPOS")
    cur = settings.get("currency", "Rs")
    rtype = report.get("report_type", "X")
    date = report.get("date", "")

    def fmt(v):
        try: return f"{cur} {float(v):,.2f}"
        except: return f"{cur} 0.00"

    subject = f"[{rname}] {rtype}-Report — {date}"

    lines = [
        f"{rname}",
        f"{rtype}-Report for {date}",
        "=" * 50,
        "",
        f"Total Sales:        {fmt(report.get('total_sales', 0))}",
        f"  Cash:             {fmt(report.get('cash_sales', 0))}",
        f"  Card:             {fmt(report.get('credit_sales', 0))}",
        f"  FoodPanda 1:      {fmt(report.get('foodpanda1_sales', 0))}",
        f"  FoodPanda 2:      {fmt(report.get('foodpanda2_sales', 0))}",
        "",
        f"Total Orders:       {report.get('total_orders', 0)}",
        f"Items Sold:         {report.get('total_items_sold', 0)}",
        f"Expenses:           {fmt(report.get('total_expenses', 0))}",
        f"Refunds:            {fmt(report.get('total_refunds', 0))}",
        f"Net Revenue:        {fmt(report.get('net_revenue', 0))}",
        "",
        "Top Items:",
    ]
    for ti in (report.get("top_items") or [])[:10]:
        lines.append(f"  - {ti.get('name')} × {ti.get('quantity')}")
    lines += ["", "—", "Sent automatically by RestoPOS"]
    plain = "\n".join(lines)

    rows_html = "".join(
        f"<tr><td style='padding:4px 12px'>{ti.get('name')}</td><td style='padding:4px 12px;text-align:right'>{ti.get('quantity')}</td></tr>"
        for ti in (report.get("top_items") or [])[:10]
    )
    html = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;color:#1A1D1A;background:#F9F8F6;padding:24px">
<div style="max-width:600px;margin:auto;background:white;border:1px solid #E5E2DC;border-radius:12px;overflow:hidden">
  <div style="background:#1E3F20;color:white;padding:20px"><h2 style="margin:0">{rname}</h2><p style="margin:4px 0 0;opacity:0.85">{rtype}-Report — {date}</p></div>
  <div style="padding:24px">
    <h3 style="margin:0 0 12px;color:#1E3F20">Sales Summary</h3>
    <table style="width:100%;border-collapse:collapse;font-size:14px">
      <tr><td style="padding:6px 0">Total Sales</td><td style="text-align:right;font-weight:bold">{fmt(report.get('total_sales',0))}</td></tr>
      <tr style="color:#5C5F5C"><td style="padding:4px 0 4px 16px">Cash</td><td style="text-align:right">{fmt(report.get('cash_sales',0))}</td></tr>
      <tr style="color:#5C5F5C"><td style="padding:4px 0 4px 16px">Card</td><td style="text-align:right">{fmt(report.get('credit_sales',0))}</td></tr>
      <tr style="color:#5C5F5C"><td style="padding:4px 0 4px 16px">FoodPanda 1</td><td style="text-align:right">{fmt(report.get('foodpanda1_sales',0))}</td></tr>
      <tr style="color:#5C5F5C"><td style="padding:4px 0 4px 16px">FoodPanda 2</td><td style="text-align:right">{fmt(report.get('foodpanda2_sales',0))}</td></tr>
      <tr><td colspan="2" style="border-top:1px solid #E5E2DC;padding:6px 0"></td></tr>
      <tr><td style="padding:4px 0">Total Orders</td><td style="text-align:right">{report.get('total_orders',0)}</td></tr>
      <tr><td style="padding:4px 0">Items Sold</td><td style="text-align:right">{report.get('total_items_sold',0)}</td></tr>
      <tr><td style="padding:4px 0;color:#A63D31">Expenses</td><td style="text-align:right;color:#A63D31">{fmt(report.get('total_expenses',0))}</td></tr>
      <tr><td style="padding:4px 0;color:#A63D31">Refunds</td><td style="text-align:right;color:#A63D31">{fmt(report.get('total_refunds',0))}</td></tr>
      <tr><td style="padding:8px 0;font-weight:bold;color:#1E3F20">Net Revenue</td><td style="text-align:right;font-weight:bold;color:#1E3F20">{fmt(report.get('net_revenue',0))}</td></tr>
    </table>
    <h3 style="margin:24px 0 8px;color:#1E3F20">Top Items</h3>
    <table style="width:100%;border-collapse:collapse;font-size:14px;border:1px solid #E5E2DC">
      <thead style="background:#F9F8F6"><tr><th style="text-align:left;padding:6px 12px">Item</th><th style="text-align:right;padding:6px 12px">Qty</th></tr></thead>
      <tbody>{rows_html or '<tr><td colspan=2 style="padding:12px;color:#5C5F5C">No items</td></tr>'}</tbody>
    </table>
  </div>
  <div style="padding:12px 24px;background:#F9F8F6;border-top:1px solid #E5E2DC;font-size:12px;color:#5C5F5C">Sent automatically by RestoPOS</div>
</div></body></html>"""
    return subject, plain, html


def _send_email_sync(host, port, user, password, use_tls, sender, to_list, subject, plain, html):
    """Synchronous SMTP send. Raises on error."""
    msg = EmailMessage()
    msg["From"] = sender or user
    msg["To"] = ", ".join(to_list)
    msg["Subject"] = subject
    msg.set_content(plain)
    if html:
        msg.add_alternative(html, subtype="html")

    if int(port) == 465:
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(host, int(port), context=ctx, timeout=20) as s:
            if user and password:
                s.login(user, password)
            s.send_message(msg)
    else:
        with smtplib.SMTP(host, int(port), timeout=20) as s:
            s.ehlo()
            if use_tls:
                s.starttls(context=ssl.create_default_context())
                s.ehlo()
            if user and password:
                s.login(user, password)
            s.send_message(msg)


async def _get_settings_doc() -> dict:
    s = await db.settings.find_one({"key": "global"}, {"_id": 0}) or {}
    return {k: s.get(k, v) for k, v in DEFAULT_SETTINGS.items()}


def _validate_smtp(s: dict):
    if not s.get("smtp_host"): raise HTTPException(status_code=400, detail="SMTP host not configured. Go to Settings → Email.")
    if not s.get("smtp_port"): raise HTTPException(status_code=400, detail="SMTP port not configured.")
    if not (s.get("smtp_user") and s.get("smtp_password")):
        raise HTTPException(status_code=400, detail="SMTP username/password not configured.")


def _filter_recipients(recipients: list, report_type: str) -> list:
    """Pick recipients who opted-in for the given report type. report_type is 'X' or 'Z'."""
    out = []
    key = "receive_x" if report_type.upper() == "X" else "receive_z"
    for r in (recipients or []):
        if not r.get("email"): continue
        # Default to True if flag missing
        if r.get(key, True):
            out.append(r["email"])
    return out


@api_router.post("/email/test")
async def email_test(request: Request, body: Dict[str, Any] = None):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    _validate_smtp(s)
    body = body or {}
    to = body.get("to") or s.get("smtp_user")
    if not to: raise HTTPException(status_code=400, detail="No recipient address")
    subject = f"[{s.get('restaurant_name','RestoPOS')}] Test Email"
    plain = "If you received this, your SMTP configuration is working correctly."
    html = f"<p>If you received this, your <b>SMTP configuration</b> is working correctly.</p><p style='color:#5C5F5C'>From: {s.get('restaurant_name','RestoPOS')}</p>"
    try:
        await asyncio.to_thread(_send_email_sync, s["smtp_host"], s["smtp_port"], s.get("smtp_user",""), s.get("smtp_password",""), bool(s.get("smtp_use_tls", True)), s.get("smtp_from") or s.get("smtp_user"), [to], subject, plain, html)
    except Exception as e:
        logger.error(f"SMTP test failed: {e}")
        raise HTTPException(status_code=500, detail=f"SMTP error: {str(e)[:200]}")
    return {"message": f"Test email sent to {to}"}


@api_router.post("/email/send-report")
async def email_send_report(request: Request, body: Dict[str, Any]):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    _validate_smtp(s)

    report_type = (body.get("report_type") or "X").upper()
    report_id_date = body.get("date")  # optional: pick a specific archived Z-report
    extra_to = body.get("extra_recipients") or []

    # Get the report data
    if report_id_date:
        rep = await db.z_reports.find_one({"date": report_id_date}, {"_id": 0})
        if not rep: raise HTTPException(status_code=404, detail="Report not found")
        rep["report_type"] = "Z"
    else:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        orders = await db.orders.find({"date": today}).to_list(10000)
        expenses = await db.expenses.find({"date": today}).to_list(500)
        refunds = await db.refunds.find({"date": today}).to_list(500)
        rep = calc_report(orders, expenses, refunds)
        rep.update({"date": today, "report_type": report_type})

    # Build recipients list
    cfg_recipients = _filter_recipients(s.get("email_recipients", []), rep.get("report_type", "X"))
    recipients = list(dict.fromkeys([*cfg_recipients, *[e for e in extra_to if e]]))
    if not recipients:
        raise HTTPException(status_code=400, detail="No recipients configured. Add at least one in Settings → Email.")

    subject, plain, html = _format_report_email(rep, s)
    try:
        await asyncio.to_thread(_send_email_sync, s["smtp_host"], s["smtp_port"], s.get("smtp_user",""), s.get("smtp_password",""), bool(s.get("smtp_use_tls", True)), s.get("smtp_from") or s.get("smtp_user"), recipients, subject, plain, html)
    except Exception as e:
        logger.error(f"SMTP send failed: {e}")
        raise HTTPException(status_code=500, detail=f"SMTP error: {str(e)[:200]}")
    return {"message": f"Report sent to {len(recipients)} recipient(s)", "recipients": recipients}


# --- WhatsApp (proxy to local Node service) ---
def _format_report_whatsapp(report: dict, settings: dict) -> str:
    rname = settings.get("restaurant_name", "RestoPOS")
    cur = settings.get("currency", "Rs")
    rtype = report.get("report_type", "X")
    date = report.get("date", "")

    def fmt(v):
        try: return f"{cur} {float(v):,.2f}"
        except: return f"{cur} 0.00"

    lines = [
        f"*{rname}*",
        f"*{rtype}-Report — {date}*",
        "",
        f"💰 Total Sales: *{fmt(report.get('total_sales', 0))}*",
        f"   • Cash: {fmt(report.get('cash_sales', 0))}",
        f"   • Card: {fmt(report.get('credit_sales', 0))}",
        f"   • FoodPanda 1: {fmt(report.get('foodpanda1_sales', 0))}",
        f"   • FoodPanda 2: {fmt(report.get('foodpanda2_sales', 0))}",
        "",
        f"🛒 Orders: {report.get('total_orders', 0)}",
        f"📦 Items Sold: {report.get('total_items_sold', 0)}",
        f"💸 Expenses: {fmt(report.get('total_expenses', 0))}",
        f"↩️ Refunds: {fmt(report.get('total_refunds', 0))}",
        f"📈 *Net Revenue: {fmt(report.get('net_revenue', 0))}*",
    ]
    top = report.get("top_items") or []
    if top:
        lines += ["", "*Top Items:*"]
        for ti in top[:5]:
            lines.append(f"  • {ti.get('name')} × {ti.get('quantity')}")
    lines += ["", "_Sent automatically by RestoPOS_"]
    return "\n".join(lines)


def _filter_whatsapp_recipients(recipients: list, report_type: str) -> list:
    out = []
    key = "receive_x" if report_type.upper() == "X" else "receive_z"
    for r in (recipients or []):
        if not r.get("phone"): continue
        if r.get(key, True):
            out.append({"name": r.get("name", ""), "phone": r["phone"]})
    return out


async def _wa_get(path: str, settings: dict, timeout: float = 10.0):
    url = (settings.get("whatsapp_service_url") or "http://127.0.0.1:3030").rstrip("/")
    async with httpx.AsyncClient(timeout=timeout) as c:
        r = await c.get(f"{url}{path}")
        return r


async def _wa_post(path: str, payload: dict, settings: dict, timeout: float = 30.0):
    url = (settings.get("whatsapp_service_url") or "http://127.0.0.1:3030").rstrip("/")
    async with httpx.AsyncClient(timeout=timeout) as c:
        r = await c.post(f"{url}{path}", json=payload)
        return r


@api_router.get("/whatsapp/status")
async def whatsapp_status(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    try:
        r = await _wa_get("/status", s, timeout=5.0)
        return r.json()
    except Exception as e:
        return {"ready": False, "phone": None, "qr_available": False, "error": f"Service not reachable: {str(e)[:160]}"}


@api_router.get("/whatsapp/qr")
async def whatsapp_qr(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    try:
        r = await _wa_get("/qr", s, timeout=5.0)
        if r.status_code == 404: return {"qr": None, "ready": False, "message": "QR not yet generated"}
        return r.json()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"WhatsApp service not reachable: {str(e)[:160]}")


@api_router.post("/whatsapp/reset")
async def whatsapp_reset(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    try:
        r = await _wa_post("/reset", {}, s, timeout=15.0)
        return r.json()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"WhatsApp service not reachable: {str(e)[:160]}")


@api_router.post("/whatsapp/test")
async def whatsapp_test(request: Request, body: Dict[str, Any]):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    to = (body or {}).get("to")
    if not to: raise HTTPException(status_code=400, detail="Phone number required (e.g., +923004928411)")
    msg = (body or {}).get("message") or f"Test message from {s.get('restaurant_name','RestoPOS')}. If you got this, WhatsApp integration is working ✅"
    try:
        r = await _wa_post("/send", {"to": to, "message": msg}, s, timeout=20.0)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"WhatsApp service not reachable: {str(e)[:160]}")
    if r.status_code >= 400:
        try: detail = r.json().get("error", "Failed")
        except: detail = "Failed"
        raise HTTPException(status_code=r.status_code, detail=detail)
    return {"message": f"Test message sent to {to}"}


@api_router.post("/whatsapp/send-report")
async def whatsapp_send_report(request: Request, body: Dict[str, Any]):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()

    report_type = (body.get("report_type") or "X").upper()
    report_id_date = body.get("date")
    extra_to = body.get("extra_recipients") or []  # list of phone strings

    if report_id_date:
        rep = await db.z_reports.find_one({"date": report_id_date}, {"_id": 0})
        if not rep: raise HTTPException(status_code=404, detail="Report not found")
        rep["report_type"] = "Z"
    else:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        orders = await db.orders.find({"date": today}).to_list(10000)
        expenses = await db.expenses.find({"date": today}).to_list(500)
        refunds = await db.refunds.find({"date": today}).to_list(500)
        rep = calc_report(orders, expenses, refunds)
        rep.update({"date": today, "report_type": report_type})

    cfg_recipients = _filter_whatsapp_recipients(s.get("whatsapp_recipients", []), rep.get("report_type", "X"))
    phones = list(dict.fromkeys([*[r["phone"] for r in cfg_recipients], *[p for p in extra_to if p]]))
    if not phones:
        raise HTTPException(status_code=400, detail="No WhatsApp recipients configured. Add at least one in Settings → WhatsApp.")

    message = _format_report_whatsapp(rep, s)
    sent_to, failed = [], []
    for phone in phones:
        try:
            r = await _wa_post("/send", {"to": phone, "message": message}, s, timeout=30.0)
            if r.status_code < 400:
                sent_to.append(phone)
            else:
                try: err = r.json().get("error", f"HTTP {r.status_code}")
                except: err = f"HTTP {r.status_code}"
                failed.append({"phone": phone, "error": err})
        except Exception as e:
            failed.append({"phone": phone, "error": f"Service unreachable: {str(e)[:120]}"})

    if not sent_to and failed:
        raise HTTPException(status_code=502, detail={"message": "All sends failed", "failed": failed})
    return {"message": f"Sent to {len(sent_to)} recipient(s)", "sent_to": sent_to, "failed": failed}


# --- Daily Auto-Send Scheduler ---
scheduler: Optional[AsyncIOScheduler] = None


async def _build_report_for_scheduler(s: dict) -> dict:
    """Generate the report to send. Default is yesterday's data."""
    use_yesterday = (s.get("daily_report_type") or "yesterday").lower() == "yesterday"
    target = (datetime.now(timezone.utc) - timedelta(days=1)) if use_yesterday else datetime.now(timezone.utc)
    target_date = target.strftime("%Y-%m-%d")

    # Prefer archived Z report if it exists for that date
    z_archived = await db.z_reports.find_one({"date": target_date}, {"_id": 0})
    if z_archived:
        z_archived["report_type"] = "Z"
        return z_archived

    orders = await db.orders.find({"date": target_date}).to_list(10000)
    expenses = await db.expenses.find({"date": target_date}).to_list(500)
    refunds = await db.refunds.find({"date": target_date}).to_list(500)
    r = calc_report(orders, expenses, refunds)
    r.update({"date": target_date, "report_type": "Z" if use_yesterday else "X"})
    return r


async def _send_whatsapp_to_recipients(rep: dict, s: dict):
    cfg = _filter_whatsapp_recipients(s.get("whatsapp_recipients", []), rep.get("report_type", "Z"))
    if not cfg: return {"sent": 0, "failed": 0}
    msg = _format_report_whatsapp(rep, s)
    sent, failed = 0, 0
    for r in cfg:
        try:
            resp = await _wa_post("/send", {"to": r["phone"], "message": msg}, s, timeout=30.0)
            if resp.status_code < 400: sent += 1
            else: failed += 1
        except Exception:
            failed += 1
    return {"sent": sent, "failed": failed}


async def _send_email_to_recipients(rep: dict, s: dict):
    cfg = _filter_recipients(s.get("email_recipients", []), rep.get("report_type", "Z"))
    if not cfg: return {"sent": 0, "skipped": True}
    if not (s.get("smtp_host") and s.get("smtp_user") and s.get("smtp_password")):
        return {"sent": 0, "skipped": True, "reason": "SMTP not configured"}
    subject, plain, html = _format_report_email(rep, s)
    try:
        await asyncio.to_thread(_send_email_sync, s["smtp_host"], s["smtp_port"], s.get("smtp_user",""), s.get("smtp_password",""), bool(s.get("smtp_use_tls", True)), s.get("smtp_from") or s.get("smtp_user"), cfg, subject, plain, html)
        return {"sent": len(cfg)}
    except Exception as e:
        logger.error(f"Scheduled email send failed: {e}")
        return {"sent": 0, "error": str(e)[:200]}


async def daily_report_job():
    """Runs at the configured time. Sends yesterday's (or today's) report via email & WhatsApp if enabled."""
    try:
        s = await _get_settings_doc()
        if not (s.get("auto_email_daily") or s.get("auto_whatsapp_daily")):
            logger.info("Daily report job: both email & whatsapp auto-send are off; skipping")
            return
        rep = await _build_report_for_scheduler(s)
        results = {"email": None, "whatsapp": None}
        if s.get("auto_email_daily"):
            results["email"] = await _send_email_to_recipients(rep, s)
        if s.get("auto_whatsapp_daily"):
            results["whatsapp"] = await _send_whatsapp_to_recipients(rep, s)
        # Log a summary record
        await db.scheduled_runs.insert_one({
            "ran_at": datetime.now(timezone.utc).isoformat(),
            "report_date": rep.get("date"),
            "report_type": rep.get("report_type"),
            "results": results,
        })
        logger.info(f"Daily report job complete: {results}")
    except Exception as e:
        logger.exception(f"Daily report job error: {e}")


async def _reschedule_daily_job():
    global scheduler
    if scheduler is None: return
    s = await _get_settings_doc()
    time_str = s.get("daily_report_time") or "02:15"
    tz_name = s.get("daily_report_timezone") or "Asia/Karachi"
    try:
        hh, mm = [int(x) for x in time_str.split(":")[:2]]
        tz = pytz.timezone(tz_name)
    except Exception as e:
        logger.error(f"Invalid schedule time/tz '{time_str}' / '{tz_name}': {e}")
        return
    # Replace existing job
    try: scheduler.remove_job("daily_report")
    except Exception: pass
    scheduler.add_job(daily_report_job, CronTrigger(hour=hh, minute=mm, timezone=tz), id="daily_report", replace_existing=True, misfire_grace_time=3600)
    logger.info(f"Daily report scheduled for {hh:02d}:{mm:02d} {tz_name}")


@api_router.get("/schedule/status")
async def schedule_status(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    info = {"daily_report_time": s.get("daily_report_time"), "daily_report_timezone": s.get("daily_report_timezone"), "auto_email_daily": bool(s.get("auto_email_daily")), "auto_whatsapp_daily": bool(s.get("auto_whatsapp_daily")), "daily_report_type": s.get("daily_report_type"), "next_run": None}
    if scheduler is not None:
        try:
            j = scheduler.get_job("daily_report")
            if j and j.next_run_time:
                info["next_run"] = j.next_run_time.isoformat()
        except Exception: pass
    return info


@api_router.post("/schedule/run-now")
async def schedule_run_now(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    asyncio.create_task(daily_report_job())
    return {"message": "Daily report job triggered. Check logs."}


@api_router.get("/schedule/timezones")
async def list_timezones(request: Request):
    await get_current_user(request)
    # A curated short list — full pytz list is too large for a dropdown
    return [
        "Asia/Karachi", "Asia/Dubai", "Asia/Riyadh", "Asia/Kolkata", "Asia/Dhaka",
        "Asia/Singapore", "Asia/Hong_Kong", "Asia/Tokyo", "Asia/Seoul",
        "Europe/London", "Europe/Paris", "Europe/Berlin", "Europe/Istanbul",
        "Africa/Cairo", "Africa/Johannesburg",
        "America/New_York", "America/Chicago", "America/Los_Angeles", "America/Toronto",
        "America/Sao_Paulo",
        "Australia/Sydney", "Pacific/Auckland",
        "UTC",
    ]


# --- Cloudflare Tunnel ---
import re as _re
_TUNNEL_URL_RE = _re.compile(r"https://[a-z0-9-]+\.trycloudflare\.com")


def _resolve_tunnel_log_path(custom: str = "") -> str:
    if custom and os.path.exists(custom):
        return custom
    # default candidates relative to project root
    root = ROOT_DIR.parent
    for name in ("cloudflared.log", "tunnel.log", "windows-setup/cloudflared.log"):
        p = root / name
        if p.exists():
            return str(p)
    return ""


def _read_tunnel_url_from_log(path: str) -> Optional[str]:
    """Find the most recent trycloudflare.com URL in the log file."""
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            # Read tail (last 64 KB is plenty)
            try:
                f.seek(0, 2); size = f.tell(); f.seek(max(0, size - 65536))
            except Exception:
                pass
            content = f.read()
        matches = _TUNNEL_URL_RE.findall(content)
        return matches[-1] if matches else None
    except Exception:
        return None


async def _maybe_notify_tunnel_url_change(new_url: str):
    """Email and WhatsApp the new tunnel URL to recipients (if enabled)."""
    try:
        s = await _get_settings_doc()
        if not s.get("tunnel_notify_on_change"): return
        rname = s.get("restaurant_name", "RestoPOS")
        msg_text = f"RestoPOS is online.\n\nRemote access URL:\n{new_url}\n\nLogin: {s.get('restaurant_email','admin@restaurant.com')} / your password\n\n— {rname}"
        msg_html = f"""<div style="font-family:Arial,sans-serif"><h2 style="color:#1E3F20">{rname} is online</h2><p>You can now access your POS from anywhere using:</p><p style="margin:16px 0"><a href="{new_url}" style="background:#1E3F20;color:white;padding:12px 24px;border-radius:8px;text-decoration:none;font-weight:bold">{new_url}</a></p><p style="color:#5C5F5C;font-size:13px">Login with the same admin email/password you use locally. This URL changes every time the Pakistan PC restarts.</p></div>"""
        # Email
        if s.get("smtp_host") and s.get("smtp_user") and s.get("smtp_password"):
            email_to = [r["email"] for r in (s.get("email_recipients") or []) if r.get("email")]
            if email_to:
                try:
                    await asyncio.to_thread(_send_email_sync, s["smtp_host"], s["smtp_port"], s.get("smtp_user",""), s.get("smtp_password",""), bool(s.get("smtp_use_tls", True)), s.get("smtp_from") or s.get("smtp_user"), email_to, f"[{rname}] Remote access URL", msg_text, msg_html)
                    logger.info(f"Notified email recipients of new tunnel URL: {new_url}")
                except Exception as e: logger.error(f"Tunnel notify email failed: {e}")
        # WhatsApp
        wa_to = [r["phone"] for r in (s.get("whatsapp_recipients") or []) if r.get("phone")]
        if wa_to:
            for phone in wa_to:
                try:
                    await _wa_post("/send", {"to": phone, "message": msg_text}, s, timeout=20.0)
                except Exception: pass
            logger.info(f"Notified WhatsApp recipients of new tunnel URL: {new_url}")
    except Exception as e:
        logger.exception(f"Tunnel notify failed: {e}")


async def tunnel_watcher_loop():
    """Background task: periodically reads cloudflared log and updates tunnel URL in DB."""
    last_seen = None
    while True:
        try:
            s = await _get_settings_doc()
            log_path = _resolve_tunnel_log_path(s.get("tunnel_log_path") or "")
            if log_path:
                url = _read_tunnel_url_from_log(log_path)
                if url and url != last_seen:
                    await db.tunnel.update_one({"key": "current"}, {"$set": {"url": url, "updated_at": datetime.now(timezone.utc).isoformat(), "log_path": log_path}}, upsert=True)
                    # always notify on change (including first detection after restart)
                    await _maybe_notify_tunnel_url_change(url)
                    logger.info(f"Tunnel URL updated: {url}")
                    last_seen = url
        except Exception as e:
            logger.error(f"Tunnel watcher error: {e}")
        await asyncio.sleep(15)


@api_router.get("/tunnel/status")
async def get_tunnel_status(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    log_path = _resolve_tunnel_log_path(s.get("tunnel_log_path") or "")
    cur = await db.tunnel.find_one({"key": "current"}, {"_id": 0}) or {}
    return {
        "url": cur.get("url"),
        "updated_at": cur.get("updated_at"),
        "log_path": log_path or s.get("tunnel_log_path") or "",
        "log_exists": bool(log_path),
        "notify_on_change": bool(s.get("tunnel_notify_on_change", True)),
    }


@api_router.post("/tunnel/refresh")
async def refresh_tunnel(request: Request):
    """Force a re-read of the tunnel log."""
    user = await get_current_user(request)
    if user.get("role") != "admin": raise HTTPException(status_code=403, detail="Admin only")
    s = await _get_settings_doc()
    log_path = _resolve_tunnel_log_path(s.get("tunnel_log_path") or "")
    if not log_path: return {"url": None, "message": "No log file found yet"}
    url = _read_tunnel_url_from_log(log_path)
    if url:
        await db.tunnel.update_one({"key": "current"}, {"$set": {"url": url, "updated_at": datetime.now(timezone.utc).isoformat(), "log_path": log_path}}, upsert=True)
    return {"url": url, "log_path": log_path}


# --- Dashboard ---
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request, date: Optional[str] = None):
    await get_current_user(request)
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Check if day was closed
    z_closed = await db.z_reports.find_one({"date": target_date})
    is_closed = z_closed is not None
    orders = await db.orders.find({"date": target_date}).to_list(10000)
    expenses = await db.expenses.find({"date": target_date}).to_list(500)
    refunds = await db.refunds.find({"date": target_date}).to_list(500)
    r = calc_report(orders, expenses, refunds)
    low_stock = await db.menu_items.aggregate([{"$match": {"$expr": {"$lte": ["$stock", "$low_stock_threshold"]}}}]).to_list(500)
    r.update({"today": target_date, "is_closed": is_closed, "low_stock_count": len(low_stock), "total_menu_items": await db.menu_items.count_documents({}), "total_categories": await db.categories.count_documents({})})
    return r

@api_router.get("/dashboard/hourly-sales")
async def get_hourly_sales(request: Request, date: Optional[str] = None):
    await get_current_user(request)
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    orders = await db.orders.find({"date": target_date}).to_list(10000)
    hourly = {f"{h:02d}:00": {"hour": f"{h:02d}:00", "cash": 0.0, "credit": 0.0, "online": 0.0, "total": 0.0, "orders": 0} for h in range(24)}
    for o in orders:
        try:
            ts = datetime.fromisoformat(o["created_at"])
            label = f"{ts.hour:02d}:00"
            amt = o.get("total", 0)
            hourly[label]["total"] = round(hourly[label]["total"] + amt, 2)
            hourly[label]["orders"] += 1
            pt = o.get("payment_type", "cash")
            if pt == "cash": hourly[label]["cash"] = round(hourly[label]["cash"] + amt, 2)
            elif pt == "credit": hourly[label]["credit"] = round(hourly[label]["credit"] + amt, 2)
            else: hourly[label]["online"] = round(hourly[label]["online"] + amt, 2)
        except: pass
    return list(hourly.values())

# --- Data Management (Admin) ---
@api_router.get("/data/export")
async def export_all_data(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    data = {"_meta": {"exported_at": datetime.now(timezone.utc).isoformat(), "version": 2}}
    # Include users (for password) and settings (for SMTP/WhatsApp/Schedule config)
    for col_name in ["users", "settings", "categories", "menu_items", "vendors", "orders", "z_reports", "expenses", "vendor_transactions", "vendor_payments", "refunds"]:
        col = db[col_name]
        docs = await col.find({}).to_list(100000)
        data[col_name] = [
            {k: (str(v) if k == "_id" else v) for k, v in d.items()}
            for d in docs
        ]
    return data

@api_router.get("/data/stats")
async def get_data_stats(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    stats = {}
    for col_name in ["orders", "z_reports", "expenses", "vendor_transactions", "vendor_payments", "refunds"]:
        stats[col_name] = await db[col_name].count_documents({})
    return stats

class DataDeleteRequest(BaseModel):
    before_date: str
    collections: List[str]

@api_router.post("/data/delete")
async def delete_old_data(req: DataDeleteRequest, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    deleted = {}
    for col_name in req.collections:
        if col_name in ["orders", "z_reports", "expenses", "vendor_transactions", "vendor_payments"]:
            result = await db[col_name].delete_many({"date": {"$lt": req.before_date}})
            deleted[col_name] = result.deleted_count
    return {"message": "Data deleted", "deleted": deleted}

@api_router.post("/data/import")
async def import_data(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    body = await request.json()
    imported = {}
    # Replace mode: wipe target collections and reinsert with original IDs preserved
    # so cross-collection refs (menu_item.category_id, vendor_transaction.vendor_id) stay valid.
    allowed = ["users", "settings", "categories", "menu_items", "vendors", "orders", "z_reports", "expenses", "vendor_transactions", "vendor_payments", "refunds"]
    for col_name in allowed:
        if col_name in body and isinstance(body[col_name], list):
            await db[col_name].delete_many({})  # clear existing
            docs = []
            for d in body[col_name]:
                doc = {k: v for k, v in d.items() if k != "_meta"}
                # Convert string _id back to ObjectId to preserve cross-collection refs
                if "_id" in doc and isinstance(doc["_id"], str):
                    try: doc["_id"] = ObjectId(doc["_id"])
                    except Exception: doc.pop("_id", None)  # invalid → let mongo assign new
                docs.append(doc)
            if docs:
                try:
                    result = await db[col_name].insert_many(docs, ordered=False)
                    imported[col_name] = len(result.inserted_ids)
                except Exception as e:
                    logger.error(f"Import {col_name} failed: {e}")
                    imported[col_name] = f"failed: {str(e)[:100]}"
            else:
                imported[col_name] = 0
    return {"message": "Data imported (existing data was replaced)", "imported": imported}

# --- Cleanup & Seed ---
async def cleanup_old_data():
    cutoff = (datetime.now(timezone.utc) - timedelta(days=60)).strftime("%Y-%m-%d")
    await db.orders.delete_many({"date": {"$lt": cutoff}})
    await db.z_reports.delete_many({"date": {"$lt": cutoff}})
    await db.expenses.delete_many({"date": {"$lt": cutoff}})
    logger.info(f"Cleaned up data older than {cutoff}")

async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@restaurant.com").lower().strip()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({"email": admin_email, "password_hash": hash_password(admin_password), "name": "Admin", "role": "admin", "permissions": ADMIN_PERMISSIONS, "created_at": datetime.now(timezone.utc).isoformat()})
        logger.info(f"Admin created: {admin_email}")
    else:
        # IMPORTANT: do NOT reset the admin's password on every boot.
        # The previous behaviour (overwrite hash if env var doesn't match) silently
        # reverted the operator's UI-changed password to the env-var default after
        # every redeploy — the symptom being "I can't log in with my password".
        # We only resync from the env var when the operator explicitly opts in with
        # FORCE_ADMIN_PASSWORD_RESET=true (set it once, deploy, unset). That gives
        # an escape hatch for a forgotten password without breaking normal operation.
        if (os.environ.get("FORCE_ADMIN_PASSWORD_RESET", "").lower() in ("1", "true", "yes")
                and not verify_password(admin_password, existing["password_hash"])):
            await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
            logger.warning(f"Admin password RESET from env var (FORCE_ADMIN_PASSWORD_RESET=true): {admin_email}")
    # Ensure admin always has the FULL set of permissions (covers cases where
    # ALL_PERMISSIONS gained new entries — e.g. online_dashboard, online_orders, etc.).
    if existing:
        current = set(existing.get("permissions") or [])
        if current != set(ADMIN_PERMISSIONS):
            await db.users.update_one({"email": admin_email}, {"$set": {"permissions": ADMIN_PERMISSIONS}})
    # Seed default settings
    if not await db.settings.find_one({"key": "global"}):
        await db.settings.insert_one({"key": "global", **DEFAULT_SETTINGS})
    # Write test credentials (safe for both cloud and local).
    # We only write the env-var password here if we KNOW it's the current truth —
    # i.e. we just created the admin row OR we just did a FORCE reset. Otherwise the
    # admin may have changed their password via the UI and the file would mislead
    # the testing agent / fork agent into trying a stale credential.
    try:
        admin_doc_now = await db.users.find_one({"email": admin_email})
        password_is_truthful = bool(admin_doc_now) and verify_password(admin_password, admin_doc_now["password_hash"])
        memory_dir = ROOT_DIR.parent / "memory"
        memory_dir.mkdir(exist_ok=True)
        with open(memory_dir / "test_credentials.md", "w") as f:
            f.write(f"# Test Credentials\n\n## Admin\n- Email: {admin_email}\n")
            if password_is_truthful:
                f.write(f"- Password: {admin_password}\n")
            else:
                f.write("- Password: <changed via UI — env-var ADMIN_PASSWORD no longer matches. Set FORCE_ADMIN_PASSWORD_RESET=true and restart to reset.>\n")
            f.write("- Role: admin\n")
    except Exception:
        pass

@app.on_event("startup")
async def startup():
    # Non-blocking: schedule heavy DB init so uvicorn binds 0.0.0.0:8080 immediately.
    # Without this, every `await db.x` below blocks the ASGI lifespan; if Mongo is
    # slow/unreachable, Fly health checks fail before the socket is ever opened.
    asyncio.create_task(_startup_background())

async def _startup_background():
    try:
        await _do_startup()
    except Exception as e:
        logger.exception(f"Background startup failed (server still listening): {e}")

async def _do_startup():
    global scheduler
    await db.users.create_index("email", unique=True)
    # Performance indexes – critical as data grows.
    # Skip silently if they already exist or the collection is empty.
    try:
        await db.orders.create_index([("date", -1), ("created_at", -1)])
        await db.orders.create_index("payment_type")
        await db.orders.create_index([("created_at", -1)])
        await db.menu_items.create_index([("sort_order", 1), ("created_at", 1)])
        # `category_id` is used both as a filter in the POS (when a tab is
        # clicked) and as the join key in /inventory. Cheap to add, big win
        # once the menu has dozens of items.
        await db.menu_items.create_index("category_id")
        await db.categories.create_index([("sort_order", 1), ("created_at", 1)])
        await db.expenses.create_index([("date", -1)])
        await db.refunds.create_index([("date", -1)])
        await db.z_reports.create_index([("date", -1)])
    except Exception as e:
        logger.warning(f"Index creation skipped: {e}")
    # Load VAPID keys from MongoDB (primary persistent store).
    # Falls back to: env vars (already loaded at module level) → vapid_keys.json → auto-generate.
    # This means keys survive every redeploy with zero manual intervention.
    if not os.environ.get("VAPID_PUBLIC_KEY"):  # env vars take priority; skip DB load if set
        db_has_keys = await _load_vapid_keys_from_db()
        if not db_has_keys:
            # Nothing in DB yet — migrate from file if present, otherwise auto-generate.
            global VAPID_PUBLIC_KEY, VAPID_PRIVATE_KEY
            if not (VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY):
                try:
                    new_pub, new_priv = _generate_vapid_keypair_raw()
                    VAPID_PUBLIC_KEY = new_pub
                    VAPID_PRIVATE_KEY = new_priv
                    logger.info("VAPID keys auto-generated on first boot.")
                except Exception as e:
                    logger.error(f"VAPID key generation failed: {e}")
            # Save whatever we have (migrated from file or newly generated) into MongoDB.
            if VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY:
                await _save_vapid_keys_to_db(VAPID_PUBLIC_KEY, VAPID_PRIVATE_KEY)
                logger.info("VAPID keys saved to MongoDB — will survive all future redeploys.")
    await seed_admin()
    await cleanup_old_data()
    # IDOR mitigation backfill — every order needs a per-order share token so
    # the public /api/track endpoint can require it. Legacy orders without one
    # are unreachable to anonymous viewers (by design: blocks enumeration of
    # historical orders' PII), but signed-in owners can still access them via
    # "My Orders". Each token is 16 url-safe bytes (~128 bits of entropy).
    try:
        legacy_count = await db.online_orders.count_documents({"track_token": {"$exists": False}})
        if legacy_count:
            logger.info(f"Backfilling track_token on {legacy_count} legacy online_orders…")
            async for doc in db.online_orders.find({"track_token": {"$exists": False}}, {"_id": 1}):
                await db.online_orders.update_one(
                    {"_id": doc["_id"]},
                    {"$set": {"track_token": secrets.token_urlsafe(16)}},
                )
            logger.info(f"track_token backfill complete on {legacy_count} orders.")
    except Exception as e:
        logger.warning(f"track_token backfill skipped: {e}")
    # Start the scheduler
    try:
        scheduler = AsyncIOScheduler()
        scheduler.start()
        await _reschedule_daily_job()
        # Expire gateway orders whose payment never completed so they don't
        # linger in awaiting_payment forever (customer closed the gateway
        # page). Runs every 10 min; window = PAYMENT_ABANDON_MINUTES (def 30).
        scheduler.add_job(_expire_abandoned_gateway_orders, "interval", minutes=10,
                          id="expire_awaiting_payment", replace_existing=True)
        # SafePay safety net: confirm paid transactions server-side even when
        # the customer never returns from the hosted checkout (see the job's
        # docstring). Every 2 min, only recent initiated SafePay txns.
        scheduler.add_job(_reconcile_safepay_payments, "interval", minutes=2,
                          id="reconcile_safepay", replace_existing=True)
        # Closed-browser POS alert: while any order sits pending (not accepted),
        # re-push admin devices every 2 min. No-op if no admin subscribed.
        scheduler.add_job(_remind_admins_pending_orders, "interval", minutes=2,
                          id="remind_admins_pending", replace_existing=True)
    except Exception as e:
        logger.error(f"Scheduler start failed: {e}")
    # Start tunnel watcher in background
    try:
        asyncio.create_task(tunnel_watcher_loop())
    except Exception as e:
        logger.error(f"Tunnel watcher start failed: {e}")
    logger.info("Restaurant POS started")

@app.on_event("shutdown")
async def shutdown_db_client():
    global scheduler
    try:
        if scheduler:
            scheduler.shutdown(wait=False)
    except Exception: pass
    client.close()

# =============================================================================
# CUSTOMER-FACING ONLINE ORDERING ENDPOINTS (Karachi Naseeb Online Website)
# =============================================================================

# --- Customer Models ---
def _normalize_and_validate_phone(v: str, *, required: bool = False, field: str = "phone") -> str:
    """Strip non-digits and require at least 11 digits."""
    if v is None:
        v = ""
    digits = "".join(ch for ch in str(v) if ch.isdigit())
    if not digits:
        if required:
            raise ValueError(f"{field} is required")
        return ""
    if len(digits) < 11:
        raise ValueError(f"{field} must contain at least 11 digits")
    return digits


class CustomerRegisterRequest(BaseModel):
    email: str
    password: str
    name: str
    phone: Optional[str] = ""

    @field_validator("phone")
    @classmethod
    def _v_phone(cls, v):
        return _normalize_and_validate_phone(v, required=False, field="phone")

class CustomerLoginRequest(BaseModel):
    email: str
    password: str

class SelectedModifier(BaseModel):
    group_id: str
    option_id: str

class OnlineOrderItem(BaseModel):
    item_id: str
    name: str
    price: float
    quantity: int
    variation_name: Optional[str] = None  # e.g. "Half" / "Full" when the item has size variations
    selected_modifiers: Optional[List[SelectedModifier]] = None
    removed_ingredients: Optional[List[str]] = None  # ingredient ids or names to omit
    line_note: Optional[str] = None

class OnlineOrderCreate(BaseModel):
    items: List[OnlineOrderItem]
    total_price: float
    customer_name: str
    phone: str
    address: Optional[str] = ""  # required only for delivery orders (enforced in endpoint)
    notes: Optional[str] = ""
    payment_method: str = "cod"
    # NEW (additive): "delivery" (default, existing behaviour) | "pickup".
    # Pickup orders skip the delivery address + delivery-fee/geo logic. Absent =
    # "delivery" so every existing client keeps working unchanged.
    order_type: str = "delivery"
    coupon_code: Optional[str] = None
    delivery_lat: Optional[float] = None
    delivery_lng: Optional[float] = None
    reward_id: Optional[str] = None  # NEW: Loyalty reward redemption
    use_wallet: Optional[bool] = False  # apply store-credit wallet (signed-in only)

    @field_validator("phone")
    @classmethod
    def _v_phone(cls, v):
        return _normalize_and_validate_phone(v, required=True, field="phone")

    @field_validator("order_type")
    @classmethod
    def _v_order_type(cls, v):
        v = (v or "delivery").strip().lower()
        return v if v in ("delivery", "pickup") else "delivery"

class OnlineOrderStatusUpdate(BaseModel):
    status: str  # pending, accepted, preparing, ready, out_for_delivery, delivered, cancelled, rejected

class OrderRejectRequest(BaseModel):
    reason: str  # "out_of_stock" | "closed" | "other" | free text

class ModifyOrderItem(BaseModel):
    item_id: Optional[str] = None
    name: str
    price: float
    quantity: int

class OrderModifyRequest(BaseModel):
    items: List[ModifyOrderItem]
    notes: Optional[str] = None

class ReviewCreate(BaseModel):
    order_id: str
    rating: int
    comment: str

class AdminReviewReply(BaseModel):
    reply: str

class OfferCreate(BaseModel):
    title: str
    description: str
    discount_percent: Optional[float] = 0
    discount_amount: Optional[float] = 0
    coupon_code: Optional[str] = ""
    image_url: Optional[str] = ""
    active: bool = True
    min_order_amount: Optional[float] = 0
    one_time_per_customer: bool = False
    valid_until: Optional[str] = None
    # Distribution controls WHERE this offer is visible.
    # "website"           → appears on /offers page
    # "app"               → appears on app Offers screen
    # "voucher_code_only" → private; never listed publicly; redeemable by code only
    # Omitting (or passing null) defaults to ["website", "app"] for backward compat.
    distribution: Optional[List[str]] = None   # e.g. ["website","app"] or ["voucher_code_only"]
    usage_limit: Optional[int] = None          # null = unlimited
    assigned_customer_id: Optional[str] = None # null = any customer with the code

class OfferUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    discount_percent: Optional[float] = None
    discount_amount: Optional[float] = None
    coupon_code: Optional[str] = None
    image_url: Optional[str] = None
    active: Optional[bool] = None
    min_order_amount: Optional[float] = None
    one_time_per_customer: Optional[bool] = None
    valid_until: Optional[str] = None
    distribution: Optional[List[str]] = None
    usage_limit: Optional[int] = None
    assigned_customer_id: Optional[str] = None

class FAQCreate(BaseModel):
    """Public-facing frequently asked questions, admin-managed.
    `sort_order` decides display order on the public /faq page (lower = higher).
    `enabled` lets admin hide entries without deleting them."""
    question: str
    answer: str
    sort_order: int = 0
    enabled: bool = True

class FAQUpdate(BaseModel):
    question: Optional[str] = None
    answer: Optional[str] = None
    sort_order: Optional[int] = None
    enabled: Optional[bool] = None

class FAQReorder(BaseModel):
    """Atomic re-order helper — admin sends the full ordered list of FAQ ids
    after a drag-and-drop or arrow-shift. We persist sort_order = list index."""
    ids: list[str]


class DeliveryAreaCreate(BaseModel):
    """Admin-managed list of areas the restaurant delivers to (e.g. "Johar Town").
    Shown on the public /delivery page and used for local-SEO area coverage.
    `note` is an optional free-text detail (e.g. "Free delivery" or a fee/time)."""
    name: str
    note: str = ""
    sort_order: int = 0
    enabled: bool = True

class DeliveryAreaUpdate(BaseModel):
    name: Optional[str] = None
    note: Optional[str] = None
    sort_order: Optional[int] = None
    enabled: Optional[bool] = None

class DeliveryAreaReorder(BaseModel):
    ids: list[str]


class EventBookingCreate(BaseModel):
    name: str
    phone: str
    event_type: str
    guests: int
    event_date: str
    message: Optional[str] = ""
    email: Optional[str] = ""

# --- Customer JWT helper ---
def create_customer_token(cid: str, email: str) -> str:
    return jwt.encode({"sub": cid, "email": email, "role": "customer", "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "customer", "iid": INSTANCE_ID}, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_customer(request: Request):
    token = request.cookies.get("customer_token")
    if not token:
        ah = request.headers.get("Authorization", "")
        if ah.startswith("Bearer "):
            token = ah[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "customer":
            raise HTTPException(status_code=401, detail="Invalid token type")
        if payload.get("iid") != INSTANCE_ID:
            raise HTTPException(status_code=401, detail="Session expired - please log in again")
        cust = await db.customers.find_one({"_id": ObjectId(payload["sub"])})
        if not cust:
            raise HTTPException(status_code=401, detail="Customer not found")
        cust["_id"] = str(cust["_id"])
        cust.pop("password_hash", None)
        return cust
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_optional_customer(request: Request):
    try:
        return await get_current_customer(request)
    except HTTPException:
        return None

# --- Customer email verification (OTP) ---
# NEW (additive): email/password signups must confirm a 6-digit code emailed to
# them before they can place orders / earn/redeem diamonds. Social logins
# (Google/Facebook) are pre-verified by the provider. Existing accounts created
# before this feature have no `email_verified` field and are grandfathered as
# verified so nobody is locked out by the rollout.
OTP_TTL_MINUTES = 10
OTP_MAX_ATTEMPTS = 5
OTP_RESEND_COOLDOWN_SECONDS = 60

def _email_verification_enabled() -> bool:
    """Feature flag (env EMAIL_VERIFICATION_REQUIRED, default ON). Set to
    false/0/no/off on the server to instantly disable the whole OTP gate in
    production without a code rollback: new signups are auto-verified, no OTP is
    sent, and ordering is never blocked on verification."""
    return os.environ.get("EMAIL_VERIFICATION_REQUIRED", "true").strip().lower() \
        not in ("0", "false", "no", "off")

def _generate_otp() -> str:
    return f"{secrets.randbelow(1000000):06d}"

def _clip(v, n: int) -> str:
    """Defensive input cap: trim + hard-truncate free-text fields so oversized
    client payloads can't bloat documents, receipts, or WhatsApp/push messages.
    Truncates (rather than rejects) so older clients never start erroring."""
    return str(v or "").strip()[:n]

def _customer_email_verified(cust: dict) -> bool:
    """True if the account may transact. Missing field = grandfathered/verified."""
    return bool(cust.get("email_verified", True))

async def _send_customer_otp_email(email: str, name: str, otp: str, purpose: str = "verify") -> bool:
    """Email a verification / password-reset code. Returns False (never raises)
    if SMTP isn't configured or sending fails, so the calling flow never
    hard-fails on email."""
    try:
        s = await _get_settings_doc()
        if not (s.get("smtp_host") and s.get("smtp_port") and s.get("smtp_user") and s.get("smtp_password")):
            logger.warning("Customer OTP email skipped: SMTP not configured")
            return False
        rname = s.get("restaurant_name", "RestoPOS")
        what = "password reset code" if purpose == "reset" else "verification code"
        subject = f"[{rname}] Your {what}: {otp}"
        plain = f"Your {rname} {what} is {otp}. It expires in {OTP_TTL_MINUTES} minutes."
        html = (f"<p>Hi {name or 'there'},</p>"
                f"<p>Your <b>{rname}</b> {what} is:</p>"
                f"<p style='font-size:26px;font-weight:800;letter-spacing:4px'>{otp}</p>"
                f"<p style='color:#5C5F5C'>It expires in {OTP_TTL_MINUTES} minutes. "
                f"If you didn't request this, you can ignore this email.</p>")
        await asyncio.to_thread(
            _send_email_sync, s["smtp_host"], s["smtp_port"], s.get("smtp_user", ""),
            s.get("smtp_password", ""), bool(s.get("smtp_use_tls", True)),
            s.get("smtp_from") or s.get("smtp_user"), [email], subject, plain, html)
        return True
    except Exception as e:
        logger.error(f"Failed to send OTP email to {email}: {e}")
        return False

async def _issue_customer_otp(customer_oid, email: str, name: str) -> bool:
    """Generate + store a fresh hashed OTP for a customer and email it."""
    otp = _generate_otp()
    now = datetime.now(timezone.utc)
    await db.customers.update_one({"_id": customer_oid}, {"$set": {
        "email_otp_hash": hash_password(otp),
        "email_otp_expires_at": (now + timedelta(minutes=OTP_TTL_MINUTES)).isoformat(),
        "email_otp_attempts": 0,
        "email_otp_sent_at": now.isoformat(),
    }})
    return await _send_customer_otp_email(email, name, otp)

class VerifyEmailRequest(BaseModel):
    otp: str

# --- Customer Auth ---
@api_router.post("/customer/register")
async def customer_register(req: CustomerRegisterRequest, response: Response):
    email = req.email.lower().strip()
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Invalid email")
    if len(req.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    _validate_password_length(req.password)
    if await db.customers.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    verify_required = _email_verification_enabled()
    if len(email) > 100:
        raise HTTPException(status_code=400, detail="Email is too long")
    doc = {
        "email": email,
        "password_hash": hash_password(req.password),
        "name": _clip(req.name, 60),
        "phone": _clip(req.phone, 20),
        # When the gate is disabled, accounts are created already verified.
        "email_verified": not verify_required,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.customers.insert_one(doc)
    cid = str(result.inserted_id)
    # Fire off the verification code only when the gate is on. If SMTP is down the
    # account still exists and the user can hit /customer/resend-otp; we surface
    # otp_sent so the UI knows whether to show the OTP step.
    otp_sent = False
    if verify_required:
        otp_sent = await _issue_customer_otp(result.inserted_id, email, req.name.strip())
    token = create_customer_token(cid, email)
    response.set_cookie(key="customer_token", value=token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")
    return {"id": cid, "email": email, "name": req.name, "phone": req.phone, "token": token,
            "email_verified": doc["email_verified"], "otp_sent": otp_sent}

@api_router.post("/customer/login")
async def customer_login(req: CustomerLoginRequest, request: Request, response: Response):
    email = req.email.lower().strip()
    ip = _client_ip(request)
    _login_throttle_check(ip, email)
    _validate_password_length(req.password)
    cust = await db.customers.find_one({"email": email})
    if not cust or not verify_password(req.password, cust["password_hash"]):
        _login_record_failure(ip, email)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    _login_record_success(ip, email)
    cid = str(cust["_id"])
    token = create_customer_token(cid, email)
    response.set_cookie(key="customer_token", value=token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")
    return {"id": cid, "email": email, "name": cust.get("name", ""), "phone": cust.get("phone", ""), "token": token,
            "email_verified": _customer_email_verified(cust)}

@api_router.get("/customer/me")
async def customer_me(request: Request):
    cust = await get_current_customer(request)
    return {"id": cust["_id"], "email": cust["email"], "name": cust.get("name", ""),
            "phone": cust.get("phone", ""), "allergens": cust.get("allergens", []),
            "email_verified": _customer_email_verified(cust),
            # Store-credit wallet (credited by refunds; spendable on future orders).
            "wallet_balance": float(cust.get("wallet_balance", 0) or 0)}

@api_router.post("/customer/verify-email")
async def customer_verify_email(req: VerifyEmailRequest, request: Request):
    """Confirm the 6-digit OTP for the signed-in customer. On success the account
    is marked email_verified and the OTP is cleared."""
    cust = await get_current_customer(request)
    if _customer_email_verified(cust) and not cust.get("email_otp_hash"):
        return {"email_verified": True}
    otp_hash = cust.get("email_otp_hash")
    exp = _parse_iso_utc(cust.get("email_otp_expires_at"))
    attempts = int(cust.get("email_otp_attempts", 0) or 0)
    if not otp_hash or exp is None:
        raise HTTPException(status_code=400, detail="No verification code pending. Please request a new one.")
    if datetime.now(timezone.utc) > exp:
        raise HTTPException(status_code=400, detail="Verification code expired. Please request a new one.")
    if attempts >= OTP_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many attempts. Please request a new code.")
    if not verify_password((req.otp or "").strip(), otp_hash):
        await db.customers.update_one({"_id": ObjectId(cust["_id"])}, {"$inc": {"email_otp_attempts": 1}})
        raise HTTPException(status_code=400, detail="Incorrect code. Please try again.")
    await db.customers.update_one(
        {"_id": ObjectId(cust["_id"])},
        {"$set": {"email_verified": True},
         "$unset": {"email_otp_hash": "", "email_otp_expires_at": "",
                    "email_otp_attempts": "", "email_otp_sent_at": ""}})
    return {"email_verified": True}

@api_router.post("/customer/resend-otp")
async def customer_resend_otp(request: Request):
    """Re-send a verification code, rate-limited to one per cooldown window."""
    cust = await get_current_customer(request)
    if _customer_email_verified(cust):
        return {"email_verified": True, "message": "Already verified"}
    sent_at = _parse_iso_utc(cust.get("email_otp_sent_at"))
    if sent_at and (datetime.now(timezone.utc) - sent_at).total_seconds() < OTP_RESEND_COOLDOWN_SECONDS:
        raise HTTPException(status_code=429, detail="Please wait a moment before requesting another code.")
    sent = await _issue_customer_otp(ObjectId(cust["_id"]), cust["email"], cust.get("name", ""))
    if not sent:
        raise HTTPException(status_code=503, detail="Could not send the verification email. Please try again shortly.")
    return {"email_verified": False, "otp_sent": True}

# --- Forgot / reset password (email OTP; additive) ---
class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    email: str
    otp: str
    new_password: str

@api_router.post("/customer/forgot-password")
async def customer_forgot_password(req: ForgotPasswordRequest):
    """Email a 6-digit reset code. ALWAYS returns the same generic response so
    the endpoint can't be used to probe which emails have accounts."""
    email = _clip(req.email, 100).lower()
    generic = {"message": "If an account exists for that email, a reset code has been sent."}
    if not email or "@" not in email:
        return generic
    cust = await db.customers.find_one({"email": email})
    if not cust:
        return generic
    # Same cooldown as verification resend, to stop email-bombing an address.
    sent_at = _parse_iso_utc(cust.get("reset_otp_sent_at"))
    if sent_at and (datetime.now(timezone.utc) - sent_at).total_seconds() < OTP_RESEND_COOLDOWN_SECONDS:
        return generic
    otp = _generate_otp()
    now = datetime.now(timezone.utc)
    await db.customers.update_one({"_id": cust["_id"]}, {"$set": {
        "reset_otp_hash": hash_password(otp),
        "reset_otp_expires_at": (now + timedelta(minutes=OTP_TTL_MINUTES)).isoformat(),
        "reset_otp_attempts": 0,
        "reset_otp_sent_at": now.isoformat(),
    }})
    await _send_customer_otp_email(email, cust.get("name", ""), otp, purpose="reset")
    return generic

@api_router.post("/customer/reset-password")
async def customer_reset_password(req: ResetPasswordRequest):
    """Set a new password after proving control of the email via the reset OTP.
    Also marks the email verified (the code proves ownership)."""
    email = _clip(req.email, 100).lower()
    if len(req.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    _validate_password_length(req.new_password)
    cust = await db.customers.find_one({"email": email})
    # Deliberately vague on which part failed — no account enumeration.
    bad = HTTPException(status_code=400, detail="Invalid or expired reset code.")
    if not cust or not cust.get("reset_otp_hash"):
        raise bad
    exp = _parse_iso_utc(cust.get("reset_otp_expires_at"))
    if exp is None or datetime.now(timezone.utc) > exp:
        raise bad
    if int(cust.get("reset_otp_attempts", 0) or 0) >= OTP_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many attempts. Please request a new code.")
    if not verify_password((req.otp or "").strip(), cust["reset_otp_hash"]):
        await db.customers.update_one({"_id": cust["_id"]}, {"$inc": {"reset_otp_attempts": 1}})
        raise bad
    await db.customers.update_one(
        {"_id": cust["_id"]},
        {"$set": {"password_hash": hash_password(req.new_password), "email_verified": True},
         "$unset": {"reset_otp_hash": "", "reset_otp_expires_at": "",
                    "reset_otp_attempts": "", "reset_otp_sent_at": ""}})
    return {"message": "Password updated. You can now sign in."}

class AllergensUpdate(BaseModel):
    allergens: List[str] = []

@api_router.put("/customer/allergens")
async def update_customer_allergens(req: AllergensUpdate, request: Request):
    cust = await get_current_customer(request)
    cleaned = []
    for a in (req.allergens or [])[:40]:
        a = _clip(a, 30)
        if a and a not in cleaned:
            cleaned.append(a)
    await db.customers.update_one({"_id": cust["_id"]}, {"$set": {"allergens": cleaned}})
    return {"allergens": cleaned}

@api_router.delete("/customer/me")
async def delete_customer_account(request: Request):
    """Permanent account deletion (required by both App Store and Play policy).
    Removes the customer document and their personal data (coupons, loyalty
    ledger, push tokens, saved allergens). Orders are BUSINESS records (receipts,
    tax) so they are kept, but anonymised: customer_id is detached so they no
    longer link to any account."""
    cust = await get_current_customer(request)
    cid_str = str(cust["_id"])
    try:
        cid_obj = ObjectId(cid_str)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid account")
    # Detach (anonymise) orders — keep them as unlinked business records. Both
    # id representations are matched because historical writes stored either.
    await db.online_orders.update_many(
        {"customer_id": {"$in": [cid_str, cid_obj]}},
        {"$set": {"customer_id": None, "account_deleted": True}})
    # Personal data: gone.
    await db.personal_coupons.delete_many({"customer_id": {"$in": [cid_str, cid_obj]}})
    await db.loyalty_transactions.delete_many({"customer_id": {"$in": [cid_str, cid_obj]}})
    # Web push subscriptions registered by this customer's browsers. (FCM tokens
    # live inside the customer doc and vanish with it below; these live in their
    # own collection and previously survived deletion — the docstring's "push
    # tokens removed" promise now actually holds for both kinds.)
    await db.push_subscriptions.delete_many({"customer_id": {"$in": [cid_str, cid_obj]}})
    await db.reviews.update_many(
        {"customer_id": {"$in": [cid_str, cid_obj]}},
        {"$set": {"customer_id": None, "customer_name": "Deleted account"}})
    await db.customers.delete_one({"_id": cid_obj})
    logger.info(f"Customer account {cid_str} deleted at their request.")
    return {"message": "Your account and personal data have been deleted."}

@api_router.post("/customer/logout")
async def customer_logout(response: Response):
    response.delete_cookie("customer_token", path="/")
    return {"message": "Logged out"}

# --- Mobile: FCM push-token registration (native app; web still uses VAPID web-push) ---
class FcmTokenRequest(BaseModel):
    token: str
    platform: Optional[str] = "android"

@api_router.post("/customer/fcm-token")
async def register_fcm_token(req: FcmTokenRequest, request: Request):
    """Register a device's FCM token so we can push order-status updates to the native app.
    Idempotent — the same token is stored at most once per customer ($addToSet)."""
    cust = await get_current_customer(request)
    tok = (req.token or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Missing FCM token")
    await db.customers.update_one(
        {"_id": ObjectId(cust["_id"])},
        {"$addToSet": {"fcm_tokens": tok}},
    )
    return {"ok": True}

@api_router.delete("/customer/fcm-token")
async def unregister_fcm_token(req: FcmTokenRequest, request: Request):
    """Remove a device's FCM token (logout / uninstall / token refresh)."""
    cust = await get_current_customer(request)
    await db.customers.update_one(
        {"_id": ObjectId(cust["_id"])},
        {"$pull": {"fcm_tokens": (req.token or "").strip()}},
    )
    return {"ok": True}

# --- V2: Social Login (Google + Facebook) ---
# REMINDER: DO NOT HARDCODE THE URL, OR ADD ANY FALLBACKS OR REDIRECT URLS, THIS BREAKS THE AUTH
# Frontend uses window.location.origin for any OAuth redirect. Backend just verifies the
# credential/access token that the client SDK returns and issues our normal customer JWT.

class GoogleLoginRequest(BaseModel):
    credential: str  # Google ID token (JWT) from @react-oauth/google GoogleLogin onSuccess

class FacebookLoginRequest(BaseModel):
    access_token: str  # Short-lived access token from Facebook JS SDK FB.login()
    user_id: Optional[str] = None  # FB user id (optional; we re-verify via Graph API)

class AppleLoginRequest(BaseModel):
    identity_token: str  # Apple identity token (JWT) from Sign in with Apple
    name: Optional[str] = ""  # Full name — Apple only provides it on FIRST sign-in


async def _social_find_or_create_customer(email: str, name: str, provider: str, provider_user_id: str):
    """Find a customer by email (case-insensitive); create one if missing. Marks the
    customer with the social provider so we can show a 'Linked with Google/Facebook' badge later."""
    email = (email or "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="Social account did not return an email")
    existing = await db.customers.find_one({"email": email})
    now_iso = datetime.now(timezone.utc).isoformat()
    if existing:
        # Idempotently record the link. The provider verified this email, so also
        # mark the account email_verified (covers a pre-existing unverified
        # email/password account whose owner now signs in with Google/Facebook).
        update = {"$set": {f"{provider}_id": provider_user_id, "last_social_login_at": now_iso, "email_verified": True}}
        await db.customers.update_one({"_id": existing["_id"]}, update)
        existing["email_verified"] = True
        return existing
    doc = {
        "email": email,
        "name": (name or "").strip() or email.split("@")[0],
        "phone": "",
        # Social-only accounts have no password — set an unguessable hash so password login fails.
        "password_hash": hash_password(_uuid.uuid4().hex + INSTANCE_ID),
        f"{provider}_id": provider_user_id,
        "signup_provider": provider,
        "email_verified": True,  # provider already verified the email address
        "created_at": now_iso,
    }
    result = await db.customers.insert_one(doc)
    doc["_id"] = result.inserted_id
    return doc


def _serialize_customer_login_response(cust: dict, token: str):
    return {
        "id": str(cust["_id"]),
        "email": cust["email"],
        "name": cust.get("name", ""),
        "phone": cust.get("phone", ""),
        "token": token,
        "email_verified": _customer_email_verified(cust),
    }


@api_router.post("/customer/google")
async def customer_google_login(req: GoogleLoginRequest, response: Response):
    """Verify Google ID token and issue a customer JWT. Find-or-creates the customer
    by email so existing email/password users get linked automatically on first social login."""
    try:
        from google.oauth2 import id_token as g_id_token
        from google.auth.transport import requests as g_requests
    except Exception as e:
        logger.error(f"Google auth libs missing: {e}")
        raise HTTPException(status_code=500, detail="Google login is not available right now")
    google_client_id = os.environ.get("GOOGLE_CLIENT_ID")
    if not google_client_id:
        raise HTTPException(status_code=503, detail="Google login is not configured")
    # Support multiple client IDs (comma-separated) — e.g. web + Android OAuth clients
    client_ids = [c.strip() for c in google_client_id.split(",") if c.strip()]
    info = None
    for cid in client_ids:
        try:
            info = g_id_token.verify_oauth2_token(req.credential, g_requests.Request(), cid)
            break
        except Exception:
            continue
    if info is None:
        logger.warning("Google token verification failed for all configured client IDs")
        raise HTTPException(status_code=401, detail="Invalid Google login")
    email = info.get("email")
    if not email or not info.get("email_verified", True):
        raise HTTPException(status_code=400, detail="Google account email is not verified")
    sub = info.get("sub") or ""
    name = info.get("name") or ""
    cust = await _social_find_or_create_customer(email, name, "google", sub)
    cid = str(cust["_id"])
    token = create_customer_token(cid, email)
    response.set_cookie(key="customer_token", value=token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")
    return _serialize_customer_login_response(cust, token)


@api_router.post("/customer/apple")
async def customer_apple_login(req: AppleLoginRequest, response: Response):
    """Sign in with Apple (required by App Store review for apps with social
    login). Verifies the identity token's RS256 signature against Apple's
    published JWKS, checks issuer + audience (our bundle id), then find-or-
    creates the customer like Google/Facebook. Apple only sends the user's name
    on the FIRST authorization, so the client passes it along when present."""
    bundle_id = os.environ.get("APPLE_BUNDLE_ID")
    if not bundle_id:
        raise HTTPException(status_code=503, detail="Apple login is not configured")
    try:
        # Apple rotates its signing keys; fetch the JWKS and pick by `kid`.
        async with httpx.AsyncClient(timeout=10) as client:
            jwks = (await client.get("https://appleid.apple.com/auth/keys")).json()
        header = jwt.get_unverified_header(req.identity_token)
        key_data = next((k for k in jwks.get("keys", []) if k.get("kid") == header.get("kid")), None)
        if not key_data:
            raise ValueError("no matching Apple signing key")
        public_key = jwt.algorithms.RSAAlgorithm.from_jwk(json.dumps(key_data))
        info = jwt.decode(
            req.identity_token,
            public_key,
            algorithms=["RS256"],
            audience=bundle_id,
            issuer="https://appleid.apple.com",
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"Apple token verification failed: {e}")
        raise HTTPException(status_code=401, detail="Invalid Apple login")
    email = info.get("email")
    sub = info.get("sub") or ""
    if not sub:
        raise HTTPException(status_code=401, detail="Invalid Apple login")
    # Apple omits `email` after the first authorization — look the account up by
    # the stable Apple user id instead.
    if not email:
        existing = await db.customers.find_one({"apple_id": sub})
        if not existing:
            raise HTTPException(status_code=400, detail="Apple did not share an email. Remove this app from your Apple ID's 'Sign in with Apple' list and try again.")
        email = existing["email"]
    cust = await _social_find_or_create_customer(email, _clip(req.name, 60), "apple", sub)
    cid = str(cust["_id"])
    token = create_customer_token(cid, email)
    response.set_cookie(key="customer_token", value=token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")
    return _serialize_customer_login_response(cust, token)


@api_router.post("/customer/facebook")
async def customer_facebook_login(req: FacebookLoginRequest, response: Response):
    """Verify Facebook access token by calling Graph API and issue a customer JWT.
    We never trust client-side user_id; the Graph API call IS the verification."""
    app_id = os.environ.get("FACEBOOK_APP_ID")
    app_secret = os.environ.get("FACEBOOK_APP_SECRET")
    if not app_id or not app_secret:
        raise HTTPException(status_code=503, detail="Facebook login is not configured")
    async with httpx.AsyncClient(timeout=10) as client:
        # 1) debug_token to confirm the access token was issued to OUR app and is valid
        try:
            app_token = f"{app_id}|{app_secret}"
            dbg = await client.get(
                "https://graph.facebook.com/debug_token",
                params={"input_token": req.access_token, "access_token": app_token},
            )
            dbg.raise_for_status()
            dbg_json = dbg.json().get("data") or {}
        except Exception as e:
            logger.warning(f"Facebook debug_token failed: {e}")
            raise HTTPException(status_code=401, detail="Invalid Facebook login")
        if not dbg_json.get("is_valid"):
            raise HTTPException(status_code=401, detail="Facebook token is not valid")
        if str(dbg_json.get("app_id")) != str(app_id):
            raise HTTPException(status_code=401, detail="Facebook token was not issued to this app")
        fb_user_id = str(dbg_json.get("user_id") or "")
        # 2) Fetch the user's profile (name + email — email scope must be requested on the client)
        try:
            prof = await client.get(
                "https://graph.facebook.com/me",
                params={"fields": "id,name,email", "access_token": req.access_token},
            )
            prof.raise_for_status()
            profile = prof.json()
        except Exception as e:
            logger.warning(f"Facebook profile fetch failed: {e}")
            raise HTTPException(status_code=401, detail="Could not fetch Facebook profile")
    email = profile.get("email")
    if not email:
        # Facebook may withhold email if the user blocked the permission. Fall back to a
        # synthetic email so we can still create an account; the user can add a real email later.
        email = f"fb_{fb_user_id}@facebook.local"
    name = profile.get("name") or ""
    cust = await _social_find_or_create_customer(email, name, "facebook", fb_user_id)
    cid = str(cust["_id"])
    token = create_customer_token(cid, email)
    response.set_cookie(key="customer_token", value=token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")
    return _serialize_customer_login_response(cust, token)

@api_router.post("/customer/facebook/deletion")
async def facebook_data_deletion(request: Request):
    """Facebook Data Deletion Callback.
    Facebook POSTs a signed_request when a user requests deletion of their data
    from Facebook's privacy settings. We anonymise the matching customer account.
    Facebook requires we return a JSON body with a url and confirmation_code."""
    import hmac as _hmac, hashlib as _hashlib, base64 as _base64, json as _json
    app_secret = os.environ.get("FACEBOOK_APP_SECRET", "")
    try:
        form = await request.form()
        signed_request = form.get("signed_request", "")
        encoded_sig, payload = signed_request.split(".", 1)
        # Pad base64url
        def _b64d(s):
            s = s.replace("-", "+").replace("_", "/")
            s += "=" * (-len(s) % 4)
            return _base64.b64decode(s)
        sig = _b64d(encoded_sig)
        data = _json.loads(_b64d(payload))
        # Verify HMAC-SHA256
        expected = _hmac.new(app_secret.encode(), payload.encode(), _hashlib.sha256).digest()
        if not _hmac.compare_digest(sig, expected):
            raise HTTPException(status_code=400, detail="Invalid signed_request")
        fb_user_id = str(data.get("user_id", ""))
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"Facebook deletion callback parse error: {e}")
        raise HTTPException(status_code=400, detail="Bad signed_request")

    # Anonymise the customer account linked to this Facebook user ID
    if fb_user_id:
        await db.customers.update_many(
            {"social_provider": "facebook", "social_id": fb_user_id},
            {"$set": {
                "name": "Deleted User",
                "email": f"deleted_fb_{fb_user_id}@deleted.local",
                "phone": "",
                "deleted": True,
                "deleted_at": datetime.utcnow(),
            }},
        )

    confirmation_code = f"knb_fb_del_{fb_user_id}"
    status_url = f"https://www.karachinaseebbiryani.com/facebook-deletion?code={confirmation_code}"
    return {"url": status_url, "confirmation_code": confirmation_code}


# --- END Social Login ---

# --- Public Menu ---
class UpsellRequest(BaseModel):
    item_ids: List[str] = []
    limit: int = 4


@api_router.post("/menu/upsell")
async def menu_upsell(req: UpsellRequest):
    """Return up-to N suggestions based on what's in the customer's cart.
    Priority: 1) explicit related_item_ids of cart items, 2) bestsellers, 3) popular, 4) fallback.
    """
    limit = max(1, min(int(req.limit or 4), 8))
    excluded = set(filter(None, req.item_ids or []))
    related_ids = []
    if excluded:
        try:
            cart_oids = [ObjectId(x) for x in excluded if len(x) == 24]
        except Exception:
            cart_oids = []
        if cart_oids:
            cart_docs = await db.menu_items.find({"_id": {"$in": cart_oids}}, {"related_item_ids": 1}).to_list(50)
            for d in cart_docs:
                for rid in (d.get("related_item_ids") or []):
                    if rid not in excluded and rid not in related_ids:
                        related_ids.append(rid)

    items = await db.menu_items.find(
        {"stock": {"$gt": 0}},
        {"name": 1, "price": 1, "image_url": 1, "is_bestseller": 1, "is_popular": 1,
         "stock": 1, "discount_type": 1, "discount_value": 1, "variations": 1},
    ).sort([("created_at", -1)]).to_list(200)
    by_id = {str(i["_id"]): i for i in items}

    ranked = []
    seen = set(excluded)
    for rid in related_ids:
        if rid in seen or rid not in by_id:
            continue
        ranked.append(by_id[rid]); seen.add(rid)
        if len(ranked) >= limit: break
    if len(ranked) < limit:
        for it in items:
            sid = str(it["_id"])
            if sid in seen: continue
            if it.get("is_bestseller"):
                ranked.append(it); seen.add(sid)
                if len(ranked) >= limit: break
    if len(ranked) < limit:
        for it in items:
            sid = str(it["_id"])
            if sid in seen: continue
            if it.get("is_popular"):
                ranked.append(it); seen.add(sid)
                if len(ranked) >= limit: break
    if len(ranked) < limit:
        for it in items:
            sid = str(it["_id"])
            if sid in seen: continue
            ranked.append(it); seen.add(sid)
            if len(ranked) >= limit: break

    def _serialize(i):
        base = float(i.get("price", 0) or 0)
        d_type = i.get("discount_type")
        d_val = float(i.get("discount_value", 0) or 0)
        sale = base
        if d_type == "percentage" and d_val > 0:
            sale = round(max(0, base * (1 - d_val / 100.0)), 2)
        elif d_type == "fixed" and d_val > 0:
            sale = round(max(0, base - d_val), 2)
        return {
            "id": str(i["_id"]),
            "name": i.get("name", ""),
            "price": sale,
            "original_price": base if sale < base else None,
            "discount_percent": int(round((1 - sale / base) * 100)) if (base and sale < base) else 0,
            "image_url": i.get("image_url", ""),
            "is_bestseller": bool(i.get("is_bestseller")),
            "is_popular": bool(i.get("is_popular")),
            "variations": i.get("variations", []),
        }

    return {"items": [_serialize(i) for i in ranked]}


@api_router.get("/menu")
async def get_public_menu(request: Request, response: Response):
    cached = _cache_get("menu")
    inm = request.headers.get("if-none-match", "")
    if cached:
        if inm and inm == cached["etag"]:
            return Response(status_code=304, headers={"ETag": cached["etag"], "Cache-Control": "public, max-age=30"})
        response.headers["ETag"] = cached["etag"]
        response.headers["Cache-Control"] = "public, max-age=30"
        return cached["value"]
    cats = await db.categories.find({}).sort([("sort_order", 1), ("created_at", 1)]).to_list(200)
    items = await db.menu_items.find({}).sort([("sort_order", 1), ("created_at", 1)]).to_list(500)
    cat_list = [{"id": str(c["_id"]), "name": c["name"], "color": c.get("color")} for c in cats]
    item_list = []
    for i in items:
        base_price = float(i.get("price", 0) or 0)
        d_type = i.get("discount_type")
        d_val = float(i.get("discount_value", 0) or 0)
        sale_price = base_price
        if d_type == "percentage" and d_val > 0:
            sale_price = round(max(0, base_price * (1 - d_val / 100.0)), 2)
        elif d_type == "fixed" and d_val > 0:
            sale_price = round(max(0, base_price - d_val), 2)
        # Apply the item-level discount to each variation's price as well so the
        # Half/Medium/Full prices reflect the discount. Without this, the "9% OFF"
        # badge on a variations-item is a lie — the customer pays the un-discounted
        # variation price. We attach `original_price` to each variation too so the
        # frontend can strike through it like the main price.
        raw_variations = i.get("variations", []) or []
        variations_out = []
        for v in raw_variations:
            v_base = float(v.get("price", 0) or 0)
            v_sale = v_base
            if d_type == "percentage" and d_val > 0:
                v_sale = round(max(0, v_base * (1 - d_val / 100.0)), 2)
            elif d_type == "fixed" and d_val > 0:
                v_sale = round(max(0, v_base - d_val), 2)
            v_out = {"name": v.get("name", ""), "price": v_sale}
            if v_sale < v_base:
                v_out["original_price"] = v_base
            variations_out.append(v_out)
        item_list.append({
            "id": str(i["_id"]),
            "name": i["name"],
            "price": sale_price,
            "original_price": base_price if sale_price < base_price else None,
            "discount_type": d_type if sale_price < base_price else None,
            "discount_value": d_val if sale_price < base_price else 0,
            "discount_percent": int(round((1 - sale_price / base_price) * 100)) if (base_price and sale_price < base_price) else 0,
            "category_id": i["category_id"],
            "stock": i.get("stock", 0),
            "image_url": i.get("image_url", ""),
            "image_type": i.get("image_type", "url"),
            "description": i.get("description", ""),
            "is_popular": i.get("is_popular", False),
            "is_bestseller": i.get("is_bestseller", False),
            "variations": variations_out,
            "variations_active": i.get("variations_active", True),
            "modifier_groups": i.get("modifier_groups", []),
            "ingredients": i.get("ingredients", []),
        })
    out = {"categories": cat_list, "items": item_list}
    entry = _cache_set("menu", out)
    response.headers["ETag"] = entry["etag"]
    response.headers["Cache-Control"] = "public, max-age=30"
    return out

@api_router.get("/menu/{item_id}")
async def get_menu_item(item_id: str):
    try:
        i = await db.menu_items.find_one({"_id": ObjectId(item_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Item not found")
    if not i:
        raise HTTPException(status_code=404, detail="Item not found")
    return {
        "id": str(i["_id"]),
        "name": i["name"],
        "price": i["price"],
        "category_id": i["category_id"],
        "stock": i.get("stock", 0),
        "image_url": i.get("image_url", ""),
        "description": i.get("description", ""),
    }

# --- Online Orders ---
def _serialize_online_order(o: dict) -> dict:
    o = dict(o)
    o["id"] = str(o.pop("_id"))
    o["receipt_no"] = o["id"][-6:].upper()
    return o

def _discounted_price(base, d_type, d_val) -> float:
    """Apply an item-level admin discount to a base price.

    MUST mirror the sale-price logic in GET /menu (percentage / fixed) so the price a
    customer sees on the menu is exactly the price charged at checkout. Returns the
    rounded sale price and never goes negative. Falls back to the base price on bad input.
    """
    try:
        base = float(base or 0)
        d_val = float(d_val or 0)
    except (TypeError, ValueError):
        return round(float(base or 0), 2)
    if d_type == "percentage" and d_val > 0:
        return round(max(0, base * (1 - d_val / 100.0)), 2)
    if d_type == "fixed" and d_val > 0:
        return round(max(0, base - d_val), 2)
    return round(base, 2)

def _validate_and_price_modifiers(db_item, line):
    """Authoritatively price a line's modifier selections + validate group rules
    and ingredient removals against the DB item. Returns
    (add_price, resolved_modifiers, resolved_removals) or raises HTTPException.
    Prices ALWAYS come from the DB — never the client."""
    groups = db_item.get("modifier_groups", []) or []
    groups_by_id = {g.get("id"): g for g in groups}
    item_name = db_item.get("name", "item")
    picks = {}
    resolved = []
    add = 0.0
    for sm in (getattr(line, "selected_modifiers", None) or []):
        gid = sm.group_id
        oid = sm.option_id
        g = groups_by_id.get(gid)
        if not g or not g.get("active", True):
            raise HTTPException(status_code=400, detail=f"Invalid option for {item_name}.")
        opt = next((o for o in (g.get("options") or []) if o.get("id") == oid), None)
        if not opt:
            raise HTTPException(status_code=400, detail=f"Invalid choice for {g.get('name', 'option')}.")
        price = float(opt.get("price", 0) or 0)
        add += price
        picks.setdefault(gid, []).append(oid)
        resolved.append({"group": g.get("name", ""), "name": opt.get("name", ""), "price": price})
    # enforce each group's rules
    for g in groups:
        if not g.get("active", True):
            continue
        n = len(picks.get(g.get("id"), []))
        gname = g.get("name", "option")
        minsel = int(g.get("min_select", 0) or 0)
        maxsel = int(g.get("max_select", 0) or 0)
        required = bool(g.get("required", False))
        if g.get("type") == "single":
            if n > 1:
                raise HTTPException(status_code=400, detail=f"Choose only one for {gname}.")
            if (required or minsel >= 1) and n < 1:
                raise HTTPException(status_code=400, detail=f"Please choose {gname}.")
        else:
            need = max(minsel, 1) if required else minsel
            if need and n < need:
                raise HTTPException(status_code=400, detail=f"Choose at least {need} for {gname}.")
            if maxsel and n > maxsel:
                raise HTTPException(status_code=400, detail=f"Choose at most {maxsel} for {gname}.")
    # ingredient removals
    ings = db_item.get("ingredients", []) or []
    ing_by_id = {i.get("id"): i for i in ings}
    ing_by_name = {str(i.get("name", "")).strip().lower(): i for i in ings}
    resolved_removals = []
    for r in (getattr(line, "removed_ingredients", None) or []):
        key = str(r).strip()
        ing = ing_by_id.get(key) or ing_by_name.get(key.lower())
        if not ing or not ing.get("removable", True):
            raise HTTPException(status_code=400, detail=f"Cannot remove '{key}' from {item_name}.")
        resolved_removals.append(str(ing.get("name", "")).strip())
    return round(add, 2), resolved, resolved_removals

@api_router.post("/online-orders")
async def create_online_order(order: OnlineOrderCreate, request: Request):
    cust = await get_optional_customer(request)
    # Business hours check — block orders when restaurant is closed.
    s_for_hours = await get_online_settings_doc()
    if s_for_hours.get("business_hours_enabled", True):
        bh = compute_business_hours_status(s_for_hours)
        if not bh.get("is_open"):
            raise HTTPException(status_code=400, detail="Restaurant is currently closed. Please order during business hours.")

    # Email verification gate: a signed-in customer with an unverified email
    # cannot place orders (blocks fake-email signups from transacting). Guests
    # and grandfathered/social accounts are unaffected. Bypassed entirely when
    # the EMAIL_VERIFICATION_REQUIRED flag is off.
    if _email_verification_enabled() and cust and not _customer_email_verified(cust):
        raise HTTPException(status_code=403, detail="Please verify your email before placing an order. We've sent a code to your inbox.")

    # Order type (additive, default "delivery" = existing behaviour). Delivery
    # requires an address; pickup does not and skips all delivery-fee/geo logic.
    order_type = order.order_type if order.order_type in ("delivery", "pickup") else "delivery"
    address_clean = (order.address or "").strip()
    if order_type == "delivery" and len(address_clean) < 6:
        raise HTTPException(status_code=400, detail="Delivery address is required for delivery orders.")

    # ===== SERVER-SIDE AUTHORITATIVE PRICING =====
    # SECURITY (payment manipulation fix): Every previous calculation below relied on
    # `order.total_price` and `order.items[*].price`, both of which come straight from
    # the client JSON. An attacker could send negative quantities, negative prices,
    # arbitrary low prices, or a manipulated total to pay Rs 1 for any cart.
    # We now ignore client-supplied price/total fields entirely and rebuild them from
    # the menu_items collection — the only source of truth for what a thing costs.
    if not order.items or len(order.items) == 0:
        raise HTTPException(status_code=400, detail="Order must contain at least one item.")
    if len(order.items) > 50:
        raise HTTPException(status_code=400, detail="Too many line items in this order.")
    validated_items: list[dict] = []
    server_subtotal: float = 0.0
    # Batch the menu lookups so a 20-item cart doesn't fan out 20 round-trips.
    requested_ids: list[ObjectId] = []
    for line in order.items:
        try:
            requested_ids.append(ObjectId(str(line.item_id)))
        except Exception:
            raise HTTPException(status_code=400, detail=f"Invalid menu item id: {line.item_id}")
    menu_docs = {}
    async for mi in db.menu_items.find({"_id": {"$in": requested_ids}}):
        menu_docs[str(mi["_id"])] = mi
    for line in order.items:
        # Quantity sanity — must be a positive integer within a sane range. Catches
        # negative quantities (the original bug: -1 made the line total negative and
        # the order rang up at Rs 0) plus zero, floats, and absurd large numbers.
        try:
            qty = int(line.quantity)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f"Invalid quantity for {getattr(line, 'name', 'item')}.")
        if qty < 1 or qty > 100:
            raise HTTPException(status_code=400, detail=f"Quantity must be between 1 and 100 (got {qty}).")
        db_item = menu_docs.get(str(line.item_id))
        if not db_item:
            raise HTTPException(status_code=400, detail=f"Menu item not found or unavailable.")
        # Optional: respect an `active` / `is_available` flag if the menu doc has one,
        # so admins can deactivate an item and have it instantly refuse new orders.
        if db_item.get("active") is False or db_item.get("is_available") is False:
            raise HTTPException(status_code=400, detail=f"'{db_item.get('name', 'Item')}' is no longer available.")
        # Server-side price. Negative DB price = data corruption, not customer's
        # fault — refuse with a 500 so it surfaces in logs / alerts.
        # Variation handling: if the client picked a size (e.g. "Half"), resolve it
        # against the item's server-side variation list so BOTH the price and the
        # display name come from the DB — never trust the client's price. This keeps
        # the anti-manipulation guarantee while preserving the chosen variation.
        # Item-level admin discount (percentage / fixed). Applied the SAME way GET /menu
        # applies it, so the discounted price the customer saw on the menu is exactly what
        # they're charged here — for both the base item and its variations.
        d_type = db_item.get("discount_type")
        d_val = db_item.get("discount_value", 0)
        base_name = db_item.get("name", "")
        line_name = base_name
        variation_name = (getattr(line, "variation_name", None) or "").strip() or None
        if variation_name:
            db_variations = db_item.get("variations") or []
            match = next((v for v in db_variations if str(v.get("name", "")).strip().lower() == variation_name.lower()), None)
            if not match:
                raise HTTPException(status_code=400, detail=f"'{variation_name}' is not a valid option for {base_name}.")
            # Normalise to the DB's exact casing/spelling for the stored name.
            variation_name = str(match.get("name", "")).strip()
            server_price = _discounted_price(match.get("price", 0), d_type, d_val)
            line_name = f"{base_name} ({variation_name})"
        else:
            server_price = _discounted_price(db_item.get("price", 0), d_type, d_val)
        if server_price < 0:
            logger.error(f"Negative price on menu_items {db_item['_id']}: {server_price}")
            raise HTTPException(status_code=500, detail="Pricing error — please contact support.")
        # Modifiers (add-ons / required choices) + ingredient removals — priced and
        # validated server-side so the client can never inflate/deflate the total.
        mod_add, resolved_mods, resolved_removals = _validate_and_price_modifiers(db_item, line)
        unit_price = round(server_price + mod_add, 2)
        line_note = _clip(getattr(line, "line_note", None), 140) or None
        server_subtotal += unit_price * qty
        validated_items.append({
            "item_id": str(db_item["_id"]),
            "name": line_name,
            "variation_name": variation_name,
            "price": unit_price,          # effective unit price incl. modifiers
            "base_price": server_price,   # item/variation price before modifiers
            "quantity": qty,
            "modifiers": resolved_mods,           # [{group, name, price}]
            "removed_ingredients": resolved_removals,  # [name]
            "line_note": line_note,
        })
    # Round to 2 decimals to avoid floating point drift propagating into discounts.
    server_subtotal = round(server_subtotal, 2)

    # Coupon validation
    discount_amount = 0.0
    coupon_used = None
    personal_coupon_id = None
    offer_id_used = None   # tracked for atomic usage increment after order saved
    if order.coupon_code:
        code_normalized = order.coupon_code.upper().strip()
        if not cust:
            raise HTTPException(status_code=401, detail="Please sign in to use a coupon. Offers and discounts are linked to your account.")
        # 1) Personal customer coupons take priority over public offers.
        personal = await db.personal_coupons.find_one({"code": code_normalized})
        if personal:
            if personal.get("used"):
                raise HTTPException(status_code=400, detail="This voucher has already been fully redeemed.")
            try:
                expires = datetime.fromisoformat(personal.get("expires_at", "").replace("Z", "+00:00"))
            except Exception:
                expires = None
            if expires and expires < datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="This voucher has expired.")
            if not cust or str(cust["_id"]) != str(personal.get("customer_id")):
                raise HTTPException(status_code=400, detail="This voucher is not valid for your account.")
            if personal.get("discount_percent"):
                discount_amount = round(server_subtotal * float(personal["discount_percent"]) / 100, 2)
            else:
                discount_amount = float(personal.get("discount_amount", 0))
            coupon_used = code_normalized
            personal_coupon_id = personal["_id"]
        else:
            offer = await db.offers.find_one({"coupon_code": code_normalized, "active": True})
            if not offer:
                raise HTTPException(status_code=400, detail="This voucher code is invalid.")
            if _offer_expired(offer):
                raise HTTPException(status_code=400, detail="This voucher has expired.")
            # Usage limit check (non-atomic here; atomic increment happens after order saved)
            usage_limit = offer.get("usage_limit")
            usage_count = int(offer.get("usage_count", 0) or 0)
            if usage_limit is not None and usage_count >= usage_limit:
                raise HTTPException(status_code=400, detail="This voucher has already been fully redeemed.")
            # Assigned-customer check
            assigned = offer.get("assigned_customer_id")
            if assigned:
                if not cust or str(cust["_id"]) != str(assigned):
                    raise HTTPException(status_code=400, detail="This voucher is not valid for your account.")
            # Minimum order amount
            min_amount = float(offer.get("min_order_amount", 0) or 0)
            if min_amount > 0 and server_subtotal < min_amount:
                raise HTTPException(status_code=400, detail=f"Minimum order of Rs. {int(min_amount)} is required to use this voucher.")
            # One-time-per-customer guard
            if offer.get("one_time_per_customer"):
                used_filter = {"coupon_code": offer["coupon_code"]}
                if cust:
                    used_filter["customer_id"] = str(cust["_id"])
                else:
                    used_filter["phone"] = order.phone
                already_used = await db.online_orders.find_one(used_filter)
                if already_used:
                    raise HTTPException(status_code=400, detail=f"This voucher can only be used once per customer.")
            if offer.get("discount_percent"):
                discount_amount = round(server_subtotal * float(offer["discount_percent"]) / 100, 2)
            elif offer.get("discount_amount"):
                discount_amount = float(offer["discount_amount"])
            coupon_used = offer["coupon_code"]
            offer_id_used = offer["_id"]
    # Delivery fee calculation (server-side, ignores any frontend value).
    # Pickup orders never incur a delivery fee and skip the service-area check.
    delivery_fee = 0.0
    distance_km = None
    if order_type == "delivery" and order.delivery_lat is not None and order.delivery_lng is not None:
        s = await get_online_settings_doc()
        distance_km = haversine_km(s["restaurant_lat"], s["restaurant_lng"], order.delivery_lat, order.delivery_lng)
        quote = calculate_delivery_fee(distance_km, s, subtotal=server_subtotal - float(discount_amount or 0))
        if not quote["in_range"]:
            raise HTTPException(status_code=400, detail=f"Delivery address is outside our {quote['max_radius_km']} km service area.")
        delivery_fee = float(quote["fee"])
    final_total = max(0, server_subtotal - discount_amount) + delivery_fee
    
    # --- LOYALTY SYSTEM: Reward Redemption & Diamond Earning ---
    loyalty_settings = await db.loyalty_settings.find_one({"key": "loyalty"}) or {}
    loyalty_enabled = loyalty_settings.get("enabled", True)
    earning_rate = float(loyalty_settings.get("earning_rate", 10.0))  # Diamonds per Rs
    min_order_for_points = float(loyalty_settings.get("min_order_for_points", 0.0))
    
    reward_applied = None
    reward_discount = 0.0
    diamonds_spent = 0
    diamonds_earned = 0
    
    # Apply reward if customer selected one
    if order.reward_id and cust and loyalty_enabled:
        try:
            reward = await db.loyalty_rewards.find_one({"_id": ObjectId(order.reward_id), "is_active": True})
            if reward:
                current_balance = cust.get("diamond_balance", 0)
                required_diamonds = reward["cost_diamonds"]
                
                if current_balance >= required_diamonds:
                    # Apply reward
                    reward_type = reward["reward_type"]
                    reward_value = reward["reward_value"]

                    # V2 Reward Stacking Rule: customers cannot combine a Coupon Discount AND a
                    # Diamond Discount reward on the same order. Free-item rewards CAN stack with
                    # a coupon discount. (See requirement #9.)
                    if reward_type in ("discount_percent", "discount_fixed") and coupon_used:
                        raise HTTPException(status_code=400, detail="A coupon discount cannot be combined with a Diamond discount reward. Please remove one to continue.")

                    if reward_type == "discount_percent":
                        reward_discount = round(final_total * float(reward_value) / 100, 2)
                    elif reward_type == "discount_fixed":
                        reward_discount = min(float(reward_value), final_total)
                    elif reward_type == "free_item":
                        # Resolve the free menu item and append it to the order so it shows up
                        # on both the customer's order summary AND the restaurant's order
                        # ticket. Price is 0 (it's free) and the name is tagged so the kitchen
                        # immediately spots it as a Diamond freebie (no charge collected).
                        try:
                            free_item_doc = await db.menu_items.find_one({"_id": ObjectId(str(reward_value))})
                        except Exception:
                            free_item_doc = None
                        if free_item_doc:
                            # Append directly to the SERVER-VALIDATED items list — the client
                            # `order.items` list is no longer used for persistence (it was the
                            # vehicle for the payment-manipulation attack). The free item's
                            # price is 0 and is set here on the server, so it's trusted.
                            validated_items.append({
                                "item_id": str(free_item_doc["_id"]),
                                "name": f"{free_item_doc.get('name', 'Free Item')} (FREE — Diamond Reward)",
                                "price": 0.0,
                                "quantity": 1,
                            })
                        else:
                            logger.warning(f"free_item reward {reward.get('_id')} references missing menu item {reward_value}")
                    
                    # Update final total
                    final_total = max(0, final_total - reward_discount)
                    diamonds_spent = required_diamonds
                    
                    # Deduct diamonds from customer (atomically). cust['_id'] may be str — coerce to ObjectId.
                    try:
                        cust_oid_for_redeem = cust["_id"] if isinstance(cust["_id"], ObjectId) else ObjectId(str(cust["_id"]))
                    except Exception:
                        cust_oid_for_redeem = None
                    if cust_oid_for_redeem is not None:
                        await db.customers.update_one(
                            {"_id": cust_oid_for_redeem},
                            {
                                "$inc": {
                                    "diamond_balance": -diamonds_spent,
                                    "lifetime_diamonds_spent": diamonds_spent
                                }
                            }
                        )
                    
                    reward_applied = {
                        "reward_id": str(reward["_id"]),
                        "title": reward["title"],
                        "diamonds_spent": diamonds_spent,
                        "discount_amount": reward_discount,
                    }
                    
                    # Increment reward redemption count
                    await db.loyalty_rewards.update_one(
                        {"_id": ObjectId(order.reward_id)},
                        {"$inc": {"total_redemptions": 1}}
                    )
                else:
                    logger.warning(f"Customer {cust['_id']} tried to redeem reward but insufficient balance")
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Reward redemption failed: {e}")
    
    # Calculate diamonds that WILL BE earned on delivery (don't award yet)
    if cust and loyalty_enabled and final_total >= min_order_for_points:
        diamonds_earned = int(final_total * earning_rate / 100)  # e.g., 10 Diamonds per Rs 100
    
    # --- END LOYALTY LOGIC ---
    
    # Determine initial payment_status
    pmethod = (order.payment_method or "cod").lower()
    if pmethod in ("cod", "pay_at_restaurant"):
        payment_status = "pending"  # collected later
    else:
        payment_status = "pending"  # awaiting card/bank flow
    # Online-gateway orders must be PAID before the restaurant sees them:
    # they start as "awaiting_payment" (excluded from the staff pending
    # queue/alert/printer) and are flipped to "pending" by the payment
    # confirmation path (_release_online_order). COD / pay-at-restaurant /
    # manual bank transfer keep the existing immediate flow.
    initial_status = "awaiting_payment" if pmethod in GATEWAY_PAYMENT_METHODS else "pending"

    # --- WALLET REDEMPTION (store credit from refunds) ---
    # Signed-in customers can apply their wallet balance to this order. The
    # deduction is ATOMIC (filtered $inc) so a double-submit or stale balance
    # can never spend more than they have — on any race the wallet is simply
    # not applied and the customer pays the full amount by their chosen method.
    # Covers-everything → order is instantly PAID (no gateway trip, straight to
    # the staff queue). Covers-part → total_price becomes the REMAINDER, which
    # flows through the normal payment path (gateway sessions already charge
    # order.total_price). Wallet money was already real revenue when the
    # original order was paid, so reduced totals keep reports conservative.
    wallet_applied = 0.0
    if getattr(order, "use_wallet", False) and cust:
        try:
            balance = float(cust.get("wallet_balance", 0) or 0)
            applied = round(min(balance, final_total), 2)
            if applied > 0:
                res = await db.customers.update_one(
                    {"_id": cust["_id"], "wallet_balance": {"$gte": applied}},
                    {"$inc": {"wallet_balance": -applied}})
                if res.modified_count:
                    wallet_applied = applied
                    final_total = round(final_total - applied, 2)
        except Exception as e:
            logger.warning(f"wallet redemption failed (order proceeds without): {e}")
    if wallet_applied > 0 and final_total <= 0:
        payment_status = "paid"
        initial_status = "pending"  # fully covered — straight to the staff queue

    now = datetime.now(timezone.utc)
    # Per-order unguessable share token. The /api/track endpoint requires this on
    # every unauthenticated request to close the IDOR — without it, anyone who
    # guesses or enumerates an order id (Mongo ObjectIds embed a sequential
    # counter so they are very predictable) could read masked PII (first name,
    # last-4-digit phone, suburb prefix, items, total). 16 url-safe bytes ≈ 128
    # bits of entropy — practically un-bruteforceable. We embed this in every
    # tracking URL we hand out (WhatsApp confirmation, status updates, etc.).
    track_token = secrets.token_urlsafe(16)
    doc = {
        # SECURITY: persist the SERVER-VALIDATED items + subtotal, never the
        # client-supplied ones. This is what payment manipulation tried to corrupt.
        "items": validated_items,
        "subtotal": server_subtotal,
        "discount_amount": discount_amount,
        "delivery_fee": delivery_fee,
        "distance_km": round(distance_km, 2) if distance_km is not None else None,
        "order_type": order_type,
        "delivery_lat": order.delivery_lat if order_type == "delivery" else None,
        "delivery_lng": order.delivery_lng if order_type == "delivery" else None,
        "total_price": final_total,
        "customer_id": cust["_id"] if cust else None,
        "customer_name": _clip(order.customer_name, 60),
        "phone": order.phone,
        "address": _clip(address_clean, 200) if order_type == "delivery" else "",
        "notes": _clip(order.notes, 300),
        "payment_method": pmethod,
        "payment_status": payment_status,
        "coupon_code": coupon_used,
        "reward_applied": reward_applied,  # NEW: Loyalty reward info
        "diamonds_earned": diamonds_earned,  # NEW: Diamonds earned from this order
        "status": initial_status,
        "wallet_applied": wallet_applied,  # store credit used on this order
        "printed": False,
        "track_token": track_token,
        "created_at": now.isoformat(),
        "date": now.strftime("%Y-%m-%d"),
    }
    result = await db.online_orders.insert_one(doc)
    doc["_id"] = result.inserted_id
    order_id = str(result.inserted_id)

    # Ledger entry for the wallet spend (auditable next to refund credits).
    if wallet_applied > 0:
        try:
            await db.wallet_transactions.insert_one({
                "customer_id": str(cust["_id"]), "type": "spend", "amount": -wallet_applied,
                "order_id": order_id, "note": f"Applied to order #{order_id[-6:].upper()}",
                "created_at": now.isoformat()})
        except Exception as e:
            logger.warning(f"wallet spend ledger failed: {e}")

    # Mark personal coupon as used so it can't be redeemed twice.
    if personal_coupon_id:
        try:
            await db.personal_coupons.update_one(
                {"_id": personal_coupon_id},
                {"$set": {"used": True, "used_on_order_id": order_id, "used_at": now.isoformat()}},
            )
        except Exception as e:
            logger.error(f"Failed to mark personal coupon {personal_coupon_id} as used: {e}")

    # Atomically increment usage_count on public/private offer vouchers.
    # Uses $inc with a conditional guard so two simultaneous last-redemption
    # attempts cannot both succeed: find_one_and_update returns None if the
    # usage_limit was already reached by a concurrent request.
    if offer_id_used:
        try:
            await db.offers.update_one(
                {
                    "_id": offer_id_used,
                    "$or": [
                        {"usage_limit": None},
                        {"usage_limit": {"$exists": False}},
                        {"$expr": {"$lt": ["$usage_count", "$usage_limit"]}},
                    ],
                },
                {"$inc": {"usage_count": 1}},
            )
        except Exception as e:
            logger.error(f"Failed to increment offer usage_count for {offer_id_used}: {e}")
    
    # Log loyalty transaction for spending only (earning happens at delivery)
    if cust and diamonds_spent > 0 and reward_applied:
        await db.loyalty_transactions.insert_one({
            "customer_id": str(cust["_id"]),
            "transaction_type": "spent",
            "diamonds": -diamonds_spent,
            "balance_after": cust.get("diamond_balance", 0) - diamonds_spent,
            "order_id": order_id,
            "reward_id": reward_applied["reward_id"],
            "notes": f"Redeemed: {reward_applied['title']}",
            "created_at": now.isoformat(),
        })
    
    serialized = _serialize_online_order(doc)
    # WhatsApp confirmation (fire-and-forget)
    try:
        tracking_url = _origin_tracking_url(request, serialized['id'], doc.get('track_token'))
        msg = _format_order_confirmation(doc, tracking_url)
        asyncio.create_task(send_whatsapp(order.phone, msg))
    except Exception as e:
        logger.warning(f"WhatsApp confirmation failed: {e}")
    # Admin push alert (fire-and-forget) — only when the order enters the staff
    # queue immediately (COD etc.). Gateway orders start awaiting_payment and
    # alert from _release_online_order once the payment confirms instead.
    if initial_status == "pending":
        try:
            asyncio.create_task(_notify_admins_new_order(doc))
        except Exception as e:
            logger.warning(f"Admin new-order push failed: {e}")
    return serialized

@api_router.get("/online-orders/me")
async def get_my_orders(request: Request):
    cust = await get_current_customer(request)
    orders = await db.online_orders.find({"customer_id": cust["_id"]}).sort("created_at", -1).to_list(200)
    return [_serialize_online_order(o) for o in orders]


@api_router.get("/personal-coupons/me")
async def get_my_personal_coupons(request: Request):
    """List the signed-in customer's active (unused + unexpired) personal coupons.
    Used to surface the second-order bonus and any future per-customer perks on
    Profile / Rewards screens and to auto-apply at checkout."""
    cust = await get_current_customer(request)
    now_iso = datetime.now(timezone.utc).isoformat()
    coupons = await db.personal_coupons.find({
        "customer_id": str(cust["_id"]),
        "used": False,
        "expires_at": {"$gt": now_iso},
    }).sort("created_at", -1).to_list(50)
    return [{
        "id": str(c["_id"]),
        "code": c["code"],
        "discount_amount": float(c.get("discount_amount", 0)),
        "discount_percent": float(c.get("discount_percent", 0)),
        "source": c.get("source", ""),
        "expires_at": c.get("expires_at"),
        "created_at": c.get("created_at"),
    } for c in coupons]

@api_router.get("/online-orders")
async def list_online_orders(request: Request, status: Optional[str] = None):
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to view online orders.")
    query = {}
    if status and status != "all":
        query["status"] = status
    else:
        # Held gateway orders (payment not completed yet) stay out of the
        # restaurant's queue — they appear only once payment releases them.
        query["status"] = {"$ne": "awaiting_payment"}
    orders = await db.online_orders.find(query).sort("created_at", -1).to_list(1000)
    return [_serialize_online_order(o) for o in orders]

@api_router.get("/online-orders/pending-print")
async def list_pending_print_orders(request: Request):
    """Endpoint for POS thermal printer agent to poll new unprinted orders."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to view online orders.")
    orders = await db.online_orders.find({"printed": False, "status": {"$nin": ["cancelled", "awaiting_payment"]}}).sort("created_at", 1).to_list(100)
    return [_serialize_online_order(o) for o in orders]

@api_router.put("/online-orders/{order_id}/printed")
async def mark_order_printed(order_id: str, request: Request):
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to update online orders.")
    await db.online_orders.update_one({"_id": ObjectId(order_id)}, {"$set": {"printed": True, "printed_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Marked as printed"}

@api_router.put("/online-orders/{order_id}/status")
async def update_online_order_status(order_id: str, body: OnlineOrderStatusUpdate, request: Request):
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to update online orders.")
    valid = {"pending", "accepted", "preparing", "ready", "out_for_delivery", "delivered", "cancelled", "rejected"}
    if body.status not in valid:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid}")
    
    # Get order before update
    order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    old_status = order.get("status")
    
    update_fields = {"status": body.status, "updated_at": datetime.now(timezone.utc).isoformat()}
    if body.status == "accepted":
        update_fields["accepted_at"] = datetime.now(timezone.utc).isoformat()

    # Auto-mint a rider token the moment the order leaves the kitchen. The token is what
    # makes the no-login /rider/{order_id}?t=... link safe — only someone with this token
    # can mark the order delivered.
    if body.status == "out_for_delivery" and not order.get("rider_token"):
        update_fields["rider_token"] = secrets.token_urlsafe(16)
    
    res = await db.online_orders.update_one({"_id": ObjectId(order_id)}, {"$set": update_fields})

    # Push notification to the customer's devices for every meaningful status flip.
    # Done before the Diamond award block so the customer's phone buzzes immediately,
    # even if the diamond credit logic below takes a beat.
    if body.status != old_status:
        try:
            refreshed_for_push = await db.online_orders.find_one({"_id": ObjectId(order_id)})
            if refreshed_for_push:
                await _notify_customer_order_status(refreshed_for_push, body.status)
        except Exception as e:
            logger.warning(f"order-status push notify failed for {order_id}: {e}")
    
    # Award Diamonds when order is delivered (only once)
    if body.status == "delivered" and old_status != "delivered":
        diamonds_to_award = order.get("diamonds_earned", 0)
        customer_id = order.get("customer_id")
        
        if diamonds_to_award > 0 and customer_id:
            try:
                # customer_id may be stored as str (legacy) or ObjectId. db.customers._id is always ObjectId
                # — coerce defensively so the credit always lands.
                if isinstance(customer_id, ObjectId):
                    cust_oid = customer_id
                else:
                    try:
                        cust_oid = ObjectId(str(customer_id))
                    except Exception:
                        cust_oid = None
                if cust_oid is None:
                    logger.warning(f"Skipping diamond credit: invalid customer_id on order {order_id}: {customer_id!r}")
                else:
                    # Update customer balance atomically
                    result = await db.customers.update_one(
                        {"_id": cust_oid},
                        {
                            "$inc": {
                                "diamond_balance": diamonds_to_award,
                                "lifetime_diamonds_earned": diamonds_to_award
                            }
                        }
                    )

                    if result.modified_count > 0:
                        # Get updated balance
                        customer = await db.customers.find_one({"_id": cust_oid})
                        new_balance = customer.get("diamond_balance", 0) if customer else 0

                        # Log transaction
                        await db.loyalty_transactions.insert_one({
                            "customer_id": str(cust_oid),
                            "transaction_type": "earned",
                            "diamonds": diamonds_to_award,
                            "balance_after": new_balance,
                            "order_id": order_id,
                            "notes": f"Earned from delivered order #{order_id[-6:].upper()}",
                            "created_at": datetime.now(timezone.utc).isoformat(),
                        })

                        logger.info(f"Awarded {diamonds_to_award} Diamonds to customer {cust_oid} for order {order_id}")
                    else:
                        logger.warning(f"Diamond credit found 0 matching customer for order {order_id} (customer_id={customer_id!r})")
            except Exception as e:
                logger.error(f"Failed to award Diamonds for order {order_id}: {e}")

        # === Second-order bonus trigger ===
        # If this delivery makes the customer's 1st-EVER delivered order, mint them a
        # personal one-time coupon worth Rs. 50 off for their 2nd order. Single-use,
        # 30-day expiry, auto-applied at checkout when present. Tracked in its own
        # collection so it never collides with the public offers table.
        if customer_id:
            try:
                cust_id_str = str(customer_id)
                delivered_count = await db.online_orders.count_documents({"customer_id": customer_id, "status": "delivered"})
                # delivered_count INCLUDES this order (we just set status=delivered above).
                already_has = await db.personal_coupons.find_one({"customer_id": cust_id_str, "source": "second_order_bonus"})
                if delivered_count == 1 and not already_has:
                    code = f"WELCOME2-{secrets.token_hex(3).upper()}"
                    while await db.personal_coupons.find_one({"code": code}):
                        code = f"WELCOME2-{secrets.token_hex(3).upper()}"
                    expires_at = datetime.now(timezone.utc) + timedelta(days=30)
                    await db.personal_coupons.insert_one({
                        "code": code,
                        "customer_id": cust_id_str,
                        "discount_amount": 50,
                        "discount_percent": 0,
                        "source": "second_order_bonus",
                        "issued_for_order_id": order_id,
                        "used": False,
                        "used_on_order_id": None,
                        "expires_at": expires_at.isoformat(),
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    })
                    logger.info(f"Issued second-order bonus coupon {code} to customer {cust_id_str}")
            except Exception as e:
                logger.error(f"Failed to issue second-order bonus for order {order_id}: {e}")
    
    # WhatsApp status update
    try:
        if body.status in {"accepted", "preparing", "out_for_delivery", "delivered", "cancelled", "rejected"}:
            tracking_url = _origin_tracking_url(request, order_id, order.get("track_token"))
            msg = _format_status_update(order, body.status, tracking_url)
            asyncio.create_task(send_whatsapp(order.get("phone", ""), msg))
    except Exception as e:
        logger.warning(f"WhatsApp status update failed: {e}")
    
    return {"message": "Status updated", "status": body.status}

# --- Smart Order Alert: Accept / Reject / Modify endpoints (extension) ---
@api_router.get("/online-orders/pending-count")
async def online_orders_pending_count(request: Request):
    """Lightweight polling endpoint: returns count of orders awaiting staff action.
    Used by POS UI to start/stop the ringing alert sound."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to view online orders.")
    count = await db.online_orders.count_documents({"status": "pending"})
    # Latest pending order id (so frontend can detect new arrivals between polls)
    latest = await db.online_orders.find({"status": "pending"}, {"_id": 1, "created_at": 1}).sort("created_at", -1).limit(1).to_list(1)
    latest_id = str(latest[0]["_id"]) if latest else None
    latest_at = latest[0]["created_at"] if latest else None
    return {"pending_count": count, "latest_id": latest_id, "latest_at": latest_at}


def _origin_tracking_url(request: Request, order_id: str, track_token: Optional[str] = None) -> str:
    """Build the public /track URL for an order. Always include the per-order
    `track_token` query parameter — the /api/track endpoint REQUIRES it for
    unauthenticated viewers (IDOR mitigation). Authenticated owners + admins
    don't need the token, but we still include it so they can share the link
    with riders / family members who aren't signed in."""
    origin = request.headers.get("origin") or ""
    if origin.endswith("/"):
        origin = origin[:-1]
    base = f"{origin}/track/{order_id}" if origin else f"/track/{order_id}"
    return f"{base}?t={track_token}" if track_token else base


@api_router.post("/online-orders/{order_id}/accept")
async def accept_online_order(order_id: str, request: Request):
    """Staff accepts a pending order. Stops ringing on the POS, notifies customer via WhatsApp."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to accept orders.")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") in {"rejected", "cancelled"}:
        raise HTTPException(status_code=400, detail=f"Cannot accept a {order.get('status')} order")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {"$set": {"status": "accepted", "accepted_at": now_iso, "accepted_by": user.get("name", ""), "updated_at": now_iso}},
    )
    refreshed = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    try:
        msg = _format_status_update(refreshed, "accepted", _origin_tracking_url(request, order_id, refreshed.get("track_token")))
        asyncio.create_task(send_whatsapp(refreshed.get("phone", ""), msg))
    except Exception as e:
        logger.warning(f"WhatsApp accept notify failed: {e}")
    return _serialize_online_order(refreshed)


@api_router.post("/online-orders/{order_id}/reject")
async def reject_online_order(order_id: str, body: OrderRejectRequest, request: Request):
    """Staff rejects a pending order with a reason. Stops ringing on the POS, notifies customer."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to reject orders.")
    reason = (body.reason or "").strip() or "other"
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") == "rejected":
        return _serialize_online_order(order)
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {"$set": {"status": "rejected", "rejection_reason": reason, "rejected_by": user.get("name", ""), "rejected_at": now_iso, "updated_at": now_iso}},
    )
    # Rejected order → any wallet credit spent on it goes straight back.
    await _restore_order_wallet(order)
    refreshed = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    try:
        msg = _format_status_update(refreshed, "rejected", _origin_tracking_url(request, order_id, refreshed.get("track_token")))
        asyncio.create_task(send_whatsapp(refreshed.get("phone", ""), msg))
    except Exception as e:
        logger.warning(f"WhatsApp reject notify failed: {e}")
    return _serialize_online_order(refreshed)


def _recalculate_totals(items: list, original: dict) -> tuple:
    """Recompute subtotal & total when items are modified.
    Preserves existing discount_amount and delivery_fee (which were already calculated)."""
    subtotal = sum(float(i.get("price", 0)) * int(i.get("quantity", 0)) for i in items)
    discount = float(original.get("discount_amount", 0) or 0)
    delivery_fee = float(original.get("delivery_fee", 0) or 0)
    total = max(0.0, subtotal - discount) + delivery_fee
    return round(subtotal, 2), round(total, 2)


@api_router.put("/online-orders/{order_id}/modify")
async def modify_online_order(order_id: str, body: OrderModifyRequest, request: Request):
    """Staff edits items (remove items, change qty, change price). Marks the order as modified
    but DOES NOT change status — staff must call /confirm-modified after contacting the customer."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to modify orders.")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") in {"rejected", "cancelled", "delivered"}:
        raise HTTPException(status_code=400, detail=f"Cannot modify a {order.get('status')} order")
    if not body.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    new_items = []
    for it in body.items:
        if int(it.quantity) <= 0:
            continue  # qty 0 means removed
        new_items.append({
            "item_id": it.item_id or "",
            "name": it.name,
            "price": round(float(it.price), 2),
            "quantity": int(it.quantity),
        })
    if not new_items:
        raise HTTPException(status_code=400, detail="At least one item must remain in the order")
    subtotal, total = _recalculate_totals(new_items, order)
    update_fields = {
        "items": new_items,
        "subtotal": subtotal,
        "total_price": total,
        "modified": True,
        "modification_pending": True,  # awaiting staff confirmation after calling customer
        "modified_at": datetime.now(timezone.utc).isoformat(),
        "modified_by": user.get("name", ""),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if body.notes is not None:
        update_fields["modification_notes"] = body.notes
    await db.online_orders.update_one({"_id": ObjectId(order_id)}, {"$set": update_fields})
    refreshed = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    return _serialize_online_order(refreshed)


# V2: Live order operations — staff can change prep time & override delivery fee
# after the order has been accepted. Customers see these updates via /api/track/{id}
# polling within ~5s.
class OrderOperationsUpdate(BaseModel):
    prep_time_min: Optional[int] = None  # 1..240. Defaults to 30 if never set.
    delivery_fee_override: Optional[float] = None  # 0 = free delivery. None = clear override.
    free_delivery: Optional[bool] = None  # True = force delivery fee to 0.


class CustomerLocationUpdate(BaseModel):
    lat: float
    lng: float
    address: Optional[str] = None
    note: Optional[str] = None


class PushSubscriptionIn(BaseModel):
    endpoint: str
    keys: Dict[str, str]  # { p256dh, auth }


async def _send_web_push(subscription_doc: dict, title: str, body: str, url: str = "/", image: Optional[str] = None) -> tuple[bool, Optional[str]]:
    """Fire a single push notification. Returns (ok, error_message).

    Returns the actual error string on failure so the admin broadcast UI can surface
    *why* a send failed (mangled VAPID key, bad endpoint, etc.) instead of a silent
    'failed' counter. Silently removes the subscription document if the push service
    returns 404/410 (subscription expired)."""
    if not (VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY):
        return False, "VAPID keys not configured on server."
    try:
        from pywebpush import webpush, WebPushException  # type: ignore
        from py_vapid import Vapid01  # type: ignore
        # Auto-repair every common env-var corruption — literal backslash-n, wrapping
        # quotes, missing newlines, etc. Idempotent on clean PEMs. See _normalize_vapid_pem.
        priv = _normalize_vapid_pem(VAPID_PRIVATE_KEY)
        # CRITICAL: pywebpush 2.x routes string input through Vapid.from_string(), which
        # strips all newlines from the PEM and tries to base64-decode the whole thing —
        # including the "-----BEGIN/END-----" markers — producing garbage bytes and an
        # "ASN.1 parsing error: invalid length". The fix is to parse the PEM ourselves
        # via Vapid01.from_pem (which extracts the body correctly) and pass the resulting
        # Vapid01 instance — pywebpush checks isinstance(Vapid01) first and skips the
        # broken from_string path. See: https://github.com/web-push-libs/pywebpush v2.x.
        vapid_obj = Vapid01.from_pem(priv.encode())
        payload_dict = {"title": title, "body": body, "url": url, "icon": "/logo-192.png", "badge": "/favicon-32x32.png"}
        if image:
            payload_dict["image"] = image
        payload = json.dumps(payload_dict)
        webpush(
            subscription_info={"endpoint": subscription_doc["endpoint"], "keys": subscription_doc["keys"]},
            data=payload,
            vapid_private_key=vapid_obj,
            vapid_claims={"sub": VAPID_EMAIL},
        )
        return True, None
    except WebPushException as e:  # type: ignore
        status = getattr(getattr(e, "response", None), "status_code", None)
        if status in (404, 410):
            try:
                await db.push_subscriptions.delete_one({"_id": subscription_doc["_id"]})
            except Exception:
                pass
        err = f"push service error: {e}"
        logger.warning(err)
        return False, err
    except Exception as e:
        err = f"send error: {type(e).__name__}: {e}"
        logger.warning(err)
        return False, err


# --- Native app push (Firebase Cloud Messaging) ----------------------------
# Inert until configured: set FIREBASE_CREDENTIALS (or GOOGLE_APPLICATION_CREDENTIALS)
# to a Firebase service-account JSON path AND `pip install firebase-admin`. Without
# either, _send_fcm_to_customer is a no-op so order-status updates never fail.
_fcm_app = None
_fcm_ready: Optional[bool] = None  # None = not yet probed


def _init_fcm() -> bool:
    """Lazily initialise the Firebase Admin app. Returns True when FCM can send.

    FIREBASE_CREDENTIALS may be EITHER a path to the service-account JSON file
    (local dev) OR the raw JSON content itself (Fly.io secrets have no
    filesystem). GOOGLE_APPLICATION_CREDENTIALS remains path-only fallback."""
    global _fcm_app, _fcm_ready
    if _fcm_ready is not None:
        return _fcm_ready
    raw = os.environ.get("FIREBASE_CREDENTIALS") or ""
    cred_source = None
    if raw.strip().startswith("{"):
        # Raw JSON content in the env var (e.g. `fly secrets set FIREBASE_CREDENTIALS=...`)
        try:
            cred_source = json.loads(raw)
        except Exception as e:
            logger.warning(f"FCM disabled: FIREBASE_CREDENTIALS is not valid JSON ({e}).")
            _fcm_ready = False
            return False
    else:
        cred_path = raw or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
        if not cred_path or not os.path.exists(cred_path):
            logger.info("FCM disabled: no service-account credentials configured.")
            _fcm_ready = False
            return False
        cred_source = cred_path
    try:
        import firebase_admin
        from firebase_admin import credentials as _fb_cred
        if not firebase_admin._apps:
            _fcm_app = firebase_admin.initialize_app(_fb_cred.Certificate(cred_source))
        else:
            _fcm_app = firebase_admin.get_app()
        _fcm_ready = True
        logger.info("FCM initialised for native-app push.")
    except Exception as e:
        logger.warning(f"FCM init failed ({type(e).__name__}: {e}); native push disabled.")
        _fcm_ready = False
    return _fcm_ready


async def _send_fcm_to_customer(customer_id, title: str, body: str, order_id: str):
    """Fan out an FCM data+notification message to every device token a customer
    registered via POST /customer/fcm-token. Best-effort; prunes dead tokens."""
    if not _init_fcm():
        return
    try:
        cust = await db.customers.find_one(
            {"_id": ObjectId(str(customer_id))}, {"fcm_tokens": 1})
    except Exception:
        return
    tokens = list((cust or {}).get("fcm_tokens") or [])
    if not tokens:
        return
    try:
        from firebase_admin import messaging as _fcm_msg
    except Exception:
        return

    def _send_all():
        # firebase-admin is synchronous; run it off the event loop.
        stale: list[str] = []
        for tok in tokens:
            try:
                _fcm_msg.send(_fcm_msg.Message(
                    token=tok,
                    notification=_fcm_msg.Notification(title=title, body=body),
                    data={"order_id": str(order_id), "type": "order_status"},
                    # channel_id must match the high-importance channel the app
                    # creates in MainActivity — that's what makes the push show
                    # as a heads-up banner and on the LOCK SCREEN instead of
                    # landing silently in the tray on FCM's default channel.
                    android=_fcm_msg.AndroidConfig(
                        priority="high",
                        notification=_fcm_msg.AndroidNotification(
                            channel_id="knb_orders", sound="default"),
                    ),
                ))
            except Exception as e:
                name = type(e).__name__
                if "NotRegistered" in name or "InvalidArgument" in name or "Unregistered" in name:
                    stale.append(tok)
        return stale

    try:
        stale = await asyncio.to_thread(_send_all)
        if stale:
            await db.customers.update_one(
                {"_id": ObjectId(str(customer_id))},
                {"$pull": {"fcm_tokens": {"$in": stale}}},
            )
    except Exception as e:
        logger.warning(f"FCM send failed for customer {customer_id}: {e}")


async def _send_fcm_broadcast(title: str, body: str, url: str = "/", customer_ids: Optional[list] = None) -> tuple[int, int]:
    """Marketing fan-out to APP devices: send an FCM notification to every
    registered fcm_token (or only those of `customer_ids` when given — used by
    the admin "Send test"). Returns (sent, failed). Best-effort, prunes dead
    tokens per customer, and is a silent no-op when FCM isn't configured —
    mirroring _send_fcm_to_customer so web broadcasts never fail because of it."""
    if not _init_fcm():
        return 0, 0
    try:
        from firebase_admin import messaging as _fcm_msg
    except Exception:
        return 0, 0
    query: dict = {"fcm_tokens": {"$exists": True, "$ne": []}}
    if customer_ids is not None:
        try:
            query["_id"] = {"$in": [ObjectId(str(c)) for c in customer_ids]}
        except Exception:
            return 0, 0
    custs = await db.customers.find(query, {"fcm_tokens": 1}).to_list(10000)
    if not custs:
        return 0, 0

    def _send_for(tokens: list) -> tuple[int, list]:
        ok, stale = 0, []
        for tok in tokens:
            try:
                _fcm_msg.send(_fcm_msg.Message(
                    token=tok,
                    notification=_fcm_msg.Notification(title=title, body=body),
                    data={"type": "broadcast", "url": str(url or "/")},
                    # Same high-importance channel as order pushes — heads-up +
                    # lock-screen visibility (see MainActivity channel setup).
                    android=_fcm_msg.AndroidConfig(
                        priority="high",
                        notification=_fcm_msg.AndroidNotification(
                            channel_id="knb_orders", sound="default"),
                    ),
                ))
                ok += 1
            except Exception as e:
                name = type(e).__name__
                if "NotRegistered" in name or "InvalidArgument" in name or "Unregistered" in name:
                    stale.append(tok)
        return ok, stale

    sent, failed = 0, 0
    for cust in custs:
        tokens = list(cust.get("fcm_tokens") or [])
        if not tokens:
            continue
        try:
            ok, stale = await asyncio.to_thread(_send_for, tokens)
            sent += ok
            failed += len(tokens) - ok
            if stale:
                await db.customers.update_one(
                    {"_id": cust["_id"]}, {"$pull": {"fcm_tokens": {"$in": stale}}})
        except Exception as e:
            failed += len(tokens)
            logger.warning(f"FCM broadcast failed for customer {cust.get('_id')}: {e}")
    return sent, failed


async def _notify_customer_order_status(order_doc: dict, new_status: str):
    """Fan out a web push notification to every device the order's customer subscribed
    from. Best-effort — failures are logged, never raised. Guests (no customer_id)
    fall back to notifying any subscription that was keyed to the same phone (if the
    customer ever signed in later)."""
    titles = {
        "accepted": "Order accepted",
        "preparing": "Your order is being prepared",
        "ready": "Order ready for pickup / delivery",
        "out_for_delivery": "Your order is on the way",
        "delivered": "Order delivered — enjoy!",
        "rejected": "Order rejected",
        "cancelled": "Order cancelled",
    }
    if new_status not in titles:
        return
    customer_id = order_doc.get("customer_id")
    if not customer_id:
        return
    receipt = str(order_doc.get("_id", ""))[-6:].upper()
    title = titles[new_status]
    body = f"Order #{receipt} — tap to view live status."
    url = f"/track/{order_doc.get('_id')}"
    try:
        subs = await db.push_subscriptions.find({"customer_id": str(customer_id)}).to_list(20)
        for s in subs:
            await _send_web_push(s, title, body, url)
    except Exception as e:
        logger.warning(f"_notify_customer_order_status failed: {e}")
    # Native app (Flutter) devices — Firebase Cloud Messaging. No-op when FCM
    # is not configured, so this never blocks the status update.
    try:
        await _send_fcm_to_customer(customer_id, title, body, str(order_doc.get("_id")))
    except Exception as e:
        logger.warning(f"_notify_customer_order_status FCM failed: {e}")


async def _notify_admins_new_order(order_doc: dict) -> None:
    """Web-push every admin device that opted in (role:"admin" subscriptions —
    see /admin/push/subscribe) when an order enters the staff queue. This is the
    closed-browser counterpart of the in-page GlobalOrderAlert ring: the OS shows
    the notification even when no admin tab is open (browser background process
    permitting). Best-effort and fire-and-forget — never blocks order creation."""
    try:
        receipt = str(order_doc.get("_id", ""))[-6:].upper()
        total = order_doc.get("total_price")
        body = f"Order #{receipt}" + (f" — Rs {total:g}" if total else "") + ". Tap to open the POS queue."
        subs = await db.push_subscriptions.find({"role": "admin"}).to_list(50)
        for s in subs:
            await _send_web_push(s, "🔔 New online order", body, "/admin/orders")
    except Exception as e:
        logger.warning(f"_notify_admins_new_order failed: {e}")


async def _remind_admins_pending_orders() -> None:
    """Scheduler job (every 2 min): re-ring admin devices via web push while any
    order is still pending (not yet accepted). Mimics the POS ringing loop for
    the closed-browser case; goes quiet the moment the queue is empty. No-op
    when nobody has opted in, so it costs nothing on default deploys."""
    try:
        count = await db.online_orders.count_documents({"status": "pending"})
        if not count:
            return
        subs = await db.push_subscriptions.find({"role": "admin"}).to_list(50)
        if not subs:
            return
        plural = "s" if count != 1 else ""
        body = f"{count} pending order{plural} waiting to be accepted."
        for s in subs:
            await _send_web_push(s, "⏰ Orders waiting", body, "/admin/orders")
    except Exception as e:
        logger.warning(f"_remind_admins_pending_orders failed: {e}")


@api_router.post("/admin/push/subscribe")
async def admin_push_subscribe(body: PushSubscriptionIn, request: Request):
    """Register a staff browser for new-order push alerts. Same collection and
    upsert semantics as the customer /push/subscribe, plus role:"admin" so the
    new-order/pending-reminder fan-outs can target staff devices only. Gated on
    the online_orders permission — the same one that guards the orders queue."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to receive order alerts.")
    if not body.endpoint or not body.keys.get("p256dh") or not body.keys.get("auth"):
        raise HTTPException(status_code=400, detail="Invalid subscription payload.")
    now = datetime.now(timezone.utc).isoformat()
    await db.push_subscriptions.update_one(
        {"endpoint": body.endpoint},
        {"$set": {
            "endpoint": body.endpoint,
            "keys": body.keys,
            "role": "admin",
            "admin_user_id": str(user["_id"]),
            "updated_at": now,
        }, "$setOnInsert": {"created_at": now}},
        upsert=True,
    )
    return {"ok": True}


# =============================================================================
# CUSTOMER REFUND REQUESTS (paid online orders)
# =============================================================================
# Customers request a refund from the tracking page (web + app); staff action it
# from Admin → Online Orders. Money moves manually in the SafePay dashboard
# (full/partial refund back to the payment method) — this system tracks the
# request lifecycle and keeps the customer informed at every step via
# WhatsApp (works for guests) + email/web-push/FCM (signed-in customers).
# State lives in an additive `refund_request` subdoc on the order:
#   {status: requested|approved|refunded|rejected, reason, amount, admin_note,
#    requested_at, updated_at, refunded_at, acted_by}
# Order `status`/`payment_status` are intentionally untouched — reports and the
# tracking timeline keep their existing meaning.

REFUND_WINDOW_DAYS = int(os.environ.get("REFUND_WINDOW_DAYS", "7"))

class RefundRequestIn(BaseModel):
    reason: str

class RefundActionIn(BaseModel):
    action: str            # approved | rejected | refunded
    note: Optional[str] = ""
    # How the money goes back when action == "refunded":
    #   gateway = sent from the SafePay dashboard (default, works for guests)
    #   wallet  = store credit into the customer's account wallet (signed-in only)
    method: Optional[str] = "gateway"

class RefundMessageIn(BaseModel):
    text: Optional[str] = ""
    image: Optional[str] = None  # data: URL — persisted to disk, never stored inline

async def _email_customer(to: str, subject: str, body_text: str) -> None:
    """Best-effort transactional email using the SMTP settings the daily-report
    emails already use. Silent no-op when SMTP isn't configured."""
    if not to:
        return
    try:
        s = await db.settings.find_one({"key": "global"}, {"_id": 0}) or {}
        host, user_, pw = s.get("smtp_host"), s.get("smtp_user"), s.get("smtp_password")
        if not (host and user_ and pw):
            return
        await asyncio.to_thread(
            _send_email_sync, host, s.get("smtp_port", 587), user_, pw,
            s.get("smtp_use_tls", True), s.get("smtp_from") or user_,
            [to], subject, body_text, None)
    except Exception as e:
        logger.warning(f"refund email to {to} failed: {e}")

async def _notify_customer_refund(order: dict, status: str, note: str = "") -> None:
    """Tell the customer where their refund stands. WhatsApp reaches everyone
    (guests included); email + web push + FCM additionally reach signed-in
    customers. Best-effort — never raises."""
    rr = order.get("refund_request") or {}
    amount = rr.get("amount") or order.get("total_price", 0)
    receipt = str(order.get("_id", ""))[-6:].upper()
    msgs = {
        "requested": ("Refund request received",
            f"We received your refund request for order #{receipt} (Rs {amount:g}). "
            "We'll review it and update you shortly."),
        "approved": ("Refund approved",
            f"Good news — your refund of Rs {amount:g} for order #{receipt} is approved and will be "
            "processed within 2-3 business days back to your original payment method. "
            "Depending on your bank, it can take a few more days to appear on your statement."),
        "refunded": ("Refund completed",
            f"Rs {amount:g} for order #{receipt} has been sent back to your payment method. "
            "Depending on your bank it may take a few days to appear on your statement."),
        "rejected": ("Refund request declined",
            f"Your refund request for order #{receipt} was declined."
            + (f" Reason: {note}" if note else "")
            + " Please call us if you'd like to discuss this."),
        "refunded_wallet": ("Refund credited to your wallet",
            f"Rs {amount:g} for order #{receipt} has been credited to your account wallet — "
            "you can use it on your next order. Open the app or sign in on the website to see your balance."),
    }
    title, body = msgs.get(status, (None, None))
    if not title:
        return
    try:
        asyncio.create_task(send_whatsapp(order.get("phone", ""), f"{title}\n{body}"))
    except Exception:
        pass
    cid = order.get("customer_id")
    if not cid:
        return
    url = f"/track/{order.get('_id')}"
    try:
        subs = await db.push_subscriptions.find({"customer_id": str(cid)}).to_list(20)
        for s in subs:
            await _send_web_push(s, title, body, url)
    except Exception as e:
        logger.warning(f"refund web push failed: {e}")
    try:
        await _send_fcm_to_customer(cid, title, body, str(order.get("_id")))
    except Exception as e:
        logger.warning(f"refund FCM failed: {e}")
    try:
        cust = await db.customers.find_one({"_id": ObjectId(str(cid))}, {"email": 1})
        if cust and cust.get("email"):
            await _email_customer(cust["email"], f"{title} — order #{receipt}", body)
    except Exception as e:
        logger.warning(f"refund email lookup failed: {e}")

async def _push_admins(title: str, body: str, url: str = "/admin/orders") -> None:
    """Push to staff devices that opted in (role:"admin" subscriptions)."""
    try:
        subs = await db.push_subscriptions.find({"role": "admin"}).to_list(50)
        for s in subs:
            await _send_web_push(s, title, body, url)
    except Exception as e:
        logger.warning(f"_push_admins failed: {e}")

@api_router.post("/online-orders/{order_id}/refund-request")
async def request_order_refund(order_id: str, body: RefundRequestIn, request: Request, t: Optional[str] = None):
    """Customer asks for their money back on a PAID online order. Authorization
    mirrors /track/{order_id}: signed-in owner/admin, or the per-order
    track_token (guests) — 404 on failure so order ids can't be enumerated."""
    try:
        o = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not o:
        raise HTTPException(status_code=404, detail="Order not found")
    # --- auth (same 2 layers as public_track_order) ---
    authorized = False
    try:
        user = await get_current_user(request)
        if user.get("role") == "admin":
            authorized = True
    except HTTPException:
        pass
    if not authorized:
        try:
            cust = await get_current_customer(request)
            if cust and o.get("customer_id") and str(o.get("customer_id")) == str(cust.get("_id")):
                authorized = True
        except HTTPException:
            pass
    if not authorized:
        tok = o.get("track_token") or ""
        if not (t and tok and secrets.compare_digest(str(t), str(tok))):
            raise HTTPException(status_code=404, detail="Order not found")
    # --- eligibility ---
    reason = (body.reason or "").strip()
    if len(reason) < 5:
        raise HTTPException(status_code=400, detail="Please describe the problem (at least a few words).")
    if o.get("payment_method") in ("cod", "pay_at_restaurant"):
        raise HTTPException(status_code=400, detail="Cash orders are settled in person — please call the restaurant.")
    if o.get("payment_status") != "paid":
        raise HTTPException(status_code=400, detail="Only paid orders can request a refund.")
    if o.get("refund_request"):
        raise HTTPException(status_code=400, detail="A refund request already exists for this order.")
    try:
        created = datetime.fromisoformat(str(o.get("created_at", "")).replace("Z", "+00:00"))
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        if (datetime.now(timezone.utc) - created).days > REFUND_WINDOW_DAYS:
            raise HTTPException(status_code=400, detail=f"Refunds can be requested within {REFUND_WINDOW_DAYS} days of the order.")
    except HTTPException:
        raise
    except Exception:
        pass  # unparseable created_at → don't block the customer on our bug
    now = datetime.now(timezone.utc).isoformat()
    rr = {
        "status": "requested",
        "reason": reason[:500],
        "amount": float(o.get("total_price", 0)),
        "admin_note": "",
        "requested_at": now,
        "updated_at": now,
    }
    await db.online_orders.update_one({"_id": o["_id"]}, {"$set": {"refund_request": rr}})
    o["refund_request"] = rr
    receipt = str(o["_id"])[-6:].upper()
    await _push_admins("💸 Refund requested",
                       f"Order #{receipt} — Rs {rr['amount']:g}. Reason: {reason[:80]}")
    await _notify_customer_refund(o, "requested")
    return {"ok": True, "refund_request": rr}

@api_router.post("/admin/online-orders/{order_id}/refund-action")
async def admin_refund_action(order_id: str, body: RefundActionIn, request: Request):
    """Staff moves a refund request through its lifecycle:
    requested → approved | rejected;  approved → refunded | rejected.
    'refunded' means the money was actually sent from the SafePay dashboard.
    Every transition notifies the customer (WhatsApp/email/push)."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to manage refunds.")
    action = (body.action or "").strip().lower()
    if action not in ("approved", "rejected", "refunded"):
        raise HTTPException(status_code=400, detail="Invalid action.")
    try:
        o = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not o or not o.get("refund_request"):
        raise HTTPException(status_code=404, detail="No refund request on this order.")
    rr = o["refund_request"]
    cur = rr.get("status")
    allowed = {"requested": {"approved", "rejected", "refunded"},
               "approved": {"refunded", "rejected"}}
    if action not in allowed.get(cur, set()):
        raise HTTPException(status_code=400, detail=f"Cannot move a '{cur}' request to '{action}'.")
    now = datetime.now(timezone.utc).isoformat()
    rr.update({
        "status": action,
        "admin_note": (body.note or "").strip()[:500],
        "acted_by": user.get("name", "") or user.get("email", ""),
        "updated_at": now,
    })
    wallet_credited = False
    if action == "refunded":
        rr["refunded_at"] = now
        rr["refund_method"] = (body.method or "gateway").strip().lower()
        # Store-credit refund: money goes into the customer's wallet instead of
        # back through the gateway. Signed-in customers only (guests have no
        # account to hold the balance).
        if rr["refund_method"] == "wallet":
            cid = o.get("customer_id")
            if not cid:
                raise HTTPException(status_code=400, detail="Guest order — wallet refunds need a customer account. Refund via SafePay instead.")
            amount = float(rr.get("amount") or 0)
            res = await db.customers.update_one(
                {"_id": ObjectId(str(cid))}, {"$inc": {"wallet_balance": amount}})
            if not res.matched_count:
                raise HTTPException(status_code=400, detail="Customer account no longer exists — refund via SafePay instead.")
            await db.wallet_transactions.insert_one({
                "customer_id": str(cid), "type": "refund_credit", "amount": amount,
                "order_id": str(o["_id"]), "note": f"Refund for order #{str(o['_id'])[-6:].upper()}",
                "created_by": user.get("name", "") or user.get("email", ""),
                "created_at": now,
            })
            wallet_credited = True
    await db.online_orders.update_one({"_id": o["_id"]}, {"$set": {"refund_request": rr}})
    o["refund_request"] = rr
    await _notify_customer_refund(o, "refunded_wallet" if wallet_credited else action, rr["admin_note"])
    return {"ok": True, "refund_request": rr}

@api_router.get("/admin/refund-requests")
async def admin_list_refund_requests(request: Request, status: str = "all"):
    """The admin Refund Requests page: every order carrying a refund_request,
    newest first, with everything staff need in one place — customer identity
    (and whether they ordered as GUEST, i.e. no account/history on their side),
    contact, items, totals, the conversation, and the current state."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to view refunds.")
    q: dict = {"refund_request": {"$exists": True, "$ne": None}}
    if status != "all":
        q["refund_request.status"] = status
    orders = await db.online_orders.find(q).sort("refund_request.requested_at", -1).to_list(300)
    out = []
    for o in orders:
        out.append({
            "order_id": str(o["_id"]),
            "receipt_no": str(o["_id"])[-6:].upper(),
            "customer_name": o.get("customer_name", ""),
            "phone": o.get("phone", ""),
            "address": o.get("address", ""),
            "is_guest": not bool(o.get("customer_id")),
            "customer_id": str(o.get("customer_id")) if o.get("customer_id") else None,
            "payment_method": o.get("payment_method", ""),
            "payment_status": o.get("payment_status", ""),
            "order_status": o.get("status", ""),
            "items": o.get("items", []),
            "total_price": o.get("total_price", 0),
            "order_created_at": o.get("created_at", ""),
            "refund_request": o.get("refund_request"),
        })
    counts = {}
    try:
        async for grp in db.online_orders.aggregate([
            {"$match": {"refund_request": {"$exists": True, "$ne": None}}},
            {"$group": {"_id": "$refund_request.status", "n": {"$sum": 1}}},
        ]):
            counts[grp["_id"]] = grp["n"]
    except Exception:
        pass
    return {"requests": out, "counts": counts}

async def _append_refund_message(o: dict, sender: str, sender_name: str, text: str, image_data_url: Optional[str]) -> dict:
    """Append one message to the refund conversation. Images arrive as data:
    URLs and are persisted to disk via the same helper the menu uses — the DB
    only ever stores a small URL."""
    text = (text or "").strip()[:1000]
    image_url = ""
    if image_data_url and str(image_data_url).startswith("data:"):
        try:
            image_url = _persist_data_url_image(image_data_url, kind="refund")
        except Exception as e:
            logger.warning(f"refund image persist failed: {e}")
    if not text and not image_url:
        raise HTTPException(status_code=400, detail="Message is empty.")
    msg = {"from": sender, "name": sender_name, "text": text,
           "image_url": image_url if image_url.startswith("/api/uploads/") else "",
           "at": datetime.now(timezone.utc).isoformat()}
    rr = o.get("refund_request") or {}
    msgs = list(rr.get("messages") or [])
    msgs.append(msg)
    rr["messages"] = msgs[-50:]  # cap the thread
    rr["updated_at"] = msg["at"]
    await db.online_orders.update_one({"_id": o["_id"]}, {"$set": {"refund_request": rr}})
    o["refund_request"] = rr
    return msg

@api_router.post("/online-orders/{order_id}/refund-message")
async def customer_refund_message(order_id: str, body: RefundMessageIn, request: Request, t: Optional[str] = None):
    """Customer side of the refund conversation (proof photos welcome).
    Same authorization as the refund request itself."""
    try:
        o = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not o or not o.get("refund_request"):
        raise HTTPException(status_code=404, detail="No refund request on this order.")
    authorized = False
    try:
        user = await get_current_user(request)
        if user.get("role") == "admin":
            authorized = True
    except HTTPException:
        pass
    if not authorized:
        try:
            cust = await get_current_customer(request)
            if cust and o.get("customer_id") and str(o.get("customer_id")) == str(cust.get("_id")):
                authorized = True
        except HTTPException:
            pass
    if not authorized:
        tok = o.get("track_token") or ""
        if not (t and tok and secrets.compare_digest(str(t), str(tok))):
            raise HTTPException(status_code=404, detail="Order not found")
    msg = await _append_refund_message(o, "customer", o.get("customer_name", "Customer"), body.text, body.image)
    receipt = str(o["_id"])[-6:].upper()
    await _push_admins("💬 Refund message", f"Order #{receipt}: {msg['text'][:80] or 'photo attached'}", "/admin/refund-requests")
    return {"ok": True, "message": msg, "refund_request": o["refund_request"]}

@api_router.post("/admin/online-orders/{order_id}/refund-message")
async def admin_refund_message(order_id: str, body: RefundMessageIn, request: Request):
    """Staff side of the refund conversation. Customer is notified so they come
    back to the thread (tracking page / app)."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to manage refunds.")
    try:
        o = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not o or not o.get("refund_request"):
        raise HTTPException(status_code=404, detail="No refund request on this order.")
    msg = await _append_refund_message(o, "staff", user.get("name", "") or "Restaurant", body.text, body.image)
    receipt = str(o["_id"])[-6:].upper()
    note = f"Message about your refund for order #{receipt}: {msg['text'][:200]}"
    try:
        asyncio.create_task(send_whatsapp(o.get("phone", ""), note))
    except Exception:
        pass
    cid = o.get("customer_id")
    if cid:
        try:
            subs = await db.push_subscriptions.find({"customer_id": str(cid)}).to_list(20)
            for s in subs:
                await _send_web_push(s, "Message from the restaurant", msg["text"][:120] or "Photo attached", f"/track/{o['_id']}")
        except Exception:
            pass
        try:
            await _send_fcm_to_customer(cid, "Message from the restaurant", msg["text"][:120] or "Photo attached", str(o["_id"]))
        except Exception:
            pass
    return {"ok": True, "message": msg, "refund_request": o["refund_request"]}

@api_router.post("/push/subscribe")
async def push_subscribe(body: PushSubscriptionIn, request: Request):
    """Register the caller's browser push subscription. Signed-in customers get
    order-status pushes (keyed by customer_id). Guests are ALSO accepted (since
    the visitor opt-in banner): stored without a customer_id, they receive only
    the admin "Send to all" marketing broadcasts. If a guest later signs in, the
    silent re-subscribe on sign-in upserts the same endpoint with their
    customer_id — the subscription upgrades automatically. Guests re-subscribing
    never SET customer_id to null, so a signed-out revisit can't unlink an
    endpoint that already belongs to a customer."""
    cust = await get_optional_customer(request)
    if not body.endpoint or not body.keys.get("p256dh") or not body.keys.get("auth"):
        raise HTTPException(status_code=400, detail="Invalid subscription payload.")
    now = datetime.now(timezone.utc).isoformat()
    fields = {
        "endpoint": body.endpoint,
        "keys": body.keys,
        "updated_at": now,
    }
    if cust:
        fields["customer_id"] = str(cust["_id"])
    await db.push_subscriptions.update_one(
        {"endpoint": body.endpoint},
        {"$set": fields, "$setOnInsert": {"created_at": now}},
        upsert=True,
    )
    return {"ok": True}


@api_router.post("/push/unsubscribe")
async def push_unsubscribe(body: dict, request: Request):
    endpoint = (body or {}).get("endpoint")
    if not endpoint:
        raise HTTPException(status_code=400, detail="endpoint required")
    await db.push_subscriptions.delete_one({"endpoint": endpoint})
    return {"ok": True}


@api_router.get("/push/vapid-public-key")
async def push_vapid_public_key():
    if not VAPID_PUBLIC_KEY:
        raise HTTPException(status_code=503, detail="Push notifications not configured on this server.")
    return {"public_key": VAPID_PUBLIC_KEY}


class AdminBroadcastIn(BaseModel):
    title: str
    body: str
    url: Optional[str] = "/"
    image: Optional[str] = None  # Large hero image URL shown on the notification (Android)
    test_only: bool = False


@api_router.post("/admin/notifications/broadcast")
async def admin_broadcast_notification(body: AdminBroadcastIn, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    title = (body.title or "").strip()
    notif_body = (body.body or "").strip()
    if not title or not notif_body:
        raise HTTPException(status_code=400, detail="Title and message are required.")
    url = body.url or "/"
    image = (body.image or "").strip() or None
    # Audience = WEB subscribers (push_subscriptions) + APP devices (fcm_tokens
    # on customer docs). Test mode narrows both to the admin's own customer
    # account so they can preview on their real browser + phone.
    fcm_customer_ids: Optional[list] = None  # None = all app devices
    if body.test_only:
        admin_email = (user.get("email") or "").lower().strip()
        cust = await db.customers.find_one({"email": admin_email}) if admin_email else None
        if not cust:
            raise HTTPException(status_code=400, detail=f"No customer account found for {admin_email}. Create one and enable notifications on your phone to receive test pushes.")
        subs = await db.push_subscriptions.find({"customer_id": str(cust["_id"])}).to_list(20)
        fcm_customer_ids = [str(cust["_id"])]
    else:
        subs = await db.push_subscriptions.find({}).to_list(10000)
    sent, failed = 0, 0
    errors: list[str] = []
    for s in subs:
        ok, err = await _send_web_push(s, title, notif_body, url, image=image)
        if ok:
            sent += 1
        else:
            failed += 1
            if err and len(errors) < 5:
                errors.append(err)
    # App devices via FCM. No-op when Firebase isn't configured, so web-only
    # deploys behave exactly as before.
    app_sent, app_failed = await _send_fcm_broadcast(title, notif_body, url, customer_ids=fcm_customer_ids)
    audience = len(subs) + app_sent + app_failed
    if audience == 0:
        raise HTTPException(status_code=400, detail="No subscribers found yet. Ask your customers to enable order alerts after placing an order, or to sign in on the app.")
    if not body.test_only:
        try:
            await db.notification_broadcasts.insert_one({
                "title": title, "body": notif_body, "url": url, "image": image,
                "sent": sent, "failed": failed, "audience_size": audience,
                "app_sent": app_sent, "app_failed": app_failed,
                "errors_sample": errors,
                "sent_by": user.get("email", ""),
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"broadcast log insert failed: {e}")
    return {"ok": True, "sent": sent, "failed": failed, "app_sent": app_sent, "app_failed": app_failed,
            "audience_size": audience, "test_only": body.test_only, "errors_sample": errors}


@api_router.get("/admin/notifications/history")
async def admin_notification_history(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    docs = await db.notification_broadcasts.find({}).sort("created_at", -1).limit(50).to_list(50)
    return [{
        "id": str(d.pop("_id")),
        **d,
    } for d in docs]


@api_router.get("/admin/notifications/stats")
async def admin_notification_stats(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    """Subscriber counts + last broadcast summary, shown on the admin notifications page.
    subscriber_count stays the combined total for backward compatibility; the web/app
    split is exposed separately so the UI can show "X web · Y app"."""
    web_count = await db.push_subscriptions.count_documents({})
    app_count = 0
    try:
        agg = await db.customers.aggregate([
            {"$match": {"fcm_tokens": {"$exists": True, "$ne": []}}},
            {"$project": {"n": {"$size": "$fcm_tokens"}}},
            {"$group": {"_id": None, "total": {"$sum": "$n"}}},
        ]).to_list(1)
        app_count = int(agg[0]["total"]) if agg else 0
    except Exception as e:
        logger.warning(f"app device count failed: {e}")
    last = await db.notification_broadcasts.find({}).sort("created_at", -1).limit(1).to_list(1)
    return {"subscriber_count": web_count + app_count, "web_subscribers": web_count, "app_devices": app_count,
            "last_broadcast": ({"id": str(last[0]["_id"]), **{k: v for k, v in last[0].items() if k != "_id"}} if last else None)}


@api_router.get("/admin/push/vapid/status")
async def admin_vapid_status(request: Request):
    """Diagnostic — shows whether the configured VAPID private key is actually parseable.
    Operators paste env vars and accidentally strip the PEM newlines all the time;
    this gives them a clear yes/no instead of a silent 'all pushes failed' counter."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return _vapid_key_health()


@api_router.post("/admin/push/vapid/regenerate")
async def admin_vapid_regenerate(request: Request):
    """Rotate the VAPID keypair. Wipes every existing push_subscription because the old
    keys are useless against the new pair — subscribers will re-register automatically
    next time they open the site (push.js is idempotent).

    Persists new keys to MongoDB so they survive redeploys — no env var update needed."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    global VAPID_PUBLIC_KEY, VAPID_PRIVATE_KEY
    try:
        new_pub, new_priv = _generate_vapid_keypair_raw()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Key generation failed: {e}")
    VAPID_PUBLIC_KEY = new_pub
    VAPID_PRIVATE_KEY = new_priv
    # Save to MongoDB — survives every redeploy without any manual env var changes.
    await _save_vapid_keys_to_db(new_pub, new_priv)
    # Also write to file as a local backup (best-effort, not critical).
    try:
        _VAPID_KEYS_PATH.write_text(json.dumps({"public_key": new_pub, "private_key": new_priv}))
    except Exception as e:
        logger.warning(f"VAPID keys not persisted to disk (non-critical, DB is source of truth): {e}")
    # Wipe stale subscriptions — they were signed against the old key and will all 410.
    wiped = await db.push_subscriptions.delete_many({})
    return {
        "ok": True,
        "public_key": new_pub,
        "private_key": new_priv,
        "subscriptions_wiped": wiped.deleted_count,
        "next_step": "Keys saved to MongoDB — no env var update needed. Subscribers will re-register automatically on their next visit.",
    }


# Broadcast image upload — admins drop a banner image (≤2MB) and we serve it back via a
# *public* URL that the user's push service (Mozilla/FCM/Apple) can fetch. The "image"
# field on a Web Push payload is rendered as a large hero on Android notifications and
# in the action panel on macOS/iOS.
@api_router.post("/admin/notifications/upload-image")
async def admin_upload_broadcast_image(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    ctype = (file.content_type or "").lower()
    if ctype not in {"image/jpeg", "image/jpg", "image/png", "image/webp"}:
        raise HTTPException(status_code=400, detail="Banner must be JPG, PNG or WebP")
    data = await file.read()
    if len(data) > 2 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Banner must be 2MB or smaller")
    ext = {"image/jpeg": "jpg", "image/jpg": "jpg", "image/png": "png", "image/webp": "webp"}.get(ctype, "jpg")
    storage_path = f"{APP_NAME}/broadcast-banners/{_uuid.uuid4().hex}.{ext}"
    try:
        result = _put_object(storage_path, data, ctype)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("broadcast banner upload failed")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)[:120]}")
    await db.uploaded_files.insert_one({
        "id": _uuid.uuid4().hex,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": ctype,
        "size": result.get("size", len(data)),
        "purpose": "broadcast_banner",
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "is_deleted": False,
    })
    # The push service (Mozilla / FCM / Apple) downloads this image from a public URL.
    # `request.base_url` gives us the *internal* cluster hostname behind our proxy
    # (Cloudflare / Fly), so we prefer X-Forwarded-Host + X-Forwarded-Proto when set —
    # those reflect the externally reachable origin the push service can actually hit.
    fwd_host = request.headers.get("x-forwarded-host") or request.headers.get("host") or ""
    fwd_proto = request.headers.get("x-forwarded-proto") or ("https" if request.url.scheme == "https" else "http")
    if fwd_host:
        base = f"{fwd_proto}://{fwd_host}"
    else:
        base = str(request.base_url).rstrip("/")
    public_url = f"{base}/api/public/broadcast-image/{result['path']}"
    return {"ok": True, "image_url": public_url, "path": result["path"]}


@api_router.get("/public/broadcast-image/{path:path}")
async def public_broadcast_image(path: str):
    """Public (no-auth) image fetch. Push services (FCM/Mozilla/Apple) need to pull
    the banner from a publicly reachable URL — they don't carry user cookies."""
    # Only serve files our admin uploaded as broadcast banners — never expose payment
    # screenshots or other private uploads through this route.
    if not path.startswith(f"{APP_NAME}/broadcast-banners/"):
        raise HTTPException(status_code=404, detail="Not found")
    try:
        data, ctype = _get_object(path)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=404, detail="Not found")
    return Response(
        content=data,
        media_type=ctype or "image/jpeg",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@api_router.get("/public/branding")
async def public_branding():
    """SEO + favicon endpoint. Returns the restaurant identity that other parts of the
    app (index.html JSON-LD, manifest.json icons, Open Graph tags) should reflect.
    Also used by the iOS install prompt to display the right logo."""
    s = await get_online_settings_doc()
    return {
        "name": s.get("restaurant_name") or "Karachi Naseeb Biryani",
        "logo_url": s.get("restaurant_logo_url") or "",
        "phone": s.get("restaurant_phone") or "",
        "whatsapp": s.get("restaurant_whatsapp") or "",
        "email": s.get("restaurant_email") or "",
        "address": s.get("restaurant_address") or "",
        "opening_hours": s.get("opening_hours") or "",
        "facebook_url": s.get("facebook_url") or "",
        "instagram_url": s.get("instagram_url") or "",
        "twitter_url": s.get("twitter_url") or "",
        "google_maps_url": s.get("google_maps_url") or "",
        "lat": s.get("restaurant_lat"),
        "lng": s.get("restaurant_lng"),
    }


@api_router.get("/public/icon")
async def public_icon():
    """Streams the configured restaurant logo. Powers /favicon.ico, the apple-touch-icon
    and the manifest.json icons so search engines + the OS shortcut all pick the right
    logo without the operator having to upload three sizes manually.

    Falls back to a 1x1 transparent PNG so browsers don't spam 404s when no logo is set."""
    s = await get_online_settings_doc()
    logo = (s.get("restaurant_logo_url") or "").strip()
    if logo:
        # Case 1: data: URL — decode and serve directly.
        if logo.startswith("data:"):
            try:
                head, b64 = logo.split(",", 1)
                ctype = head.split(";")[0].replace("data:", "") or "image/png"
                import base64 as _b64x
                return Response(
                    content=_b64x.b64decode(b64),
                    media_type=ctype,
                    headers={"Cache-Control": "public, max-age=3600"},
                )
            except Exception:
                pass
        # Case 2: external URL — proxy it so we control caching + same-origin.
        if logo.startswith("http://") or logo.startswith("https://"):
            try:
                async with httpx.AsyncClient(timeout=10) as cx:
                    r = await cx.get(logo)
                if r.status_code == 200 and r.content:
                    return Response(
                        content=r.content,
                        media_type=r.headers.get("content-type", "image/png"),
                        headers={"Cache-Control": "public, max-age=3600"},
                    )
            except Exception:
                pass
    # Fallback — transparent 1x1 PNG so the browser caches *something* and stops
    # hammering this endpoint. Operator hasn't uploaded a logo yet.
    import base64 as _b64x
    transparent_png = _b64x.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNkYAAAAAYAAjCB0C8AAAAASUVORK5CYII="
    )
    return Response(
        content=transparent_png,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=300"},
    )


@api_router.get("/rider/orders/{order_id}")
async def rider_get_order(order_id: str, token: str):
    """Token-protected, no-login rider view. The token is auto-generated on the order
    when it transitions to `out_for_delivery` (or when an admin assigns a rider). The
    rider opens a WhatsApp link like /rider/<order_id>?t=<token> on their phone — no
    sign-in needed, but the token is unguessable so randoms can't peek at orders."""
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order.get("rider_token") or token != order.get("rider_token"):
        raise HTTPException(status_code=403, detail="Invalid or expired rider link.")
    return _serialize_online_order(order)


@api_router.post("/rider/orders/{order_id}/delivered")
async def rider_mark_delivered(order_id: str, token: str, request: Request):
    """One-tap "Mark Delivered" for the rider. Same token check as the GET above."""
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order.get("rider_token") or token != order.get("rider_token"):
        raise HTTPException(status_code=403, detail="Invalid or expired rider link.")
    if order.get("status") in {"delivered", "cancelled", "rejected"}:
        raise HTTPException(status_code=400, detail=f"Order already {order.get('status')}.")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {"$set": {"status": "delivered", "delivered_at": now_iso, "updated_at": now_iso, "delivered_by_rider": True}},
    )
    refreshed = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    try:
        await _notify_customer_order_status(refreshed, "delivered")
    except Exception as e:
        logger.warning(f"rider deliver push notify failed: {e}")
    return _serialize_online_order(refreshed)


@api_router.post("/online-orders/{order_id}/customer-location")
async def update_customer_location(order_id: str, body: CustomerLocationUpdate, request: Request):
    """Customer-initiated location update from the tracking page. Lets the customer
    re-share their GPS coords (e.g. after they moved, the rider got lost, or the
    initial pin was wrong). We append every update to `customer_location_history`
    so the restaurant can see the trail and reach the customer at the latest spot."""
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") in {"delivered", "cancelled", "rejected"}:
        raise HTTPException(status_code=400, detail=f"This order is already {order.get('status')} — location updates are no longer accepted.")
    cust = await get_optional_customer(request)
    # If the order belongs to a signed-in customer, verify ownership before letting
    # someone update its location. Guest orders are protected by the order_id (which
    # is unguessable enough to act as a token).
    if order.get("customer_id"):
        if not cust or str(cust["_id"]) != str(order.get("customer_id")):
            raise HTTPException(status_code=403, detail="You don't have permission to update this order's location.")
    now_iso = datetime.now(timezone.utc).isoformat()
    entry = {
        "lat": float(body.lat),
        "lng": float(body.lng),
        "address": (body.address or "").strip() or None,
        "note": (body.note or "").strip() or None,
        "updated_at": now_iso,
    }
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {
            "$push": {"customer_location_history": entry},
            "$set": {
                "customer_lat": entry["lat"],
                "customer_lng": entry["lng"],
                "customer_address_updated": entry["address"] or order.get("customer_address_updated"),
                "updated_at": now_iso,
            },
        },
    )
    return {"ok": True, "history_count": len((order.get("customer_location_history") or [])) + 1, "last_entry": entry}


@api_router.put("/online-orders/{order_id}/operations")
async def update_order_operations(order_id: str, body: OrderOperationsUpdate, request: Request):
    """Staff-only: update preparation time and/or override the delivery fee on a live order.
    Customers see the new values within a few seconds on the tracking page."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to manage orders.")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") in {"delivered", "cancelled", "rejected"}:
        raise HTTPException(status_code=400, detail=f"Cannot modify a {order.get('status')} order")

    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    recalc_total = False
    new_delivery_fee = order.get("delivery_fee", 0)

    if body.prep_time_min is not None:
        pt = int(body.prep_time_min)
        if pt < 1 or pt > 240:
            raise HTTPException(status_code=400, detail="Preparation time must be between 1 and 240 minutes")
        update_fields["prep_time_min"] = pt
        update_fields["prep_time_updated_at"] = datetime.now(timezone.utc).isoformat()

    if body.free_delivery is True:
        new_delivery_fee = 0.0
        update_fields["delivery_fee_override"] = 0.0
        update_fields["delivery_fee_overridden"] = True
        update_fields["delivery_fee"] = 0.0
        recalc_total = True
    elif body.delivery_fee_override is not None:
        fee = float(body.delivery_fee_override)
        if fee < 0:
            raise HTTPException(status_code=400, detail="Delivery fee cannot be negative")
        new_delivery_fee = fee
        update_fields["delivery_fee_override"] = fee
        update_fields["delivery_fee_overridden"] = True
        update_fields["delivery_fee"] = fee
        recalc_total = True

    if recalc_total:
        # Preserve the discount we already applied; just rebuild total with the new delivery fee.
        subtotal = float(order.get("subtotal", 0) or 0)
        discount = float(order.get("discount_amount", 0) or 0)
        reward_discount = float((order.get("reward_applied") or {}).get("discount_amount", 0) or 0)
        new_total = max(0.0, subtotal - discount - reward_discount) + float(new_delivery_fee or 0)
        update_fields["total_price"] = round(new_total, 2)

    await db.online_orders.update_one({"_id": ObjectId(order_id)}, {"$set": update_fields})
    refreshed = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    return _serialize_online_order(refreshed)


@api_router.post("/online-orders/{order_id}/confirm-modified")
async def confirm_modified_online_order(order_id: str, request: Request):
    """Called after staff has phoned the customer to confirm the modified order.
    Moves the order to 'accepted' and notifies the customer."""
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to confirm modified orders.")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order.get("modified"):
        raise HTTPException(status_code=400, detail="Order has not been modified yet")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {"$set": {
            "status": "accepted",
            "accepted_at": now_iso,
            "accepted_by": user.get("name", ""),
            "modification_pending": False,
            "modification_confirmed_at": now_iso,
            "updated_at": now_iso,
        }},
    )
    refreshed = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    try:
        # Special "modified+confirmed" message
        msg = _format_status_update(refreshed, "modified", _origin_tracking_url(request, order_id, refreshed.get("track_token")))
        asyncio.create_task(send_whatsapp(refreshed.get("phone", ""), msg))
    except Exception as e:
        logger.warning(f"WhatsApp modified-confirm notify failed: {e}")
    return _serialize_online_order(refreshed)

@api_router.get("/online-orders/{order_id}")
async def get_online_order_detail(order_id: str, request: Request):
    """Return order detail. Customers can only view their own; admins can view any."""
    try:
        o = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not o:
        raise HTTPException(status_code=404, detail="Order not found")
    # Try staff/admin first — anyone with online_orders permission can view order details.
    try:
        user = await get_current_user(request)
        if _has_perm(user, "online_orders"):
            return _serialize_online_order(o)
    except HTTPException:
        pass
    cust = await get_optional_customer(request)
    if cust and o.get("customer_id") == cust["_id"]:
        return _serialize_online_order(o)
    raise HTTPException(status_code=403, detail="Not allowed")

# --- Reviews ---
@api_router.post("/reviews")
async def create_review(req: ReviewCreate, request: Request):
    cust = await get_current_customer(request)
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(req.order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("customer_id") != cust["_id"]:
        raise HTTPException(status_code=403, detail="You can only review your own orders")
    if order.get("status") != "delivered":
        raise HTTPException(status_code=400, detail="Reviews are allowed only for delivered orders")
    if await db.reviews.find_one({"order_id": req.order_id, "customer_id": cust["_id"]}):
        raise HTTPException(status_code=400, detail="You have already reviewed this order")
    if not (1 <= req.rating <= 5):
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    now = datetime.now(timezone.utc)
    doc = {
        "order_id": req.order_id,
        "customer_id": cust["_id"],
        "customer_name": cust.get("name", "Customer"),
        "rating": int(req.rating),
        "comment": _clip(req.comment, 500),
        "created_at": now.isoformat(),
    }
    result = await db.reviews.insert_one(doc)
    return {"id": str(result.inserted_id), "rating": doc["rating"], "comment": doc["comment"], "customer_name": doc["customer_name"], "created_at": doc["created_at"]}


# --- Public review (no auth) — used by the QR code on the printed receipt ---
class PublicReviewCreate(BaseModel):
    rating: int
    comment: Optional[str] = ""
    customer_name: Optional[str] = ""


@api_router.get("/reviews/order/{order_id}")
async def get_review_by_order(order_id: str):
    """Public: lets the review page check if a review already exists for this order
    and pre-fill restaurant info. No auth required."""
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    review = await db.reviews.find_one({"order_id": order_id})
    return {
        "order": {
            "id": str(order["_id"]),
            "receipt_no": str(order["_id"])[-6:].upper(),
            "customer_name": order.get("customer_name", ""),
            "items": order.get("items", []),
            "total_price": order.get("total_price", 0),
            "status": order.get("status", "pending"),
            "created_at": order.get("created_at", ""),
        },
        "review": (
            {
                "id": str(review["_id"]),
                "rating": review.get("rating", 0),
                "comment": review.get("comment", ""),
                "customer_name": review.get("customer_name", "Customer"),
                "created_at": review.get("created_at", ""),
            }
            if review else None
        ),
    }


@api_router.post("/reviews/public/{order_id}")
async def create_public_review(order_id: str, body: PublicReviewCreate):
    """Public review submission — used by the QR code on the printed receipt
    where the customer is unlikely to be logged in."""
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if not (1 <= int(body.rating) <= 5):
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    if await db.reviews.find_one({"order_id": order_id}):
        raise HTTPException(status_code=400, detail="A review has already been submitted for this order")
    now = datetime.now(timezone.utc)
    name = (body.customer_name or order.get("customer_name") or "Customer").strip()[:60]
    doc = {
        "order_id": order_id,
        "customer_id": order.get("customer_id"),
        "customer_name": name,
        "rating": int(body.rating),
        "comment": (body.comment or "").strip()[:500],
        "source": "public_qr",
        "created_at": now.isoformat(),
    }
    result = await db.reviews.insert_one(doc)
    return {
        "id": str(result.inserted_id),
        "rating": doc["rating"],
        "comment": doc["comment"],
        "customer_name": doc["customer_name"],
        "created_at": doc["created_at"],
    }

@api_router.get("/reviews")
async def list_reviews(limit: int = 50):
    revs = await db.reviews.find({}).sort("created_at", -1).to_list(max(1, min(limit, 200)))
    return [{
        "id": str(r["_id"]),
        "rating": r["rating"],
        "comment": r["comment"],
        "customer_name": r.get("customer_name", "Customer"),
        "created_at": r.get("created_at", ""),
        "admin_reply": r.get("admin_reply", ""),
        "replied_at": r.get("replied_at", ""),
        "replied_by": r.get("replied_by", ""),
        "is_feedback": bool(r.get("is_feedback", False)),
    } for r in revs]


# --- Public feedback (no order required) ---
class PublicFeedbackCreate(BaseModel):
    rating: int
    comment: str
    customer_name: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""

    @field_validator("phone")
    @classmethod
    def _v_phone(cls, v):
        return _normalize_and_validate_phone(v, required=False, field="phone")


@api_router.post("/feedback")
async def submit_feedback(req: PublicFeedbackCreate, request: Request):
    """Public: customer submits general feedback (not tied to an order). Stored in `reviews`
    with is_feedback=true so admins see it in the existing review module."""
    if not (1 <= int(req.rating) <= 5):
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    cust = await get_optional_customer(request)
    now = datetime.now(timezone.utc)
    doc = {
        "order_id": None,
        "customer_id": cust["_id"] if cust else None,
        "customer_name": (cust.get("name") if cust else None) or (req.customer_name or "Anonymous"),
        "customer_email": (cust.get("email") if cust else None) or (req.email or ""),
        "customer_phone": req.phone or "",
        "rating": int(req.rating),
        "comment": req.comment.strip(),
        "is_feedback": True,
        "created_at": now.isoformat(),
    }
    if cust:
        doc["customer_id"] = cust["_id"]
    result = await db.reviews.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Thank you for your feedback!"}

# --- Offers ---
def _parse_iso_utc(s):
    """Parse an ISO-8601 string to an aware UTC datetime, or None if unparseable."""
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None

def _offer_expired(o: dict, now: datetime | None = None) -> bool:
    """True when an offer's server-controlled valid_until has passed."""
    vu = _parse_iso_utc(o.get("valid_until"))
    if vu is None:
        return False
    return (now or datetime.now(timezone.utc)) > vu

def _offer_distribution(o: dict) -> list:
    """Return the distribution list for an offer. Docs without the field default
    to ["website","app"] so existing public offers stay visible (backward compat)."""
    d = o.get("distribution")
    if not d:
        return ["website", "app"]
    return d

def _offer_serial(o: dict, now: datetime | None = None) -> dict:
    """Serialise an offer doc for API responses (admin or public)."""
    now = now or datetime.now(timezone.utc)
    usage_limit = o.get("usage_limit")
    usage_count = int(o.get("usage_count", 0) or 0)
    remaining = (usage_limit - usage_count) if usage_limit is not None else None
    expired = _offer_expired(o, now)
    fully_redeemed = (usage_limit is not None and usage_count >= usage_limit)
    if not o.get("active", True):
        computed_status = "INACTIVE"
    elif expired:
        computed_status = "EXPIRED"
    elif fully_redeemed:
        computed_status = "FULLY_REDEEMED"
    else:
        computed_status = "ACTIVE"
    return {
        "id": str(o["_id"]),
        "title": o["title"],
        "description": o.get("description", ""),
        "discount_percent": o.get("discount_percent", 0),
        "discount_amount": o.get("discount_amount", 0),
        "coupon_code": o.get("coupon_code", ""),
        "image_url": o.get("image_url", ""),
        "active": o.get("active", True),
        "min_order_amount": float(o.get("min_order_amount", 0) or 0),
        "one_time_per_customer": bool(o.get("one_time_per_customer", False)),
        "valid_until": o.get("valid_until"),
        "created_at": o.get("created_at", ""),
        "server_now": now.isoformat(),
        # New fields
        "distribution": _offer_distribution(o),
        "usage_limit": usage_limit,
        "usage_count": usage_count,
        "remaining_uses": remaining,
        "assigned_customer_id": o.get("assigned_customer_id"),
        "share_token": o.get("share_token"),
        "computed_status": computed_status,
    }

@api_router.get("/offers")
@api_router.get("/admin/offers")
async def admin_list_offers(request: Request, active_only: bool = False):
    """Admin offer list — returns ALL offers, INCLUDING private
    'voucher_code_only' vouchers that GET /offers intentionally hides from the
    public. Admin-only. This is what the admin Offers page must call so private
    vouchers stay visible (and shareable) after they are created."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    q = {"active": True} if active_only else {}
    offers = await db.offers.find(q).sort("created_at", -1).to_list(200)
    now = datetime.now(timezone.utc)
    return [_offer_serial(o, now) for o in offers]

@api_router.get("/offers/lookup")
async def lookup_offer(code: str, request: Request):
    """Resolve a single offer by its coupon/voucher code for checkout validation.
    Unlike GET /offers (the public list), this DOES return private
    'voucher_code_only' offers, so a customer handed a private code can apply it."""
    code_normalized = (code or "").upper().strip()
    if not code_normalized:
        raise HTTPException(status_code=400, detail="Please enter a coupon code.")
    offer = await db.offers.find_one({"coupon_code": code_normalized, "active": True})
    if not offer:
        raise HTTPException(status_code=404, detail="This voucher code is invalid.")
    now = datetime.now(timezone.utc)
    if _offer_expired(offer, now):
        raise HTTPException(status_code=400, detail="This voucher has expired.")
    usage_limit = offer.get("usage_limit")
    usage_count = int(offer.get("usage_count", 0) or 0)
    if usage_limit is not None and usage_count >= usage_limit:
        raise HTTPException(status_code=400, detail="This voucher has already been fully redeemed.")
    assigned = offer.get("assigned_customer_id")
    if assigned:
        cust = await get_optional_customer(request)
        if not cust or str(cust["_id"]) != str(assigned):
            raise HTTPException(status_code=400, detail="This voucher is not valid for your account.")
    return _offer_serial(offer, now)
        @api_router.post("/offers")
async def create_offer(offer: OfferCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    # Validate coupon code uniqueness if one was supplied
    code = (offer.coupon_code or "").upper().strip()
    if code:
        existing = await db.offers.find_one({"coupon_code": code})
        if existing:
            raise HTTPException(status_code=400, detail=f"Coupon code {code} already exists.")
    dist = offer.distribution if offer.distribution else ["website", "app"]
    doc = {
        "title": offer.title,
        "description": offer.description,
        "discount_percent": offer.discount_percent or 0,
        "discount_amount": offer.discount_amount or 0,
        "coupon_code": code,
        "image_url": offer.image_url or "",
        "active": offer.active,
        "min_order_amount": float(offer.min_order_amount or 0),
        "one_time_per_customer": bool(offer.one_time_per_customer),
        "valid_until": (_parse_iso_utc(offer.valid_until).isoformat()
                        if _parse_iso_utc(offer.valid_until) else None),
        "distribution": dist,
        "usage_limit": offer.usage_limit,
        "usage_count": 0,
        "assigned_customer_id": offer.assigned_customer_id or None,
        "share_token": secrets.token_urlsafe(16),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.offers.insert_one(doc)
    doc["_id"] = result.inserted_id
    return _offer_serial(doc)

@api_router.put("/offers/{offer_id}")
async def update_offer(offer_id: str, offer: OfferUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    ud = {k: v for k, v in offer.model_dump().items() if v is not None}
    if "coupon_code" in ud:
        new_code = ud["coupon_code"].upper().strip()
        # Uniqueness check — allow keeping the same code on this offer
        existing = await db.offers.find_one({"coupon_code": new_code, "_id": {"$ne": ObjectId(offer_id)}})
        if existing:
            raise HTTPException(status_code=400, detail=f"Coupon code {new_code} is already used by another offer.")
        ud["coupon_code"] = new_code
    if "valid_until" in ud:
        parsed = _parse_iso_utc(ud["valid_until"])
        ud["valid_until"] = parsed.isoformat() if parsed else None
    # Allow explicitly clearing assigned_customer_id and usage_limit
    raw = offer.model_dump()
    if raw.get("assigned_customer_id") is None and "assigned_customer_id" in raw:
        ud["assigned_customer_id"] = None
    if raw.get("usage_limit") is None and "usage_limit" in raw:
        ud["usage_limit"] = None
    if ud:
        await db.offers.update_one({"_id": ObjectId(offer_id)}, {"$set": ud})
    updated = await db.offers.find_one({"_id": ObjectId(offer_id)})
    return _offer_serial(updated) if updated else {"message": "Updated"}

@api_router.delete("/offers/{offer_id}")
async def delete_offer(offer_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    await db.offers.delete_one({"_id": ObjectId(offer_id)})
    return {"message": "Deleted"}


# ----- Voucher share endpoints -----
# /api/public/voucher/{share_token}  — JSON for the React VoucherPage
# /api/v/{share_token}               — Full HTML page (OG meta + branded card)
#                                       Vercel proxies /v/* here so WhatsApp/social
#                                       crawlers get server-rendered OG tags.
# /api/v/{share_token}/og-image.png  — Dynamically generated 1200×630 OG image

@api_router.get("/public/voucher/{share_token}")
async def get_voucher_by_token(share_token: str):
    """JSON data for the React VoucherPage (/v/:token on the frontend).
    Returns enough to render the page; never leaks assigned customer PII."""
    offer = await db.offers.find_one({"share_token": share_token})
    if not offer:
        raise HTTPException(status_code=404, detail="Voucher not found.")
    now = datetime.now(timezone.utc)
    serial = _offer_serial(offer, now)
    # Strip private fields before sending to anonymous visitors
    return {
        "id": serial["id"],
        "title": serial["title"],
        "description": serial["description"],
        "discount_percent": serial["discount_percent"],
        "discount_amount": serial["discount_amount"],
        "coupon_code": serial["coupon_code"],
        "min_order_amount": serial["min_order_amount"],
        "valid_until": serial["valid_until"],
        "usage_limit": serial["usage_limit"],
        "remaining_uses": serial["remaining_uses"],
        "computed_status": serial["computed_status"],
        "share_token": share_token,
        # Do NOT expose assigned_customer_id or customer name here
    }


def _voucher_og_image_bytes(title: str, discount_txt: str, code: str, valid_until: str | None) -> bytes:
    """Generate a 1200×630 branded OG image for the voucher. Returns PNG bytes."""
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io, textwrap
        W, H = 1200, 630
        GREEN   = (21, 94, 63)
        YELLOW  = (254, 201, 2)
        WHITE   = (255, 255, 255)
        DARK    = (26, 29, 26)
        FONT_B  = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        FONT_R  = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"

        img = Image.new("RGB", (W, H), GREEN)
        d = ImageDraw.Draw(img)

        # Background diamond watermark
        for i in range(0, W + 200, 120):
            d.polygon([(i, H//2-40),(i+40,H//2),(i,H//2+40),(i-40,H//2)],
                      fill=(25, 100, 68))

        # White card
        cx, cy, cw, ch = 100, 80, W - 200, H - 160
        d.rounded_rectangle([cx, cy, cx+cw, cy+ch], radius=32, fill=WHITE)

        # KNB brand bar (left strip)
        d.rounded_rectangle([cx, cy, cx+12, cy+ch], radius=6, fill=YELLOW)

        # "KARACHI NASEEB BIRYANI" header
        f_brand = ImageFont.truetype(FONT_B, 28)
        d.text((cx + 40, cy + 36), "KARACHI NASEEB BIRYANI", font=f_brand, fill=GREEN)

        # 🎁 label
        f_label = ImageFont.truetype(FONT_R, 22)
        d.text((cx + 40, cy + 78), "SPECIAL VOUCHER", font=f_label, fill=(92, 95, 92))

        # Discount amount — big
        f_disc = ImageFont.truetype(FONT_B, 90)
        disc_w = d.textlength(discount_txt, font=f_disc)
        d.text((cx + cw/2 - disc_w/2, cy + 130), discount_txt, font=f_disc, fill=GREEN)

        # Title (wrapped)
        f_title = ImageFont.truetype(FONT_B, 34)
        lines = textwrap.wrap(title, width=32)
        ty = cy + 255
        for line in lines[:2]:
            lw = d.textlength(line, font=f_title)
            d.text((cx + cw/2 - lw/2, ty), line, font=f_title, fill=DARK)
            ty += 44

        # Code pill
        f_code = ImageFont.truetype(FONT_B, 42)
        pill_w = d.textlength(code, font=f_code) + 60
        pill_x = cx + cw/2 - pill_w/2
        pill_y = ty + 14
        d.rounded_rectangle([pill_x, pill_y, pill_x+pill_w, pill_y+62], radius=31, fill=YELLOW)
        cw2 = d.textlength(code, font=f_code)
        d.text((pill_x + (pill_w - cw2)/2, pill_y + 8), code, font=f_code, fill=DARK)

        # Valid until
        if valid_until:
            try:
                vu = _parse_iso_utc(valid_until)
                vu_str = f"Valid until {vu.strftime('%d %b %Y')}" if vu else ""
            except Exception:
                vu_str = ""
            if vu_str:
                f_small = ImageFont.truetype(FONT_R, 22)
                sw = d.textlength(vu_str, font=f_small)
                d.text((cx + cw/2 - sw/2, pill_y + 78), vu_str, font=f_small, fill=(92, 95, 92))

        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        return buf.getvalue()
    except Exception as e:
        logger.warning(f"OG image generation failed: {e}")
        return b""


@api_router.get("/v/{share_token}/og-image.png", response_class=Response)
async def voucher_og_image(share_token: str):
    """Dynamically generated 1200×630 PNG for WhatsApp/social OG preview."""
    offer = await db.offers.find_one({"share_token": share_token})
    if not offer:
        raise HTTPException(status_code=404, detail="Not found")
    if offer.get("discount_amount"):
        disc = f"Rs. {int(offer['discount_amount'])} OFF"
    elif offer.get("discount_percent"):
        disc = f"{int(offer['discount_percent'])}% OFF"
    else:
        disc = "Special Offer"
    png = _voucher_og_image_bytes(
        title=offer.get("title", "Special Voucher"),
        discount_txt=disc,
        code=offer.get("coupon_code", ""),
        valid_until=offer.get("valid_until"),
    )
    if not png:
        raise HTTPException(status_code=500, detail="Image generation failed")
    return Response(content=png, media_type="image/png",
                    headers={"Cache-Control": "public, max-age=3600"})


@api_router.get("/v/{share_token}", response_class=Response)
async def voucher_share_page(share_token: str, request: Request):
    """Full server-rendered HTML voucher page. WhatsApp/social crawlers land here
    and get proper OG meta tags. Real users get a branded mobile-friendly card with
    a [Copy Code] button and a direct [Order Now] link."""
    offer = await db.offers.find_one({"share_token": share_token})
    base = _abs_origin(request)
    site_url = os.environ.get("SITE_URL", "https://www.karachinaseebbiryani.com")

    if not offer:
        html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Voucher Not Found — Karachi Naseeb Biryani</title></head>
<body style="background:#155E3F;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;font-family:sans-serif">
<div style="background:#fff;border-radius:20px;padding:40px;text-align:center;max-width:400px">
<p style="font-size:48px">🤔</p>
<h1 style="color:#1A1D1A">Voucher Not Found</h1>
<p style="color:#5C5F5C">This voucher link is invalid or has been removed.</p>
<a href="{site_url}" style="display:inline-block;margin-top:20px;background:#155E3F;color:#fff;padding:14px 28px;border-radius:30px;text-decoration:none;font-weight:bold">Order Online</a>
</div></body></html>"""
        return Response(content=html, media_type="text/html; charset=utf-8", status_code=404)

    now = datetime.now(timezone.utc)
    serial = _offer_serial(offer, now)
    status = serial["computed_status"]
    code = offer.get("coupon_code", "")
    title = offer.get("title", "Special Voucher")
    desc = offer.get("description", "")

    if offer.get("discount_amount"):
        disc_txt = f"Rs. {int(offer['discount_amount'])} OFF"
    elif offer.get("discount_percent"):
        disc_txt = f"{int(offer['discount_percent'])}% OFF"
    else:
        disc_txt = "Special Offer"

    og_img_url = f"{base}/api/v/{share_token}/og-image.png"
    og_title = f"Karachi Naseeb Biryani — {disc_txt}"
    og_desc = f"Private voucher • Code: {code} • {desc}" if desc else f"Private voucher • Code: {code}"

    valid_str = ""
    if offer.get("valid_until"):
        try:
            vu = _parse_iso_utc(offer["valid_until"])
            valid_str = vu.strftime("%d %b %Y") if vu else ""
        except Exception:
            pass

    remaining = serial.get("remaining_uses")
    min_amt = float(offer.get("min_order_amount", 0) or 0)

    # Status banner content
    if status == "FULLY_REDEEMED":
        banner = '<div style="background:#C41E3A;color:#fff;padding:18px;border-radius:12px;text-align:center;font-weight:bold;font-size:18px">🚫 Voucher Fully Redeemed<br><span style="font-size:14px;font-weight:normal">This voucher is no longer available.</span></div>'
        show_code = False
    elif status == "EXPIRED":
        banner = f'<div style="background:#C41E3A;color:#fff;padding:18px;border-radius:12px;text-align:center;font-weight:bold;font-size:18px">⌛ Voucher Expired<br><span style="font-size:14px;font-weight:normal">This voucher expired on {valid_str}.</span></div>'
        show_code = False
    elif status == "INACTIVE":
        banner = '<div style="background:#888;color:#fff;padding:18px;border-radius:12px;text-align:center;font-weight:bold;font-size:18px">⛔ Voucher Unavailable<br><span style="font-size:14px;font-weight:normal">This voucher is currently inactive.</span></div>'
        show_code = False
    else:
        banner = ""
        show_code = True

    code_section = ""
    if show_code and code:
        code_section = f"""
<div style="text-align:center;margin:24px 0">
  <p style="color:#5C5F5C;font-size:13px;margin:0 0 8px">VOUCHER CODE</p>
  <div style="display:inline-flex;align-items:center;gap:10px;background:#F9F8F6;border:2px dashed #FEC902;border-radius:12px;padding:14px 24px">
    <span id="code" style="font-size:28px;font-weight:900;letter-spacing:3px;color:#1A1D1A">{code}</span>
    <button onclick="navigator.clipboard.writeText('{code}');this.textContent='✓ Copied!';setTimeout(()=>this.textContent='Copy',2000)"
      style="background:#FEC902;border:none;border-radius:8px;padding:8px 16px;font-weight:bold;cursor:pointer;font-size:13px">Copy</button>
  </div>
  <p style="color:#5C5F5C;font-size:13px;margin-top:8px">Enter this code at checkout</p>
</div>
<div style="text-align:center;margin-top:8px">
  <a href="{site_url}/menu" style="display:inline-block;background:#155E3F;color:#fff;padding:16px 40px;border-radius:30px;text-decoration:none;font-weight:bold;font-size:16px">🍽️ Order Now</a>
</div>"""

    details_rows = ""
    if min_amt > 0:
        details_rows += f'<tr><td style="color:#5C5F5C;padding:6px 0">Minimum Order</td><td style="font-weight:600;text-align:right">Rs. {int(min_amt)}</td></tr>'
    if valid_str:
        details_rows += f'<tr><td style="color:#5C5F5C;padding:6px 0">Valid Until</td><td style="font-weight:600;text-align:right">{valid_str}</td></tr>'
    if remaining is not None:
        details_rows += f'<tr><td style="color:#5C5F5C;padding:6px 0">Remaining Uses</td><td style="font-weight:600;text-align:right">{remaining}</td></tr>'
    details = f'<table style="width:100%;border-collapse:collapse;margin-top:16px">{details_rows}</table>' if details_rows else ""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{og_title}</title>
<meta name="description" content="{og_desc}">
<meta property="og:type" content="website">
<meta property="og:url" content="{site_url}/v/{share_token}">
<meta property="og:title" content="{og_title}">
<meta property="og:description" content="{og_desc}">
<meta property="og:image" content="{og_img_url}">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:site_name" content="Karachi Naseeb Biryani">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="{og_title}">
<meta name="twitter:description" content="{og_desc}">
<meta name="twitter:image" content="{og_img_url}">
<meta name="robots" content="noindex,nofollow">
<link rel="preconnect" href="https://fonts.googleapis.com">
</head>
<body style="margin:0;background:#155E3F;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:16px;box-sizing:border-box;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif">
<div style="background:#fff;border-radius:24px;max-width:440px;width:100%;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,0.3)">
  <!-- Header -->
  <div style="background:#155E3F;padding:24px;text-align:center">
    <div style="width:52px;height:52px;background:#C41E3A;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;font-size:22px;font-weight:900;color:#fff;margin-bottom:10px">K</div>
    <p style="color:#FEC902;font-size:12px;font-weight:700;letter-spacing:2px;margin:0">KARACHI NASEEB BIRYANI</p>
    <p style="color:#fff;opacity:.7;font-size:12px;margin:4px 0 0">🎁 Special Voucher</p>
  </div>
  <!-- Body -->
  <div style="padding:28px 24px">
    {banner}
    <div style="text-align:center;margin-bottom:16px">
      <div style="font-size:52px;font-weight:900;color:#155E3F;line-height:1">{disc_txt}</div>
      <h1 style="font-size:18px;font-weight:700;color:#1A1D1A;margin:8px 0 4px">{title}</h1>
      {'<p style="color:#5C5F5C;font-size:14px;margin:0">' + desc + '</p>' if desc else ''}
    </div>
    {code_section}
    {details}
  </div>
  <div style="padding:16px 24px;background:#F9F8F6;text-align:center;border-top:1px solid #E5E2DC">
    <p style="margin:0;color:#5C5F5C;font-size:12px">Karachi Naseeb Biryani & Murg Pulao · 68 Chatri Chowk, D Block, Lahore</p>
  </div>
</div>
</body>
</html>"""
    return Response(content=html, media_type="text/html; charset=utf-8",
                    headers={"Cache-Control": "public, max-age=60"})
# Public list (only enabled) drives the /faq page and the FAQ JSON-LD schema.
# Admin endpoints (full CRUD + reorder) live behind get_current_user role=admin.

def _serialize_faq(f: dict) -> dict:
    return {
        "id": str(f["_id"]),
        "question": f.get("question", ""),
        "answer": f.get("answer", ""),
        "sort_order": int(f.get("sort_order", 0)),
        "enabled": bool(f.get("enabled", True)),
        "created_at": f.get("created_at", ""),
        "updated_at": f.get("updated_at", ""),
    }


@api_router.get("/faqs")
async def list_faqs_public():
    """Public FAQ feed — only enabled entries, ordered for display. Cached by the
    frontend; backs both the /faq page UI and the FAQPage schema.org JSON-LD."""
    faqs = await db.faqs.find({"enabled": True}).sort([("sort_order", 1), ("created_at", 1)]).to_list(500)
    return [_serialize_faq(f) for f in faqs]


@api_router.get("/admin/faqs")
async def list_faqs_admin(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    faqs = await db.faqs.find({}).sort([("sort_order", 1), ("created_at", 1)]).to_list(1000)
    return [_serialize_faq(f) for f in faqs]


@api_router.post("/admin/faqs")
async def create_faq(body: FAQCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    q = (body.question or "").strip()
    a = (body.answer or "").strip()
    if not q or not a:
        raise HTTPException(status_code=400, detail="Both question and answer are required.")
    now_iso = datetime.now(timezone.utc).isoformat()
    # If no explicit sort_order, append to the end so new FAQs go last by default.
    if body.sort_order == 0:
        last = await db.faqs.find({}).sort("sort_order", -1).limit(1).to_list(1)
        next_order = (int(last[0].get("sort_order", 0)) + 1) if last else 0
    else:
        next_order = int(body.sort_order)
    doc = {
        "question": q,
        "answer": a,
        "sort_order": next_order,
        "enabled": bool(body.enabled),
        "created_at": now_iso,
        "updated_at": now_iso,
    }
    result = await db.faqs.insert_one(doc)
    doc["_id"] = result.inserted_id
    return _serialize_faq(doc)


@api_router.put("/admin/faqs/{faq_id}")
async def update_faq(faq_id: str, body: FAQUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    try:
        oid = ObjectId(faq_id)
    except Exception:
        raise HTTPException(status_code=404, detail="FAQ not found")
    ud = {k: v for k, v in body.model_dump().items() if v is not None}
    if not ud:
        raise HTTPException(status_code=400, detail="Nothing to update")
    ud["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.faqs.update_one({"_id": oid}, {"$set": ud})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="FAQ not found")
    refreshed = await db.faqs.find_one({"_id": oid})
    return _serialize_faq(refreshed)


@api_router.delete("/admin/faqs/{faq_id}")
async def delete_faq(faq_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    try:
        oid = ObjectId(faq_id)
    except Exception:
        raise HTTPException(status_code=404, detail="FAQ not found")
    res = await db.faqs.delete_one({"_id": oid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="FAQ not found")
    return {"message": "Deleted"}


@api_router.post("/admin/faqs/reorder")
async def reorder_faqs(body: FAQReorder, request: Request):
    """Persist a new ordering for all FAQs in one call. Sent by the admin UI
    after the operator drags entries / clicks the up/down arrows. Each id's
    sort_order is set to its index in the submitted list. Unlisted ids stay put."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    if not body.ids or not isinstance(body.ids, list):
        raise HTTPException(status_code=400, detail="ids list required")
    now_iso = datetime.now(timezone.utc).isoformat()
    for idx, faq_id in enumerate(body.ids):
        try:
            oid = ObjectId(faq_id)
        except Exception:
            continue  # skip malformed ids — don't fail the whole batch
        await db.faqs.update_one({"_id": oid}, {"$set": {"sort_order": idx, "updated_at": now_iso}})
    return {"message": "Reordered", "count": len(body.ids)}


# ===== DELIVERY AREA ENDPOINTS =====
# Mirrors the FAQ pattern: public list (enabled only) drives the /delivery page;
# admin endpoints (full CRUD + reorder) live behind get_current_user role=admin.
# Listing real area names (Johar Town, DHA, Model Town, ...) on a crawlable page
# helps rank for "biryani delivery in <area>" local searches.

def _serialize_area(a: dict) -> dict:
    return {
        "id": str(a["_id"]),
        "name": a.get("name", ""),
        "note": a.get("note", ""),
        "sort_order": int(a.get("sort_order", 0)),
        "enabled": bool(a.get("enabled", True)),
        "created_at": a.get("created_at", ""),
        "updated_at": a.get("updated_at", ""),
    }


@api_router.get("/delivery-areas")
async def list_delivery_areas_public():
    """Public delivery-area feed — only enabled entries, ordered for display.
    Backs the /delivery page grid of areas served."""
    areas = await db.delivery_areas.find({"enabled": True}).sort([("sort_order", 1), ("created_at", 1)]).to_list(500)
    return [_serialize_area(a) for a in areas]


@api_router.get("/admin/delivery-areas")
async def list_delivery_areas_admin(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    areas = await db.delivery_areas.find({}).sort([("sort_order", 1), ("created_at", 1)]).to_list(1000)
    return [_serialize_area(a) for a in areas]


@api_router.post("/admin/delivery-areas")
async def create_delivery_area(body: DeliveryAreaCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Area name is required.")
    now_iso = datetime.now(timezone.utc).isoformat()
    if body.sort_order == 0:
        last = await db.delivery_areas.find({}).sort("sort_order", -1).limit(1).to_list(1)
        next_order = (int(last[0].get("sort_order", 0)) + 1) if last else 0
    else:
        next_order = int(body.sort_order)
    doc = {
        "name": name,
        "note": (body.note or "").strip(),
        "sort_order": next_order,
        "enabled": bool(body.enabled),
        "created_at": now_iso,
        "updated_at": now_iso,
    }
    result = await db.delivery_areas.insert_one(doc)
    doc["_id"] = result.inserted_id
    return _serialize_area(doc)


@api_router.put("/admin/delivery-areas/{area_id}")
async def update_delivery_area(area_id: str, body: DeliveryAreaUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    try:
        oid = ObjectId(area_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Delivery area not found")
    ud = {k: v for k, v in body.model_dump().items() if v is not None}
    if not ud:
        raise HTTPException(status_code=400, detail="Nothing to update")
    ud["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.delivery_areas.update_one({"_id": oid}, {"$set": ud})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Delivery area not found")
    refreshed = await db.delivery_areas.find_one({"_id": oid})
    return _serialize_area(refreshed)


@api_router.delete("/admin/delivery-areas/{area_id}")
async def delete_delivery_area(area_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    try:
        oid = ObjectId(area_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Delivery area not found")
    res = await db.delivery_areas.delete_one({"_id": oid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Delivery area not found")
    return {"message": "Deleted"}


@api_router.post("/admin/delivery-areas/reorder")
async def reorder_delivery_areas(body: DeliveryAreaReorder, request: Request):
    """Persist a new ordering for all delivery areas in one call — mirrors
    /admin/faqs/reorder. Each id's sort_order is set to its index."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    if not body.ids or not isinstance(body.ids, list):
        raise HTTPException(status_code=400, detail="ids list required")
    now_iso = datetime.now(timezone.utc).isoformat()
    for idx, area_id in enumerate(body.ids):
        try:
            oid = ObjectId(area_id)
        except Exception:
            continue
        await db.delivery_areas.update_one({"_id": oid}, {"$set": {"sort_order": idx, "updated_at": now_iso}})
    return {"message": "Reordered", "count": len(body.ids)}


# ===== SEO endpoints: sitemap.xml + robots.txt =====
# These are served via the /api prefix so the Kubernetes ingress routes them to
# this backend. The frontend (Vercel) rewrites /sitemap.xml and /robots.txt
# (see vercel.json) so search engines / AI crawlers find them at the root.

def _abs_origin(request: Request) -> str:
    """Best-effort canonical https://hostname/ for sitemap URLs. Falls back to
    the production domain so a sitemap fetched directly from the fly.io host
    still emits the public-facing karachinaseebbiryani.com URLs (which is what
    Google indexes)."""
    public = os.environ.get("PUBLIC_SITE_URL", "https://www.karachinaseebbiryani.com").rstrip("/")
    return public


@api_router.get("/sitemap.xml", response_class=Response)
async def sitemap_xml(request: Request):
    """Dynamic XML sitemap. Includes every public route + active menu category
    landing fragments + active offer fragments + FAQ slugs. Google + Bing +
    Perplexity all use this to discover what to crawl. Updated last-mod is
    important for re-crawl priority — we pull the most recent updated_at off
    each entity so the sitemap stays fresh without a manual rebuild."""
    base = _abs_origin(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    urls: list[dict] = [
        {"loc": f"{base}/",                  "changefreq": "daily",   "priority": "1.0", "lastmod": today},
        {"loc": f"{base}/menu",              "changefreq": "weekly",  "priority": "0.9", "lastmod": today},
        {"loc": f"{base}/offers",            "changefreq": "weekly",  "priority": "0.7", "lastmod": today},
        {"loc": f"{base}/events",            "changefreq": "monthly", "priority": "0.6", "lastmod": today},
        {"loc": f"{base}/about",             "changefreq": "monthly", "priority": "0.7", "lastmod": today},
        {"loc": f"{base}/contact",           "changefreq": "monthly", "priority": "0.7", "lastmod": today},
        {"loc": f"{base}/faq",               "changefreq": "weekly",  "priority": "0.7", "lastmod": today},
        {"loc": f"{base}/delivery",          "changefreq": "monthly", "priority": "0.6", "lastmod": today},
        {"loc": f"{base}/rewards-program",   "changefreq": "monthly", "priority": "0.6", "lastmod": today},
        {"loc": f"{base}/privacy",           "changefreq": "monthly", "priority": "0.5", "lastmod": today},
        {"loc": f"{base}/terms",             "changefreq": "monthly", "priority": "0.5", "lastmod": today},
        {"loc": f"{base}/refunds",           "changefreq": "monthly", "priority": "0.5", "lastmod": today},
        {"loc": f"{base}/ownership",         "changefreq": "monthly", "priority": "0.5", "lastmod": today},
        {"loc": f"{base}/feedback",          "changefreq": "monthly", "priority": "0.4", "lastmod": today},
        # /login and /register intentionally excluded — thin authenticated pages
        # that consume crawl budget without adding indexable content.
    ]
    # Offers and categories are discoverable from /offers and /menu respectively.
    # Fragment (#code, #slug) URLs are INVALID in sitemaps — the spec forbids
    # fragment identifiers; Google logs a parse error and discards the sitemap.
    # Removed the #fragment dynamic entries; the parent pages already cover them.

    body_parts = ['<?xml version="1.0" encoding="UTF-8"?>',
                  '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for u in urls:
        body_parts.append("  <url>")
        body_parts.append(f"    <loc>{u['loc']}</loc>")
        body_parts.append(f"    <lastmod>{u['lastmod']}</lastmod>")
        body_parts.append(f"    <changefreq>{u['changefreq']}</changefreq>")
        body_parts.append(f"    <priority>{u['priority']}</priority>")
        body_parts.append("  </url>")
    body_parts.append("</urlset>")
    xml = "\n".join(body_parts)
    return Response(content=xml, media_type="application/xml",
                    headers={"Cache-Control": "public, max-age=3600"})


@api_router.get("/robots.txt", response_class=Response)
async def robots_txt(request: Request):
    """robots.txt: open the site for search engines + AI crawlers (GPTBot,
    PerplexityBot, Google-Extended, ClaudeBot) and explicitly disallow the
    transactional / admin / customer-PII routes that should never appear in
    search results."""
    base = _abs_origin(request)
    body = "\n".join([
        "User-agent: *",
        "Allow: /",
        "Disallow: /admin",
        "Disallow: /admin/",
        "Disallow: /api/admin/",
        "Disallow: /checkout",
        "Disallow: /cart",
        "Disallow: /track/",
        "Disallow: /rider/",
        "Disallow: /order/",
        "Disallow: /profile",
        "Disallow: /orders",
        "",
        # Explicitly invite the major AI crawlers (some respect User-agent blocks
        # globally, so spelling them out lets us allow them even if we ever flip
        # the default-deny). Today we explicitly allow them everywhere.
        "User-agent: GPTBot",
        "Allow: /",
        "",
        "User-agent: ChatGPT-User",
        "Allow: /",
        "",
        "User-agent: PerplexityBot",
        "Allow: /",
        "",
        "User-agent: Google-Extended",
        "Allow: /",
        "",
        "User-agent: ClaudeBot",
        "Allow: /",
        "",
        f"Sitemap: {base}/sitemap.xml",
        "",
    ])
    return Response(content=body, media_type="text/plain",
                    headers={"Cache-Control": "public, max-age=3600"})


@api_router.get("/llms.txt", response_class=Response)
async def llms_txt(request: Request):
    """llms.txt — a Markdown briefing for AI assistants (ChatGPT, Perplexity,
    Claude, Gemini) following the https://llmstxt.org convention. When someone
    nearby asks an AI 'best biryani delivery in Lahore', this file gives the
    model clean, current facts (menu + prices, delivery areas, hours, contact)
    to recommend and cite us. Generated from live DB data so it never goes
    stale; served at the site root via a Vercel rewrite (see vercel.json)."""
    base = _abs_origin(request)
    s = await db.settings.find_one({"key": "global"}, {"_id": 0}) or {}
    name = s.get("restaurant_name") or DEFAULT_SETTINGS["restaurant_name"]
    address = s.get("restaurant_address") or DEFAULT_SETTINGS["restaurant_address"]
    phone = s.get("restaurant_phone") or DEFAULT_SETTINGS["restaurant_phone"]

    # Live opening hours from the admin-managed weekly schedule (the same data
    # that actually gates online ordering) — never a hardcoded copy that drifts.
    hours_line = "See website for opening hours"
    try:
        osett = await get_online_settings_doc()
        sched = osett.get("weekly_schedule") or {}
        spans = {}
        for day in DAY_KEYS:
            d = sched.get(day) or {}
            key = "Closed" if d.get("closed") else f"{d.get('open', '?')}–{d.get('close', '?')}"
            spans.setdefault(key, []).append(day.capitalize())
        if len(spans) == 1:
            only = next(iter(spans))
            hours_line = f"{only}, every day" if only != "Closed" else "Temporarily closed"
        else:
            hours_line = "; ".join(f"{'/'.join(days)}: {span}" for span, days in spans.items())
        if "?" in hours_line:  # malformed/missing schedule — don't publish garbage
            hours_line = "See website for opening hours"
    except Exception:
        pass

    lines = [
        f"# {name.title()}",
        "",
        f"> Authentic Karachi-style biryani, murgh pulao, BBQ and karahi in Lahore, "
        f"Pakistan — order online for delivery or pickup at {base}/menu. "
        "Cash on Delivery, live order tracking, loyalty rewards.",
        "",
        f"- Address: {address}",
        f"- Phone / WhatsApp: {phone}",
        f"- Hours (Pakistan time): {hours_line}",
        f"- Order online: {base}/menu",
        "- Payment: Cash on Delivery, cards and Pakistani wallets (SafePay), bank transfer",
        "",
    ]

    # Real customer rating — only stated when there are enough reviews to be
    # meaningful (mirrors the aggregateRating rule for schema.org markup).
    try:
        revs = await db.reviews.find({}, {"rating": 1}).to_list(1000)
        ratings = [r.get("rating") for r in revs if isinstance(r.get("rating"), (int, float))]
        if len(ratings) >= 3:
            avg = sum(ratings) / len(ratings)
            lines += [f"- Customer rating: {avg:.1f}/5 from {len(ratings)} reviews", ""]
    except Exception:
        pass

    # Menu snapshot: category -> a few dishes with PKR prices.
    try:
        cats = await db.categories.find({}).sort("sort_order", 1).to_list(50)
        items = await db.menu_items.find({}).sort("sort_order", 1).to_list(500)
        by_cat: dict = {}
        for i in items:
            by_cat.setdefault(str(i.get("category_id") or i.get("category") or ""), []).append(i)
        lines += ["## Menu highlights", ""]
        for c in cats[:12]:
            cat_items = by_cat.get(str(c["_id"]), []) or by_cat.get(c.get("name", ""), [])
            if not cat_items:
                continue
            dishes = ", ".join(
                f"{i.get('name')} (Rs {int(float(i.get('price') or 0))})"
                for i in cat_items[:5]
                if i.get("name") and float(i.get("price") or 0) > 0
            )
            if not dishes:
                continue
            lines.append(f"- **{c.get('name', 'Menu')}**: {dishes}")
        lines += ["", f"Full menu with photos and live prices: {base}/menu", ""]
    except Exception:
        pass

    # Delivery areas — the strongest "near me" signal for AI answers.
    try:
        areas = await db.delivery_areas.find({"enabled": True}).sort("sort_order", 1).to_list(100)
        area_names = [a.get("name") for a in areas if a.get("name")]
        if area_names:
            lines += ["## Delivery areas in Lahore", "",
                      ", ".join(area_names), "",
                      f"Details: {base}/delivery", ""]
    except Exception:
        pass

    # Top FAQs — direct answers AI assistants can quote.
    try:
        faqs = await db.faqs.find({"enabled": True}).sort("sort_order", 1).to_list(10)
        if faqs:
            lines += ["## FAQ", ""]
            for f in faqs:
                q = (f.get("question") or "").strip()
                a = " ".join((f.get("answer") or "").split())[:300]
                if q and a:
                    lines.append(f"- **{q}** {a}")
            lines += ["", f"More: {base}/faq", ""]
    except Exception:
        pass

    lines += [
        "## Key pages",
        "",
        f"- [Menu & online ordering]({base}/menu)",
        f"- [Current offers & deals]({base}/offers)",
        f"- [Delivery areas]({base}/delivery)",
        f"- [About us]({base}/about)",
        f"- [Contact]({base}/contact)",
        "",
    ]
    return Response(content="\n".join(lines), media_type="text/plain; charset=utf-8",
                    headers={"Cache-Control": "public, max-age=3600"})

# --- Event Bookings ---
@api_router.post("/event-bookings")
async def create_event_booking(req: EventBookingCreate, request: Request):
    cust = await get_optional_customer(request)
    now = datetime.now(timezone.utc)
    doc = {
        "name": req.name,
        "phone": req.phone,
        "email": req.email or "",
        "event_type": req.event_type,
        "guests": int(req.guests),
        "event_date": req.event_date,
        "message": req.message or "",
        "customer_id": cust["_id"] if cust else None,
        "status": "pending",
        "created_at": now.isoformat(),
    }
    result = await db.event_bookings.insert_one(doc)
    return {"id": str(result.inserted_id), "status": "pending", "message": "Booking received! We'll contact you shortly."}

@api_router.get("/event-bookings")
async def list_event_bookings(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    bookings = await db.event_bookings.find({}).sort("created_at", -1).to_list(500)
    return [{
        "id": str(b["_id"]),
        "name": b["name"],
        "phone": b["phone"],
        "email": b.get("email", ""),
        "event_type": b["event_type"],
        "guests": b["guests"],
        "event_date": b["event_date"],
        "message": b.get("message", ""),
        "status": b.get("status", "pending"),
        "created_at": b.get("created_at", ""),
    } for b in bookings]

@api_router.put("/event-bookings/{booking_id}/status")
async def update_event_booking_status(booking_id: str, body: dict, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    status = body.get("status", "pending")
    if status not in {"pending", "confirmed", "cancelled", "completed"}:
        raise HTTPException(status_code=400, detail="Invalid status")
    await db.event_bookings.update_one({"_id": ObjectId(booking_id)}, {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Status updated", "status": status}

# --- Seed sample data for online site ---
SAMPLE_MENU_DATA = {
    "Biryani": [
        {"name": "Chicken Biryani (Half)", "price": 350, "stock": 100, "image_url": "https://images.pexels.com/photos/23830980/pexels-photo-23830980.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940", "description": "Aromatic basmati rice with tender chicken, signature spices.", "is_popular": True},
        {"name": "Chicken Biryani (Full)", "price": 600, "stock": 100, "image_url": "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Family-size chicken biryani, serves 2-3.", "is_popular": True},
        {"name": "Beef Biryani", "price": 450, "stock": 80, "image_url": "https://images.pexels.com/photos/23830980/pexels-photo-23830980.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940", "description": "Slow-cooked beef biryani with caramelized onions.", "is_popular": False},
    ],
    "Murg Pulao": [
        {"name": "Murg Pulao (Half)", "price": 320, "stock": 100, "image_url": "https://images.unsplash.com/photo-1589302168068-964664d93dc0?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Karachi-style murg pulao, fragrant and rich.", "is_popular": True},
        {"name": "Murg Pulao (Full)", "price": 580, "stock": 100, "image_url": "https://images.unsplash.com/photo-1589302168068-964664d93dc0?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Family-size pulao, serves 2-3.", "is_popular": True},
    ],
    "BBQ & Grill": [
        {"name": "Mixed BBQ Platter", "price": 850, "stock": 60, "image_url": "https://images.unsplash.com/photo-1555939594-58d7cb561ad1?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Seekh kebab, chicken tikka, malai boti combo.", "is_popular": True},
        {"name": "Chicken Tikka", "price": 380, "stock": 80, "image_url": "https://images.unsplash.com/photo-1555939594-58d7cb561ad1?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Tandoori spiced chicken tikka.", "is_popular": False},
        {"name": "Seekh Kebab (4 pcs)", "price": 320, "stock": 100, "image_url": "https://images.unsplash.com/photo-1555939594-58d7cb561ad1?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Char-grilled minced beef kebabs.", "is_popular": False},
    ],
    "Sides & Drinks": [
        {"name": "Raita", "price": 60, "stock": 200, "image_url": "https://images.unsplash.com/photo-1694579740719-0e601c5d2437?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Cooling yogurt with cucumber and mint.", "is_popular": False},
        {"name": "Salad", "price": 80, "stock": 150, "image_url": "https://images.unsplash.com/photo-1694579740719-0e601c5d2437?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Fresh garden salad.", "is_popular": False},
        {"name": "Soft Drink (500ml)", "price": 90, "stock": 200, "image_url": "https://images.unsplash.com/photo-1694579740719-0e601c5d2437?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Coke, Sprite, or Fanta.", "is_popular": False},
        {"name": "Zarda (Sweet Rice)", "price": 200, "stock": 80, "image_url": "https://images.unsplash.com/photo-1694579740719-0e601c5d2437?crop=entropy&cs=srgb&fm=jpg&q=85", "description": "Traditional sweet saffron rice.", "is_popular": True},
    ],
}

SAMPLE_OFFERS = [
    {"title": "Family Feast — 15% OFF", "description": "Get 15% off on orders above Rs. 1500. Use code FAMILY15.", "discount_percent": 15, "discount_amount": 0, "coupon_code": "FAMILY15", "image_url": "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?crop=entropy&cs=srgb&fm=jpg&q=85", "active": True, "one_time_per_customer": False},
    {"title": "First Order — Rs. 100 OFF", "description": "Welcome offer! Flat Rs. 100 off on your first order. Code: WELCOME100.", "discount_percent": 0, "discount_amount": 100, "coupon_code": "WELCOME100", "image_url": "https://images.pexels.com/photos/23830980/pexels-photo-23830980.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940", "active": True, "one_time_per_customer": True},
    {"title": "Weekend BBQ Bonanza", "description": "Free Raita + Salad with any BBQ Platter on weekends.", "discount_percent": 0, "discount_amount": 0, "coupon_code": "", "image_url": "https://images.unsplash.com/photo-1555939594-58d7cb561ad1?crop=entropy&cs=srgb&fm=jpg&q=85", "active": True, "one_time_per_customer": False},
]

async def seed_online_data():
    # Seed menu only if empty
    if await db.menu_items.count_documents({}) == 0:
        sort_idx_cat = 0
        sort_idx_item = 0
        for cat_name, items in SAMPLE_MENU_DATA.items():
            cat_doc = {"name": cat_name, "color": "#D92D20", "sort_order": sort_idx_cat, "created_at": datetime.now(timezone.utc).isoformat()}
            cat_res = await db.categories.insert_one(cat_doc)
            cid = str(cat_res.inserted_id)
            sort_idx_cat += 1
            for it in items:
                await db.menu_items.insert_one({
                    "name": it["name"],
                    "price": it["price"],
                    "category_id": cid,
                    "stock": it["stock"],
                    "low_stock_threshold": 10,
                    "image_url": it["image_url"],
                    "description": it["description"],
                    "is_popular": it.get("is_popular", False),
                    "sort_order": sort_idx_item,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
                sort_idx_item += 1
        logger.info(f"Seeded {sort_idx_item} menu items in {sort_idx_cat} categories")
    # Seed offers only if empty
    if await db.offers.count_documents({}) == 0:
        for o in SAMPLE_OFFERS:
            await db.offers.insert_one({**o, "created_at": datetime.now(timezone.utc).isoformat()})
        logger.info(f"Seeded {len(SAMPLE_OFFERS)} offers")
    # Seed a starter list of Lahore delivery areas only if empty. Admin can then
    # edit / reorder / disable these under Admin → Delivery Areas; because we only
    # seed when the collection is empty, we never overwrite the operator's edits.
    if await db.delivery_areas.count_documents({}) == 0:
        _starter_areas = [
            ("Johar Town", "Free delivery · ~35 min"),
            ("Model Town", "Free delivery · ~35 min"),
            ("Faisal Town", "~30 min"),
            ("Garden Town", "~30 min"),
            ("Iqbal Town (Allama Iqbal Town)", "~35 min"),
            ("Township", "~40 min"),
            ("Gulberg", "~40 min"),
            ("Muslim Town", "~30 min"),
            ("Samanabad", "~40 min"),
            ("Wapda Town", "~40 min"),
            ("Valencia Town", "~45 min"),
            ("DHA (Defence)", "Delivery fee may apply"),
        ]
        now_iso = datetime.now(timezone.utc).isoformat()
        for idx, (name, note) in enumerate(_starter_areas):
            await db.delivery_areas.insert_one({
                "name": name,
                "note": note,
                "sort_order": idx,
                "enabled": True,
                "created_at": now_iso,
                "updated_at": now_iso,
            })
        logger.info(f"Seeded {len(_starter_areas)} delivery areas")
    # Seed dine-in tables only if empty (T1-T6 across two sections). Never
    # overwrites the operator's own floor layout (Admin → Tables / Floor).
    if await db.restaurant_tables.count_documents({}) == 0:
        default_tables = [
            {"name": "T1", "section": "Main Hall", "capacity": 2},
            {"name": "T2", "section": "Main Hall", "capacity": 4},
            {"name": "T3", "section": "Main Hall", "capacity": 4},
            {"name": "T4", "section": "Main Hall", "capacity": 6},
            {"name": "T5", "section": "Outdoor", "capacity": 4},
            {"name": "T6", "section": "Outdoor", "capacity": 2},
        ]
        for t in default_tables:
            await db.restaurant_tables.insert_one({**t, "status": "available", "created_at": datetime.now(timezone.utc).isoformat()})
        logger.info(f"Seeded {len(default_tables)} dine-in tables")
    # Backfill: any pre-existing offer whose code starts with WELCOME or FIRST is treated
    # as one-time-per-customer by default. Without this, the WELCOME100 code in production
    # could be reused indefinitely by the same customer — straight revenue leak.
    await db.offers.update_many(
        {
            "$or": [{"coupon_code": {"$regex": "^WELCOME", "$options": "i"}}, {"coupon_code": {"$regex": "^FIRST", "$options": "i"}}],
            "one_time_per_customer": {"$exists": False},
        },
        {"$set": {"one_time_per_customer": True}},
    )
    # Backfill: correct the restaurant location ONLY when it still holds the old
    # wrong default (31.4520, 74.2680) — ~14 km west of the real location.
    try:
        _loc = await db.online_settings.find_one({"key": "online"}, {"restaurant_lat": 1, "restaurant_lng": 1})
        if _loc:
            _lat = _loc.get("restaurant_lat")
            _lng = _loc.get("restaurant_lng")
            if (_lat is not None and _lng is not None
                    and abs(float(_lat) - 31.4520) < 1e-6
                    and abs(float(_lng) - 74.2680) < 1e-6):
                await db.online_settings.update_one(
                    {"key": "online"},
                    {"$set": {"restaurant_lat": 31.4761875, "restaurant_lng": 74.4163125}},
                )
                logger.info("Backfilled restaurant location from old default to real coords")
    except Exception as e:
        logger.warning(f"Restaurant location backfill skipped: {e}")
    # Indexes
    try:
        await db.customers.create_index("email", unique=True)
        await db.online_orders.create_index([("created_at", -1)])
        await db.online_orders.create_index("customer_id")
        await db.online_orders.create_index("status")
        await db.online_orders.create_index([("coupon_code", 1), ("customer_id", 1)])
        await db.online_orders.create_index([("coupon_code", 1), ("phone", 1)])
        await db.personal_coupons.create_index("code", unique=True)
        await db.personal_coupons.create_index([("customer_id", 1), ("used", 1)])
        await db.reviews.create_index([("created_at", -1)])
        await db.offers.create_index("coupon_code")
    except Exception as e:
        logger.warning(f"Online data index creation skipped: {e}")

@app.on_event("startup")
async def startup_online_seed():
    # Non-blocking: schedule online seeding so uvicorn binds 0.0.0.0:8080 immediately.
    asyncio.create_task(_startup_online_seed_background())

async def _startup_online_seed_background():
    try:
        await seed_online_data()
    except Exception as e:
        logger.exception(f"Online seed failed (server still listening): {e}")

# =============================================================================
# END CUSTOMER ENDPOINTS
# =============================================================================

# =============================================================================
# ONLINE SETTINGS, DELIVERY ZONES, AND PAYMENT INTEGRATION
# =============================================================================

import math
from emergentintegrations.payments.stripe.checkout import (
    StripeCheckout, CheckoutSessionRequest
)

DEFAULT_ONLINE_SETTINGS = {
    "restaurant_lat": 31.4761875,
    "restaurant_lng": 74.4163125,
    "delivery_free_radius_km": 2.0,
    "delivery_base_fee": 100.0,
    "delivery_per_km_fee": 15.0,
    "delivery_max_radius_km": 15.0,
    "free_delivery_min_subtotal": 0.0,  # 0 = disabled. Set e.g. 1500 for "free delivery on orders Rs.1500+"
    "payment_methods": {
        "cod": True,
        "pay_at_restaurant": True,
        "bank_transfer": True,
        "card": True,
    },
    "bank_account_title": "Karachi Naseeb Biryani",
    "bank_account_number": "0123456789012",
    "bank_name": "Meezan Bank",
    "iban": "PK00MEZN0000000000000000",
    "easypaisa_number": "03014117048",
    "easypaisa_account_title": "Karachi Naseeb",
    "jazzcash_number": "03014117048",
    "jazzcash_account_title": "Karachi Naseeb",
    # Business hours — Mon..Sun. closed=true blocks ordering for that day.
    # "open"/"close" are HH:MM strings in the configured timezone.
    "business_hours_enabled": True,
    "business_hours_timezone": "Asia/Karachi",
    "weekly_schedule": {
        "mon": {"open": "10:00", "close": "23:00", "closed": False},
        "tue": {"open": "10:00", "close": "23:00", "closed": False},
        "wed": {"open": "10:00", "close": "23:00", "closed": False},
        "thu": {"open": "10:00", "close": "23:00", "closed": False},
        "fri": {"open": "10:00", "close": "23:00", "closed": False},
        "sat": {"open": "10:00", "close": "23:00", "closed": False},
        "sun": {"open": "10:00", "close": "23:00", "closed": False},
    },
}

async def get_online_settings_doc():
    s = await db.online_settings.find_one({"key": "online"}, {"_id": 0})
    if not s:
        return DEFAULT_ONLINE_SETTINGS.copy()
    merged = DEFAULT_ONLINE_SETTINGS.copy()
    for k, v in s.items():
        if k != "key":
            merged[k] = v
    return merged

# =============================================================================
# PAYMENT GATEWAY SETTINGS (DB-configured, admin-managed)
# =============================================================================
# Configuration for the hosted wallet gateways (EasyPaisa, JazzCash) lives in
# the db.payment_gateway_settings singleton doc {"key": "gateways"} and is
# managed from the admin "Payment Gateways" page. The DB value wins; env vars
# are a per-field fallback so existing env-driven deploys keep working.
#
# PAYFAST: the live PayFast integration (further down this file) still reads
# ONLY its env vars (PAYFAST_MERCHANT_ID / PAYFAST_SECURED_KEY / PAYFAST_ENV).
# The "payfast" subdoc here is saved by the admin page but is INERT — it
# pre-stages the switch to DB config. To migrate PayFast onto this framework
# later: (1) add a PayfastGateway driver reading get_gateway_config("payfast")
# (env fallback below already maps its fields), (2) register it in
# GATEWAY_REGISTRY, (3) change "payfast_enabled" in /public/settings to
# `await gateway_ready("payfast")`, (4) retire the literal /payments/payfast/*
# routes only after all clients use the generic /payments/{gateway}/* paths.

DEFAULT_GATEWAY_SETTINGS = {
    "easypaisa": {
        "enabled": False,
        "mode": "sandbox",
        "store_id": "",
        "hash_key": "",           # 16-char AES-128 key from the Easypay portal
        # Optional REST order-inquiry credentials. When set, a successful
        # postback is confirmed server-to-server and the order goes straight
        # to "paid"; without them, successes land in "pending_verification"
        # for manual admin approval (Easypay's postback is unsigned).
        "inquiry_username": "",
        "inquiry_password": "",
    },
    "jazzcash": {
        "enabled": False,
        "mode": "sandbox",
        "merchant_id": "",
        "password": "",
        "integrity_salt": "",
    },
    # Inert until PayFast joins GATEWAY_REGISTRY — see migration note above.
    "payfast": {
        "enabled": False,
        "mode": "sandbox",
        "merchant_id": "",
        "secured_key": "",
    },
    # SafePay is redirect-based (no signed callback to pre-register). The
    # existing /payments/safepay/* endpoints read this config (env fallback
    # preserved); "enabled" additionally gates the WEBSITE checkout option.
    "safepay": {
        "enabled": False,
        "mode": "sandbox",
        "api_key": "",
        "secret": "",
    },
}

# DB field -> env var fallback (DB non-empty wins).
_GATEWAY_ENV_FALLBACK = {
    "easypaisa": {
        "store_id": "EASYPAISA_STORE_ID",
        "hash_key": "EASYPAISA_HASH_KEY",
        "inquiry_username": "EASYPAISA_INQUIRY_USERNAME",
        "inquiry_password": "EASYPAISA_INQUIRY_PASSWORD",
    },
    "jazzcash": {
        "merchant_id": "JAZZCASH_MERCHANT_ID",
        "password": "JAZZCASH_PASSWORD",
        "integrity_salt": "JAZZCASH_INTEGRITY_SALT",
    },
    "payfast": {
        "merchant_id": "PAYFAST_MERCHANT_ID",
        "secured_key": "PAYFAST_SECURED_KEY",
    },
    "safepay": {
        "api_key": "SAFEPAY_API_KEY",
        "secret": "SAFEPAY_SECRET",
    },
}

# Credentials without which a gateway cannot take payments.
_GATEWAY_REQUIRED = {
    "easypaisa": ("store_id", "hash_key"),
    "jazzcash": ("merchant_id", "password", "integrity_salt"),
    "payfast": ("merchant_id", "secured_key"),
    "safepay": ("api_key",),
}

_GATEWAY_SECRET_FIELDS = {"hash_key", "inquiry_password", "password", "integrity_salt", "secured_key", "api_key", "secret"}

async def get_gateway_settings_doc() -> dict:
    stored = await db.payment_gateway_settings.find_one({"key": "gateways"}, {"_id": 0}) or {}
    merged = {}
    for gw, defaults in DEFAULT_GATEWAY_SETTINGS.items():
        sub = defaults.copy()
        for k, v in (stored.get(gw) or {}).items():
            if k in sub:
                sub[k] = v
        merged[gw] = sub
    return merged

async def get_gateway_config(name: str) -> dict:
    """Effective config for one gateway: stored values with per-field env
    fallback applied (a non-empty DB value wins over the env var)."""
    cfg = (await get_gateway_settings_doc()).get(name, {}).copy()
    for field, env_name in _GATEWAY_ENV_FALLBACK.get(name, {}).items():
        if not (cfg.get(field) or "").strip():
            cfg[field] = (os.environ.get(env_name) or "").strip()
    return cfg

async def gateway_ready(name: str) -> bool:
    cfg = await get_gateway_config(name)
    return bool(cfg.get("enabled")) and all(
        (cfg.get(f) or "").strip() for f in _GATEWAY_REQUIRED.get(name, ())
    )

def _api_public_base(request: Request) -> str:
    """Public origin of THIS backend, used to build gateway callback URLs.
    Prod (Vercel frontend + Fly backend) must set PUBLIC_API_BASE, e.g.
    https://knb-backend.fly.dev; locally the request host is correct."""
    env = (os.environ.get("PUBLIC_API_BASE") or "").strip().rstrip("/")
    if env:
        return env
    host = request.headers.get("host", "localhost:8001")
    scheme = "http" if host.split(":")[0] in ("localhost", "127.0.0.1") else "https"
    return f"{scheme}://{host}"

class GatewayConfigUpdate(BaseModel):
    enabled: Optional[bool] = None
    mode: Optional[str] = None            # "sandbox" | "live"
    store_id: Optional[str] = None        # easypaisa
    hash_key: Optional[str] = None        # easypaisa secret (16 chars = AES-128 key)
    inquiry_username: Optional[str] = None
    inquiry_password: Optional[str] = None  # secret
    merchant_id: Optional[str] = None     # jazzcash + payfast
    password: Optional[str] = None        # jazzcash secret
    integrity_salt: Optional[str] = None  # jazzcash secret
    secured_key: Optional[str] = None     # payfast secret
    api_key: Optional[str] = None         # safepay secret
    secret: Optional[str] = None          # safepay optional secret

class PaymentGatewaysUpdate(BaseModel):
    easypaisa: Optional[GatewayConfigUpdate] = None
    jazzcash: Optional[GatewayConfigUpdate] = None
    payfast: Optional[GatewayConfigUpdate] = None
    safepay: Optional[GatewayConfigUpdate] = None

def _masked_gateway_view(settings: dict, request: Request) -> dict:
    """Admin-facing view: secrets are reduced to <field>_set + <field>_last4."""
    out = {}
    for gw, cfg in settings.items():
        view = {}
        for k, v in cfg.items():
            if k in _GATEWAY_SECRET_FIELDS:
                sval = (v or "").strip()
                view[f"{k}_set"] = bool(sval)
                view[f"{k}_last4"] = sval[-4:] if sval else ""
            else:
                view[k] = v
        if gw in ("easypaisa", "jazzcash"):
            view["callback_url"] = f"{_api_public_base(request)}/api/payments/{gw}/return"
        out[gw] = view
    out["payfast"]["note"] = (
        "PayFast currently runs from server environment variables. These saved "
        "credentials are not used yet — they pre-stage the switch to database "
        "configuration."
    )
    return out

@api_router.get("/admin/payment-gateways")
async def admin_get_payment_gateways(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return _masked_gateway_view(await get_gateway_settings_doc(), request)

@api_router.put("/admin/payment-gateways")
async def admin_update_payment_gateways(req: PaymentGatewaysUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    current = await get_gateway_settings_doc()
    updates = {}
    for gw in ("easypaisa", "jazzcash", "payfast", "safepay"):
        upd = getattr(req, gw)
        if upd is None:
            continue
        merged = current[gw].copy()
        for field, value in upd.model_dump().items():
            if value is None or field not in DEFAULT_GATEWAY_SETTINGS[gw]:
                continue
            if isinstance(value, str):
                value = value.strip()[:128]
                # Blank secret means "keep the stored one" — changing a secret
                # requires typing a new value.
                if field in _GATEWAY_SECRET_FIELDS and not value:
                    continue
            merged[field] = value
        if merged["mode"] not in ("sandbox", "live"):
            raise HTTPException(status_code=400, detail=f"{gw}: mode must be 'sandbox' or 'live'")
        if gw == "easypaisa" and merged["hash_key"] and len(merged["hash_key"]) != 16:
            raise HTTPException(
                status_code=400,
                detail="EasyPaisa Hash Key must be exactly 16 characters (AES-128)")
        if merged.get("enabled"):
            # Validate against the EFFECTIVE config (env fallback counts).
            effective = merged.copy()
            for field, env_name in _GATEWAY_ENV_FALLBACK.get(gw, {}).items():
                if not (effective.get(field) or "").strip():
                    effective[field] = (os.environ.get(env_name) or "").strip()
            missing = [f for f in _GATEWAY_REQUIRED[gw]
                       if not (effective.get(f) or "").strip()]
            if missing:
                raise HTTPException(
                    status_code=400,
                    detail=f"{gw}: enable requires credentials: {', '.join(missing)}")
        updates[gw] = merged
    if updates:
        await db.payment_gateway_settings.update_one(
            {"key": "gateways"}, {"$set": updates}, upsert=True)
    return _masked_gateway_view(await get_gateway_settings_doc(), request)

# ===== END PAYMENT GATEWAY SETTINGS ==========================================

# Payment methods where the customer pays ONLINE through a gateway before the
# restaurant should see the order. Orders placed with these start with status
# "awaiting_payment" (kept out of the staff pending queue, alert sound and
# printer feed) and are released to "pending" only when the payment confirms
# (paid or pending_verification). COD / pay-at-restaurant / manual bank
# transfer are unaffected. "card" = Stripe.
GATEWAY_PAYMENT_METHODS = {"easypaisa", "jazzcash", "safepay", "card"}

async def _release_online_order(order_id) -> None:
    """Flip an awaiting_payment order to pending so it enters the staff queue
    (alert + printer). Idempotent — a released/accepted order is untouched."""
    try:
        oid = ObjectId(order_id) if not isinstance(order_id, ObjectId) else order_id
        res = await db.online_orders.update_one(
            {"_id": oid, "status": "awaiting_payment"},
            {"$set": {"status": "pending",
                      "released_at": datetime.now(timezone.utc).isoformat()}},
        )
        if res.modified_count:
            logger.info(f"Order {order_id} released to restaurant queue after payment")
            # Gateway-paid order just entered the staff queue — ring admin
            # devices via push (same alert COD orders get at creation).
            try:
                released = await db.online_orders.find_one({"_id": oid}, {"total_price": 1})
                asyncio.create_task(_notify_admins_new_order(released or {"_id": oid}))
            except Exception as e:
                logger.warning(f"Admin release push failed: {e}")
    except Exception as e:
        logger.error(f"Failed to release order {order_id}: {e}")

async def _restore_order_wallet(o: dict) -> None:
    """Give back wallet credit spent on an order that will never be fulfilled
    (rejected / cancelled / abandoned checkout). Idempotent via the
    wallet_restored flag, so multiple cancel paths can call it safely."""
    try:
        amt = float(o.get("wallet_applied") or 0)
        cid = o.get("customer_id")
        if amt <= 0 or not cid or o.get("wallet_restored"):
            return
        res = await db.online_orders.update_one(
            {"_id": o["_id"], "wallet_restored": {"$ne": True}},
            {"$set": {"wallet_restored": True}})
        if not res.modified_count:
            return  # another path already restored it
        await db.customers.update_one({"_id": ObjectId(str(cid))}, {"$inc": {"wallet_balance": amt}})
        await db.wallet_transactions.insert_one({
            "customer_id": str(cid), "type": "restore", "amount": amt,
            "order_id": str(o["_id"]), "note": "Order not fulfilled — credit returned",
            "created_at": datetime.now(timezone.utc).isoformat()})
    except Exception as e:
        logger.error(f"wallet restore failed for order {o.get('_id')}: {e}")

async def _expire_abandoned_gateway_orders() -> None:
    """Cancel awaiting_payment orders whose gateway checkout was abandoned
    (no payment confirmation within PAYMENT_ABANDON_MINUTES, default 30).
    Keeps the DB clean and frees any redeemed coupon/reward pressure; the
    customer was never charged (a late confirmation would find the order
    cancelled and land in pending_verification for the admin to restore)."""
    minutes = int(os.environ.get("PAYMENT_ABANDON_MINUTES", "30"))
    cutoff = (datetime.now(timezone.utc) - timedelta(minutes=minutes)).isoformat()
    try:
        q = {"status": "awaiting_payment", "created_at": {"$lt": cutoff},
             "payment_status": {"$nin": ["paid", "pending_verification"]}}
        # Give back any wallet credit applied to these doomed orders BEFORE
        # cancelling them — abandoned checkouts must never eat store credit.
        async for wo in db.online_orders.find({**q, "wallet_applied": {"$gt": 0}}):
            await _restore_order_wallet(wo)
        res = await db.online_orders.update_many(
            q,
            {"$set": {"status": "cancelled",
                      "cancel_reason": "payment_not_completed",
                      "cancelled_at": datetime.now(timezone.utc).isoformat()}},
        )
        if res.modified_count:
            logger.info(f"Expired {res.modified_count} abandoned gateway order(s)")
    except Exception as e:
        logger.error(f"Abandoned-order expiry failed: {e}")

async def _reconcile_safepay_payments() -> None:
    """Server-side confirmation safety net for SafePay. The hosted checkout
    confirms via the browser redirect -> result-page polling, but if the
    customer closes the browser before redirecting (SafePay's mobile-mode
    interstitial even tells them to), nothing would ever flip the order to
    paid. This job polls SafePay for recent initiated transactions so paid
    orders reach the restaurant within a couple of minutes regardless."""
    cfg = await _safepay_cfg()
    api_key = (cfg.get("api_key") or "").strip()
    if not api_key:
        return
    api_base, _ = _safepay_bases_for(cfg.get("mode", "sandbox"))
    lookback = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()
    try:
        txns = await db.payment_transactions.find({
            "gateway": "safepay", "payment_status": "initiated",
            "created_at": {"$gt": lookback},
        }).to_list(50)
    except Exception as e:
        logger.error(f"SafePay reconcile query failed: {e}")
        return
    for txn in txns:
        tracker = txn.get("tracker")
        if not tracker:
            continue
        try:
            async with httpx.AsyncClient(timeout=15) as client:
                r = await client.get(f"{api_base}/order/v1/{tracker}")
            if r.status_code >= 400:
                continue
            data = (r.json() or {}).get("data") or {}
            state = str(data.get("state") or data.get("status") or "").lower()
            if state not in ("paid", "tracker_ended", "completed", "succeeded"):
                continue
            now = datetime.now(timezone.utc).isoformat()
            await db.payment_transactions.update_one(
                {"_id": txn["_id"], "payment_status": {"$ne": "paid"}},
                {"$set": {"payment_status": "paid", "paid_at": now,
                          "confirmed_by": "reconcile_job"}},
            )
            order_id = txn.get("order_id")
            if order_id:
                await db.online_orders.update_one(
                    {"_id": ObjectId(order_id), "payment_status": {"$ne": "paid"}},
                    {"$set": {"payment_status": "paid", "paid_at": now}},
                )
                await _release_online_order(order_id)
                logger.info(f"SafePay reconcile: order {order_id} confirmed paid ({tracker})")
        except Exception as e:
            logger.warning(f"SafePay reconcile error for {tracker}: {e}")

DAY_KEYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]

def _parse_hhmm(value: str, default_minutes: int) -> int:
    """Parse a time string into minutes since midnight.

    Accepts:
      - 24-hour: "13:00", "9:30", "23", "13"
      - 12-hour: "1:00 PM", "01:00 pm", "1 PM", "11pm", "12:30 AM"
    Returns default_minutes on parse failure.
    """
    try:
        if value is None:
            return default_minutes
        s = str(value).strip().lower()
        if not s:
            return default_minutes
        ampm = None
        if s.endswith("am") or s.endswith("pm"):
            ampm = s[-2:]
            s = s[:-2].strip()
        parts = s.replace(".", ":").split(":")
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 and parts[1] != "" else 0
        if ampm == "am":
            if h == 12:
                h = 0
        elif ampm == "pm":
            if h < 12:
                h += 12
        total = h * 60 + m
        return max(0, min(24 * 60, total))
    except Exception:
        return default_minutes


def _fmt_12h(minutes: int) -> str:
    """Format minutes-since-midnight as a 12-hour string like '1:00 PM'."""
    minutes = max(0, min(24 * 60 - 1, int(minutes)))
    h, m = divmod(minutes, 60)
    suffix = "AM" if h < 12 else "PM"
    hh = h % 12
    if hh == 0:
        hh = 12
    return f"{hh}:{m:02d} {suffix}"

def compute_business_hours_status(settings: dict) -> dict:
    """Return current open/closed status using the configured weekly_schedule + timezone.
    Falls back to "always open" when business_hours_enabled is False.
    Supports overnight wrap (e.g. open 22:00 close 02:00) when close <= open."""
    enabled = settings.get("business_hours_enabled", True)
    tz_name = settings.get("business_hours_timezone", "Asia/Karachi") or "Asia/Karachi"
    schedule = settings.get("weekly_schedule") or DEFAULT_ONLINE_SETTINGS["weekly_schedule"]
    try:
        tz = pytz.timezone(tz_name)
    except Exception:
        tz = pytz.timezone("Asia/Karachi")
    now_local = datetime.now(tz)
    today_key = DAY_KEYS[now_local.weekday()]
    current_minutes = now_local.hour * 60 + now_local.minute
    today = schedule.get(today_key) or {}
    if not enabled:
        return {
            "is_open": True,
            "enabled": False,
            "timezone": tz_name,
            "today": {"day": today_key, **today},
            "now": now_local.strftime("%H:%M"),
            "next_open_at": None,
            "next_close_at": None,
            "weekly_schedule": schedule,
        }
    is_open = False
    next_close_at = None
    next_open_at = None
    # Check yesterday's window for overnight-wrap (e.g. open 22:00, close 02:00)
    yesterday_key = DAY_KEYS[(now_local.weekday() - 1) % 7]
    y = schedule.get(yesterday_key) or {}
    if not y.get("closed", False):
        y_open = _parse_hhmm(y.get("open"), 10 * 60)
        y_close = _parse_hhmm(y.get("close"), 23 * 60)
        if y_close <= y_open:  # wrapped overnight
            if current_minutes < y_close:
                is_open = True
                next_close_at = y.get("close")
    if not is_open and not today.get("closed", False):
        open_min = _parse_hhmm(today.get("open"), 10 * 60)
        close_min = _parse_hhmm(today.get("close"), 23 * 60)
        if close_min > open_min:
            # Same-day window
            if open_min <= current_minutes < close_min:
                is_open = True
                next_close_at = today.get("close", "23:00")
            elif current_minutes < open_min:
                next_open_at = today.get("open", "10:00")
        elif close_min <= open_min:
            # Overnight window starting today (close is tomorrow)
            if current_minutes >= open_min:
                is_open = True
                next_close_at = today.get("close", "23:00")
            elif current_minutes < open_min:
                next_open_at = today.get("open", "10:00")
    if not is_open and next_open_at is None:
        # Look ahead up to 7 days for the next open day
        for offset in range(1, 8):
            day_idx = (now_local.weekday() + offset) % 7
            d = schedule.get(DAY_KEYS[day_idx]) or {}
            if d.get("closed", False):
                continue
            next_dt = (now_local + timedelta(days=offset)).replace(hour=0, minute=0, second=0, microsecond=0)
            open_min = _parse_hhmm(d.get("open"), 10 * 60)
            next_dt = next_dt + timedelta(minutes=open_min)
            next_open_at = next_dt.isoformat()
            break
    # Derive a friendly 12-hour display for next_open_at when it's a simple HH:MM
    next_open_display = None
    if next_open_at and ":" in str(next_open_at) and "T" not in str(next_open_at):
        next_open_display = _fmt_12h(_parse_hhmm(next_open_at, 0))
    next_close_display = None
    if next_close_at:
        next_close_display = _fmt_12h(_parse_hhmm(next_close_at, 0))
    return {
        "is_open": is_open,
        "enabled": True,
        "timezone": tz_name,
        "today": {"day": today_key, **today},
        "now": now_local.strftime("%H:%M"),
        "now_display": _fmt_12h(current_minutes),
        "next_open_at": next_open_at,
        "next_open_display": next_open_display,
        "next_close_at": next_close_at,
        "next_close_display": next_close_display,
        "weekly_schedule": schedule,
    }

def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0  # Earth radius in km
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

def calculate_delivery_fee(distance_km: float, settings: dict, subtotal: float = 0.0) -> dict:
    free_radius = float(settings.get("delivery_free_radius_km", 2.0))
    base_fee = float(settings.get("delivery_base_fee", 100.0))
    per_km = float(settings.get("delivery_per_km_fee", 15.0))
    max_radius = float(settings.get("delivery_max_radius_km", 15.0))
    free_min_subtotal = float(settings.get("free_delivery_min_subtotal", 0) or 0)
    in_range = distance_km <= max_radius
    if not in_range:
        return {"distance_km": round(distance_km, 2), "fee": 0, "in_range": False, "free_delivery": False, "max_radius_km": max_radius, "free_delivery_min_subtotal": free_min_subtotal}
    if distance_km <= free_radius:
        return {"distance_km": round(distance_km, 2), "fee": 0, "in_range": True, "free_delivery": True, "max_radius_km": max_radius, "free_delivery_min_subtotal": free_min_subtotal, "free_delivery_reason": "in-free-radius"}
    if free_min_subtotal > 0 and subtotal >= free_min_subtotal:
        return {"distance_km": round(distance_km, 2), "fee": 0, "in_range": True, "free_delivery": True, "max_radius_km": max_radius, "free_delivery_min_subtotal": free_min_subtotal, "free_delivery_reason": "subtotal-threshold"}
    extra_km = distance_km - free_radius
    fee = base_fee + (extra_km * per_km)
    return {"distance_km": round(distance_km, 2), "fee": round(fee, 0), "in_range": True, "free_delivery": False, "max_radius_km": max_radius, "free_delivery_min_subtotal": free_min_subtotal}

@api_router.get("/public/business-hours")
async def get_business_hours():
    """Public: current open/closed state + weekly schedule for the customer site."""
    s = await get_online_settings_doc()
    return compute_business_hours_status(s)


@api_router.get("/public/settings")
async def get_public_settings():
    s = await get_online_settings_doc()
    return {
        "restaurant_lat": s["restaurant_lat"],
        "restaurant_lng": s["restaurant_lng"],
        "delivery_free_radius_km": s["delivery_free_radius_km"],
        "delivery_base_fee": s["delivery_base_fee"],
        "delivery_per_km_fee": s["delivery_per_km_fee"],
        "delivery_max_radius_km": s["delivery_max_radius_km"],
        "free_delivery_min_subtotal": float(s.get("free_delivery_min_subtotal", 0) or 0),
        "payment_methods": s["payment_methods"],
        # PayFast (wallet gateway) readiness — lets clients route Easypaisa/
        # JazzCash to the hosted checkout instead of the manual transfer flow.
        # Purely additive: old clients ignore it.
        "payfast_enabled": payfast_configured(),
        # Hosted wallet gateways (DB-configured via the admin Payment Gateways
        # page). True only when enabled AND credentialed — checkout shows the
        # option only then. Additive: old clients ignore these.
        "easypaisa_gateway_enabled": await gateway_ready("easypaisa"),
        "jazzcash_gateway_enabled": await gateway_ready("jazzcash"),
        "safepay_gateway_enabled": await gateway_ready("safepay"),
        "bank_account_title": s["bank_account_title"],
        "bank_account_number": s["bank_account_number"],
        "bank_name": s["bank_name"],
        "iban": s.get("iban", ""),
        "easypaisa_number": s["easypaisa_number"],
        "easypaisa_account_title": s.get("easypaisa_account_title", ""),
        "jazzcash_number": s["jazzcash_number"],
        "jazzcash_account_title": s.get("jazzcash_account_title", ""),
    }

class OnlineSettingsUpdate(BaseModel):
    # Restaurant Info
    restaurant_name: Optional[str] = None
    restaurant_phone: Optional[str] = None
    restaurant_whatsapp: Optional[str] = None
    restaurant_email: Optional[str] = None
    restaurant_address: Optional[str] = None
    restaurant_logo_url: Optional[str] = None
    # Social Links
    facebook_url: Optional[str] = None
    instagram_url: Optional[str] = None
    twitter_url: Optional[str] = None
    # Opening Hours
    opening_hours: Optional[str] = None  # e.g., "Mon-Sun: 10AM - 11PM"
    # Invoice/Receipt
    invoice_footer_text: Optional[str] = None
    # Location
    restaurant_lat: Optional[float] = None
    restaurant_lng: Optional[float] = None
    google_maps_url: Optional[str] = None
    # Delivery Settings
    delivery_free_radius_km: Optional[float] = None
    delivery_base_fee: Optional[float] = None
    delivery_per_km_fee: Optional[float] = None
    delivery_max_radius_km: Optional[float] = None
    free_delivery_min_subtotal: Optional[float] = None  # 0 = disabled. e.g. 500 means: free delivery if cart >= Rs.500
    # Payment Methods
    payment_methods: Optional[Dict[str, bool]] = None
    bank_account_title: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_name: Optional[str] = None
    iban: Optional[str] = None
    easypaisa_number: Optional[str] = None
    easypaisa_account_title: Optional[str] = None
    jazzcash_number: Optional[str] = None
    jazzcash_account_title: Optional[str] = None
    twilio_whatsapp_from: Optional[str] = None
    # Business hours
    business_hours_enabled: Optional[bool] = None
    business_hours_timezone: Optional[str] = None
    weekly_schedule: Optional[Dict[str, Dict[str, Any]]] = None

@api_router.get("/admin/online-settings")
async def admin_get_online_settings(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return await get_online_settings_doc()

@api_router.put("/admin/online-settings")
async def admin_update_online_settings(req: OnlineSettingsUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    ud = {k: v for k, v in req.model_dump().items() if v is not None}
    if ud:
        await db.online_settings.update_one({"key": "online"}, {"$set": ud}, upsert=True)
    return await get_online_settings_doc()

class DeliveryQuoteRequest(BaseModel):
    lat: float
    lng: float
    subtotal: Optional[float] = 0.0

@api_router.post("/delivery/quote")
async def delivery_quote(req: DeliveryQuoteRequest):
    s = await get_online_settings_doc()
    distance = haversine_km(s["restaurant_lat"], s["restaurant_lng"], req.lat, req.lng)
    return calculate_delivery_fee(distance, s, subtotal=float(req.subtotal or 0))

# --- Stripe Payment Integration ---
class StripeSessionRequest(BaseModel):
    order_id: str
    origin_url: str

@api_router.post("/payments/stripe/create-session")
async def create_stripe_session(req: StripeSessionRequest, http_request: Request):
    """Create Stripe checkout session for an existing order. Amount comes from server-side order data."""
    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="Stripe not configured")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(req.order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Order already paid")
    # Server-side amount (no frontend manipulation)
    amount = float(order.get("total_price", 0))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Invalid order amount")
    origin = req.origin_url.rstrip("/")
    success_url = f"{origin}/payment/success?session_id={{CHECKOUT_SESSION_ID}}&order_id={req.order_id}"
    cancel_url = f"{origin}/payment/cancel?order_id={req.order_id}"
    host_url = str(http_request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    stripe = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
    metadata = {"order_id": req.order_id, "receipt_no": str(order["_id"])[-6:].upper()}
    # Convert PKR amount to USD (approx 280 PKR = 1 USD) for Stripe test mode
    # In production, Stripe supports PKR directly with proper account
    usd_amount = round(amount / 280.0, 2)
    if usd_amount < 0.50:
        usd_amount = 0.50  # Stripe minimum
    sess_req = CheckoutSessionRequest(
        amount=usd_amount,
        currency="usd",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata=metadata,
    )
    session = await stripe.create_checkout_session(sess_req)
    # Create payment_transactions record
    await db.payment_transactions.insert_one({
        "session_id": session.session_id,
        "order_id": req.order_id,
        "amount_pkr": amount,
        "amount_usd": usd_amount,
        "currency": "usd",
        "metadata": metadata,
        "payment_status": "initiated",
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"url": session.url, "session_id": session.session_id}

@api_router.get("/payments/stripe/status/{session_id}")
async def stripe_status(session_id: str, http_request: Request):
    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="Stripe not configured")
    txn = await db.payment_transactions.find_one({"session_id": session_id}, {"_id": 0})
    host_url = str(http_request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    stripe = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
    try:
        status = await stripe.get_checkout_status(session_id)
    except Exception as e:
        logger.warning(f"Stripe status retrieval failed: {e}")
        # Fall back to DB record
        if txn:
            return {
                "session_id": session_id,
                "status": "unknown",
                "payment_status": txn.get("payment_status", "unpaid"),
                "amount_total": int(round(float(txn.get("amount_usd", 0)) * 100)),
                "currency": txn.get("currency", "usd"),
                "order_id": txn.get("order_id"),
            }
        return {"session_id": session_id, "status": "unknown", "payment_status": "unpaid", "amount_total": 0, "currency": "usd", "order_id": None}
    if txn and txn.get("payment_status") != "paid" and status.payment_status == "paid":
        # Idempotent update
        await db.payment_transactions.update_one(
            {"session_id": session_id, "payment_status": {"$ne": "paid"}},
            {"$set": {"payment_status": "paid", "paid_at": datetime.now(timezone.utc).isoformat(), "stripe_status": status.status}}
        )
        order_id = (txn or {}).get("order_id")
        if order_id:
            await db.online_orders.update_one(
                {"_id": ObjectId(order_id), "payment_status": {"$ne": "paid"}},
                {"$set": {"payment_status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
            )
            await _release_online_order(order_id)
    return {
        "session_id": session_id,
        "status": status.status,
        "payment_status": status.payment_status,
        "amount_total": status.amount_total,
        "currency": status.currency,
        "order_id": (txn or {}).get("order_id"),
    }

@app.post("/api/webhook/stripe")
async def stripe_webhook(request: Request):
    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        return {"received": False}
    host_url = str(request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    stripe = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
    body = await request.body()
    sig = request.headers.get("Stripe-Signature", "")
    try:
        event = await stripe.handle_webhook(body, sig)
    except Exception as e:
        logger.warning(f"Stripe webhook verify failed: {e}")
        return {"received": False}
    if event.payment_status == "paid":
        await db.payment_transactions.update_one(
            {"session_id": event.session_id, "payment_status": {"$ne": "paid"}},
            {"$set": {"payment_status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
        )
        order_id = (event.metadata or {}).get("order_id")
        if order_id:
            await db.online_orders.update_one(
                {"_id": ObjectId(order_id), "payment_status": {"$ne": "paid"}},
                {"$set": {"payment_status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
            )
            await _release_online_order(order_id)
    return {"received": True}

# --- SafePay (Pakistan) hosted-checkout gateway ----------------------------
# Config comes from the admin Payment Gateways page (db-backed) with env-var
# fallback (SAFEPAY_API_KEY, optional SAFEPAY_SECRET) via get_gateway_config
# ("safepay"). Mode: sandbox|live (legacy SAFEPAY_ENV values still accepted
# through the fallback). Without an api key both endpoints return 503 and
# clients hide the option. Mirrors the Stripe flow: server-side amount,
# a payment_transactions record keyed by tracker, and an idempotent order->paid
# update once SafePay confirms the charge.
#   Docs / dashboard: https://sandbox.api.getsafepay.com  (sandbox)
class SafepaySessionRequest(BaseModel):
    order_id: str
    origin_url: str

def _safepay_bases_for(mode: str) -> tuple[str, str]:
    """Return (api_base, checkout_base) for the configured environment."""
    if (mode or "").strip().lower() in ("production", "prod", "live"):
        return "https://api.getsafepay.com", "https://getsafepay.com"
    return "https://sandbox.api.getsafepay.com", "https://sandbox.api.getsafepay.com"

async def _safepay_cfg() -> dict:
    """Effective SafePay config (DB wins, env fallback). Legacy SAFEPAY_ENV
    overrides mode only when the DB doc has never been saved."""
    cfg = await get_gateway_config("safepay")
    stored = await db.payment_gateway_settings.find_one(
        {"key": "gateways"}, {"safepay": 1})
    if not (stored or {}).get("safepay"):
        env = (os.environ.get("SAFEPAY_ENV") or "").strip().lower()
        if env:
            cfg["mode"] = "live" if env in ("production", "prod", "live") else "sandbox"
    return cfg

@api_router.post("/payments/safepay/create-session")
async def create_safepay_session(req: SafepaySessionRequest, http_request: Request):
    """Create a SafePay tracker for an existing order and return the hosted
    checkout URL. Amount is taken from server-side order data (never the client)."""
    cfg = await _safepay_cfg()
    api_key = (cfg.get("api_key") or "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="SafePay not configured")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(req.order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Order already paid")
    amount = float(order.get("total_price", 0))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Invalid order amount")

    api_base, checkout_base = _safepay_bases_for(cfg.get("mode", "sandbox"))
    env = "production" if cfg.get("mode") == "live" else "sandbox"
    # 1) Create a payment tracker (SafePay v1 init). PKR amount is in rupees.
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            init = await client.post(
                f"{api_base}/order/v1/init",
                json={
                    "client": api_key,
                    "amount": round(amount, 2),
                    "currency": "PKR",
                    "environment": env,
                },
            )
        if init.status_code >= 400:
            logger.warning(f"SafePay init failed {init.status_code}: {init.text[:300]}")
            raise HTTPException(status_code=502, detail="Could not start SafePay checkout")
        tracker = (((init.json() or {}).get("data") or {}).get("token") or "").strip()
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"SafePay init error: {type(e).__name__}: {e}")
        raise HTTPException(status_code=502, detail="Could not reach SafePay")
    if not tracker:
        raise HTTPException(status_code=502, detail="SafePay did not return a tracker")

    # 2) Build the hosted-checkout URL. On completion SafePay redirects to
    #    redirect_url — the app watches for our origin; the website result
    #    page polls /payments/safepay/status/{tracker} via gateway+ref params.
    origin = req.origin_url.rstrip("/")
    # source=mobile makes SafePay show "return to your mobile application"
    # instead of redirecting the browser — correct ONLY for the app's webview
    # (sentinel origin; in mobile mode SafePay navigates to its own
    # /mobile?action=complete|cancelled URL, which the app's webview watches).
    # Website checkouts must use source=custom — SafePay's documented value
    # for custom integrations; undocumented values (e.g. "web") can leave the
    # customer stranded on SafePay's page with the redirect_url ignored.
    source = "mobile" if origin.startswith("https://knb.payment.return") else "custom"
    from urllib.parse import urlencode
    qs = urlencode({
        "beacon": tracker,
        "env": env,
        "source": source,
        "order_id": req.order_id,
        "redirect_url": f"{origin}/payment/success?gateway=safepay&ref={tracker}&order_id={req.order_id}",
        "cancel_url": f"{origin}/payment/cancel?gateway=safepay&ref={tracker}&order_id={req.order_id}",
    })
    checkout_url = f"{checkout_base}/checkout/pay?{qs}"

    await db.payment_transactions.insert_one({
        "gateway": "safepay",
        "tracker": tracker,
        "order_id": req.order_id,
        "amount_pkr": amount,
        "currency": "PKR",
        "environment": env,
        "payment_status": "initiated",
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"url": checkout_url, "tracker": tracker}

@api_router.get("/payments/safepay/status/{tracker}")
async def safepay_status(tracker: str):
    """Report SafePay payment status for a tracker. Best-effort verification with
    a DB fallback (same shape as the Stripe status endpoint)."""
    # Same effective config as create-session and the reconcile job (DB wins,
    # env fallback). Regression fix: this endpoint previously read only the
    # SAFEPAY_API_KEY env var and called the removed _safepay_bases() helper,
    # so it 500/503'd and payments could never confirm from the client side.
    cfg = await _safepay_cfg()
    api_key = (cfg.get("api_key") or "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="SafePay not configured")
    txn = await db.payment_transactions.find_one({"tracker": tracker}, {"_id": 0})
    api_base, _ = _safepay_bases_for(cfg.get("mode", "sandbox"))
    verified_paid = False
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(f"{api_base}/order/v1/{tracker}")
        if r.status_code < 400:
            data = (r.json() or {}).get("data") or {}
            state = str(data.get("state") or data.get("status") or "").lower()
            verified_paid = state in ("paid", "tracker_ended", "completed", "succeeded")
    except Exception as e:
        logger.info(f"SafePay status check fell back to DB: {e}")

    if verified_paid and txn and txn.get("payment_status") != "paid":
        now = datetime.now(timezone.utc).isoformat()
        await db.payment_transactions.update_one(
            {"tracker": tracker, "payment_status": {"$ne": "paid"}},
            {"$set": {"payment_status": "paid", "paid_at": now}},
        )
        order_id = txn.get("order_id")
        if order_id:
            await db.online_orders.update_one(
                {"_id": ObjectId(order_id), "payment_status": {"$ne": "paid"}},
                {"$set": {"payment_status": "paid", "paid_at": now}},
            )
            await _release_online_order(order_id)
    payment_status = "paid" if verified_paid else ((txn or {}).get("payment_status") or "pending")
    return {
        "tracker": tracker,
        "payment_status": payment_status,
        "order_id": (txn or {}).get("order_id"),
    }

# --- PayFast Pakistan (gopayfast / APPS) hosted-checkout gateway -------------
# Charges Easypaisa / JazzCash wallets (and cards) natively — SafePay does not
# list those wallets, so wallet payments route here instead. Same shape as the
# SafePay/Stripe integrations: inert until PAYFAST_MERCHANT_ID +
# PAYFAST_SECURED_KEY are set (endpoints return 503 and the app falls back to
# the manual transfer flow); server-side amounts; a payment_transactions record
# keyed by basket_id; idempotent order->paid update on IPN / status check.
# Flow (PayFast redirect model):
#   1) POST /Ecommerce/api/Transaction/GetAccessToken (merchant_id+secured_key
#      +basket_id+txnamt) -> one-time ACCESS TOKEN bound to this transaction.
#   2) Client opens a WebView that form-POSTs to /Ecommerce/api/Transaction/
#      PostTransaction with TOKEN + order fields; PayFast shows its hosted page
#      (wallet MPIN / OTP) and redirects to SUCCESS_URL/FAILURE_URL. err_code=000
#      on the redirect is treated as provisional-paid.
#   3) GET /payments/payfast/status/{basket_id} confirms/refreshes from our DB
#      (redirect params are recorded by POST /payments/payfast/return).
#   Docs: https://gopayfast.com/docs/  (UAT host: ipguat.apps.net.pk)
class PayfastSessionRequest(BaseModel):
    order_id: str
    origin_url: str

def _payfast_base() -> str:
    env = (os.environ.get("PAYFAST_ENV") or "sandbox").strip().lower()
    if env in ("production", "prod", "live"):
        return os.environ.get("PAYFAST_API_BASE", "https://ipg1.apps.net.pk")
    return "https://ipguat.apps.net.pk"

def payfast_configured() -> bool:
    return bool(os.environ.get("PAYFAST_MERCHANT_ID") and os.environ.get("PAYFAST_SECURED_KEY"))

@api_router.post("/payments/payfast/create-session")
async def create_payfast_session(req: PayfastSessionRequest):
    """Mint a PayFast transaction token for an existing order and return the
    hosted-checkout form parameters. Amount comes from server-side order data."""
    if not payfast_configured():
        raise HTTPException(status_code=503, detail="PayFast not configured")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(req.order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Order already paid")
    amount = float(order.get("total_price", 0))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Invalid order amount")

    base = _payfast_base()
    # basket_id must be unique per attempt — retries after a failed/abandoned
    # checkout need a fresh token, so suffix the order id with a nonce.
    basket_id = f"{req.order_id}-{secrets.token_hex(4)}"
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            tok = await client.post(
                f"{base}/Ecommerce/api/Transaction/GetAccessToken",
                data={
                    "MERCHANT_ID": os.environ["PAYFAST_MERCHANT_ID"],
                    "SECURED_KEY": os.environ["PAYFAST_SECURED_KEY"],
                    "BASKET_ID": basket_id,
                    "TXNAMT": f"{amount:.2f}",
                    "CURRENCY_CODE": "PKR",
                },
            )
        if tok.status_code >= 400:
            logger.warning(f"PayFast token failed {tok.status_code}: {tok.text[:300]}")
            raise HTTPException(status_code=502, detail="Could not start PayFast checkout")
        token = ((tok.json() or {}).get("ACCESS_TOKEN") or "").strip()
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"PayFast token error: {type(e).__name__}: {e}")
        raise HTTPException(status_code=502, detail="Could not reach PayFast")
    if not token:
        raise HTTPException(status_code=502, detail="PayFast did not return a token")

    origin = req.origin_url.rstrip("/")
    now = datetime.now(timezone.utc)
    await db.payment_transactions.insert_one({
        "gateway": "payfast",
        "tracker": basket_id,          # same field name as safepay for admin queries
        "basket_id": basket_id,
        "order_id": req.order_id,
        "amount_pkr": amount,
        "currency": "PKR",
        "environment": (os.environ.get("PAYFAST_ENV") or "sandbox").strip().lower(),
        "payment_status": "initiated",
        "created_at": now.isoformat(),
    })
    # The client renders an auto-submitting form with exactly these fields.
    # PayFast's hosted page handles wallet number + MPIN/OTP itself.
    return {
        "action_url": f"{base}/Ecommerce/api/Transaction/PostTransaction",
        "basket_id": basket_id,
        "fields": {
            "MERCHANT_ID": os.environ["PAYFAST_MERCHANT_ID"],
            "MERCHANT_NAME": (await get_online_settings_doc()).get("restaurant_name", "Restaurant"),
            "TOKEN": token,
            "PROCCODE": "00",
            "TXNAMT": f"{amount:.2f}",
            "CURRENCY_CODE": "PKR",
            "CUSTOMER_MOBILE_NO": str(order.get("phone", "") or ""),
            "CUSTOMER_EMAIL_ADDRESS": "",
            "SIGNATURE": secrets.token_hex(8),
            "VERSION": "MERCHANT-CART-0.1",
            "TXNDESC": f"Order {req.order_id[-6:].upper()}",
            "SUCCESS_URL": f"{origin}/success?basket_id={basket_id}",
            "FAILURE_URL": f"{origin}/cancel?basket_id={basket_id}",
            "BASKET_ID": basket_id,
            "ORDER_DATE": now.strftime("%Y-%m-%d %H:%M:%S"),
            "CHECKOUT_URL": os.environ.get("PUBLIC_SITE_URL", "https://www.karachinaseebbiryani.com").rstrip("/") + "/api/payments/payfast/return",
        },
    }

def _payfast_validation_hash(basket_id: str, err_code: str) -> str:
    """Response-validation hash per PayFast (APPS) docs:
    sha256(basket_id|secured_key|merchant_id|err_code)."""
    raw = f"{basket_id}|{os.environ.get('PAYFAST_SECURED_KEY','')}|" \
          f"{os.environ.get('PAYFAST_MERCHANT_ID','')}|{err_code}"
    return _hashlib.sha256(raw.encode()).hexdigest()

@api_router.post("/payments/payfast/return")
async def payfast_return(request: Request):
    """Record the redirect/IPN result for a PayFast attempt. Accepts BOTH
    JSON (our app forwarding the redirect params) and form-encoded bodies
    (PayFast's own server-to-server CHECKOUT_URL callback).

    SECURITY: this endpoint is unauthenticated (the webview calls it after the
    gateway redirect), so the reported err_code alone MUST NOT be able to flip
    an order to paid — otherwise anyone with their own basket_id could fake a
    success. An order is marked *paid* only when the request carries PayFast's
    validation_hash and it verifies against our SECURED_KEY. A success report
    WITHOUT a valid hash lands in `pending_verification` — visible to the admin
    exactly like a manual bank transfer, never silently trusted. Idempotent:
    a paid order/txn is never downgraded."""
    ctype = (request.headers.get("content-type") or "").lower()
    try:
        if "application/json" in ctype:
            raw = await request.json()
        else:
            raw = dict(await request.form())
    except Exception:
        raise HTTPException(status_code=400, detail="Unreadable payload")
    def _pick(*names):
        for n in names:
            v = raw.get(n)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ""
    basket_id = _pick("basket_id", "BASKET_ID")
    if not basket_id:
        raise HTTPException(status_code=400, detail="basket_id required")
    err_code = _pick("err_code", "ERR_CODE")
    err_msg = _pick("err_msg", "ERR_MSG")
    transaction_id = _pick("transaction_id", "TRANSACTION_ID", "txn_id")
    validation_hash = _pick("validation_hash", "VALIDATION_HASH", "Response_Key")

    txn = await db.payment_transactions.find_one({"gateway": "payfast", "basket_id": basket_id})
    if not txn:
        raise HTTPException(status_code=404, detail="Unknown transaction")
    ok = err_code in ("000", "00", "0")
    hash_ok = bool(
        validation_hash
        and secrets.compare_digest(
            validation_hash.lower(),
            _payfast_validation_hash(basket_id, err_code),
        )
    )
    now = datetime.now(timezone.utc).isoformat()
    if ok and txn.get("payment_status") != "paid":
        new_status = "paid" if hash_ok else "pending_verification"
        set_fields = {
            "payment_status": new_status,
            "gateway_txn_id": transaction_id[:64],
            "hash_verified": hash_ok,
        }
        if hash_ok:
            set_fields["paid_at"] = now
        await db.payment_transactions.update_one(
            {"_id": txn["_id"], "payment_status": {"$nin": ["paid"]}},
            {"$set": set_fields},
        )
        await db.online_orders.update_one(
            {"_id": ObjectId(txn["order_id"]), "payment_status": {"$nin": ["paid"]}},
            {"$set": {"payment_status": new_status,
                      **({"paid_at": now} if hash_ok else
                         {"payment_reference": f"PayFast {(transaction_id or basket_id)[:64]}",
                          "payment_submitted_at": now})}},
        )
        # Release held (awaiting_payment) orders into the staff queue now that
        # the gateway flow completed — app wallet payments route through here.
        await _release_online_order(txn["order_id"])
        return {"payment_status": new_status}
    if not ok and txn.get("payment_status") == "initiated":
        await db.payment_transactions.update_one(
            {"_id": txn["_id"], "payment_status": "initiated"},
            {"$set": {"payment_status": "failed",
                      "gateway_err": f"{err_code[:8]}: {err_msg[:200]}"}},
        )
        return {"payment_status": "failed"}
    return {"payment_status": txn.get("payment_status") or "pending"}

@api_router.get("/payments/payfast/status/{basket_id}")
async def payfast_status(basket_id: str):
    """Authoritative payment status for a PayFast attempt (DB-backed — populated
    by /payments/payfast/return). Same response shape as the SafePay/Stripe
    status endpoints so clients can share polling code."""
    if not payfast_configured():
        raise HTTPException(status_code=503, detail="PayFast not configured")
    txn = await db.payment_transactions.find_one(
        {"gateway": "payfast", "basket_id": basket_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Unknown transaction")
    return {
        "tracker": basket_id,
        "payment_status": txn.get("payment_status") or "pending",
        "order_id": txn.get("order_id"),
    }

# --- Bank Transfer manual verification ---
class BankPaymentReference(BaseModel):
    transaction_id: str
    payer_name: Optional[str] = ""
    payment_via: Optional[str] = "bank"  # bank | easypaisa | jazzcash

@api_router.post("/online-orders/{order_id}/bank-payment")
async def submit_bank_payment(order_id: str, req: BankPaymentReference):
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    # Hardening: this endpoint is unauthenticated by design (guest checkout), but we
    # must not let an attacker overwrite payment details on an already-verified order
    # or on a stale order they don't own. Reject if payment is already settled or if
    # the submission window has elapsed.
    if order.get("payment_status") in {"paid", "refunded", "failed"}:
        raise HTTPException(status_code=400, detail="Payment can no longer be modified for this order")
    if not _order_within_payment_window(order):
        raise HTTPException(status_code=400, detail="Payment submission window has expired for this order")
    # Lightweight input length caps so an attacker can't stuff arbitrary blobs into the DB.
    txn_id = (req.transaction_id or "").strip()[:128]
    payer = (req.payer_name or "").strip()[:128]
    via = (req.payment_via or "bank").strip().lower()
    if via not in {"bank", "easypaisa", "jazzcash"}:
        raise HTTPException(status_code=400, detail="Invalid payment method")
    if not txn_id:
        raise HTTPException(status_code=400, detail="Transaction reference is required")
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {"$set": {
            "payment_method": via,
            "payment_reference": txn_id,
            "payer_name": payer,
            "payment_status": "pending_verification",
            "payment_submitted_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    # A gateway order that fell back to manual transfer is still held in
    # awaiting_payment — the submitted reference releases it for verification.
    await _release_online_order(order_id)
    return {"message": "Payment reference submitted. Awaiting verification.", "payment_status": "pending_verification"}

@api_router.put("/online-orders/{order_id}/payment-status")
async def admin_update_payment_status(order_id: str, body: dict, request: Request):
    user = await get_current_user(request)
    if not _has_perm(user, "online_orders"):
        raise HTTPException(status_code=403, detail="You don't have permission to update payment status.")
    new_status = body.get("payment_status")
    if new_status not in {"pending", "pending_verification", "paid", "failed", "refunded"}:
        raise HTTPException(status_code=400, detail="Invalid payment_status")
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {"$set": {"payment_status": new_status, "payment_updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if new_status in ("paid", "pending_verification"):
        # Admin confirming a payment also releases a held gateway order.
        await _release_online_order(order_id)
    return {"message": "Updated", "payment_status": new_status}

# =============================================================================
# HOSTED PAYMENT GATEWAYS (DB-configured): EasyPaisa (Easypay) + JazzCash
# =============================================================================
# Driver/registry pattern so future gateways drop in: implement a HostedGateway
# subclass, add it to GATEWAY_REGISTRY, and the generic
# /payments/{gateway}/create-session | /return | /status routes below serve it.
# Credentials come from the admin Payment Gateways page (db-backed, env
# fallback) via get_gateway_config() — see the settings block near
# get_online_settings_doc() for the schema and the PayFast migration notes.
#
# ROUTE ORDERING INVARIANT: these generic /payments/{gateway}/... routes MUST
# stay physically AFTER the literal /payments/stripe|safepay|payfast routes.
# Routes register on api_router in file order (include_router runs once at the
# bottom of this file), so the literals win. GATEWAY_REGISTRY membership is the
# second guard: stripe/safepay/payfast are NOT in it, so even a reorder could
# not silently reroute them here.
#
# Trust model (same as PayFast above): the return endpoints are unauthenticated
# browser-redirect targets, so a reported success alone can NEVER mark an order
# paid. JazzCash responses carry an HMAC we verify -> "paid"; EasyPaisa's
# postback is unsigned -> "paid" only after a server-to-server inquiry
# confirms, else "pending_verification" (admin verifies, like a manual
# transfer). All state updates are idempotent — a paid order is never
# downgraded.

class GatewaySessionRequest(BaseModel):
    order_id: str
    origin_url: str

class HostedGateway:
    """Interface for hosted-redirect payment gateways."""
    name = ""

    async def create_session(self, order: dict, order_id: str, origin_url: str,
                             cfg: dict, request: Request) -> dict:
        """Insert a payment_transactions record and return
        {action_url, method, ref, fields} for a client auto-submit form."""
        raise NotImplementedError

    async def handle_return(self, request: Request, params: dict, cfg: dict):
        """Process the gateway's browser return/callback and redirect the
        customer to the frontend result page."""
        raise NotImplementedError

def _gateway_result_redirect(origin_url: str, outcome: str, gateway: str,
                             ref: str, order_id: str = "") -> RedirectResponse:
    """303 (POST->GET) to the frontend result page after a gateway return."""
    base = (origin_url or os.environ.get("PUBLIC_SITE_URL", "")).rstrip("/")
    url = f"{base}/payment/{outcome}?gateway={gateway}&ref={ref}"
    if order_id:
        url += f"&order_id={order_id}"
    return RedirectResponse(url, status_code=303)

async def _mark_gateway_paid(txn: dict, now: str, gateway_txn_id: str = ""):
    """Idempotently flip txn + order to paid (never downgrades)."""
    await db.payment_transactions.update_one(
        {"_id": txn["_id"], "payment_status": {"$nin": ["paid"]}},
        {"$set": {"payment_status": "paid", "paid_at": now,
                  "gateway_txn_id": (gateway_txn_id or "")[:64],
                  "hash_verified": True}},
    )
    await db.online_orders.update_one(
        {"_id": ObjectId(txn["order_id"]), "payment_status": {"$nin": ["paid"]}},
        {"$set": {"payment_status": "paid", "paid_at": now}},
    )
    await _release_online_order(txn["order_id"])

async def _mark_gateway_pending_verification(txn: dict, now: str, reference: str):
    """Success reported but not provider-verified: surface to the admin
    exactly like a manual bank transfer, never silently trust it."""
    await db.payment_transactions.update_one(
        {"_id": txn["_id"], "payment_status": {"$nin": ["paid"]}},
        {"$set": {"payment_status": "pending_verification", "hash_verified": False}},
    )
    await db.online_orders.update_one(
        {"_id": ObjectId(txn["order_id"]), "payment_status": {"$nin": ["paid"]}},
        {"$set": {"payment_status": "pending_verification",
                  "payment_reference": reference[:64],
                  "payment_submitted_at": now}},
    )
    # The customer DID complete the gateway flow — release to the restaurant
    # while the admin verifies (same trust level as a manual transfer receipt).
    await _release_online_order(txn["order_id"])

async def _mark_gateway_failed(txn: dict, err: str):
    """Mark a still-open attempt failed (paid/pending are never downgraded)."""
    await db.payment_transactions.update_one(
        {"_id": txn["_id"], "payment_status": {"$in": ["initiated", "token_received"]}},
        {"$set": {"payment_status": "failed", "gateway_err": err[:220]}},
    )

# --- JazzCash (hosted Page Redirection) --------------------------------------
# Docs: https://sandbox.jazzcash.com.pk/SandboxDocumentation/ — merchant
# form-POSTs pp_* fields to the CustomerPortal merchant form; JazzCash shows
# its hosted page (wallet MPIN / card) and browser-POSTs the signed response
# to pp_ReturnURL. Credentials: Merchant ID, Password, Integrity Salt.

def _jazzcash_base(mode: str) -> str:
    host = "https://payments.jazzcash.com.pk" if mode == "live" \
        else "https://sandbox.jazzcash.com.pk"
    return host + "/CustomerPortal/transactionmanagement/merchantform/"

def _jazzcash_secure_hash(fields: dict, salt: str) -> str:
    """pp_SecureHash per JazzCash spec — used for BOTH signing our request and
    verifying their response so the two can never drift: take every pp_*/ppmpf_*
    field, sort alphabetically by key, DROP empty values and pp_SecureHash
    itself, join the VALUES with '&', prepend the Integrity Salt + '&', then
    HMAC-SHA256 keyed with the salt, uppercase hex. Getting inclusion or
    ordering wrong fails every transaction with 'invalid secure hash'."""
    vals = [str(v) for k, v in sorted(fields.items())
            if k.lower().startswith("pp") and k != "pp_SecureHash"
            and str(v).strip() != ""]
    msg = salt + "&" + "&".join(vals)
    return hmac.new(salt.encode("utf-8"), msg.encode("utf-8"),
                    _hashlib.sha256).hexdigest().upper()

class JazzcashGateway(HostedGateway):
    name = "jazzcash"

    async def create_session(self, order, order_id, origin_url, cfg, request):
        amount = float(order.get("total_price", 0))
        # JazzCash validates txn/expiry times against Pakistan time — UTC
        # timestamps make transactions instantly expired.
        pkt = datetime.now(pytz.timezone("Asia/Karachi"))
        txn_dt = pkt.strftime("%Y%m%d%H%M%S")
        expiry = (pkt + timedelta(hours=1)).strftime("%Y%m%d%H%M%S")
        # Fresh ref per attempt (19 chars, limit 20) so retries never collide.
        ref = "T" + txn_dt + secrets.token_hex(2)
        # Amount is integer PAISA. round() first — int(x*100) float-truncates.
        paisa = str(int(round(amount * 100)))
        fields = {
            "pp_Version": "1.1",
            # Empty TxnType lets the customer choose wallet/card on the hosted
            # page; the non-empty rule correctly excludes it from the hash.
            "pp_TxnType": "",
            "pp_Language": "EN",
            "pp_MerchantID": cfg["merchant_id"],
            "pp_Password": cfg["password"],
            "pp_TxnRefNo": ref,
            "pp_Amount": paisa,
            "pp_TxnCurrency": "PKR",
            "pp_TxnDateTime": txn_dt,
            "pp_TxnExpiryDateTime": expiry,
            "pp_BillReference": ("ORD" + order_id[-8:]).upper(),
            "pp_Description": "Food order",
            "pp_ReturnURL": f"{_api_public_base(request)}/api/payments/jazzcash/return",
        }
        fields["pp_SecureHash"] = _jazzcash_secure_hash(fields, cfg["integrity_salt"])
        await db.payment_transactions.insert_one({
            "gateway": "jazzcash",
            "tracker": ref,
            "order_id": order_id,
            "amount_pkr": amount,
            "amount_paisa": int(paisa),
            "currency": "PKR",
            "environment": cfg["mode"],
            "payment_status": "initiated",
            "origin_url": origin_url.rstrip("/"),
            "api_base": _api_public_base(request),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        return {"action_url": _jazzcash_base(cfg["mode"]), "method": "POST",
                "ref": ref, "fields": fields}

    async def handle_return(self, request, params, cfg):
        ref = str(params.get("pp_TxnRefNo") or "").strip()
        txn = await db.payment_transactions.find_one(
            {"gateway": "jazzcash", "tracker": ref}) if ref else None
        if not txn:
            # No txn -> no stored origin; send them to the site root.
            return RedirectResponse(
                os.environ.get("PUBLIC_SITE_URL", "/") or "/", status_code=303)
        code = str(params.get("pp_ResponseCode") or "").strip()
        msg = str(params.get("pp_ResponseMessage") or "").strip()
        rrn = str(params.get("pp_RetreivalReferenceNo") or "").strip()  # sic, JazzCash's spelling
        their_hash = str(params.get("pp_SecureHash") or "").strip()
        received = {k: v for k, v in params.items() if k.lower().startswith("pp")}
        computed = _jazzcash_secure_hash(received, cfg["integrity_salt"])
        hash_ok = bool(their_hash) and hmac.compare_digest(their_hash.upper(), computed)
        # Defense in depth: the signed response must be for OUR amount, not a
        # replayed success from a cheaper transaction.
        amount_ok = str(params.get("pp_Amount", "")).strip() == str(txn.get("amount_paisa", ""))
        ok = code == "000"
        now = datetime.now(timezone.utc).isoformat()
        origin = txn.get("origin_url") or ""
        if ok and hash_ok and amount_ok:
            await _mark_gateway_paid(txn, now, rrn or ref)
            return _gateway_result_redirect(origin, "success", "jazzcash", ref)
        if ok:
            logger.warning(f"JazzCash success without valid hash for {ref} "
                           f"(hash_ok={hash_ok} amount_ok={amount_ok})")
            await _mark_gateway_pending_verification(txn, now, f"JazzCash {rrn or ref}")
            return _gateway_result_redirect(origin, "success", "jazzcash", ref)
        await _mark_gateway_failed(txn, f"{code[:8]}: {msg[:200]}")
        return _gateway_result_redirect(origin, "cancel", "jazzcash", ref,
                                        txn.get("order_id", ""))

# --- EasyPaisa (Easypay hosted checkout, two-step) ---------------------------
# Docs: Easypay Merchant Integration Guide v4 — step 1 form-POSTs to
# /easypay/Index.jsf with an AES-hashed request; after payment Easypay sends
# the browser back to postBackURL with an auth_token, which we form-POST to
# /easypay/Confirm.jsf; the final redirect carries status/orderRefNumber.
# Credentials: Store ID + 16-char Hash Key (+ optional inquiry API creds).

def _easypay_host(mode: str) -> str:
    return "https://easypay.easypaisa.com.pk" if mode == "live" \
        else "https://easypaystg.easypaisa.com.pk"

def _easypay_hashed_req(fields: dict, hash_key: str) -> str:
    """merchantHashedReq per Easypay v4: every posted field (except the hash
    itself), sorted alphabetically by key, joined as 'k=v' with '&', then
    AES-128-ECB with the 16-char merchant Hash Key, PKCS7 padding, base64.
    The plaintext values must byte-match what is POSTed (esp. the amount)."""
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.primitives import padding as _crypto_pad
    import base64 as _b64
    plaintext = "&".join(f"{k}={v}" for k, v in sorted(fields.items())
                         if str(v) != "")
    padder = _crypto_pad.PKCS7(128).padder()
    padded = padder.update(plaintext.encode("utf-8")) + padder.finalize()
    enc = Cipher(algorithms.AES(hash_key.encode("utf-8")), modes.ECB()).encryptor()
    return _b64.b64encode(enc.update(padded) + enc.finalize()).decode()

async def _easypay_inquire(ref: str, cfg: dict) -> Optional[bool]:
    """Server-to-server order-status inquiry. Returns True (provider says
    paid), False (provider says not paid), or None (inquiry unavailable —
    creds missing or the call failed)."""
    import base64 as _b64
    user = (cfg.get("inquiry_username") or "").strip()
    pw = (cfg.get("inquiry_password") or "").strip()
    if not user or not pw:
        return None
    try:
        creds = _b64.b64encode(f"{user}:{pw}".encode()).decode()
        async with httpx.AsyncClient(timeout=20) as client:
            resp = await client.post(
                f"{_easypay_host(cfg['mode'])}/easypay-service/rest/v4/inquire-transaction",
                headers={"Credentials": creds},
                json={"orderId": ref, "storeId": cfg["store_id"], "accountNum": ""},
            )
        if resp.status_code >= 400:
            logger.warning(f"Easypay inquiry HTTP {resp.status_code} for {ref}")
            return None
        data = resp.json() or {}
        if str(data.get("responseCode", "")).strip() != "0000":
            return None
        status = str(data.get("transactionStatus", "")).strip().upper()
        return status in ("PAID", "SUCCESS")
    except Exception as e:
        logger.warning(f"Easypay inquiry error for {ref}: {type(e).__name__}: {e}")
        return None

class EasypaisaGateway(HostedGateway):
    name = "easypaisa"

    async def create_session(self, order, order_id, origin_url, cfg, request):
        if len(cfg.get("hash_key") or "") != 16:
            raise HTTPException(
                status_code=503,
                detail="EasyPaisa hash key misconfigured (must be 16 characters)")
        amount = float(order.get("total_price", 0))
        pkt = datetime.now(pytz.timezone("Asia/Karachi"))
        ref = "EP" + pkt.strftime("%y%m%d%H%M%S") + secrets.token_hex(2)
        # The exact amount string is hashed — store it so verification and
        # debugging can byte-match what was sent.
        amount_str = f"{amount:.1f}"
        api_base = _api_public_base(request)
        # ref rides in the query so the step-1 postback (which only carries
        # auth_token) can be correlated back to this transaction.
        postback = f"{api_base}/api/payments/easypaisa/return?ref={ref}"
        fields = {
            "storeId": cfg["store_id"],
            "amount": amount_str,
            "postBackURL": postback,
            "orderRefNum": ref,
            "expiryDate": (pkt + timedelta(hours=1)).strftime("%Y%m%d %H%M%S"),
            "autoRedirect": "1",
            "paymentMethod": "MA_PAYMENT_METHOD",  # mobile-account flow — the option the customer picked
            "mobileNum": str(order.get("phone", "") or ""),
            "timeStamp": pkt.strftime("%Y-%m-%dT%H:%M:%S"),
        }
        fields["merchantHashedReq"] = _easypay_hashed_req(fields, cfg["hash_key"])
        await db.payment_transactions.insert_one({
            "gateway": "easypaisa",
            "tracker": ref,
            "order_id": order_id,
            "amount_pkr": amount,
            "amount_str": amount_str,
            "currency": "PKR",
            "environment": cfg["mode"],
            "payment_status": "initiated",
            "origin_url": origin_url.rstrip("/"),
            "api_base": api_base,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        return {"action_url": f"{_easypay_host(cfg['mode'])}/easypay/Index.jsf",
                "method": "POST", "ref": ref, "fields": fields}

    async def handle_return(self, request, params, cfg):
        import html as _html
        # Defensive: some Easypay deployments naively append '?auth_token=' to
        # a postBackURL that already has a query string, yielding '?ref=..?auth_token=..'.
        raw_q = str(request.url.query or "")
        if "?" in raw_q:
            from urllib.parse import parse_qsl
            for k, v in parse_qsl(raw_q.replace("?", "&")):
                params.setdefault(k, v)
        ref = str(params.get("orderRefNumber") or params.get("ref") or "").strip()
        txn = await db.payment_transactions.find_one(
            {"gateway": "easypaisa", "tracker": ref}) if ref else None
        if not txn:
            return RedirectResponse(
                os.environ.get("PUBLIC_SITE_URL", "/") or "/", status_code=303)
        origin = txn.get("origin_url") or ""
        order_id = txn.get("order_id", "")
        status = str(params.get("status") or "").strip()
        auth_token = str(params.get("auth_token") or "").strip()

        # Phase 1: Easypay handed back an auth token — confirm it by form-
        # POSTing to Confirm.jsf (browser stays in the loop by design).
        if auth_token and not status:
            await db.payment_transactions.update_one(
                {"_id": txn["_id"], "payment_status": "initiated"},
                {"$set": {"auth_token": auth_token[:128],
                          "payment_status": "token_received"}},
            )
            confirm_url = f"{_easypay_host(cfg['mode'])}/easypay/Confirm.jsf"
            postback2 = f"{txn.get('api_base', _api_public_base(request))}/api/payments/easypaisa/return?ref={ref}"
            esc = lambda v: _html.escape(str(v), quote=True)
            return HTMLResponse(
                "<html><body>Completing payment&hellip;"
                f"<form method=\"post\" action=\"{esc(confirm_url)}\">"
                f"<input type=\"hidden\" name=\"auth_token\" value=\"{esc(auth_token)}\">"
                f"<input type=\"hidden\" name=\"postBackURL\" value=\"{esc(postback2)}\">"
                "</form><script>document.forms[0].submit()</script></body></html>")

        # Phase 2: final status postback. It is UNSIGNED — only a server-to-
        # server inquiry may mark the order paid.
        ok = status in ("0000", "SUCCESS")
        now = datetime.now(timezone.utc).isoformat()
        if ok:
            confirmed = await _easypay_inquire(ref, cfg)
            if confirmed is True:
                await _mark_gateway_paid(txn, now, str(params.get("transactionId") or ref))
                return _gateway_result_redirect(origin, "success", "easypaisa", ref)
            if confirmed is False:
                await _mark_gateway_failed(txn, "inquiry: provider reports not paid")
                return _gateway_result_redirect(origin, "cancel", "easypaisa", ref, order_id)
            # Inquiry unavailable — never trust the browser, never strand the
            # customer on a provider outage: admin verifies manually.
            await _mark_gateway_pending_verification(txn, now, f"EasyPaisa {ref}")
            return _gateway_result_redirect(origin, "success", "easypaisa", ref)
        desc = str(params.get("desc") or "").strip()
        await _mark_gateway_failed(txn, f"{status[:8]}: {desc[:200]}")
        return _gateway_result_redirect(origin, "cancel", "easypaisa", ref, order_id)

# stripe/safepay/payfast intentionally absent — their literal routes above serve them.
GATEWAY_REGISTRY: Dict[str, HostedGateway] = {
    "easypaisa": EasypaisaGateway(),
    "jazzcash": JazzcashGateway(),
}

@api_router.post("/payments/{gateway}/create-session")
async def create_gateway_session(gateway: str, req: GatewaySessionRequest, request: Request):
    gw = GATEWAY_REGISTRY.get(gateway)
    if not gw:
        raise HTTPException(status_code=404, detail="Unknown gateway")
    if not await gateway_ready(gateway):
        raise HTTPException(status_code=503, detail=f"{gateway} not configured")
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(req.order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Order already paid")
    if float(order.get("total_price", 0)) <= 0:
        raise HTTPException(status_code=400, detail="Invalid order amount")
    cfg = await get_gateway_config(gateway)
    return await gw.create_session(order, req.order_id, req.origin_url, cfg, request)

@api_router.api_route("/payments/{gateway}/return", methods=["GET", "POST"])
async def gateway_return(gateway: str, request: Request):
    """Unauthenticated browser-redirect target for the hosted gateways. The
    per-gateway verification (HMAC / inquiry / pending_verification) is what
    protects 'paid' — see the trust-model note at the top of this section."""
    gw = GATEWAY_REGISTRY.get(gateway)
    if not gw:
        raise HTTPException(status_code=404, detail="Unknown gateway")
    cfg = await get_gateway_config(gateway)
    params: Dict[str, Any] = dict(request.query_params)
    if request.method == "POST":
        ctype = (request.headers.get("content-type") or "").lower()
        try:
            if "application/json" in ctype:
                body = await request.json()
                if isinstance(body, dict):
                    params.update(body)
            else:
                params.update(dict(await request.form()))
        except Exception:
            pass
    return await gw.handle_return(request, params, cfg)

@api_router.get("/payments/{gateway}/status/{ref}")
async def gateway_status(gateway: str, ref: str):
    """DB-backed status (populated by the return endpoint). Same response
    shape as the Stripe/SafePay/PayFast status endpoints so clients share
    polling code."""
    if gateway not in GATEWAY_REGISTRY:
        raise HTTPException(status_code=404, detail="Unknown gateway")
    txn = await db.payment_transactions.find_one(
        {"gateway": gateway, "tracker": ref}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Unknown transaction")
    return {
        "tracker": ref,
        "payment_status": txn.get("payment_status") or "pending",
        "order_id": txn.get("order_id"),
    }

# ===== END HOSTED PAYMENT GATEWAYS ===========================================

# =============================================================================
# END ONLINE SETTINGS / DELIVERY / PAYMENT
# =============================================================================

# =============================================================================
# OBJECT STORAGE, WHATSAPP NOTIFICATIONS, LIVE TRACKING
# =============================================================================
import requests as _requests
from fastapi import Header, Query
from fastapi.responses import Response

OBJ_STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
APP_NAME = os.environ.get("APP_NAME", "karachi-naseeb")
_obj_storage_key = None
LOCAL_UPLOAD_DIR = os.environ.get(
    "LOCAL_UPLOAD_DIR",
    # Auto-detect the persistent Fly.io volume mount when present so uploaded files
    # survive every redeploy. Falls back to a folder next to server.py for local dev.
    "/app/uploads" if os.path.isdir("/app/uploads") else os.path.join(os.path.dirname(__file__), "uploads"),
)

# =============================================================================
# Inline-image migration helpers
#
# Historic menu items stored their photo as a base64 `data:` URL inside the
# `image_url` MongoDB field. With 50+ items × ~120 KB each, every public menu
# request shipped ~6 MB of JSON which made the menu visibly slow (~60s on
# cold Cloudflare cache). The helpers below convert those data URLs into real
# files on the persistent volume and serve them via /api/uploads/* with
# long-lived Cache-Control. Content-hashed filenames make cache invalidation
# automatic when the image changes.
# =============================================================================

import hashlib as _hashlib
import base64 as _b64lib
import re as _re_img

_DATA_URL_RE = _re_img.compile(
    r"^data:(image/(jpeg|jpg|png|webp|gif));base64,(.+)$", _re_img.IGNORECASE | _re_img.DOTALL
)


def _persist_data_url_image(image_url: str, kind: str = "menu") -> str:
    """If `image_url` is a base64 data: URL, decode it, write it to
    LOCAL_UPLOAD_DIR/{kind}/<sha16>.<ext>, and return the public URL
    (/api/uploads/{kind}/<sha16>.<ext>). Otherwise return the input unchanged.
    Idempotent: same bytes → same path → no extra writes."""
    if not image_url or not isinstance(image_url, str) or not image_url.startswith("data:"):
        return image_url or ""
    m = _DATA_URL_RE.match(image_url)
    if not m:
        return image_url  # unrecognised data: shape — leave for manual review
    ext_raw = m.group(2).lower()
    b64data = m.group(3)
    ext = "jpg" if ext_raw == "jpeg" else ext_raw
    try:
        binary = _b64lib.b64decode(b64data, validate=False)
    except Exception as e:
        logger.warning(f"_persist_data_url_image: base64 decode failed ({e}); keeping data URL")
        return image_url
    digest = _hashlib.sha256(binary).hexdigest()[:16]
    rel_path = f"{kind}/{digest}.{ext}"
    abs_dir = os.path.abspath(LOCAL_UPLOAD_DIR)
    abs_target = os.path.join(abs_dir, rel_path)
    try:
        os.makedirs(os.path.dirname(abs_target), exist_ok=True)
        if not os.path.exists(abs_target):
            with open(abs_target, "wb") as f:
                f.write(binary)
    except Exception as e:
        logger.error(f"_persist_data_url_image: write failed for {rel_path}: {e}")
        return image_url
    return f"/api/uploads/{rel_path}"


@api_router.get("/uploads/{path:path}")
async def serve_public_upload(path: str):
    """Public, unauthenticated read of LOCAL_UPLOAD_DIR contents.
    Used for menu/category images that were migrated out of inline base64.
    Browser/CDN cacheable for 1 year because URLs are content-hashed."""
    abs_dir = os.path.abspath(LOCAL_UPLOAD_DIR)
    abs_target = os.path.abspath(os.path.join(abs_dir, path))
    # Block path traversal — abs_target MUST be inside abs_dir.
    if not (abs_target == abs_dir or abs_target.startswith(abs_dir + os.sep)):
        raise HTTPException(status_code=404, detail="Not found")
    if not os.path.isfile(abs_target):
        raise HTTPException(status_code=404, detail="Not found")
    ext = abs_target.rsplit(".", 1)[-1].lower() if "." in abs_target else ""
    ctype = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
        "webp": "image/webp", "gif": "image/gif", "pdf": "application/pdf",
    }.get(ext, "application/octet-stream")
    try:
        with open(abs_target, "rb") as f:
            data = f.read()
    except Exception as e:
        logger.error(f"serve_public_upload: read failed for {path}: {e}")
        raise HTTPException(status_code=500, detail="Read failed")
    return Response(
        content=data,
        media_type=ctype,
        headers={"Cache-Control": "public, max-age=31536000, immutable"},
    )


@api_router.post("/admin/migrate-menu-images")
async def migrate_menu_images(request: Request, limit: int = 1000, dry_run: bool = False):
    """One-shot migration: every menu_item whose image_url is a base64 data URL
    gets persisted to LOCAL_UPLOAD_DIR/menu/<sha16>.<ext> and its image_url is
    rewritten to /api/uploads/menu/<sha16>.<ext>. Idempotent — re-running is
    safe; already-migrated docs are excluded by the query. Admin only.

    Query params:
      - limit:   max docs to process this call (default 1000)
      - dry_run: when true, report changes without writing
    """
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    migrated, skipped, failed = 0, 0, 0
    examples: list = []
    try:
        cursor = db.menu_items.find({"image_url": {"$regex": "^data:"}}).limit(int(limit))
        async for doc in cursor:
            try:
                new_url = _persist_data_url_image(doc.get("image_url", ""), kind="menu")
                if isinstance(new_url, str) and new_url.startswith("/api/uploads/"):
                    if not dry_run:
                        await db.menu_items.update_one(
                            {"_id": doc["_id"]},
                            {"$set": {"image_url": new_url, "image_type": "url"}},
                        )
                    migrated += 1
                    if len(examples) < 5:
                        examples.append({
                            "id": str(doc["_id"]),
                            "name": doc.get("name", ""),
                            "new_url": new_url,
                        })
                else:
                    skipped += 1
            except Exception as e:
                failed += 1
                logger.exception(f"migrate_menu_images: failed for {doc.get('_id')}: {e}")
    except Exception as e:
        logger.exception(f"migrate_menu_images: cursor failed: {e}")
        raise HTTPException(status_code=500, detail=f"Migration cursor failed: {str(e)[:120]}")
    return {
        "migrated": migrated,
        "skipped": skipped,
        "failed": failed,
        "examples": examples,
        "dry_run": dry_run,
        "upload_dir": LOCAL_UPLOAD_DIR,
    }



def _init_obj_storage():
    global _obj_storage_key
    if _obj_storage_key:
        return _obj_storage_key
    try:
        emkey = os.environ.get("EMERGENT_LLM_KEY", "")
        if not emkey:
            return None
        resp = _requests.post(f"{OBJ_STORAGE_URL}/init", json={"emergent_key": emkey}, timeout=30)
        resp.raise_for_status()
        _obj_storage_key = resp.json().get("storage_key")
        logger.info("Object storage initialized (cloud)")
        return _obj_storage_key
    except Exception as e:
        logger.warning(f"Object storage init failed: {e}")
        return None

def _put_object(path: str, data: bytes, content_type: str):
    """Upload file to cloud storage or fallback to local filesystem."""
    key = _init_obj_storage()
    
    # Try cloud storage first
    if key:
        try:
            resp = _requests.put(
                f"{OBJ_STORAGE_URL}/objects/{path}",
                headers={"X-Storage-Key": key, "Content-Type": content_type},
                data=data, timeout=120,
            )
            if resp.status_code == 403:
                # Re-init and retry once
                global _obj_storage_key
                _obj_storage_key = None
                key = _init_obj_storage()
                if key:
                    resp = _requests.put(
                        f"{OBJ_STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": key, "Content-Type": content_type},
                        data=data, timeout=120,
                    )
            resp.raise_for_status()
            logger.info(f"File uploaded to cloud storage: {path}")
            return resp.json()
        except Exception as e:
            logger.warning(f"Cloud storage upload failed, falling back to local: {e}")
    
    # Fallback to local filesystem
    try:
        local_path = os.path.join(LOCAL_UPLOAD_DIR, path)
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, "wb") as f:
            f.write(data)
        logger.info(f"File uploaded to local storage: {path}")
        return {"path": path, "size": len(data), "storage": "local"}
    except Exception as e:
        logger.error(f"Local storage upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Storage failed: {str(e)[:100]}")

def _get_object(path: str):
    """Retrieve file from cloud storage or local filesystem."""
    key = _init_obj_storage()
    
    # Try cloud storage first
    if key:
        try:
            resp = _requests.get(
                f"{OBJ_STORAGE_URL}/objects/{path}",
                headers={"X-Storage-Key": key}, timeout=60,
            )
            if resp.status_code == 403:
                global _obj_storage_key
                _obj_storage_key = None
                key = _init_obj_storage()
                if key:
                    resp = _requests.get(
                        f"{OBJ_STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": key}, timeout=60,
                    )
            resp.raise_for_status()
            return resp.content, resp.headers.get("Content-Type", "application/octet-stream")
        except Exception as e:
            logger.warning(f"Cloud storage retrieval failed, trying local: {e}")
    
    # Fallback to local filesystem
    try:
        local_path = os.path.join(LOCAL_UPLOAD_DIR, path)
        if not os.path.exists(local_path):
            raise HTTPException(status_code=404, detail="File not found")
        with open(local_path, "rb") as f:
            data = f.read()
        # Guess content type from extension
        ext = path.rsplit(".", 1)[-1].lower() if "." in path else ""
        content_type = {
            "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
            "webp": "image/webp", "pdf": "application/pdf"
        }.get(ext, "application/octet-stream")
        return data, content_type
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Local storage retrieval failed: {e}")
        raise HTTPException(status_code=500, detail=f"Storage read failed: {str(e)[:100]}")

@app.on_event("startup")
async def _startup_storage():
    # Non-blocking: object-storage probe should never block the listener.
    asyncio.create_task(_startup_storage_background())

async def _startup_storage_background():
    try:
        _init_obj_storage()
    except Exception as e:
        logger.exception(f"Object storage init failed (server still listening): {e}")

ALLOWED_UPLOAD_MIMES = {"image/jpeg", "image/jpg", "image/png", "image/webp", "application/pdf"}
ALLOWED_UPLOAD_EXTS = {"jpg", "jpeg", "png", "webp", "pdf"}
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB

def _order_within_payment_window(order: dict) -> bool:
    """A guest can only attach a payment reference / screenshot for PAYMENT_SUBMIT_WINDOW_SEC
    after the order was created. Past that window the order is treated as immutable from
    the guest side (admin can still update via authenticated endpoints)."""
    try:
        created = datetime.fromisoformat(str(order.get("created_at", "")).replace("Z", "+00:00"))
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
    except Exception:
        return False
    elapsed = (datetime.now(timezone.utc) - created).total_seconds()
    return 0 <= elapsed <= PAYMENT_SUBMIT_WINDOW_SEC

@api_router.post("/online-orders/{order_id}/payment-screenshot")
async def upload_payment_screenshot(order_id: str, file: UploadFile = File(...)):
    """Customer uploads payment proof screenshot/PDF for bank/easypaisa/jazzcash.
    Allowed only within PAYMENT_SUBMIT_WINDOW_SEC after order creation and only while
    payment is still 'pending' / 'pending_verification' — prevents an attacker from
    overwriting screenshots on paid or stale orders."""
    try:
        order = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception as e:
        logger.warning(f"Payment upload - invalid order ID {order_id}: {e}")
        raise HTTPException(status_code=404, detail="Order not found")
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("payment_status") in {"paid", "refunded", "failed"}:
        raise HTTPException(status_code=400, detail="Payment screenshot can no longer be submitted for this order")
    if not _order_within_payment_window(order):
        raise HTTPException(status_code=400, detail="Payment submission window has expired for this order")

    ctype = (file.content_type or "").lower()
    if ctype not in ALLOWED_UPLOAD_MIMES:
        raise HTTPException(status_code=400, detail=f"Unsupported file type. Allowed: JPG, PNG, WebP, PDF")

    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large (max 5MB)")

    # Whitelist the extension. Never trust the raw filename — it can contain path
    # separators, NUL bytes, or unicode tricks that would let the storage_path escape
    # the upload jail (e.g. "x.jpg/../../etc/passwd").
    raw_filename = file.filename or ""
    raw_ext = raw_filename.rsplit(".", 1)[-1].lower() if "." in raw_filename else ""
    # strip everything except a-z0-9
    safe_ext = "".join(ch for ch in raw_ext if ch.isalnum())[:5]
    if safe_ext not in ALLOWED_UPLOAD_EXTS:
        # Fall back from content-type
        safe_ext = {"image/jpeg": "jpg", "image/jpg": "jpg", "image/png": "png",
                    "image/webp": "webp", "application/pdf": "pdf"}.get(ctype, "bin")
    storage_path = f"{APP_NAME}/payments/{order_id}/{_uuid.uuid4().hex}.{safe_ext}"
    
    try:
        result = _put_object(storage_path, data, ctype)
        logger.info(f"Payment screenshot uploaded for order {order_id}: {result.get('storage', 'cloud')} storage")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Payment screenshot upload failed for order {order_id}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)[:100]}")
    
    file_record = {
        "id": _uuid.uuid4().hex,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": ctype,
        "size": result.get("size", len(data)),
        "order_id": order_id,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "is_deleted": False,
    }
    await db.uploaded_files.insert_one(file_record)
    await db.online_orders.update_one(
        {"_id": ObjectId(order_id)},
        {"$set": {
            "payment_screenshot_path": result["path"],
            "payment_screenshot_uploaded_at": file_record["uploaded_at"],
        }},
    )
    return {"ok": True, "path": result["path"], "id": file_record["id"]}

@api_router.get("/files/{path:path}")
async def serve_file(path: str, request: Request):
    """Serves uploaded file. Admin only (Bearer token)."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    record = await db.uploaded_files.find_one({"storage_path": path, "is_deleted": False})
    if not record:
        raise HTTPException(status_code=404, detail="File not found")
    try:
        data, ctype = _get_object(path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Read failed: {str(e)[:100]}")
    return Response(content=data, media_type=record.get("content_type", ctype))

# --- WhatsApp via Twilio ---
def _normalize_pk_phone(p: str) -> str:
    p = (p or "").strip().replace(" ", "").replace("-", "")
    if p.startswith("+"):
        return p
    if p.startswith("0"):
        return "+92" + p[1:]
    if p.startswith("92"):
        return "+" + p
    if len(p) == 10:
        return "+92" + p
    return p

async def send_whatsapp(to_phone: str, body: str):
    """Send WhatsApp message via Twilio. Silently no-ops if creds missing."""
    sid = os.environ.get("TWILIO_ACCOUNT_SID")
    tok = os.environ.get("TWILIO_AUTH_TOKEN")
    if not sid or not tok or not to_phone:
        return False
    try:
        s = await get_online_settings_doc()
        from_num = s.get("twilio_whatsapp_from") or os.environ.get("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")
        if not from_num.startswith("whatsapp:"):
            from_num = f"whatsapp:{from_num}"
        to_num = _normalize_pk_phone(to_phone)
        if not to_num.startswith("whatsapp:"):
            to_num = f"whatsapp:{to_num}"
        from twilio.rest import Client
        client = Client(sid, tok)
        msg = await asyncio.to_thread(
            lambda: client.messages.create(from_=from_num, to=to_num, body=body)
        )
        logger.info(f"WhatsApp sent: {msg.sid}")
        return True
    except Exception as e:
        logger.warning(f"WhatsApp send failed: {e}")
        return False

def _format_order_confirmation(order: dict, tracking_url: str) -> str:
    receipt = str(order.get("_id", ""))[-6:].upper() if "_id" in order else order.get("receipt_no", "")
    items_lines = "\n".join([f"  • {i['quantity']}x {i['name']} — Rs.{i['price']*i['quantity']}" for i in order.get("items", [])])
    total = order.get("total_price", 0)
    pay = (order.get("payment_method") or "cod").upper()
    return (
        f"🍛 *Karachi Naseeb Biryani*\n"
        f"Order #{receipt} confirmed!\n\n"
        f"{items_lines}\n\n"
        f"*Total: Rs. {int(total)}*\n"
        f"Payment: {pay}\n"
        f"ETA: 30–45 min\n\n"
        f"Track your order:\n{tracking_url}\n\n"
        f"Questions? Call +92 300 4928411"
    )

def _format_status_update(order: dict, status: str, tracking_url: str) -> str:
    receipt = str(order.get("_id", ""))[-6:].upper() if "_id" in order else order.get("receipt_no", "")
    rejection_reason = order.get("rejection_reason") or ""
    reason_pretty = {
        "out_of_stock": "We've run out of stock for one or more items.",
        "closed": "Sorry, the kitchen is currently closed.",
        "other": "Please contact us for details.",
    }.get(rejection_reason.lower().strip(), rejection_reason or "Please contact us for details.")
    labels = {
        "accepted": "✅ Your order has been accepted and is being prepared!",
        "preparing": "👨‍🍳 We've started preparing your order!",
        "ready": "✅ Your order is ready!",
        "out_for_delivery": "🛵 Your order is on the way!",
        "delivered": "🎉 Your order has been delivered. Enjoy!",
        "cancelled": "❌ Your order has been cancelled.",
        "rejected": f"❌ Your order was rejected: {reason_pretty}",
        "modified": "✏️ Your order has been updated and confirmed. It is now being prepared!",
    }
    msg = labels.get(status, f"Status updated to: {status}")
    return f"🍛 *Karachi Naseeb Biryani*\nOrder #{receipt}\n\n{msg}\n\nTrack: {tracking_url}"

# --- Live Order Tracking (public, no auth) ---
@api_router.get("/public/restaurant-info")
async def public_restaurant_info():
    """Public: restaurant info needed for receipts, reviews, and customer pages."""
    o = await db.online_settings.find_one({"key": "online"}, {"_id": 0}) or {}
    g = await db.settings.find_one({"key": "global"}, {"_id": 0}) or {}
    
    # Prefer online_settings, fallback to global settings
    lat = o.get("restaurant_lat", 31.4520)
    lng = o.get("restaurant_lng", 74.2680)
    
    return {
        "name": o.get("restaurant_name") or g.get("restaurant_name", "Karachi Naseeb Biryani & Murg Pulao"),
        "phone": o.get("restaurant_phone") or g.get("restaurant_phone", "+923004928411"),
        "whatsapp": o.get("restaurant_whatsapp") or o.get("restaurant_phone") or g.get("restaurant_phone", "+923004928411"),
        "email": o.get("restaurant_email", "info@karachinaseeb.com"),
        "address": o.get("restaurant_address") or g.get("restaurant_address", "68 Chatri Chowk, Punjab Small Industry, D Block, Lahore"),
        "logo_url": o.get("restaurant_logo_url", ""),
        "facebook_url": o.get("facebook_url", ""),
        "instagram_url": o.get("instagram_url", ""),
        "twitter_url": o.get("twitter_url", ""),
        "opening_hours": o.get("opening_hours", "Mon-Sun: 10AM - 11PM"),
        "currency": g.get("currency", "Rs"),
        "lat": lat,
        "lng": lng,
        "google_maps_url": o.get("google_maps_url") or f"https://www.google.com/maps/search/?api=1&query={lat},{lng}",
    }


@api_router.get("/track/{order_id}")
async def public_track_order(order_id: str, request: Request, t: Optional[str] = None):
    """Public tracking endpoint reachable via the order's WhatsApp / receipt link.

    IDOR mitigation (CRITICAL): order ids are MongoDB ObjectIds and embed a
    sequential counter — enumerable in seconds. Without a secondary secret an
    attacker who sees ONE order id (e.g. their own receipt) can iterate the
    counter byte and harvest PII (first name, last-4 phone, suburb prefix,
    items, total) from every neighbouring order. We require a per-order
    `track_token` query parameter that is issued at order creation and only
    embedded in links we hand directly to the customer. Without a valid token
    we return 404 — never 401/403 — so attackers cannot probe which ids exist.

    Authenticated owners + admins bypass the token check (their identity
    already proves authorization). The owner also continues to see full PII;
    everyone else (token-only viewers, e.g. a family member the customer
    shared the link with) gets the masked response."""
    try:
        o = await db.online_orders.find_one({"_id": ObjectId(order_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Order not found")
    if not o:
        raise HTTPException(status_code=404, detail="Order not found")

    # Authorization layer 1: signed-in admin OR the order's customer.
    show_full = False
    is_authorized = False
    try:
        user = await get_current_user(request)
        if user.get("role") == "admin":
            show_full = True
            is_authorized = True
    except HTTPException:
        pass
    if not is_authorized:
        try:
            cust = await get_current_customer(request)
            if cust and o.get("customer_id") and str(o.get("customer_id")) == str(cust.get("_id")):
                show_full = True
                is_authorized = True
        except HTTPException:
            pass

    # Authorization layer 2: per-order share token. Constant-time compare so we
    # don't leak token bytes via timing. Only checked if the caller is not
    # already an authenticated owner / admin (saves a DB hit on the hot path
    # for signed-in users polling their own orders every 5s).
    if not is_authorized:
        order_token = o.get("track_token") or ""
        if not (t and order_token and secrets.compare_digest(str(t), str(order_token))):
            # Return 404, not 401/403 — denying existence prevents enumeration of
            # which ids are real even after the token requirement is enforced.
            raise HTTPException(status_code=404, detail="Order not found")

    def _mask_phone(p: str) -> str:
        p = str(p or "")
        if len(p) <= 4:
            return "***"
        return "*" * (len(p) - 4) + p[-4:]

    def _mask_address(a: str) -> str:
        a = str(a or "")
        if len(a) <= 18:
            return a[:6] + ("…" if len(a) > 6 else "")
        return a[:18] + "…"

    return {
        "id": str(o["_id"]),
        "receipt_no": str(o["_id"])[-6:].upper(),
        "status": o.get("status", "pending"),
        "payment_status": o.get("payment_status", "pending"),
        "payment_method": o.get("payment_method", "cod"),
        "items": o.get("items", []),
        "subtotal": o.get("subtotal", 0),
        "total_price": o.get("total_price", 0),
        "delivery_fee": o.get("delivery_fee", 0),
        "delivery_fee_overridden": bool(o.get("delivery_fee_overridden", False)),
        "discount_amount": o.get("discount_amount", 0),
        "customer_name": o.get("customer_name", "") if show_full else (str(o.get("customer_name", "") or "").split(" ")[0] or "Customer"),
        "phone": o.get("phone", "") if show_full else _mask_phone(o.get("phone", "")),
        "address": o.get("address", "") if show_full else _mask_address(o.get("address", "")),
        "created_at": o.get("created_at", ""),
        "updated_at": o.get("updated_at", ""),
        "modified": bool(o.get("modified", False)),
        "modification_pending": bool(o.get("modification_pending", False)),
        "rejection_reason": o.get("rejection_reason", ""),
        "accepted_at": o.get("accepted_at", ""),
        # Refund lifecycle (customer-requested; see /online-orders/{id}/refund-request).
        # Visible to token viewers too — status/amount only, nothing more personal
        # than the rest of this masked response already reveals.
        "refund_request": o.get("refund_request"),
        # V2: live prep time + countdown helpers
        "prep_time_min": int(o.get("prep_time_min") or 30),  # default 30 min
        "prep_time_updated_at": o.get("prep_time_updated_at", ""),
        "response_deadline_seconds": _response_deadline_seconds(o),
    }


def _response_deadline_seconds(order: dict) -> int:
    """How many seconds remain in the 2-minute restaurant response window for pending orders.
    Returns 0 (or negative) once the deadline has passed. Used by the countdown timer
    shown to both customers and staff."""
    if order.get("status") != "pending":
        return 0
    try:
        created = datetime.fromisoformat(str(order.get("created_at", "")).replace("Z", "+00:00"))
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
    except Exception:
        return 0
    window = int(os.environ.get("ORDER_RESPONSE_WINDOW_SEC", "120"))
    elapsed = (datetime.now(timezone.utc) - created).total_seconds()
    return max(0, int(window - elapsed))

# Hook tracking + WhatsApp into existing endpoints. We do this via Mongo-side "after-write"
# triggers by wrapping the existing endpoints — but to keep changes minimal, we'll add a
# "post" endpoint that the frontend calls right after order creation. The cleaner path is
# to update the create_online_order and update_online_order_status implementations directly.
# That is done above; we just need them to call these helpers. We use a small middleware-like
# function that gets invoked from those endpoints (added by patching them with monkey-patch
# is risky). Instead we add explicit notify calls in the existing handlers above by editing.
# (See create_online_order patches below.)

# Add online_settings field for twilio_whatsapp_from (one-time migration)
async def _ensure_twilio_setting():
    s = await db.online_settings.find_one({"key": "online"})
    if not s or "twilio_whatsapp_from" not in s:
        await db.online_settings.update_one(
            {"key": "online"},
            {"$set": {"twilio_whatsapp_from": os.environ.get("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")}},
            upsert=True,
        )

@app.on_event("startup")
async def _startup_twilio_setting():
    # Non-blocking: Twilio settings load should never block the listener.
    asyncio.create_task(_startup_twilio_background())

async def _startup_twilio_background():
    try:
        await _ensure_twilio_setting()
    except Exception as e:
        logger.exception(f"Twilio settings load failed (server still listening): {e}")


@app.on_event("startup")
async def _startup_menu_image_migration():
    """Self-healing: any menu_item whose image_url is an inline base64 data URL
    gets persisted to the volume and rewritten to a short /api/uploads/... URL.
    Idempotent — items already migrated are excluded by the query. Runs in
    the background after a small delay so it never blocks Fly's port binding
    or the first health check."""
    asyncio.create_task(_auto_migrate_menu_images_background())

async def _auto_migrate_menu_images_background():
    try:
        # Brief wait so primary connection is healthy and indexes are warm.
        await asyncio.sleep(2)
        migrated = 0
        skipped = 0
        failed = 0
        cursor = db.menu_items.find({"image_url": {"$regex": "^data:"}}).limit(2000)
        async for doc in cursor:
            try:
                new_url = _persist_data_url_image(doc.get("image_url", ""), kind="menu")
                if isinstance(new_url, str) and new_url.startswith("/api/uploads/"):
                    await db.menu_items.update_one(
                        {"_id": doc["_id"]},
                        {"$set": {"image_url": new_url, "image_type": "url"}},
                    )
                    migrated += 1
                else:
                    skipped += 1
            except Exception as e:
                failed += 1
                logger.warning(f"auto-migrate item {doc.get('_id')} failed: {e}")
        if migrated or failed:
            logger.info(
                f"Auto-migrated menu images: migrated={migrated} skipped={skipped} failed={failed}"
            )
    except Exception as e:
        logger.exception(f"Auto-migration of menu images crashed (server still listening): {e}")

# =============================================================================
# END OBJECT STORAGE / WHATSAPP / TRACKING
# =============================================================================

# =============================================================================
# ADMIN: REVIEW MANAGEMENT
# =============================================================================

@api_router.get("/admin/reviews")
async def get_all_reviews(request: Request, status: str = "all", limit: int = 100):
    """Admin: Get all reviews with optional status filter (all/pending/replied)"""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if status == "pending":
        query["admin_reply"] = {"$exists": False}
    elif status == "replied":
        query["admin_reply"] = {"$exists": True}
    
    reviews = await db.reviews.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    
    result = []
    for r in reviews:
        # V2 Fix: Some entries (general feedback via POST /api/feedback) have order_id=None.
        # Calling ObjectId(None) crashes the endpoint → "Failed to load reviews".
        # Skip the order lookup gracefully when there's no order_id.
        order = None
        oid = r.get("order_id")
        if oid:
            try:
                order = await db.online_orders.find_one({"_id": ObjectId(oid)})
            except Exception:
                order = None

        result.append({
            "id": str(r["_id"]),
            "order_id": oid or "",
            "order_receipt_no": (order.get("receipt_no") if order else None) or (oid[-6:].upper() if oid else "FEEDBACK"),
            "customer_id": str(r.get("customer_id", "") or ""),
            "customer_name": r.get("customer_name", "Anonymous"),
            "customer_email": r.get("customer_email", ""),
            "customer_phone": r.get("customer_phone", ""),
            "rating": r.get("rating", 0),
            "comment": r.get("comment", ""),
            "admin_reply": r.get("admin_reply", ""),
            "replied_by": r.get("replied_by", ""),
            "replied_at": r.get("replied_at", ""),
            "is_feedback": bool(r.get("is_feedback", False)),
            "created_at": r.get("created_at", ""),
        })
    
    return result

@api_router.post("/admin/reviews/{review_id}/reply")
async def reply_to_review(review_id: str, body: AdminReviewReply, request: Request):
    """Admin: Reply to a customer review (and email the customer if SMTP is configured)."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        review = await db.reviews.find_one({"_id": ObjectId(review_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Review not found")
    
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    
    await db.reviews.update_one(
        {"_id": ObjectId(review_id)},
        {"$set": {
            "admin_reply": body.reply,
            "replied_by": user.get("name", user.get("email", "")),
            "replied_at": datetime.now(timezone.utc).isoformat(),
        }}
    )

    # Best-effort: email the customer to let them know the restaurant has replied.
    email_sent = False
    try:
        cust_id = review.get("customer_id")
        cust_email = None
        cust_name = review.get("customer_name") or "Valued Customer"
        if cust_id:
            cust = await db.customers.find_one({"_id": cust_id if isinstance(cust_id, ObjectId) else ObjectId(cust_id)})
            if cust:
                cust_email = cust.get("email")
                cust_name = cust.get("name") or cust_name
        if cust_email:
            s = await _get_settings_doc()
            if s.get("smtp_host") and s.get("smtp_port") and s.get("smtp_user") and s.get("smtp_password"):
                rname = s.get("restaurant_name", "Karachi Naseeb")
                stars = "★" * int(review.get("rating", 0)) + "☆" * (5 - int(review.get("rating", 0)))
                subject = f"[{rname}] We replied to your review"
                plain = (
                    f"Hi {cust_name},\n\n"
                    f"Thank you for your review! The {rname} team has just replied:\n\n"
                    f"Your rating: {stars}\n"
                    f"Your comment: {review.get('comment','')}\n\n"
                    f"Our reply:\n{body.reply}\n\n"
                    f"— {rname}\n"
                )
                html = f"""<div style='font-family:Manrope,Arial,sans-serif;max-width:560px;margin:0 auto;border:1px solid #E5E2DC;border-radius:12px;overflow:hidden'>
                  <div style='padding:16px 24px;background:#1E3F20;color:#fff'><h2 style='margin:0'>{rname}</h2></div>
                  <div style='padding:20px 24px;color:#1A1A1A'>
                    <p style='margin:0 0 12px'>Hi <strong>{cust_name}</strong>,</p>
                    <p style='margin:0 0 12px'>Thank you for your review — the {rname} team has just posted a reply.</p>
                    <div style='background:#F9F8F6;border:1px solid #E5E2DC;border-radius:8px;padding:12px 16px;margin:12px 0'>
                      <div style='font-size:13px;color:#5C5F5C'>Your rating</div>
                      <div style='font-size:18px;color:#D29C2C;letter-spacing:2px'>{stars}</div>
                      <div style='font-size:13px;color:#5C5F5C;margin-top:8px'>Your comment</div>
                      <div style='color:#1A1A1A'>{review.get('comment','')}</div>
                    </div>
                    <div style='background:#FFF4E5;border-left:4px solid #D29C2C;padding:12px 16px;border-radius:4px'>
                      <div style='font-size:13px;color:#5C5F5C;margin-bottom:6px'>Our reply</div>
                      <div style='color:#1A1A1A;white-space:pre-wrap'>{body.reply}</div>
                    </div>
                    <p style='margin:16px 0 0;color:#5C5F5C;font-size:13px'>Thank you for being part of {rname}.</p>
                  </div>
                </div>"""
                try:
                    await asyncio.to_thread(
                        _send_email_sync,
                        s["smtp_host"], s["smtp_port"], s.get("smtp_user", ""), s.get("smtp_password", ""),
                        bool(s.get("smtp_use_tls", True)),
                        s.get("smtp_from") or s.get("smtp_user"),
                        [cust_email], subject, plain, html,
                    )
                    email_sent = True
                except Exception as e:
                    logger.warning(f"Review reply email failed: {e}")
    except Exception as e:
        logger.warning(f"Review reply email lookup failed: {e}")
    
    return {"ok": True, "message": "Reply posted", "email_sent": email_sent}

@api_router.delete("/admin/reviews/{review_id}")
async def delete_review(review_id: str, request: Request):
    """Admin: Delete a review (moderation)"""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        result = await db.reviews.delete_one({"_id": ObjectId(review_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Review not found")
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Review not found")
    
    return {"ok": True, "message": "Review deleted"}

# =============================================================================
# LOYALTY / DIAMOND REWARD SYSTEM
# =============================================================================

# --- Admin: Loyalty Settings ---
@api_router.get("/admin/loyalty/settings")
async def get_loyalty_settings(request: Request):
    await get_current_user(request)
    settings = await db.loyalty_settings.find_one({"key": "loyalty"}, {"_id": 0}) or {}
    return {
        "enabled": settings.get("enabled", True),
        "earning_rate": settings.get("earning_rate", 10.0),
        "min_order_for_points": settings.get("min_order_for_points", 0.0),
        "points_expiry_days": settings.get("points_expiry_days", None),
    }

@api_router.post("/admin/loyalty/settings")
async def update_loyalty_settings(settings: LoyaltySettingsUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = settings.dict(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = str(user["_id"])
    
    await db.loyalty_settings.update_one(
        {"key": "loyalty"},
        {"$set": update_data},
        upsert=True
    )
    return {"ok": True, "message": "Loyalty settings updated"}

# --- Admin: Rewards Management ---
@api_router.get("/admin/loyalty/rewards")
async def get_all_rewards(request: Request):
    await get_current_user(request)
    rewards = await db.loyalty_rewards.find().to_list(500)
    return [{
        "id": str(r.pop("_id")),
        **r
    } for r in rewards]

@api_router.post("/admin/loyalty/rewards")
async def create_reward(reward: LoyaltyRewardCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    doc = {
        **reward.dict(),
        "total_redemptions": 0,
        "created_by": str(user["_id"]),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.loyalty_rewards.insert_one(doc)
    return {"id": str(result.inserted_id), **reward.dict()}

@api_router.put("/admin/loyalty/rewards/{reward_id}")
async def update_reward(reward_id: str, reward: LoyaltyRewardUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        existing = await db.loyalty_rewards.find_one({"_id": ObjectId(reward_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Reward not found")
    
    if not existing:
        raise HTTPException(status_code=404, detail="Reward not found")
    
    update_data = reward.dict(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = str(user["_id"])
    
    await db.loyalty_rewards.update_one(
        {"_id": ObjectId(reward_id)},
        {"$set": update_data}
    )
    return {"ok": True, "message": "Reward updated"}

@api_router.delete("/admin/loyalty/rewards/{reward_id}")
async def delete_reward(reward_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        result = await db.loyalty_rewards.delete_one({"_id": ObjectId(reward_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Reward not found")
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reward not found")
    
    return {"ok": True, "message": "Reward deleted"}

# --- Admin: Customer Loyalty View ---
@api_router.get("/admin/loyalty/customers")
async def get_customers_with_loyalty(request: Request, limit: int = 100):
    await get_current_user(request)
    customers = await db.customers.find({}).sort("diamond_balance", -1).limit(limit).to_list(limit)
    return [{
        "id": str(c.pop("_id")),
        "email": c.get("email"),
        "name": c.get("name"),
        "diamond_balance": c.get("diamond_balance", 0),
        "lifetime_diamonds_earned": c.get("lifetime_diamonds_earned", 0),
        "lifetime_diamonds_spent": c.get("lifetime_diamonds_spent", 0),
    } for c in customers]

# --- Admin: Manual Balance Adjustment ---
@api_router.post("/admin/loyalty/adjust")
async def adjust_customer_balance(adjustment: LoyaltyAdjustRequest, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        customer = await db.customers.find_one({"_id": ObjectId(adjustment.customer_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    current_balance = customer.get("diamond_balance", 0)
    new_balance = max(0, current_balance + adjustment.diamonds)  # Can't go negative
    
    # Update customer balance
    await db.customers.update_one(
        {"_id": ObjectId(adjustment.customer_id)},
        {"$set": {"diamond_balance": new_balance}}
    )
    
    # Log transaction
    await db.loyalty_transactions.insert_one({
        "customer_id": adjustment.customer_id,
        "transaction_type": "adjusted",
        "diamonds": adjustment.diamonds,
        "balance_after": new_balance,
        "notes": adjustment.notes,
        "adjusted_by": str(user["_id"]),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    
    return {"ok": True, "new_balance": new_balance}

# --- Customer: View Balance & Rewards ---
@api_router.get("/loyalty/balance")
async def get_my_loyalty_balance(request: Request):
    customer = await get_current_customer(request)
    return {
        "diamond_balance": customer.get("diamond_balance", 0),
        "lifetime_diamonds_earned": customer.get("lifetime_diamonds_earned", 0),
        "lifetime_diamonds_spent": customer.get("lifetime_diamonds_spent", 0),
    }

@api_router.get("/loyalty/transactions")
async def get_my_loyalty_transactions(request: Request, limit: int = 50):
    customer = await get_current_customer(request)
    transactions = await db.loyalty_transactions.find(
        {"customer_id": str(customer["_id"])}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    return [{
        "id": str(t.pop("_id")),
        **t
    } for t in transactions]

@api_router.get("/loyalty/rewards")
async def get_available_rewards(request: Request):
    # Public or authenticated - shows active rewards
    rewards = await db.loyalty_rewards.find({"is_active": True}).to_list(500)
    out = []
    for r in rewards:
        rid = str(r.pop("_id"))
        item = {"id": rid, **{k: v for k, v in r.items() if k not in ["created_by", "updated_by"]}}
        # Enrich free_item rewards with the linked menu item's name + image so the
        # frontend can render a proper "FREE — <item name>" line in the cart / checkout
        # summary instead of just the reward title. (Customers complained they couldn't
        # tell which item they'd get until the order was placed.)
        if item.get("reward_type") == "free_item" and item.get("reward_value"):
            try:
                mi = await db.menu_items.find_one({"_id": ObjectId(str(item["reward_value"]))}, {"name": 1, "image_url": 1, "price": 1})
                if mi:
                    item["free_item_name"] = mi.get("name", "")
                    item["free_item_image"] = mi.get("image_url", "")
                    item["free_item_value"] = float(mi.get("price", 0) or 0)
            except Exception:
                pass
        out.append(item)
    return out

# =============================================================================
# END LOYALTY SYSTEM
# =============================================================================

app.include_router(api_router)
# CORS hardening: in production set CORS_ORIGINS to an explicit comma-separated list
# (e.g. "https://karachinaseeb.com,https://www.karachinaseeb.com"). If the env var is
# missing we DEFAULT TO CLOSED instead of "*" so a misconfigured production deploy fails
# safe rather than silently accepting cross-origin requests from anywhere.
_cors_origins_raw = os.environ.get("CORS_ORIGINS", "")
_cors_origins = [o.strip() for o in _cors_origins_raw.split(',') if o.strip()] if _cors_origins_raw else []
_cors_allow_credentials = bool(_cors_origins) and ("*" not in _cors_origins)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=_cors_allow_credentials,
    allow_methods=["*"],
    allow_headers=["*"],
)

# GZip compression: cuts JSON menu payloads (heavy with base64 images / repetitive
# text) by 3-5x with zero application-level code change. minimum_size avoids
# wasting CPU compressing already-small responses.
app.add_middleware(GZipMiddleware, minimum_size=1024)


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    """Adds baseline HTTP security headers to every response. The headers picked here are
    OWASP-recommended and safe defaults for an API + SPA stack — they do not break the
    POS or the online ordering frontend. Strict-Transport-Security is only meaningful
    when served over HTTPS (Fly.io edge already enforces force_https=true)."""
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        # Don't override headers already set by an upstream proxy.
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("X-Frame-Options", "DENY")
        response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        response.headers.setdefault("Permissions-Policy", "geolocation=(self), microphone=(self), camera=()")
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
        return response


app.add_middleware(SecurityHeadersMiddleware)

@app.get("/api/health")
async def health():
    """Deploy verification: confirms which code version is running."""
    return {"status": "ok", "version": "1.4.2-cors-credentials", "cors_credentials_enabled": _cors_allow_credentials}

# --- Serve built frontend (production / local Windows install) ---
# When `frontend/build` exists, serve it from the backend at "/".
# This gives single-origin (no CORS, no proxy needed), fast load (no dev
# compilation), and makes Cloudflare tunneling trivial (one port).
# Skipped automatically in dev environments where build doesn't exist.
try:
    from fastapi.staticfiles import StaticFiles
    from fastapi.responses import FileResponse
    _frontend_build = ROOT_DIR.parent / "frontend" / "build"
    if _frontend_build.exists() and (_frontend_build / "index.html").exists():
        # Static assets (JS/CSS/images)
        app.mount("/static", StaticFiles(directory=str(_frontend_build / "static")), name="static")
        _build_real = os.path.realpath(str(_frontend_build))

        @app.get("/{full_path:path}")
        async def spa_fallback(full_path: str):
            # API routes are handled by api_router (mounted earlier).
            # Anything else returns index.html so React Router can take over.
            if full_path.startswith("api/"):
                raise HTTPException(status_code=404, detail="API route not found")
            # Reject any path that contains traversal sequences or NUL bytes outright.
            if ".." in full_path.split("/") or "\x00" in full_path:
                return FileResponse(str(_frontend_build / "index.html"))
            asset = _frontend_build / full_path
            # Resolve symlinks and verify the resolved file is still inside the build dir
            # to prevent reading arbitrary files via crafted full_path values.
            try:
                asset_real = os.path.realpath(str(asset))
            except Exception:
                asset_real = ""
            if asset_real and (asset_real == _build_real or asset_real.startswith(_build_real + os.sep)) and os.path.isfile(asset_real):
                return FileResponse(asset_real)
            return FileResponse(str(_frontend_build / "index.html"))
        logger.info(f"Serving built frontend from: {_frontend_build}")
except Exception as e:
    logger.warning(f"Frontend static serve disabled: {e}")

