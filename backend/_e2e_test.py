"""End-to-end test for menu import/export + online-order cancel.
Runs against the test server on :8899 (test DB epos_test_scratch). Read-only to real data."""
import io, sys, time
import requests
from openpyxl import load_workbook, Workbook

BASE = "http://127.0.0.1:8899/api"
S = requests.Session()
PASS, FAIL = [], []

def check(name, cond, extra=""):
    (PASS if cond else FAIL).append(name)
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}" + (f"  — {extra}" if extra else ""))

# 0. wait for server
for _ in range(40):
    try:
        r = S.get(BASE.replace('/api','') + "/api/menu", timeout=2)
        if r.status_code in (200, 401, 404): break
    except Exception:
        time.sleep(0.5)
else:
    print("Server never came up"); sys.exit(2)

# 1. login as admin
r = S.post(f"{BASE}/auth/login", json={"email": "admin@restaurant.com", "password": "admin123"}, timeout=10)
check("admin login", r.status_code == 200, f"status={r.status_code}")
if r.status_code != 200:
    print(r.text[:300]); sys.exit(2)

# 2. download template
r = S.get(f"{BASE}/menu-items/template", timeout=10)
check("GET /menu-items/template returns 200", r.status_code == 200, f"status={r.status_code}")
check("template is xlsx", r.content[:2] == b"PK" and len(r.content) > 500, f"{len(r.content)} bytes")
wb = load_workbook(io.BytesIO(r.content))
check("template has 'Menu' sheet", "Menu" in wb.sheetnames, str(wb.sheetnames))
hdr = [c.value for c in wb["Menu"][1]]
EXPECTED = ["name","category","price","description","variations","discount_type","discount_value","stock","is_popular","is_bestseller"]
check("template headers match", hdr == EXPECTED, str(hdr))

# 3. build an import file with a new category + variations + discount
wb2 = Workbook(); ws = wb2.active; ws.title = "Menu"
ws.append(EXPECTED)
ws.append(["E2E Test Biryani", "E2E Test Cat", 500, "auto test item", "Half=250; Full=450", "percentage", 10, 77, "yes", "no"])
ws.append(["E2E Test Water", "E2E Test Cat", 80, "", "", "", "", 200, "no", "no"])
ws.append(["", "NoName Cat", 100, "should be skipped", "", "", "", "", "no", "no"])  # missing name -> error row
buf = io.BytesIO(); wb2.save(buf); buf.seek(0)

r = S.post(f"{BASE}/menu-items/import", files={"file": ("import.xlsx", buf.getvalue(),
          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, timeout=15)
check("POST /menu-items/import returns 200", r.status_code == 200, f"status={r.status_code} {r.text[:200]}")
res = r.json() if r.status_code == 200 else {}
check("import created 2 items", res.get("created") == 2, str(res))
check("import created category", res.get("categories_created") == 1, str(res))
check("import reported 1 error row (missing name)", res.get("error_count") == 1, str(res.get("errors")))

# 4. verify items landed correctly via /menu-items
r = S.get(f"{BASE}/menu-items", timeout=10)
items = {i["name"]: i for i in r.json()}
b = items.get("E2E Test Biryani", {})
check("imported item present", bool(b))
check("variations parsed", isinstance(b.get("variations"), list) and len(b.get("variations", [])) == 2, str(b.get("variations")))
check("discount parsed", b.get("discount_type") == "percentage" and float(b.get("discount_value") or 0) == 10, f"{b.get('discount_type')}/{b.get('discount_value')}")
check("stock parsed", int(b.get("stock", -1)) == 77, str(b.get("stock")))
check("is_popular parsed", b.get("is_popular") is True, str(b.get("is_popular")))

# 5. export -> should round-trip our items
r = S.get(f"{BASE}/menu-items/export", timeout=10)
check("GET /menu-items/export returns 200", r.status_code == 200, f"status={r.status_code}")
check("export is xlsx", r.content[:2] == b"PK", f"{len(r.content)} bytes")
wbx = load_workbook(io.BytesIO(r.content)); wsx = wbx["Menu"]
xhdr = [c.value for c in wsx[1]]
check("export headers match template", xhdr == EXPECTED, str(xhdr))
rows = {row[0]: row for row in wsx.iter_rows(min_row=2, values_only=True)}
check("export contains imported item", "E2E Test Biryani" in rows, "found" if "E2E Test Biryani" in rows else "missing")
if "E2E Test Biryani" in rows:
    row = rows["E2E Test Biryani"]
    check("export variations round-trip format", "Half=250" in str(row[4]) and "Full=450" in str(row[4]), str(row[4]))
    check("export category name resolved", row[1] == "E2E Test Cat", str(row[1]))
    check("export is_popular = yes", str(row[8]).lower() == "yes", str(row[8]))

# 6. re-import the exported file with an edited price -> should UPDATE, not duplicate
wbx["Menu"].cell(row=[i for i,rw in enumerate(wsx.iter_rows(min_row=2,values_only=True),start=2) if rw[0]=="E2E Test Biryani"][0], column=3, value=999)
buf2 = io.BytesIO(); wbx.save(buf2); buf2.seek(0)
r = S.post(f"{BASE}/menu-items/import", files={"file": ("reimport.xlsx", buf2.getvalue(),
          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, timeout=15)
res2 = r.json() if r.status_code == 200 else {}
check("re-import updates (0 created)", res2.get("created") == 0, str(res2))
check("re-import updated existing", (res2.get("updated") or 0) >= 2, str(res2))
r = S.get(f"{BASE}/menu-items", timeout=10)
items2 = {i["name"]: i for i in r.json()}
check("price updated via round-trip", float(items2.get("E2E Test Biryani", {}).get("price", 0)) == 999,
      str(items2.get("E2E Test Biryani", {}).get("price")))
check("no duplicate created", sum(1 for n in items2 if n == "E2E Test Biryani") == 1)

# 7. online order cancel: create an order, accept, then cancel via status route
#    (mirrors the new frontend Cancel button -> PUT /online-orders/{id}/status {cancelled})
cat_id = b.get("category_id")
order_payload = {
    "customer_name": "E2E Tester", "customer_phone": "03001234567",
    "delivery_address": "Test St", "order_type": "delivery",
    "items": [{"item_id": b.get("id"), "name": "E2E Test Biryani", "quantity": 1, "price": 999}],
    "payment_method": "cash",
}
r = S.post(f"{BASE}/online-orders", json=order_payload, timeout=10)
if r.status_code not in (200, 201):
    check("create online order", False, f"status={r.status_code} {r.text[:200]}")
else:
    oid = r.json().get("id") or r.json().get("order", {}).get("id")
    check("create online order", bool(oid), f"id={oid}")
    if oid:
        S.post(f"{BASE}/online-orders/{oid}/accept", timeout=10)
        r = S.put(f"{BASE}/online-orders/{oid}/status", json={"status": "cancelled"}, timeout=10)
        check("PUT status=cancelled returns 200", r.status_code == 200, f"status={r.status_code} {r.text[:200]}")
        r = S.get(f"{BASE}/online-orders/{oid}", timeout=10)
        if r.status_code == 200:
            check("order status is cancelled", r.json().get("status") == "cancelled", str(r.json().get("status")))

print(f"\n==== {len(PASS)} passed, {len(FAIL)} failed ====")
if FAIL:
    print("FAILURES:", FAIL); sys.exit(1)
print("ALL GREEN")
